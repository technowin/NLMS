from django.conf import settings
from django.shortcuts import render
from services.file_storage_service import file_storage_service
# Create your views here.
import string
from django.http import Http404, JsonResponse
from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
import requests
from Reports.models import *
from Account.models import *
import Db 
import json
from dateutil.relativedelta import relativedelta
from calendar import monthrange
from mysql.connector.errors import InterfaceError
import calendar
import pandas as pd
import xlwt
from django.http import HttpResponse
import os
import time
import xlsxwriter
import io
import os
# Create your views here.
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.contrib import messages
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from flask import Flask, render_template
from django.shortcuts import render
from django.db import transaction
from NLMS.settings import MEDIA_ROOT
app = Flask(__name__)
# Create your views here.
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph
from rest_framework import serializers
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from django.http import FileResponse
from xhtml2pdf import pisa
from django.template.loader import get_template
import traceback
from Account.db_utils import callproc
from django.utils import timezone
from NLMS.encryption import *
from L01.models import *
from django.db.models import *
from django.core.files.storage import default_storage
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar
from NLMS.access_control import no_direct_access
from django.http import HttpResponse  # ✅ Add this import

# @login_required
# def payment_report(request):
#     try:
#         if request.user.is_authenticated ==True:                
#             global user
#             user = request.user.id
#             library_code = request.session.get('library_db', None)
#             username = request.session.get('username', None)
#             user_id = request.session["user_id"]
#             role_id = request.session["role_id"]
            
#         if request.method == "GET":
#             library_name = tbl_librarymasterL01.objects.using(library_code).filter(is_active=True).first()
#             library_name_mar = library_name.library_name_mar if library_name else ''

#             reports = PaymentReport.objects.all().order_by('-generated_date')

#             for rep in reports:
#                 # encrypted id for links
#                 rep.report_encrypted_id = enc(str(rep.id))

#                 # Determine if the record is "filled" for the four fields. notepad 387
#                 # Consider a field filled if:
#                 #  - receipt_no: not None and not empty string
#                 #  - deposit_amount: not None and not zero (Decimal('0') considered empty)
#                 #  - receipt_upload: not None and not empty string
#                 #  - deposit_date: not None
#                 receipt_no_filled = bool(rep.receipt_no and str(rep.receipt_no).strip() != "")
                
#                 # deposit_amount is DecimalField default 0. Treat 0 as NOT filled.
#                 try:
#                     deposit_amount_val = rep.deposit_amount if rep.deposit_amount is not None else Decimal('0')
#                 except Exception:
#                     deposit_amount_val = Decimal('0')
#                 deposit_amount_filled = (deposit_amount_val is not None) and (Decimal(str(deposit_amount_val)) != Decimal('0'))

#                 receipt_upload_filled = bool(rep.receipt_upload and str(rep.receipt_upload).strip() != "")
#                 deposit_date_filled = bool(rep.deposit_date is not None)

#                 # If ALL are filled => do NOT allow edit (can_edit = False)
#                 rep.can_edit = not (receipt_no_filled and deposit_amount_filled and receipt_upload_filled and deposit_date_filled)

#             return render(request, 'Reports/payment_report.html', {
#                 'payment_reports': reports,
#                 'library_name_mar': library_name_mar
#             })
            
#         if request.method == "POST":
            
#             report_type = request.POST.get("report_type")
#             from_month = request.POST.get("from_month")
#             to_month = request.POST.get("to_month")

#             # 🛑 Guard clause: Proceed only if both months are provided and type is valid
#             if report_type not in ["pdf", "excel"] or not from_month or not to_month:
#                 # create post
#                 from_date = request.POST.get("from_date")
#                 to_date = request.POST.get("to_date")
#                 total_amount = request.POST.get("total_amount")
#                 summary_data_json = request.POST.get("summary_data")

#                 # Convert to Python date objects
#                 from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
#                 to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()

#                 # Check for overlap
#                 overlap_exists = PaymentReport.objects.filter(
#                     from_date__lte=to_date_obj,
#                     to_date__gte=from_date_obj
#                 ).exists()

#                 if overlap_exists:
#                     messages.error(request, "A report already exists for this date range! कृपया दुसरी दिनांक निवडा.")
#                     return redirect("payment_report")

#                 try:
#                     with transaction.atomic():  # 👈 Ensures all or nothing

#                         # Create main report
#                         report = PaymentReport.objects.create(
#                             from_date=from_date_obj,
#                             to_date=to_date_obj,
#                             generated_date=timezone.now().date(),
#                             total_amount=total_amount,
#                             created_by=user_id,
#                             updated_by=user_id
#                         )

#                         # Parse and insert key-value pairs
#                         if summary_data_json:
#                             summary_data = json.loads(summary_data_json)
#                             for item in summary_data:
#                                 PaymentReportKeyValue.objects.create(
#                                     payment_report=report,
#                                     key=item.get("key"),
#                                     value=item.get("value")
#                                 )

#                     messages.success(request, "Payment summary generated successfully!")

#                 except Exception as e:
#                     transaction.set_rollback(True)
#                     messages.error(request, f"Error while generating report: {str(e)}")

#                 messages.success(request, "Payment summary generated successfully!")
                
#                 return redirect("payment_report")

#             else:
#                 # ✅ Parse months safely
#                 try:
#                     from_date = datetime.strptime(from_month + "-01", "%Y-%m-%d")
#                     year, month = map(int, to_month.split("-"))
#                     to_date = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
#                 except Exception as e:
#                     return HttpResponse("Invalid date format received.", content_type="text/plain")

#                 # ✅ Fetch reports between date range
#                 reports = PaymentReport.objects.filter(
#                     generated_date__gte=from_date,
#                     generated_date__lt=to_date
#                 )

#                 if not reports.exists():
#                     return HttpResponse("No reports found for the selected months.", content_type="text/plain")

#                 # ✅ Collect all distinct keys for dynamic columns
#                 all_keys = PaymentReportKeyValue.objects.values_list('key', flat=True).distinct()

#                 # ✅ Build header (added Deposit Amount)
#                 # header = ["Receipt No", "Generated Date", "Deposit Date"] + list(all_keys) + ["Deposit Amount"]
#                 # data = [header]
#                 header = ["पावती क्रमांक", "अहवाल दिनांक", "ठेव दिनांक"] + list(all_keys) + ["ठेव रक्कम"]
#                 data = [header]

#                 for report in reports:
#                     key_value_map = {kv.key: kv.value for kv in report.key_values.all()}

#                     # ✅ Show only date part (no time)
#                     generated_date = report.generated_date.strftime("%Y-%m-%d") if report.generated_date else "-"
#                     deposit_date = report.deposit_date.strftime("%Y-%m-%d") if report.deposit_date else "-"

#                     # ✅ Build row with deposit amount
#                     row = [
#                         report.receipt_no or "-",
#                         generated_date,
#                         deposit_date,
#                     ]

#                     for key in all_keys:
#                         row.append(key_value_map.get(key, "0.00"))

#                     # ✅ Add deposit amount at end
#                     row.append(f"{report.deposit_amount:.2f}" if report.deposit_amount else "0.00")

#                     data.append(row)
                    
#                 from_year, from_month_num = map(int, from_month.split("-"))
#                 to_year, to_month_num = map(int, to_month.split("-"))

#                 from_month_name = calendar.month_name[from_month_num]
#                 to_month_name = calendar.month_name[to_month_num]

#                 # Optional: Marathi month names (if you want them instead of English)
#                 marathi_months = {
#                     1: "जानेवारी", 2: "फेब्रुवारी", 3: "मार्च", 4: "एप्रिल", 5: "मे", 6: "जून",
#                     7: "जुलै", 8: "ऑगस्ट", 9: "सप्टेंबर", 10: "ऑक्टोबर", 11: "नोव्हेंबर", 12: "डिसेंबर"
#                 }
#                 from_month_name_mar = marathi_months.get(from_month_num, from_month_name)
#                 to_month_name_mar = marathi_months.get(to_month_num, to_month_name)

#                 # Create report range string
#                 report_range = f"{from_month_name_mar} {from_year} ते {to_month_name_mar} {to_year}"

#                 # ==================== PDF Export ====================
                
#                 if report_type == "pdf":
#                     from weasyprint import HTML, CSS
#                     from django.http import HttpResponse
#                     import tempfile, os

#                     # 🔹 Fetch library info
#                     library = tbl_librarymasterL01.objects.filter(library_code=library_code).first()
#                     library_name = library.library_name_mar if library else "ग्रंथालयाचे नाव उपलब्ध नाही"

#                     # 🔹 Build Marathi title data
#                     today_str = datetime.now().strftime("%d-%m-%Y")
#                     report_title = "भरणा अहवाल"
#                     file_name = f"Payment_Report_{today_str}.pdf"

#                     # ✅ Font path (your local/static font file)
#                     font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSansDevanagari-Regular.ttf")

#                     # ✅ Build HTML dynamically
#                     html_content = f"""
#                     <!DOCTYPE html>
#                     <html lang="mr">
#                         <head>
#                             <meta charset="UTF-8">
#                             <title>{report_title}</title>
#                             <style>
#                                 @font-face {{
#                                     font-family: 'MarathiFont';
#                                     src: url('file://{font_path}') format('truetype');
#                                 }}
#                                 @page {{
#                                     size: A4 landscape;
#                                     margin: 1cm;
#                                     @bottom-right {{
#                                         content: "अहवाल तयार दिनांक : {today_str}";
#                                         font-size: 10px;
#                                         color: #444;
#                                         font-family: 'MarathiFont', sans-serif;
#                                     }}
#                                 }}
#                                 body {{
#                                     font-family: 'MarathiFont', sans-serif;
#                                     font-size: 11px;
#                                     line-height: 1.5;
#                                     color: #000;
#                                     border: 1px solid #bbb;
#                                     padding: 15px;
#                                     background: #fff;
#                                 }}
#                                 h1, h2, h3 {{
#                                     text-align: center;
#                                     margin: 0;
#                                 }}
#                                 h1 {{ font-size: 18px; }}
#                                 h2 {{ font-size: 15px; margin-top: 3px; }}
#                                 h3 {{ font-size: 13px; margin: 4px 0 10px 0; }}
#                                 table {{
#                                     width: 100%;
#                                     border-collapse: collapse;
#                                     table-layout: fixed;
#                                     margin-top: 10px;
#                                 }}
#                                 th, td {{
#                                     border: 0.5px solid #ccc;
#                                     padding: 6px;
#                                     text-align: center;
#                                     vertical-align: middle;
#                                     word-wrap: break-word;
#                                     white-space: normal;
#                                     font-size: 11px;
#                                 }}
#                                 th {{
#                                     background-color: #727cf5;
#                                     color: white;
#                                     font-weight: 600;
#                                 }}
#                                 tr:nth-child(even) {{
#                                     background: #f9f9f9;
#                                 }}
#                             </style>
#                         </head>
#                         <body>
#                             <h1>नवी मुंबई महानगरपालिका</h1>
#                             <h2>{library_name}</h2>
#                             <h3>{report_title} कालावधी: {report_range}</h3>

#                             <table>
#                                 <thead>
#                                     <tr>
#                                         {''.join(f'<th>{h}</th>' for h in data[0])}
#                                     </tr>
#                                 </thead>
#                                 <tbody>
#                                     {''.join('<tr>' + ''.join(f'<td>{c}</td>' for c in row) + '</tr>' for row in data[1:])}
#                                 </tbody>
#                             </table>
#                         </body>
#                     </html>
#                     """

#                     # ✅ Generate PDF (direct, no temp file needed)
#                     pdf_file = HTML(string=html_content).write_pdf(
#                         stylesheets=[CSS(string="@page { size: A4 landscape; margin: 1cm; }")]
#                     )

#                     # ✅ Return PDF as download
#                     response = HttpResponse(pdf_file, content_type="application/pdf")
#                     response["Content-Disposition"] = f'attachment; filename="{file_name}"'
#                     return response
                
#                 # ==================== Excel Export ====================
#                 elif report_type == "excel":
#                     from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
#                     # 🔹 Fetch library info
#                     library = tbl_librarymasterL01.objects.filter(library_code=library_code).first()
#                     library_name = library.library_name_mar if library else "ग्रंथालयाचे नाव उपलब्ध नाही"

#                     # 🔹 Build workbook and sheet
#                     wb = openpyxl.Workbook()
#                     ws = wb.active
#                     ws.title = "Payment Report"

#                     # ================= HEADER SECTION ==================
#                     # Titles at the top
#                     ws.merge_cells("A1:{}1".format(get_column_letter(len(header))))
#                     ws["A1"] = "नवी मुंबई महानगरपालिका"
#                     ws["A1"].font = Font(size=16, bold=True)
#                     ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

#                     ws.merge_cells("A2:{}2".format(get_column_letter(len(header))))
#                     ws["A2"] = library_name
#                     ws["A2"].font = Font(size=13, bold=True)
#                     ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
                    
#                     from datetime import timedelta

#                     ws.merge_cells("A3:{}3".format(get_column_letter(len(header))))
#                     ws["A3"] = f"भरणा अहवाल ({from_date.strftime('%B %Y')} ते {to_date.replace(day=1) - timedelta(days=1):%B %Y})"
#                     ws["A3"].font = Font(size=12, bold=True)
#                     ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

#                     ws.append([])  # Empty row before table

#                     # ================= TABLE HEADER ==================
#                     start_row = ws.max_row + 1
#                     header_fill = PatternFill(start_color="727cf5", end_color="727cf5", fill_type="solid")
#                     border_style = Border(
#                         left=Side(border_style="thin", color="000000"),
#                         right=Side(border_style="thin", color="000000"),
#                         top=Side(border_style="thin", color="000000"),
#                         bottom=Side(border_style="thin", color="000000"),
#                     )

#                     for col_num, col_name in enumerate(header, 1):
#                         cell = ws.cell(row=start_row, column=col_num, value=col_name)
#                         cell.font = Font(bold=True, color="FFFFFF")
#                         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#                         cell.fill = header_fill
#                         cell.border = border_style
#                         ws.column_dimensions[get_column_letter(col_num)].width = 20

#                     ws.row_dimensions[start_row].height = 25

#                     # ================= TABLE DATA ==================
#                     for row_num, row_data in enumerate(data[1:], start_row + 1):
#                         for col_num, value in enumerate(row_data, 1):
#                             cell = ws.cell(row=row_num, column=col_num, value=value)
#                             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#                             cell.border = border_style

#                     # ================= FOOTER ==================
#                     footer_row = ws.max_row + 2
#                     ws.merge_cells(f"A{footer_row}:{get_column_letter(len(header))}{footer_row}")
#                     ws[f"A{footer_row}"] = f"अहवाल तयार दिनांक : {datetime.now().strftime('%d-%m-%Y')}"
#                     ws[f"A{footer_row}"].alignment = Alignment(horizontal="right", vertical="center")
#                     ws[f"A{footer_row}"].font = Font(size=10, italic=True, color="555555")

#                     # ================= SAVE & RETURN ==================
#                     buffer = io.BytesIO()
#                     wb.save(buffer)
#                     buffer.seek(0)
                    
#                     from django.http import HttpResponse

#                     filename = f"Payment_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#                     response = HttpResponse(
#                         buffer,
#                         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                     )
#                     response["Content-Disposition"] = f'attachment; filename="{filename}"'
#                     return response

        
#     except Exception as e:
#         tb = traceback.extract_tb(e.__traceback__)
#         fun = tb[0].name
#         callproc("stp_error_log",[fun,str(e),request.user.id])  
#         messages.error(request, 'Oops...! Something went wrong!')

from django.http import HttpResponse  # ✅ Add this import

@no_direct_access 
@login_required 
def payment_report(request):
    try:
        if request.user.is_authenticated == True:                
            global user
            user = request.user.id
            library_code = request.session.get('library_db', None)
            username = request.session.get('username', None)
            user_id = request.session["user_id"]
            role_id = request.session["role_id"]
            
        if not request.user.is_authenticated:
            # Clear any session flags
            if '_session_expired' in request.session:
                request.session.pop('_session_expired')
            messages.warning(request, "Your session has expired. Please log in again.")
            return redirect('library_list')
            
        if request.method == "GET":
            library_name = tbl_librarymasterL01.objects.using(library_code).filter(is_active=True).first()
            library_name_mar = library_name.library_name_mar if library_name else ''

            reports = PaymentReport.objects.all().order_by('-generated_date')

            for rep in reports:
                # encrypted id for links
                rep.report_encrypted_id = enc(str(rep.id))

                # Determine if the record is "filled" for the four fields
                receipt_no_filled = bool(rep.receipt_no and str(rep.receipt_no).strip() != "")
                
                try:
                    deposit_amount_val = rep.deposit_amount if rep.deposit_amount is not None else Decimal('0')
                except Exception:
                    deposit_amount_val = Decimal('0')
                deposit_amount_filled = (deposit_amount_val is not None) and (Decimal(str(deposit_amount_val)) != Decimal('0'))

                receipt_upload_filled = bool(rep.receipt_upload and str(rep.receipt_upload).strip() != "")
                deposit_date_filled = bool(rep.deposit_date is not None)

                # If ALL are filled => do NOT allow edit (can_edit = False)
                rep.can_edit = not (receipt_no_filled and deposit_amount_filled and receipt_upload_filled and deposit_date_filled)

            return render(request, 'Reports/payment_report.html', {
                'payment_reports': reports,
                'library_name_mar': library_name_mar
            })
            
        if request.method == "POST":
            
            report_type = request.POST.get("report_type")
            from_month = request.POST.get("from_month")
            to_month = request.POST.get("to_month")
            single_date = request.POST.get("single_date")  # ✅ Single date field

            # 🛑 Guard clause: If not PDF/Excel, create/update report
            if report_type not in ["pdf", "excel"]:
                # create post
                from_date = request.POST.get("from_date")
                total_amount = request.POST.get("total_amount")
                summary_data_json = request.POST.get("summary_data")

                # Convert to Python date objects
                from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()

                # Check if report already exists for same from_date
                existing_report = PaymentReport.objects.filter(
                    from_date=from_date_obj
                ).first()

                try:
                    with transaction.atomic():

                        if existing_report:
                            # ------------ UPDATE EXISTING REPORT ------------
                            existing_report.to_date = from_date_obj
                            existing_report.total_amount = total_amount
                            existing_report.updated_by = user_id
                            existing_report.generated_date = timezone.now().date()
                            existing_report.save()

                            # DELETE old key-value pairs
                            PaymentReportKeyValue.objects.filter(payment_report=existing_report).delete()

                            # INSERT new key-value pairs
                            if summary_data_json:
                                summary_data = json.loads(summary_data_json)
                                for item in summary_data:
                                    PaymentReportKeyValue.objects.create(
                                        payment_report=existing_report,
                                        key=item.get("key"),
                                        value=item.get("value")
                                    )

                            messages.success(request, "Payment summary updated successfully!")

                        else:
                            # ------------ CREATE NEW REPORT ------------
                            report = PaymentReport.objects.create(
                                from_date=from_date_obj,
                                to_date=from_date_obj,
                                generated_date=timezone.now().date(),
                                total_amount=total_amount,
                                created_by=user_id,
                                updated_by=user_id
                            )

                            # Insert key-value pairs
                            if summary_data_json:
                                summary_data = json.loads(summary_data_json)
                                for item in summary_data:
                                    PaymentReportKeyValue.objects.create(
                                        payment_report=report,
                                        key=item.get("key"),
                                        value=item.get("value")
                                    )

                            messages.success(request, "Payment summary generated successfully!")

                except Exception as e:
                    transaction.set_rollback(True)
                    messages.error(request, f"Error while saving report: {str(e)}")
                    return redirect("payment_report")

                return redirect("payment_report")

            else:
                # ✅ Handle PDF/Excel export (month range or single date)
                try:
                    # Determine if it's single date or month range
                    if single_date:
                        # ---------- SINGLE DATE REPORT ----------
                        selected_date = datetime.strptime(single_date, "%Y-%m-%d").date()
                        
                        # ✅ Fix: Use exact date comparison for DateField
                        reports = PaymentReport.objects.filter(
                            generated_date=selected_date
                        )
                        
                        if not reports.exists():
                            return HttpResponse("या दिनांकासाठी कोणताही अहवाल उपलब्ध नाही. / No reports found for the selected date.", content_type="text/plain")
                        
                        # Set date range for display
                        from_date = selected_date
                        to_date = selected_date
                        
                        # Create month/year display strings
                        from_year = selected_date.year
                        from_month_num = selected_date.month
                        to_year = selected_date.year
                        to_month_num = selected_date.month
                        
                    else:
                        # ---------- MONTH RANGE REPORT ----------
                        if not from_month or not to_month:
                            return HttpResponse("कृपया महिना श्रेणी निवडा. / Please select month range.", content_type="text/plain")
                        
                        # Parse months
                        from_date = datetime.strptime(from_month + "-01", "%Y-%m-%d").date()
                        year, month = map(int, to_month.split("-"))
                        
                        # Calculate to_date (first day of next month)
                        if month == 12:
                            to_date = datetime(year + 1, 1, 1).date()
                        else:
                            to_date = datetime(year, month + 1, 1).date()
                        
                        # ✅ Fetch reports between date range (using DateField)
                        reports = PaymentReport.objects.filter(
                            generated_date__gte=from_date,
                            generated_date__lt=to_date
                        )
                        
                        if not reports.exists():
                            return HttpResponse("निवडलेल्या महिन्यांसाठी कोणताही अहवाल उपलब्ध नाही. / No reports found for the selected months.", content_type="text/plain")
                        
                        # Parse month/year for display
                        from_year, from_month_num = map(int, from_month.split("-"))
                        to_year, to_month_num = map(int, to_month.split("-"))
                    
                    # ✅ Collect all distinct keys for dynamic columns
                    all_keys = PaymentReportKeyValue.objects.values_list('key', flat=True).distinct()

                    # ✅ Build header
                    header = ["पावती क्रमांक", "अहवाल दिनांक", "ठेव दिनांक"] + list(all_keys) + ["ठेव रक्कम"]
                    data = [header]

                    for report in reports:
                        key_value_map = {kv.key: kv.value for kv in report.key_values.all()}

                        # ✅ Show only date part (no time)
                        generated_date = report.generated_date.strftime("%Y-%m-%d") if report.generated_date else "-"
                        deposit_date = report.deposit_date.strftime("%Y-%m-%d") if report.deposit_date else "-"

                        # ✅ Build row with deposit amount
                        row = [
                            report.receipt_no or "-",
                            generated_date,
                            deposit_date,
                        ]

                        for key in all_keys:
                            row.append(key_value_map.get(key, "0.00"))

                        # ✅ Add deposit amount at end
                        row.append(f"{report.deposit_amount:.2f}" if report.deposit_amount else "0.00")

                        data.append(row)
                    
                    # Create report range string for display
                    marathi_months = {
                        1: "जानेवारी", 2: "फेब्रुवारी", 3: "मार्च", 4: "एप्रिल", 5: "मे", 6: "जून",
                        7: "जुलै", 8: "ऑगस्ट", 9: "सप्टेंबर", 10: "ऑक्टोबर", 11: "नोव्हेंबर", 12: "डिसेंबर"
                    }
                    
                    if single_date:
                        month_name_mar = marathi_months.get(from_month_num, "")
                        report_range = f"{month_name_mar} {from_year}"
                    else:
                        from_month_name_mar = marathi_months.get(from_month_num, "")
                        to_month_name_mar = marathi_months.get(to_month_num, "")
                        report_range = f"{from_month_name_mar} {from_year} ते {to_month_name_mar} {to_year}"

                    # ==================== PDF Export ====================
                    if report_type == "pdf":
                        from weasyprint import HTML, CSS
                        import tempfile, os
                        from django.conf import settings
                        
                        # 🔹 Fetch library info
                        library = tbl_librarymasterL01.objects.filter(library_code=library_code).first()
                        library_name = library.library_name_mar if library else "ग्रंथालयाचे नाव उपलब्ध नाही"
                        
                        # 🔹 Build Marathi title data
                        today_str = datetime.now().strftime("%d-%m-%Y")
                        report_title = "भरणा अहवाल"
                        file_name = f"Payment_Report_{today_str}.pdf"
                        
                        # ✅ Font path - try multiple possible locations
                        font_paths = [
                            os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSansDevanagari-Regular.ttf"),
                            os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSansDevanagari-Regular.otf"),
                            os.path.join(settings.BASE_DIR, "static", "fonts", "Mangal.ttf"),
                            os.path.join(settings.BASE_DIR, "static", "fonts", "Nirmala.ttf"),
                            "/usr/share/fonts/truetype/noto/NotoSansDevanagari-Regular.ttf",
                            "/System/Library/Fonts/Supplemental/Arial.ttf",
                        ]
                        
                        font_path = None
                        for path in font_paths:
                            if os.path.exists(path):
                                font_path = path
                                break
                        
                        # If no font found, use system default
                        font_face_style = ""
                        if font_path:
                            font_face_style = f"""
                            @font-face {{
                                font-family: 'MarathiFont';
                                src: url('file://{font_path}') format('truetype');
                            }}
                            """
                        
                        # ✅ Build HTML with inline styles
                        html_content = f"""
                        <!DOCTYPE html>
                        <html lang="mr">
                            <head>
                                <meta charset="UTF-8">
                                <title>{report_title}</title>
                                <style>
                                    {font_face_style}
                                    @page {{
                                        size: A4 landscape;
                                        margin: 1.5cm;
                                        @bottom-center {{
                                            content: "पृष्ठ " counter(page) " / " counter(pages);
                                            font-size: 9px;
                                            color: #666;
                                            font-family: {font_face_style and "'MarathiFont'" or "'DejaVu Sans', 'Arial', sans-serif"};
                                        }}
                                        @bottom-right {{
                                            content: "अहवाल तयार दिनांक : {today_str}";
                                            font-size: 8px;
                                            color: #444;
                                            font-family: {font_face_style and "'MarathiFont'" or "'DejaVu Sans', 'Arial', sans-serif"};
                                        }}
                                    }}
                                    body {{
                                        font-family: {font_face_style and "'MarathiFont'" or "'DejaVu Sans', 'Arial', sans-serif"};
                                        font-size: 9px;
                                        line-height: 1.4;
                                        color: #333;
                                        margin: 0;
                                        padding: 0;
                                        background: #fff;
                                    }}
                                    .report-container {{
                                        width: 100%;
                                    }}
                                    .header {{
                                        text-align: center;
                                        margin-bottom: 15px;
                                        padding: 10px;
                                        border-bottom: 2px solid #727cf5;
                                    }}
                                    .header h1 {{
                                        font-size: 18px;
                                        margin: 0 0 5px 0;
                                        color: #333;
                                        font-weight: bold;
                                    }}
                                    .header h2 {{
                                        font-size: 14px;
                                        margin: 0 0 3px 0;
                                        color: #555;
                                        font-weight: normal;
                                    }}
                                    .header h3 {{
                                        font-size: 12px;
                                        margin: 5px 0 0 0;
                                        color: #727cf5;
                                        font-weight: normal;
                                    }}
                                    table {{
                                        width: 100%;
                                        border-collapse: collapse;
                                        margin-top: 10px;
                                        font-size: 8px;
                                    }}
                                    th, td {{
                                        border: 0.8px solid #ddd;
                                        padding: 6px 4px;
                                        text-align: center;
                                        vertical-align: middle;
                                        word-wrap: break-word;
                                    }}
                                    th {{
                                        background-color: #727cf5;
                                        color: white;
                                        font-weight: bold;
                                        font-size: 9px;
                                        padding: 8px 4px;
                                    }}
                                    tr:nth-child(even) {{
                                        background-color: #f9f9f9;
                                    }}
                                    .footer {{
                                        margin-top: 15px;
                                        padding-top: 8px;
                                        text-align: right;
                                        font-size: 8px;
                                        color: #666;
                                        border-top: 1px solid #ddd;
                                    }}
                                </style>
                            </head>
                            <body>
                                <div class="report-container">
                                    <div class="header">
                                        <h1>नवी मुंबई महानगरपालिका</h1>
                                        <h2>{library_name}</h2>
                                        <h3>{report_title} - कालावधी: {report_range}</h3>
                                    </div>
                                    
                                    <table>
                                        <thead>
                                            <tr>
                                                {''.join(f'<th>{h}</th>' for h in data[0])}
                                            </tr>
                                        </thead>
                                        <tbody>
                                            {''.join('</tr>' + ''.join(f'<td>{c}</td>' for c in row) + '</tr>' for row in data[1:])}
                                        </tbody>
                                    </table>
                                    
                                    <div class="footer">
                                        <div>अहवाल तयार दिनांक : {today_str}</div>
                                        <div>एकूण रेकॉर्ड : {len(data) - 1}</div>
                                    </div>
                                </div>
                            </body>
                        </html>
                        """
                        
                        # ✅ Generate PDF
                        pdf_file = HTML(string=html_content).write_pdf(
                            stylesheets=[CSS(string="@page { size: A4 landscape; margin: 1.5cm; }")],
                            presentational_hints=True
                        )
                        
                        # ✅ Return PDF as download
                        response = HttpResponse(pdf_file, content_type="application/pdf")
                        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
                        return response
                    
                    # ==================== Excel Export ====================
                    elif report_type == "excel":
                        import openpyxl
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                        from openpyxl.utils import get_column_letter
                        
                        # 🔹 Fetch library info
                        library = tbl_librarymasterL01.objects.filter(library_code=library_code).first()
                        library_name = library.library_name_mar if library else "ग्रंथालयाचे नाव उपलब्ध नाही"

                        # 🔹 Build workbook and sheet
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Payment Report"

                        # ================= HEADER SECTION ==================
                        ws.merge_cells(f"A1:{get_column_letter(len(header))}1")
                        ws["A1"] = "नवी मुंबई महानगरपालिका"
                        ws["A1"].font = Font(size=16, bold=True)
                        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

                        ws.merge_cells(f"A2:{get_column_letter(len(header))}2")
                        ws["A2"] = library_name
                        ws["A2"].font = Font(size=13, bold=True)
                        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
                        
                        ws.merge_cells(f"A3:{get_column_letter(len(header))}3")
                        ws["A3"] = f"भरणा अहवाल ({report_range})"
                        ws["A3"].font = Font(size=12, bold=True)
                        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

                        ws.append([])

                        # ================= TABLE HEADER ==================
                        start_row = ws.max_row + 1
                        header_fill = PatternFill(start_color="727cf5", end_color="727cf5", fill_type="solid")
                        border_style = Border(
                            left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000"),
                        )

                        for col_num, col_name in enumerate(header, 1):
                            cell = ws.cell(row=start_row, column=col_num, value=col_name)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.fill = header_fill
                            cell.border = border_style
                            ws.column_dimensions[get_column_letter(col_num)].width = 20

                        ws.row_dimensions[start_row].height = 25

                        # ================= TABLE DATA ==================
                        for row_num, row_data in enumerate(data[1:], start_row + 1):
                            for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                cell.border = border_style

                        # ================= FOOTER ==================
                        footer_row = ws.max_row + 2
                        ws.merge_cells(f"A{footer_row}:{get_column_letter(len(header))}{footer_row}")
                        ws[f"A{footer_row}"] = f"अहवाल तयार दिनांक : {datetime.now().strftime('%d-%m-%Y')}"
                        ws[f"A{footer_row}"].alignment = Alignment(horizontal="right", vertical="center")
                        ws[f"A{footer_row}"].font = Font(size=10, italic=True, color="555555")

                        # ================= SAVE & RETURN ==================
                        buffer = io.BytesIO()
                        wb.save(buffer)
                        buffer.seek(0)
                        
                        filename = f"Payment_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        response = HttpResponse(
                            buffer,
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                        response["Content-Disposition"] = f'attachment; filename="{filename}"'
                        return response
                        
                except Exception as e:
                    print(f"Error in PDF/Excel generation: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    return HttpResponse(f"Error generating report: {str(e)}", content_type="text/plain")
        
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
        return redirect("payment_report")

# @login_required
# def get_payment_preview(request):
#     try:
#         from_date = request.POST.get("from")
#         to_date = request.POST.get("to")

#         if not from_date or not to_date:
#             return JsonResponse({"error": "Missing date range"}, status=400)

#         # Step 1️⃣ - Get payments in selected date range
#         payments = PaymentDetails.objects.filter(payment_date__range=[from_date, to_date])

#         # Step 2️⃣ - Get all membership types from master (so we show all even if count = 0)
#         all_memberships = MembershipMaster.objects.all().values("membership_type_en", "membership_type")

#         # Step 3️⃣ - Initialize dictionary with all membership types (English + Marathi)
#         membership_summary = {
#             f"{m['membership_type_en']} ({m['membership_type']})": {"count": 0, "total": 0}
#             for m in all_memberships
#         }

#         # Step 4️⃣ - Add totals from payments if available
#         for pay in payments.select_related("membership__membership"):
#             mem = pay.membership
#             if mem and mem.membership:
#                 mtype_en = mem.membership.membership_type_en
#                 mtype_mr = mem.membership.membership_type
#                 label = f"{mtype_en} ({mtype_mr})"

#                 if label in membership_summary:
#                     membership_summary[label]["count"] += 1
#                     membership_summary[label]["total"] += float(pay.total_subscription_amount or 0)

#         # Step 5️⃣ - Overall totals
#         total_deposit = payments.aggregate(total=Sum("deposit_amount")).get("total") or 0
#         total_fine = payments.aggregate(total=Sum("fine_amount")).get("total") or 0
#         total_subscription = sum(v["total"] for v in membership_summary.values())

#         # Step 6️⃣ - Convert to list for frontend
#         membership_list = [
#             {
#                 "membership_type": k,
#                 "total_amount": round(v["total"], 2),
#                 "count": v["count"]
#             }
#             for k, v in membership_summary.items()
#         ]

#         # Step 7️⃣ - Final data
#         data = {
#             "deposit_total": round(total_deposit, 2),
#             "fine_total": round(total_fine, 2),
#             "total_subscription_amount": round(total_subscription, 2),
#             "membership_summary": membership_list
#         }

#         return JsonResponse(data)
    
#     except Exception as e:
#         tb = traceback.extract_tb(e.__traceback__)
#         fun = tb[-1].name if tb else "get_payment_preview"
#         callproc("stp_error_log", [fun, str(e), request.user.id])
#         messages.error(request, "Oops...! Something went wrong!")
#         return JsonResponse({"error": "Something went wrong. Please try again later."}, status=500)

@login_required
def get_payment_preview(request):
    try:
        from_date = request.POST.get("from")
        # to_date = request.POST.get("to")

        if not from_date:
            return JsonResponse({"error": "Missing date range"}, status=400)

        # Step 1️⃣ - Get payments in selected date range
        payments = PaymentDetails.objects.filter(payment_date=from_date)

        # Step 2️⃣ - Get all membership types from master (so we show all even if count = 0)
        all_memberships = MembershipMaster.objects.all().values("membership_type_en", "membership_type")

        # Step 3️⃣ - Initialize dictionary with all membership types (English + Marathi)
        membership_summary = {
            f"{m['membership_type_en']} ({m['membership_type']})": {"count": 0, "total": 0}
            for m in all_memberships
        }

        # Step 4️⃣ - Add totals from payments if available
        for pay in payments.select_related("membership__membership"):
            mem = pay.membership
            if mem and mem.membership:
                mtype_en = mem.membership.membership_type_en
                mtype_mr = mem.membership.membership_type
                label = f"{mtype_en} ({mtype_mr})"

                if label in membership_summary:
                    membership_summary[label]["count"] += 1
                    membership_summary[label]["total"] += float(pay.total_subscription_amount or 0)

        # Step 5️⃣ - Overall totals
        entry_fees = payments.aggregate(total=Sum("entry_fee_amount")).get("total") or 0
        total_deposit = payments.aggregate(total=Sum("deposit_amount")).get("total") or 0
        total_fine = payments.aggregate(total=Sum("fine_amount")).get("total") or 0
        total_subscription = sum(v["total"] for v in membership_summary.values())

        # Step 6️⃣ - Convert to list for frontend
        membership_list = [
            {
                "membership_type": k,
                "total_amount": round(v["total"], 2),
                "count": v["count"]
            }
            for k, v in membership_summary.items()
        ]

        # Step 7️⃣ - Final data
        data = {
            "entry_fees":round(entry_fees, 2),
            "deposit_total": round(total_deposit, 2),
            "fine_total": round(total_fine, 2),
            "total_subscription_amount": round(total_subscription, 2),
            "membership_summary": membership_list
        }

        return JsonResponse(data)
    
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[-1].name if tb else "get_payment_preview"
        callproc("stp_error_log", [fun, str(e), request.user.id])
        messages.error(request, "Oops...! Something went wrong!")
        return JsonResponse({"error": "Something went wrong. Please try again later."}, status=500)

@no_direct_access 
@login_required 
def view_payment_report(request):
    try:
        if not request.user.is_authenticated:
            # Clear any session flags
            if '_session_expired' in request.session:
                request.session.pop('_session_expired')
            messages.warning(request, "Your session has expired. Please log in again.")
            return redirect('library_list')
        
        report_encrypted_id = request.GET.get("report_encrypted_id")
        report_id = dec(report_encrypted_id)
        report = get_object_or_404(PaymentReport, id=report_id)
        key_values = PaymentReportKeyValue.objects.filter(payment_report=report)
       
        if report.receipt_upload:
            report.report_encrypted_receipt_upload = enc(str(report.receipt_upload))

        return render(request, 'Reports/view_payment_report.html', {
            'report': report,
            'key_values': key_values,
            "MEDIA_URL": settings.MEDIA_URL,
        })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
        return redirect("payment_report")

@login_required 
def edit_payment_report(request):
    try:
        
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id
            library_code = request.session.get('library_db', None)
            username = request.session.get('username', None)
            user_id = request.session["user_id"]
            role_id = request.session["role_id"]
        
        if not request.user.is_authenticated:
            # Clear any session flags
            if '_session_expired' in request.session:
                request.session.pop('_session_expired')
            messages.warning(request, "Your session has expired. Please log in again.")
            return redirect('library_list')
            
        if request.method == "GET":
            
            report_encrypted_id = request.GET.get("report_encrypted_id")
            report_id = dec(report_encrypted_id)
            report = get_object_or_404(PaymentReport, id=report_id)
            key_values = PaymentReportKeyValue.objects.filter(payment_report=report)

            return render(request, 'Reports/edit_payment_report.html', {
                'report': report,
                'key_values': key_values,
                'report_encrypted_id': report_encrypted_id
            })

        # --- POST Method: Save updates ---
        elif request.method == "POST":
            report_encrypted_id = request.POST.get("report_encrypted_id")
            report_id = dec(report_encrypted_id)
            report = get_object_or_404(PaymentReport, id=report_id)

            receipt_no = request.POST.get("receipt_no")
            deposit_amount = request.POST.get("deposit_amount")
            remarks = request.POST.get("remarks")
            deposit_date = request.POST.get("deposit_date")

            report.receipt_no = receipt_no
            report.deposit_amount = deposit_amount
            report.remarks = remarks

            # Parse and store deposit date
            if deposit_date:
                from datetime import datetime
                report.deposit_date = datetime.strptime(deposit_date, "%Y-%m-%d")

            # --- Handle File Upload (Receipt PDF/Image/Excel) ---
            # receipt_file = request.FILES.get("receipt_upload")
            # if receipt_file:
            #     from django.core.files.storage import default_storage
            #     import os
            #     base_folder = os.path.join(settings.MEDIA_ROOT, library_code, "Receipts", "PaymentReceipts")
            #     receipt_folder = os.path.join(base_folder, str(receipt_no))
            #     os.makedirs(receipt_folder, exist_ok=True)
            #     original_name, ext = os.path.splitext(receipt_file.name)
            #     date_suffix = deposit_date.replace("-", "") if deposit_date else datetime.now().strftime("%Y%m%d")
            #     file_name = f"{original_name}_{receipt_no}{ext}"
            #     file_path = os.path.join(receipt_folder, file_name)
            #     with default_storage.open(file_path, "wb+") as destination:
            #         for chunk in receipt_file.chunks():
            #             destination.write(chunk)
            #     relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
            #     report.receipt_upload = relative_path
            receipt_file = request.FILES.get("receipt_upload")
            if receipt_file:
                import os
                from datetime import datetime

                # Extract filename + extension
                original_name, ext = os.path.splitext(receipt_file.name)

                # Optional: sanitize filename (recommended)
                safe_name = original_name.replace(" ", "_")

                # Use deposit_date or fallback
                date_suffix = (
                    deposit_date.replace("-", "")
                    if deposit_date
                    else datetime.now().strftime("%Y%m%d")
                )

                # Final file name
                safe_receipt_no = receipt_no.replace("/", "-")

                file_name = f"{safe_name}_{safe_receipt_no}{ext}"

                save_path = (
                    f"{library_code}/Receipts/PaymentReceipts/"
                    f"{safe_receipt_no}/{file_name}"
                )

                normalized_path = file_storage_service.save_file(
                    file=receipt_file,
                    save_path=save_path
                )

                report.receipt_upload = normalized_path
            

            report.updated_by = user_id
            report.save()

            messages.success(request, "पावती अहवाल यशस्वीपणे अद्ययावत केला गेला आहे!")
            return redirect("payment_report")
        
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log", [fun, str(e), request.user.id])
        messages.error(request, 'Oops...! Something went wrong!')
        return redirect("payment_report")

@login_required
def view_secure_receipt(request, enc_id):
    try:
        # decrypt relative file path
        relative_path = dec(enc_id)

        # get secure file URL from storage service
        file_url = file_storage_service.get_file_url(relative_path)

        if not file_url:
            raise Http404("File not found.")

        # redirect to secure/signed URL
        return redirect(file_url)

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log", [fun, str(e), request.user.id])
        messages.error(request, 'Oops...! Something went wrong!')
        return redirect("payment_report")

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
        return redirect("payment_report")
   
# pdf marathi download payment    
def to_marathi_digits(date_obj):
    if not date_obj:
        return ""
    text = str(date_obj)
    return text.translate(str.maketrans("0123456789", "०१२३४५६७८९"))

from weasyprint import HTML, CSS

@login_required
def download_secure_receipt(request):
    try:
        report_encrypted_id = request.GET.get("report_encrypted_id")
        report_id = dec(report_encrypted_id)
        report = PaymentReport.objects.get(id=report_id)
        key_values = PaymentReportKeyValue.objects.filter(payment_report=report).order_by("id")

        library_code = request.session.get("library_db")
        library = tbl_librarymasterL01.objects.filter(library_code=library_code).first()

        today_str = datetime.now().strftime("%d-%m-%Y")
        period = f"{report.from_date.strftime('%d-%m-%Y')} ते {report.to_date.strftime('%d-%m-%Y')}"
        file_name = f"{report.from_date.strftime('%d-%m-%Y')} ते {report.to_date.strftime('%d-%m-%Y')}.pdf"

        # Path to your Marathi font
        font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSansDevanagari-Regular.ttf")

        html_content = f"""
        <!DOCTYPE html>
        <html lang="mr">
            <head>
                <meta charset="UTF-8">
                <title>भरणा अहवाल</title>
                <style>
                    @font-face {{
                        font-family: 'MarathiFont';
                        src: url('file://{font_path}') format('truetype');
                    }}
                    body {{
                        font-family: 'MarathiFont', sans-serif;
                        background: #f8f9fa;
                        padding: 40px;
                    }}
                    h1, h2, h3 {{
                        text-align: center;
                        margin: 0;
                    }}
                    p {{
                        font-size: 18px;
                        margin: 5px 0;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin-top: 25px;
                    }}
                    th, td {{
                        border: 1px solid #ccc;
                        padding: 8px 10px;
                        font-size: 18px;
                        text-align: left;
                    }}
                    th {{
                        background: #d7e3fc;
                        text-align: center;
                    }}
                    tr:nth-child(even) {{
                        background: #f8fafd;
                    }}
                </style>
            </head>
            <body>
                <h1>नवी मुंबई महानगरपालिका</h1>
                <h2>{library.library_name_mar or library.library_name} - {library.location}</h2>
                <h3>भरणा अहवाल</h3>
                <p>दिनांक: {today_str}</p>
                <p>कालावधी: {period}</p>

                <table>
                    <tr>
                        <th>क्र.</th>
                        <th>माहिती</th>
                        <th>तपशील</th>
                    </tr>
            """

        for idx, kv in enumerate(key_values, start=1):
            html_content += f"""
                <tr>
                    <td style="text-align:center;">{idx}</td>
                    <td>{kv.key or '-'}</td>
                    <td>{kv.value or '-'}</td>
                </tr>
            """

        html_content += """
            </table>
        </body>
        </html>
        """

        pdf_file = HTML(string=html_content).write_pdf(
            stylesheets=[CSS(string="@page { size: A4; margin: 1in; }")]
        )

        response = HttpResponse(pdf_file, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        return response

    except Exception as e:
        print("PDF Error:", e)
        messages.error(request, "Something went wrong generating the PDF.")
        return redirect("payment_report")

# Report section Previous
@login_required
def common_html(request):
    title,note ='',''
    try:
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id
            entity =request.GET.get('entity', '')  
            title,note ='',''
            if request.method=="GET":
                forms = callproc("stp_get_forms",['report',user]) 
                if entity == '' or None:
                   entity =  forms[0][0]         
                filter_name = callproc("stp_get_filter_names",[entity])          
                column_name = callproc("stp_get_column_names",[entity])        
                result = callproc("stp_get_report_title", [entity])
                if result and result[0]:
                    items = result[0] 
                    if isinstance(items, tuple):
                        if len(items) == 2:
                            title, note = items
                        elif len(items) == 1:
                            title = items[0]
                            note = ''
                    else:
                        title = items
                        note = ''
                saved_names = callproc("stp_get_saved_filters",[entity,user])        

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        return render(request,'Reports/common_reports.html', {'forms':forms,'filter_name':filter_name,'column_name':column_name,'saved_names':saved_names,'entity':entity,'title':title,'note':note})
    
@login_required      
def get_filter(request):
    try:
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id
            if request.method=="GET":
                entity =request.GET.get('entity', '')
                data4 = callproc("stp_get_filter_names",[entity])
                drop_down=[]
                data5 = []
                for items in data4:
                    data5=[]
                    data5=list(items)
                    unit = common_model(id1=data5[0], name=data5[1])
                    drop_down.append(common_dict(unit))
                if len(drop_down) == 0:
                    drop_down = 0
    except Exception  as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        return JsonResponse(drop_down, safe=False)
    
def common_dict(unit):
    return {
        'id1': unit.id1,
        'name': unit.name,
    }  
    
@login_required    
def get_sub_filter(request):
    try:
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id 
            if request.method=="GET":
                filter_id =request.GET.get('filter_id', '')
                data4 = callproc("stp_get_sub_filter",[filter_id,user])
                drop_down=[]
                data5 = []
                for items in data4:
                    data5=[]
                    data5=list(items)
                    unit = common_model(id1=data5[0], name=data5[1])
                    drop_down.append(common_dict(unit))
                if len(drop_down) == 0:
                    drop_down = 0 
    except Exception  as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        return JsonResponse(drop_down, safe=False)  

@login_required
def add_new_filter(request):
    try:
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id
            if request.method == "GET":
                filter_count =str(request.GET.get('filter_count', ''))
                entity =str(request.GET.get('entity', ''))
                filter_name = callproc("stp_get_filter_names",[entity])                
                fId = filter_count + 'filterId'           
                sfId = filter_count + 'subFilterId'           
                context = {'filter_name':filter_name,'fId':fId,'sfId':sfId,'fcount':filter_count}
                html = render_to_string('Reports/_add_new_filter.html', context)
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        data = {'html': html}
        return JsonResponse(data, safe=False)
    
@login_required
def partial_report(request):
    try:
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id
            if request.method == "GET":
                columnName =str(request.GET.get('columnName', ''))
                filterid =str(request.GET.get('filterid', ''))
                subFilterId =str(request.GET.get('subFilterId', ''))
                sft =str(request.GET.get('sft', ''))
                entity =str(request.GET.get('entity', ''))
                filterid1 = filterid.split(',')
                SubFilterId1 = subFilterId.split(',')
                sft1 = sft.split(',')
                data = common_fun(columnName,filterid1,SubFilterId1,sft1,entity,user,'0')
                headers = data['headers']
                emptycheck = data['emptycheck']
                data_list = data['data_list']
                display_name_list = data['display_name_list']
                # entityName = entity
                context = {'emptycheck':emptycheck,'columns':display_name_list,'rows':data_list,'entity':entity}
                html = render_to_string('Reports/_partial_report.html', context)
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        data = {'html': html}
        return JsonResponse(data, safe=False)
    
def common_fun(columnName,filterid,SubFilterId,sft,entity,user,is_export):
    try:
        report_filters= []
        report_columns= []
        column_join_list= []
        mandatory_arr= []
        result_data = callproc("stp_get_report_filters", [entity])
        if result_data and result_data[0]:
            for row in result_data:
                report_filters.append(list(row))
        result_data =callproc("stp_get_report_columns", [entity])
        if result_data and result_data[0]:
            for row in result_data:
                report_columns.append(list(row))
        result_data = callproc("stp_get_column_join", [entity])
        if result_data and result_data[0]:
            for row in result_data:
                column_join_list.append(list(row))
        result_data = callproc("stp_get_mandatory", [entity])
        mandf = ''
        if result_data and result_data[0]: 
            mandf = result_data[0][0]
        if mandf != '':
            mandatory_arr = mandf.split(',')
      
        from_clause = ""
        language = ""
        where_clause = ""
        where_extra = ""
        join_query = ""
        join_clause = ""
        from_clause1 = ""
        where_clause1 = [""] * len(filterid)
        join_query1 = [""] * len(filterid)
        order_by = ""
        group_by = ""
        columns = ""
        b = 0
        for fid in filterid:
            from_clause = next((f[4] for f in report_filters if f[0] == int(fid)), '')
            if from_clause != '':
                from_clause1 = from_clause

            where_clause1[b] = next((f[5] for f in report_filters if f[0] == int(fid)), '')
            join_query1[b] = next((f[6] for f in report_filters if f[0] == int(fid)), '')
            group_by = next((f[7] for f in report_filters if f[0] == int(fid)), '')
            order_by = next((f[8] for f in report_filters if f[0] == int(fid)), '')
            
            where_clause1[b] = where_clause1[b] if where_clause1[b] is not None else ''
            join_query1[b] = join_query1[b] if join_query1[b] is not None else ''
            group_by = group_by if group_by is not None else ''
            order_by = order_by if order_by is not None else ''
            b += 1
        from_clause = from_clause1
        header_filter = []
        header_sub_filter = []
        filter_name = ""
        cnt1 = 0
        for i in range(len(filterid)):
                if sft[i] and sft[i].strip() not in ("", "0"):
                    filter_name = next((f[2] for f in report_filters if f[0] == int(filterid[i])), '')
                    if filter_name in header_filter:
                       idx = header_filter.index(filter_name)
                       header_sub_filter[idx] += '|' + sft[i]
                    else:
                       header_filter.append(filter_name)
                       header_sub_filter.append(sft[i])  
        b =0
        for sub in range(len(SubFilterId)):
            SubFilterId[sub] = SubFilterId[sub].replace("|", "','")
            if not SubFilterId[sub] or SubFilterId[sub] in ("0", "", " "):
                where_clause1[b] = ""
            else:
                where_clause1[b] = where_clause1[b].replace("BindPara1", SubFilterId[sub])
            b += 1
        
        column_name = callproc("stp_get_dispay_names",[entity])        
        
                
        if columnName == '': 
            column_name_arr = [col[0] for col in column_name] 
            display_name_arr = [col[1] for col in column_name]
            columns = " , ".join(column_name_arr)
        else :
            column_name_arr = columnName.split(',')
            for i, col in enumerate(column_name_arr):
                column_name_arr[i] = col.replace('|', ',')
            display_name_arr = []

            for item in column_name:
                if item[0] in column_name_arr:
                    display_name_arr.append(item[1])

            columns = " , ".join(column_name_arr)

        display_names = " , ".join(display_name_arr)

        for dr in column_join_list:
            check = dr[0]
            if check in columns:
                replace = dr[1]
                columns = columns.replace(check, replace)
            join_clause += dr[2] + " "
        for z in range(len(filterid)):
            if where_clause1[z] not in where_clause:
                if not where_clause:
                    where_clause = " where " + where_clause1[z]
                else:
                    where_clause += " and " + where_clause1[z]
                            
        if join_query1[z] not in join_clause:
            join_clause += join_query1[z]

        sql_query = "Select " + columns + " " + from_clause + " " + join_clause + " " + where_clause + " " + where_extra + " " + group_by + " " + order_by
        
        ch = 0
        for value in mandatory_arr:
            if value not in filterid:
                ch = 1
                break
        if ch == 0:
            if not all(value.strip() for value in SubFilterId[:len(mandatory_arr)]):
                ch = 1
            elif len(filterid) != len(SubFilterId):
                ch = 1
                
        data_list= []
        if ch == 0:
            result_data = callproc("stp_get_execute_report_query", [sql_query])
            if result_data and result_data[0]:
                data_list = preprocess_data_list(result_data,is_export)
      
        display_name_list = list(display_name_arr)

        if len(data_list) > 0:
            emptycheck = 0
        else : emptycheck = 1
        
        hl = []
        for filter_key, filter_value in zip(header_filter, header_sub_filter):
            if "|" in filter_value:
                values = filter_value.split("|")
                hl.append(f"{filter_key} :- ({','.join(values)})")
            else:
                hl.append(f"{filter_key} :- {filter_value}")
        hl_r = " , ".join(hl)

        data = {
               'headers': hl_r,
               'emptycheck': emptycheck,
               'data_list': data_list,
               'display_name_list': display_name_list,
               'sql_query': sql_query,
               'display_names': display_names
            }
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),user])  
    finally:
          return data

def dl_file(request, file_id):
    try:
        form_file = form_file.objects.get(id=dec(file_id))
        file_path = os.path.join(MEDIA_ROOT, form_file.file_path)
        if not os.path.exists(file_path):
            raise Http404("File not found.")
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=form_file.uploaded_name)
    except form_file.DoesNotExist:
        raise Http404("File not found.")

def preprocess_data_list(result_data, is_export):
    data_list = []
    for row in result_data:
        processed_row = []
        for value in row:
            if isinstance(value, str) and 'tdmsformfiles_' in value:
                file_ids = [v.replace('tdmsformfiles_', '') for v in value.split(',') if v.startswith('tdmsformfiles_')]
                uploaded_names = []
                if is_export == '1':
                    for file_id in file_ids:
                        try:
                            form_file = form_file.objects.get(id=file_id)
                            uploaded_names.append(form_file.uploaded_name)
                        except form_file.DoesNotExist:
                            continue
                    processed_row.append(', '.join(uploaded_names))
                else:
                    file_links = []
                    for file_id in file_ids:
                        try:
                            form_file = form_file.objects.get(id=file_id)
                            # form_file_id1 = callproc("stp_get_check_file",[form_file.id,user])
                            # if form_file_id1 and form_file_id1[0][0] == form_file.id:
                            file_path = os.path.join(MEDIA_ROOT, form_file.file_path)
                            file_exists = os.path.exists(file_path)
                            file_links.append({
                                'file_name': form_file.uploaded_name,
                                'exists': file_exists,
                                'id': enc(str(file_id)),
                            })
                            # else:
                            #     uploaded_names.append(form_file.uploaded_name)
                            #     processed_row.append(', '.join(uploaded_names))

                        except form_file.DoesNotExist:
                            continue
                    if file_links:
                        processed_row.append({'file_links': file_links})
            else:
                processed_row.append(value)
        data_list.append(processed_row)
    return data_list

def render_to_pdf(html):
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if pdf.err:
        return HttpResponse("Invalid PDF", status_code=400, content_type='text/plain')
    return HttpResponse(result.getvalue(), content_type='application/pdf')

@login_required
def report_pdf(request):
    response = ''
    html_string = ''
    try:
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id
            if request.method == "POST":
                columnName = str(request.POST.get('columnName', ''))
                filterid = str(request.POST.get('filterid', ''))
                subFilterId = str(request.POST.get('subFilterId', ''))
                sft = str(request.POST.get('sft', ''))
                entity = str(request.POST.get('entity', ''))
                filterid1 = filterid.split(',')
                SubFilterId1 = subFilterId.split(',')
                sft1 = sft.split(',')
                data = common_fun(columnName, filterid1, SubFilterId1, sft1, entity, user, '1')

                headers = data['headers']
                emptycheck = data['emptycheck']
                data_list = data['data_list']
                column_list = data['display_name_list']

                result_data = callproc("stp_get_report_title", [entity])
                title = ''
                if result_data and result_data[0]:
                    for items in result_data:  
                        title = items[0]
                        
                html_string = render_to_string('Reports/report_template.html', {
                    'title': title,
                    'headers': headers,
                    'column_list': column_list,
                    'data_list': data_list,
                })
                pdf = render_to_pdf(html_string)
                filename = title+'.pdf'
                if pdf:
                    response = FileResponse(pdf, content_type='application/pdf')
                    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        return response
    
@login_required
def report_xlsx(request):
    response = ''
    try:
        if request.user.is_authenticated:                
            global user
            user = request.user.id
            if request.method == "POST":
                columnName = str(request.POST.get('columnName', ''))
                filterid = str(request.POST.get('filterid', ''))
                subFilterId = str(request.POST.get('subFilterId', ''))
                sft = str(request.POST.get('sft', ''))
                entity = str(request.POST.get('entity', ''))
                filterid1 = filterid.split(',')
                SubFilterId1 = subFilterId.split(',')
                sft1 = sft.split(',')
                data = common_fun(columnName, filterid1, SubFilterId1, sft1, entity, user, '1')

                headers = data['headers']
                emptycheck = data['emptycheck']
                data_list = data['data_list']
                column_list = data['display_name_list']

                result_data = callproc("stp_get_report_title", [entity])
                title = ''
                if result_data and result_data[0]:
                    for items in result_data:  
                        title = items[0]

                output = io.BytesIO()
                workbook = xlsxwriter.Workbook(output)
                worksheet = workbook.add_worksheet(str(entity))
            
                # Insert logo
                worksheet.insert_image('A1', 'static/images/technologo.png', {
                    'x_offset': 1, 'y_offset': 1, 'x_scale': 0.04, 'y_scale': 0.04
                })
                
                # Formats
                header_format = workbook.add_format({'align': 'center','valign': 'vcenter', 'bold': True,'font_size': 14})
                data_format = workbook.add_format({'border': 1})
                filter_format = workbook.add_format({'bold': True})
                column_header_format = workbook.add_format({'bold': True, 'bg_color': '#7f9cf0', 'font_color': 'black'})
                from xlsxwriter.utility import xl_range
                merged_range = xl_range(1, 0, 1, len(column_list) - 1)  # A4 to ??4
                worksheet.merge_range(merged_range, title, header_format)
            
                # Write filters/info row
                worksheet.write(3, 0, headers, filter_format)
            
                # Column headers
                for i, column_name in enumerate(column_list):
                    worksheet.write(5, i, column_name, column_header_format)
            
                # Data rows
                for row_num, row_data in enumerate(data_list, start=6):
                    for col_num, col_data in enumerate(row_data):
                        worksheet.write(row_num, col_num, str(col_data), data_format)
            
                # Auto column width
                for col_num in range(len(column_list)):
                    col_cells = [str(row[col_num]) for row in data_list] + [column_list[col_num]]
                    max_width = max(len(cell) for cell in col_cells)
                    worksheet.set_column(col_num, col_num, max_width)
            
                workbook.close()
            
                # Response
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
                output.seek(0)
                response.write(output.read())
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log", [fun, str(e), request.user.id])
        messages.error(request, f'Oops...! Something went wrong! {str(e)}')
    finally:
        return response
    
def add_page_number(canvas, doc):
    canvas.saveState()
    page_num = canvas.getPageNumber()
    text = "Page %s" % page_num
    canvas.drawRightString(200*2.54, 1*2.54*2.54, text)
    canvas.restoreState()    

@login_required
def save_filters(request):
    try:
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id
            if request.method == "GET":
                columnName =str(request.GET.get('columnName', ''))
                filterid =str(request.GET.get('filterid', ''))
                subFilterId =str(request.GET.get('subFilterId', ''))
                sft =str(request.GET.get('sft', ''))
                entity =str(request.GET.get('entity', ''))
                saved_name =str(request.GET.get('save_filter_name', ''))
                f_count =str(request.GET.get('f_count', ''))
                filterid1 = filterid.split(',')
                SubFilterId1 = subFilterId.split(',')
                sft1 = sft.split(',')
                data = common_fun(columnName,filterid1,SubFilterId1,sft1,entity,user,'0')
                sql_query = data['sql_query'] 
                display_names = data['display_names']       
                datalist = callproc("stp_save_report_filters",[saved_name,entity,filterid,subFilterId,columnName,f_count,display_names,sql_query,user])
                response_data = {'result': datalist[0][0]}                       
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
        response_data = {'result': 'fail'}       
    finally:
        return JsonResponse(response_data,safe=False)

@login_required
def delete_filters(request):
    try:
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id
            if request.method == "GET":
                entity =str(request.GET.get('entity', ''))
                saved_id =str(request.GET.get('save_filter_name', ''))
                datalist = callproc("stp_delete_report_filters",[saved_id,entity,user])
                response_data = {'result': datalist[0][0]}                       
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
        response_data = {'result': 'fail','messages ':'something went wrong !'}       
    finally:
        return JsonResponse(response_data,safe=False)
    
@login_required
def saved_filters(request):
    try:
        if request.user.is_authenticated ==True:                
            global user
            user = request.user.id
            if request.method == "GET":
                entity =str(request.GET.get('entity', ''))
                saved_id =str(request.GET.get('saved_id', ''))
                result_data  = callproc("stp_get_saved_report_filters",[saved_id,entity,user])
                filters, sub_filters, selected_columns, f_count, display_name, sql_query = ('',) * 6  # Initialize variables
                if result_data and result_data[0]: 
                    for items in result_data: 
                        filters, sub_filters, selected_columns, f_count, display_name, sql_query = items 

                display_name_arr = display_name.split(',')
                display_name_list = list(display_name_arr)
                fil_arr = filters.split(',')
                sub_fil_arr = sub_filters.split(',')
                sel_col_arr = selected_columns.split(',')
                f_id = ''
                s_fid = ''
                if len(fil_arr) > 0:
                    f_id = fil_arr[0]
                    fil_arr = fil_arr[1:]
                if len(sub_fil_arr) > 0:
                    s_fid = sub_fil_arr[0]
                    sub_fil_arr = sub_fil_arr[1:]

                data_list= []
                result_data  = callproc("stp_get_execute_report_query", [sql_query])
                if result_data and result_data[0]:
                    for row in result_data:
                        data_list.append(list(row))

                if len(data_list) > 0:
                    emptycheck = 0
                else : emptycheck = 1

                table = render_to_string('Reports/_partial_report.html', {'emptycheck':emptycheck,'columns':display_name_list,'rows':data_list})

                context = {'result': 'success','filters':fil_arr,'sub_filters':sub_fil_arr,'sel_col_arr':sel_col_arr,
                           'sel_col':selected_columns,'f_count':f_count,'table':table,'f_id':f_id,'s_fid':s_fid}                       
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log",[fun,str(e),request.user.id])  
        messages.error(request, 'Oops...! Something went wrong!')
        context = {'result': 'fail'}       
    finally:
        return JsonResponse(context,safe=False)
 
 