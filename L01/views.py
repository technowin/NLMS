# L01/views.py
from django.http import HttpResponse
from L01.models import *
from weasyprint import HTML
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from weasyprint import CSS
from administration.models import *
from openpyxl import Workbook
from django.conf import settings
import traceback
from django.contrib import messages
import Db
import requests
from django.db.models import Count, Sum, Q, F, Avg
from django.http import HttpResponse, JsonResponse
from NLMS.encryption import *
from Account.db_utils import callproc
import re
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
import json
from django.shortcuts import render, get_object_or_404
from django.shortcuts import render, get_object_or_404
from django.conf import settings
from .models import CompetitiveExamMaster, Sections, Subjects, Topics, Chapters 
from .models import BookCatalog
from django.shortcuts import render, redirect
from django.db import transaction
import os
from django.http import HttpResponse, JsonResponse
from NLMS.encryption import *
import re
import json
from django.utils import timezone
import datetime
import uuid
from django.contrib.auth.hashers import make_password
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from django.http import Http404, FileResponse
from pathlib import Path
from django.core.files.storage import default_storage
import logging
from Masters.models import *
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
import pathlib
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from io import BytesIO
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from datetime import datetime, date, timedelta
from django.http import JsonResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404
import barcode
from barcode.writer import ImageWriter
import base64
from django.core.exceptions import ObjectDoesNotExist
logger = logging.getLogger(__name__)
from django.db import transaction as db_transaction
from PIL import Image, ImageDraw, ImageFont
from django.views import View
from barcode import Code128
from reportlab.lib.utils import ImageReader
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.db.models import Value, CharField
from django.db.models.functions import Concat
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from PIL import Image, ImageFont, ImageDraw
from django.core.paginator import Paginator
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Max, Count, Avg, Exists, OuterRef
from django.template.loader import get_template
from xhtml2pdf import pisa
import io
import csv
from django.utils.dateparse import parse_date, parse_datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from django.views.decorators.http import require_POST
from services.file_storage_service import file_storage_service
import boto3
from botocore.exceptions import ClientError
from django.http import HttpResponse, Http404, StreamingHttpResponse
import mimetypes
# Part First While Filling Membership Form

def index(request):
    # Get the library code from session
    library_code = request.session.get('library_db', None)

    if library_code:
        library_details = LibraryMaster.objects.using('default').filter(
            is_active=1, 
            library_code=library_code
        )

        for lilo in library_details:

            encrypted_library_code = enc(lilo.library_code)
            lilo.libraries = encrypted_library_code
            
            image_urls = []

            if lilo.image_url:
                image_paths = [p.strip() for p in lilo.image_url.split(",") if p.strip()]
                for path in image_paths:
                    image_urls.append(file_storage_service.get_file_url(path))

            # ✅ attach list to object
            lilo.image_urls = image_urls

            # ✅ first image (for big preview)
            lilo.main_image = image_urls[0] if image_urls else ""

            # -------------------------------------------
            # FETCH RECENT 5 BOOKS
            # -------------------------------------------
            books = BookCatalog.objects.filter(status_id=1).order_by('-cat_ref_num')[:5]

            lilo.books = [
                {
                    "title": book.title or "Untitled",
                    "subtitle": book.subtitle or "",
                    "author": book.author or "Unknown",
                    "publisher": book.publisher or "",
                    "isbn_issn": book.isbn_issn or "",
                    "edition": book.edition or "",
                    "publication_place": book.publication_place or "",
                    "year_of_publication": book.year_of_publication or "N/A",
                    "pages": book.pages or "",
                    "language": book.language or "Unknown",
                    "front_page_photo": file_storage_service.get_file_url(book.front_page_photo) if book.front_page_photo else "",
                    "last_page_photo": file_storage_service.get_file_url(book.last_page_photo) if book.last_page_photo else "",
                    "remarks": book.remarks or "",
                    "description": book.remarks or "No description available.",
                    "ebook_available": book.ebook_available or "No",   # ⭐ ADDED LINE
                    "encrypted_cat_ref_num": enc(str(book.cat_ref_num)),
                }
                for book in books
            ]

            # -------------------------------------------
            # FETCH RECENT 5 EBOOKS
            # -------------------------------------------
            ebooks = LibraryEbook.objects.filter(eb_status_id=1).order_by('-ebook_id')[:5]

            lilo.ebooks = [
                {
                    "eb_title": ebook.eb_title or "Untitled",
                    "eb_subtitle": ebook.eb_subtitle or "",
                    "eb_author": ebook.eb_author or "Unknown",
                    "eb_publisher": ebook.eb_publisher or "",
                    "eb_isbn_issn": ebook.eb_isbn_issn or "",
                    "eb_edition": ebook.eb_edition or "",
                    "eb_publication_place": ebook.eb_publication_place or "",
                    "eb_year_of_publication": ebook.eb_year_of_publication or "N/A",
                    "eb_pages": ebook.eb_pages or "",
                    "eb_language": ebook.eb_language or "Unknown",
                    "eb_front_page_photo": file_storage_service.get_file_url(ebook.eb_front_page_photo) if ebook.eb_front_page_photo else "",
                    "eb_last_page_photo": file_storage_service.get_file_url(ebook.eb_last_page_photo) if ebook.eb_last_page_photo else "",
                    "remarks": ebook.remarks or "",
                    "description": ebook.eb_keywords or "No description available.",
                    "eb_pdf_url": ebook.eb_pdf_url or "",
                    "encrypted_ebook_id": enc(str(ebook.ebook_id)),
                    "encrypted_pdf_url": enc(ebook.eb_pdf_url or ""),
                }
                for ebook in ebooks
            ]

        library = library_details.first()
        library_name = library_details.first().library_name if library_details.exists() else ""
        # library_name =(library.library_name_mar if library and library.library_name_mar else library.library_name if library else "")
        library_name_mar = library_details.first().library_name_mar if library_details.exists() else ""

    else:
        # -------------------------------------------
        # WHEN NO LIBRARY IS SELECTED
        # -------------------------------------------
        library_details = LibraryMaster.objects.using('default').filter(is_active=1)

        for lilo in library_details:

            # --- Books fetch ---
            books = BookCatalog.objects.filter(status_id=1).order_by('-cat_ref_num')[:5]

            lilo.books = [
                {
                    "title": book.title or "Untitled",
                    "subtitle": book.subtitle or "",
                    "author": book.author or "Unknown",
                    "publisher": book.publisher or "",
                    "isbn_issn": book.isbn_issn or "",
                    "edition": book.edition or "",
                    "publication_place": book.publication_place or "",
                    "year_of_publication": book.year_of_publication or "N/A",
                    "pages": book.pages or "",
                    "language": book.language or "Unknown",
                    "front_page_photo":  file_storage_service.get_file_url(book.front_page_photo) if book.front_page_photo else "",
                    "last_page_photo":  file_storage_service.get_file_url(book.last_page_photo) if book.last_page_photo else "",
                    "remarks": book.remarks or "",
                    "description": book.remarks or "No description available.",
                    "ebook_available": book.ebook_available or "No",   # ⭐ ADDED LINE
                    "encrypted_cat_ref_num": enc(str(book.cat_ref_num)),
                }
                for book in books
            ]

            # --- eBooks fetch ---
            ebooks = LibraryEbook.objects.filter(eb_status_id=1).order_by('-ebook_id')[:5]

            lilo.ebooks = [
                {
                    "eb_title": ebook.eb_title or "Untitled",
                    "eb_subtitle": ebook.eb_subtitle or "",
                    "eb_author": ebook.eb_author or "Unknown",
                    "eb_publisher": ebook.eb_publisher or "",
                    "eb_isbn_issn": ebook.eb_isbn_issn or "",
                    "eb_edition": ebook.eb_edition or "",
                    "eb_publication_place": ebook.eb_publication_place or "",
                    "eb_year_of_publication": ebook.eb_year_of_publication or "N/A",
                    "eb_pages": ebook.eb_pages or "",
                    "eb_language": ebook.eb_language or "Unknown",
                    "eb_front_page_photo":  file_storage_service.get_file_url(ebook.eb_front_page_photo) if ebook.eb_front_page_photo else "",
                    "eb_last_page_photo":  file_storage_service.get_file_url(ebook.eb_last_page_photo) if ebook.eb_last_page_photo else "",
                    "remarks": ebook.remarks or "",
                    "description": ebook.eb_keywords or "No description available.",
                    "eb_pdf_url": ebook.eb_pdf_url or "",
                    "encrypted_ebook_id": enc(str(ebook.ebook_id)),
                    "encrypted_pdf_url": enc(ebook.eb_pdf_url or ""),
                }
                for ebook in ebooks
            ]

        library_name = ""

    return render(request, "L01/index.html", {
        'libraries': library_details,
        'library_name': library_name,
        'library_name_mar':library_name_mar,
        'MEDIA_URL': settings.MEDIA_URL
    })
    
def check_user_id(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
       user_id = request.GET.get("user_id", "").strip()
       exists = MembershipDetails.objects.filter(user_id=user_id).exists()
       return JsonResponse({"exists": exists})
    except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name if tb else 'library_list'
            cursor.callproc("stp_error_log", [fun, str(e), ''])
            print(f"error: {e}")
            messages.error(request, 'Oops...! Something went wrong!')

def check_aadhar(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        aadhar_no = request.GET.get("aadhar_number", "").strip()

        # Aadhar number format validation (12 digits)
        if not re.match(r'^\d{12}$', aadhar_no):
            return JsonResponse({"exists": False, "error": "Invalid Aadhar format (must be 12 digits)."})

        # Check if Aadhar already exists in the database
        exists = MembershipDetails.objects.filter(aadhar_no=aadhar_no).exists()

        return JsonResponse({"exists": exists})

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else 'check_aadhar'
        cursor.callproc("stp_error_log", [fun, str(e), ''])
        print(f"error: {e}")
        return JsonResponse({"error": "Oops... Something went wrong!"}, status=500)

def get_pincodes(request):
    wardname = request.GET.get("wardname")
    pincodes = []
    if wardname:
        try:
            # ward = WardMaster.objects.using('default').get(ward_name=wardname, is_active=1)
            ward = WardMaster.objects.get(ward_name=wardname, is_active=1)
            if ward.pincode:
                pincodes = ward.pincode.split(",")  # split comma-separated
                pincodes = [p.strip() for p in pincodes]  # remove extra spaces
        except WardMaster.DoesNotExist:
            pass
    return JsonResponse({"pincodes": pincodes})

# Notepad++ 485
def get_membership_details(request):
    if request.method == "GET":
        enc_id = request.GET.get("id")
        membership_detail_id = request.GET.get("membership_detail_id", "")
        try:
            membership_id = enc_id  # decrypt if needed

            membership = MembershipMaster.objects.get(
                id=membership_id,
                isactive=1
            )

            data = {
                "deposit": str(membership.deposit) if hasattr(membership, 'deposit') and membership.deposit is not None else "0",
                "entry_fees": str(membership.entry_fees) if hasattr(membership, 'entry_fees') and membership.entry_fees is not None else "0",
                "subscription_fees": str(membership.subscription_fees) if hasattr(membership, 'subscription_fees') and membership.subscription_fees is not None else "0",
                "fine_membership": str(membership.fine_membership) if hasattr(membership, 'fine_membership') and membership.fine_membership is not None else "0",
            }
            
            data1 = None
            
            if membership_detail_id:
                membershipDetails = MembershipDetails.objects.filter(
                    id=membership_detail_id,
                    isactive=1
                ).first()
                
                if membershipDetails:
                    data1 = {
                        "from_date": membershipDetails.from_date.strftime("%Y-%m-%d") if membershipDetails.from_date else None,
                        "membership_duration": membershipDetails.membership_duration,
                        "to_date": membershipDetails.to_date.strftime("%Y-%m-%d") if membershipDetails.to_date else None,
                        
                        # Note: MembershipDetails has 'subscription' field, MembershipMaster has 'subscription_fees'
                        "deposit": str(membershipDetails.deposit) if membershipDetails.deposit is not None else str(membership.deposit) if hasattr(membership, 'deposit') and membership.deposit is not None else "0",  
                        "entry_fees": str(membershipDetails.entry_fees) if membershipDetails.entry_fees is not None else str(membership.entry_fees) if hasattr(membership, 'entry_fees') and membership.entry_fees is not None else "0",
                        "subscription": str(membershipDetails.subscription) if membershipDetails.subscription is not None else str(membership.subscription_fees) if hasattr(membership, 'subscription_fees') and membership.subscription_fees is not None else "0",
                    }
                else:
                    data1 = None

            return JsonResponse({
                "success": True,
                "data": data,
                "data1": data1
            })

        except MembershipMaster.DoesNotExist:
            return JsonResponse({
                "success": False,
                "error": "Membership not found"
            })
        except Exception as e: 
            import traceback
            print(f"Error in get_membership_details: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({
                "success": False,
                "error": str(e)
            })
            
def registration(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        library_code = request.session.get('library_db', None)

        if request.method == "GET":
            
            formfilledsuccessfullyVariable = request.GET.get("formfilledsuccessfully", None)
            
            # you will get throgh post function formfilledsuccessfullyVariable encrypted value
            # decrypt it here
            if formfilledsuccessfullyVariable != None:
                
                formfilledsuccessfullyDecrypted = dec(formfilledsuccessfullyVariable)
                
                template = "L01/FormFilledSuccessfully.html"
                context = {'MEDIA_URL': settings.MEDIA_URL,
                           "library_code": request.session.get("library_db"),
                        }
                
            else:
                if library_code:
                    library_details = tbl_librarymasterL01.objects.filter(is_active=1, library_code=library_code)
                    for lilo in library_details:
                        encrypted_library_code = enc(lilo.library_code)
                        lilo.libraries = encrypted_library_code
                    print(library_details)
                
                    membership_options = parameter_master_L01.objects.filter(isactive=1,parameter_name='MembershipForm').order_by('parameter_id')
                    
                    # ward_details = WardMaster.objects.using('default').filter(is_active=1)
                    ward_details = WardMaster.objects.filter(is_active=1)
                    
                    membership_Master = MembershipMaster.objects.filter(isactive=1)
                    if membership_Master.exists():
                        for mema in membership_Master:
                            encrypted_membership_id = enc(str(mema.id))
                            mema.membership_id_enc = encrypted_membership_id
                        print(membership_Master)

                template = "L01/registration.html"
                context = {
                    'MEDIA_URL': settings.MEDIA_URL,
                    'library_details': library_details,
                    'membership_options': membership_options,
                    'ward_details': ward_details,
                    'membership_Master': membership_Master,
                }
                
            return render(request, template, context)

        elif request.method == "POST":
            try:
                with transaction.atomic():
                    
                    membertype = request.POST.get("membertype")
                    
                    user_id = request.POST.get("user_id") or None

                    # Default education field
                    education_value = request.POST.get("education") or None

                    # Override if membertype == 5
                    if membertype == "5":
                        education_value = request.POST.get("practitioner_details") or None
                    # === Collect Membership Data ===
                    data = {
                        "user_id": request.POST.get("user_id") or None,
                        "aadhar_no": request.POST.get("aadhar_no") or None,
                        "first_name": request.POST.get("first_name") or None,
                        "middle_name": request.POST.get("middle_name") or None,
                        "last_name": request.POST.get("last_name") or None,
                        "first_name_mar": request.POST.get("first_name_mar") or None,
                        "middle_name_mar": request.POST.get("middle_name_mar") or None,
                        "last_name_mar": request.POST.get("last_name_mar") or None,
                        "ward": request.POST.get("ward_name") or None,
                        "other_ward": request.POST.get("custom_ward_name") if request.POST.get("ward_name") == "Other" else None,
                        "pincode": (request.POST.get("custom_pincode") if request.POST.get("ward_name") == "Other" else request.POST.get("pincode")) or None,
                        "library_name": request.POST.get("library_name") or None,
                        "library_name_mar": request.POST.get("library_name_mar") or None,
                        "local_address": request.POST.get("local_address") or None,
                        "mobile_no": request.POST.get("mobile_number") or None,
                        "occupation": request.POST.get("occupation_details") or None,
                        "office_phone": request.POST.get("phone_number") or None,
                        "education": education_value,  # ✅ dynamic based on membertype
                        "institute_name": request.POST.get("institution_name") or None,
                        "recommender_details": request.POST.get("referrer_details") or None,
                        "member_type_id": membertype or None,
                        "is_resident_of_nmmc": 1 if request.POST.get("is_nmmc") == "yes" else 0,
                        "address_same_as_aadhar": 1 if request.POST.get("same_aadhar") == "yes" else 0,
                        "membership_duration": request.POST.get("months") or None,
                        "from_date": request.POST.get("fromDate") or None,
                        "to_date": request.POST.get("toDate") or None,
                        "deposit": request.POST.get("deposit") or None,
                        "entry_fees": request.POST.get("entry_fees") or None,
                        "subscription": request.POST.get("subscription") or None,
                        "email": request.POST.get("email") or None,
                        "dob": request.POST.get("dob") or None,
                        "membership_id": request.POST.get("membershiptype") or None,
                        "created_by": user_id,
                        "created_at": timezone.now(),
                        "status_id": 1,   # pending
                        "isactive": 1,
                        "actionperformed": f"Applied for membership",
                        "reviewed": None,
                        "reviewed_at": None,
                    }
                    
                    raw_password = request.POST.get("repeat_password")
                    username  = request.POST.get("user_id")

                    print("===== Membership Form Data =====")
                    for k, v in data.items():
                        print(f"{k}: {v}")
                    print(f"Password (future use): {raw_password}")

                    # === Save Membership ===
                    membership = MembershipDetails.objects.create(**data)
                    
                    # === Handle Document Uploads ===
                    library_code = request.session.get("library_db", "default")
                    documents_info = []

                    for field_name, doc_id in [
                        ("photo_upload", request.POST.get("document_photo_upload")),
                        ("id_upload", request.POST.get("document_id_upload")),
                        ("agreement_copy", request.POST.get("document_agreement_copy")),
                        ("nagarsevak_letter", request.POST.get("document_nagarsevak_letter")),
                        ("employee_letter", request.POST.get("employee_letter")),
                    ]:
                        file = request.FILES.get(field_name)
                        if file and doc_id:
                            # Extract original filename + extension
                            filename, ext = os.path.splitext(file.name)

                            # Generate unique filename
                            timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
                            short_uuid = str(uuid.uuid4())[:8]
                            unique_filename = f"{filename}_{timestamp}_{short_uuid}{ext}"

                            # Build save path -> library_code/Document - {doc_id}/unique_filename
                            save_dir = os.path.join(library_code, username, f"Document - {doc_id}")
                            save_path = os.path.join(save_dir, unique_filename)
                            
                            # ✅ USE STORAGE SERVICE TO SAVE FILE
                            saved_file_path = file_storage_service.save_file(file, save_path)

                            # ✅ Save document record
                            DocumentDetails.objects.create(
                                membership=membership,
                                document_id=doc_id,
                                file_name=unique_filename,
                                file_path=saved_file_path,  # This will be the path
                                created_by=request.POST.get("user_id"),
                                created_at=timezone.now()
                            )

                            documents_info.append({
                                "document_id": doc_id,
                                "file_name": unique_filename,
                                "file_path": saved_file_path,
                                "file_url": file_storage_service.get_file_url(saved_file_path),  # Get URL
                            })

                    print("\n===== Documents Uploaded =====")
                    for d in documents_info:
                        print(d)
                        
                    if CustomUser.objects.filter(username=data.get("user_id")).exists():
                        raise ValueError(f"❌ User ID '{data.get('user_id')}' already exists!")

                    # === Create CustomUser ===

                    custom_user = CustomUser.objects.create(
                        first_name=data.get("first_name"),
                        last_name=data.get("last_name"),
                        full_name=f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
                        email=data.get("email"),
                        phone=data.get("mobile_no"),
                        role_id=3,  # ✅ Hardcoded
                        address=data.get("local_address"),
                        govt_id=data.get("aadhar_no"),
                        is_active=False,
                        is_staff=True,
                        user_type="member",
                        username=data.get("user_id"),
                        password=make_password(raw_password),  # ✅ Hash password
                        first_time_login=1,
                    )

                    # === Save Raw Password in password_storage ===
                    password_storage.objects.create(
                        user=custom_user,
                        passwordText=raw_password
                    )
                    
                    formfilledsuccessfullyVariable = enc('1')

                    # return redirect('/registration/?formfilledsuccessfully=1')
                    return redirect(f'/{library_code}/registration?formfilledsuccessfully={formfilledsuccessfullyVariable}')

            except Exception as e:
                tb = traceback.extract_tb(e.__traceback__)
                fun = tb[0].name if tb else "registration"
                cursor.callproc("stp_error_log", [fun, str(e), request.session.get("library_db", "default_lib")])
                messages.error(request, f"❌ Error while saving: {str(e)}")
                return HttpResponse(f"Error: {str(e)}", status=500)
                
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else 'library_list'
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, 'Oops...! Something went wrong!')

# Part Second After Filling Membership Form Approval by Admin

from collections import defaultdict

def membership_approval(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    library_code = request.session.get('library_db', None)

    # Main 5 Tabs (Static but grouping uses DB statuses)
    TAB_GROUPS = {
        "Application Submitted": ["APP_SUB", "APPROVED"],
        "Payment Offline": ["PAY_OFFLINE", "PAY_SUCCESS", "PAY_PENDING", "PAY_PROGRESS", "PAY_FAIL"],
        "Renewal Submitted": ["APP_RENEW", "APP_RENEW_REJECT"],
        "Application Cancelled": ["APP_CANCEL_REQ", "APP_CANCEL"],
        "Rejected": ["REJECT"],
    }

    try:
        # Load all membership records
        memberships = (
            MembershipDetails.objects
            .select_related("status")
            .all()
            .order_by("-created_at")  # sort by created_at
        )

        # DEBUG: Print total records and their statuses
        print(f"Total memberships: {memberships.count()}")
        for mem in memberships:
            print(f"ID: {mem.id}, Name: {mem.first_name}, Status: {mem.status.status_code}")
        
        # Encrypt IDs
        for mem in memberships:
            mem.membership_id_enc = enc(str(mem.id))

        # Group by status_code
        memberships_by_status = defaultdict(list)
        for mem in memberships:
            memberships_by_status[mem.status.status_code].append(mem)

        # DEBUG: Print grouped data
        print("\nGrouped by status:")
        for status_code, mem_list in memberships_by_status.items():
            print(f"Status {status_code}: {len(mem_list)} records")

        # ---------- Prepare final tab data ----------
        tabs = []

        for tab_title, status_codes in TAB_GROUPS.items():
            # 1. Combine all records for all status codes
            combined_members = []
            for code in status_codes:
                combined_members += memberships_by_status.get(code, [])

            # 2. Sort inside group using created_at desc
            combined_members = sorted(combined_members, key=lambda x: x.created_at, reverse=True)

            # 3. Append final structured tab
            tabs.append({
                "tab_title": tab_title,
                "status_codes": status_codes,  # Add this for debugging
                "memberships": combined_members,
            })
            
            # DEBUG: Print tab info
            print(f"\nTab: {tab_title}")
            print(f"  Status codes: {status_codes}")
            print(f"  Total records: {len(combined_members)}")
            for mem in combined_members[:3]:  # Show first 3
                print(f"  - {mem.first_name} {mem.last_name}: {mem.status.status_code}")

        # Get library name
        library = tbl_librarymasterL01.objects.filter(
            library_code=library_code, is_active=1
        ).first()
        library_name_mar = library.library_name_mar if library else ""

        return render(request, "L01/membership_approval/membership_approval.html", {
            "tabs": tabs,
            "library_name_mar": library_name_mar,
            "MEDIA_URL": settings.MEDIA_URL,
        })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "membership_approval"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        messages.error(request, "Oops...! Something went wrong!")
        return render(request, "L01/membership_approval/membership_approval.html", {})

@login_required
def membership_form_create(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        library_code = request.session.get('library_db', None)
        user_code = request.session["user_id"]

        if request.method == "GET":
            
            if library_code:
                    library_details = tbl_librarymasterL01.objects.filter(is_active=1, library_code=library_code)
                    for lilo in library_details:
                        encrypted_library_code = enc(lilo.library_code)
                        lilo.libraries = encrypted_library_code
                    print(library_details)
                
                    membership_options = parameter_master_L01.objects.filter(isactive=1,parameter_name='MembershipForm').order_by('parameter_id')
                    
                    # ward_details = WardMaster.objects.using('default').filter(is_active=1)
                    ward_details = WardMaster.objects.filter(is_active=1)
                    
                    membership_Master = MembershipMaster.objects.filter(isactive=1)
                    if membership_Master.exists():
                        for mema in membership_Master:
                            encrypted_membership_id = enc(str(mema.id))
                            mema.membership_id_enc = encrypted_membership_id
                        print(membership_Master)
                        
            template = "L01/membership_approval/membership_form_create.html"
            context = {
                'MEDIA_URL': settings.MEDIA_URL,
                'library_details': library_details,
                'membership_options': membership_options,
                'ward_details': ward_details,
                'membership_Master': membership_Master,
                'library_code': library_code,
            }
                
            return render(request, template, context)

        elif request.method == "POST":
            try:
                with transaction.atomic():
                    
                    membertype = request.POST.get("membertype")

                    # Default education field
                    education_value = request.POST.get("education") or None

                    # Override if membertype == 5
                    if membertype == "5":
                        education_value = request.POST.get("practitioner_details") or None
                        
                    # === Collect Membership Data ===
                    data = {
                        "user_id": request.POST.get("user_id") or None,
                        "aadhar_no": request.POST.get("aadhar_no") or None,
                        "first_name": request.POST.get("first_name") or None,
                        "middle_name": request.POST.get("middle_name") or None,
                        "last_name": request.POST.get("last_name") or None,
                        "first_name_mar": request.POST.get("first_name_mar") or None,
                        "middle_name_mar": request.POST.get("middle_name_mar") or None,
                        "last_name_mar": request.POST.get("last_name_mar") or None,
                        "ward": request.POST.get("ward_name") or None,
                        "other_ward": request.POST.get("custom_ward_name") if request.POST.get("ward_name") == "Other" else None,
                        "pincode": (request.POST.get("custom_pincode") if request.POST.get("ward_name") == "Other" else request.POST.get("pincode")) or None,
                        "library_name": request.POST.get("library_name") or None,
                        "library_name_mar": request.POST.get("library_name_mar") or None,
                        "local_address": request.POST.get("local_address") or None,
                        "mobile_no": request.POST.get("mobile_number") or None,
                        "occupation": request.POST.get("occupation_details") or None,
                        "office_phone": request.POST.get("phone_number") or None,
                        "education": education_value,  # ✅ dynamic based on membertype
                        "institute_name": request.POST.get("institution_name") or None,
                        "recommender_details": request.POST.get("referrer_details") or None,
                        "member_type_id": membertype or None,
                        "is_resident_of_nmmc": 1 if request.POST.get("is_nmmc") == "yes" else 0,
                        "address_same_as_aadhar": 1 if request.POST.get("same_aadhar") == "yes" else 0,
                        "membership_duration": request.POST.get("months") or None,
                        "from_date": request.POST.get("fromDate") or None,
                        "to_date": request.POST.get("toDate") or None,
                        "deposit": request.POST.get("deposit") or None,
                        "entry_fees": request.POST.get("entry_fees") or None,
                        "subscription": request.POST.get("subscription") or None,
                        "email": request.POST.get("email") or None,
                        "dob": request.POST.get("dob") or None,
                        "membership_id": request.POST.get("membershiptype") or None,
                        "created_by": request.POST.get("user_id"),
                        "created_at": timezone.now(),
                        "status_id": 1,   # pending
                        "isactive": 1,
                    }
                    
                    raw_password = request.POST.get("repeat_password")
                    username  = request.POST.get("user_id")

                    print("===== Membership Form Data =====")
                    for k, v in data.items():
                        print(f"{k}: {v}")
                    print(f"Password (future use): {raw_password}")
                    
                    # === Save Membership ===
                    membership = MembershipDetails.objects.create(**data)

                    # === Handle Document Uploads using FileStorageService ===
                    documents_info = []
                    
                    # Map form fields to document IDs
                    document_mapping = [
                        ("photo_upload", request.POST.get("document_photo_upload")),
                        ("id_upload", request.POST.get("document_id_upload")),
                        ("agreement_copy", request.POST.get("document_agreement_copy")),
                        ("nagarsevak_letter", request.POST.get("document_nagarsevak_letter")),
                        ("employee_letter", request.POST.get("employee_letter")),
                    ]
                    
                    for field_name, doc_id in document_mapping:
                        file = request.FILES.get(field_name)
                        if file and doc_id:
                            try:
                                # Extract original filename + extension
                                filename, ext = os.path.splitext(file.name)
                                
                                # Generate unique filename
                                timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
                                short_uuid = str(uuid.uuid4())[:8]
                                unique_filename = f"{filename}_{timestamp}_{short_uuid}{ext}"
                                
                                # Build save path -> library_code/username/Document - {doc_id}/unique_filename
                                # ✅ NORMALIZE PATH: Use forward slashes for consistency
                                save_dir = f"{library_code}/{username}/Document - {doc_id}"
                                save_path = f"{save_dir}/{unique_filename}"
                                
                                # ✅ USE STORAGE SERVICE TO SAVE FILE
                                saved_file_path = file_storage_service.save_file(file, save_path)
                                
                                # ✅ Save document record
                                document = DocumentDetails.objects.create(
                                    membership=membership,
                                    document_id=doc_id,
                                    file_name=unique_filename,
                                    file_path=saved_file_path,  # Normalized path from storage service
                                    created_by=user_code,
                                    created_at=timezone.now()
                                )
                                
                                documents_info.append({
                                    "document_id": doc_id,
                                    "file_name": unique_filename,
                                    "file_path": saved_file_path,
                                    "file_url": file_storage_service.get_file_url(saved_file_path),  # Get URL for reference
                                    "status": "Success"
                                })
                                
                                print(f"✅ Document '{field_name}' saved: {saved_file_path}")
                                
                            except Exception as e:
                                print(f"❌ Error saving document '{field_name}': {str(e)}")
                                documents_info.append({
                                    "document_id": doc_id,
                                    "file_name": file.name,
                                    "error": str(e),
                                    "status": "Failed"
                                })
                    
                    print("\n===== Documents Upload Summary =====")
                    for d in documents_info:
                        print(d)
                    

                    if CustomUser.objects.filter(username=data.get("user_id")).exists():
                        messages.error(request, f"❌ User ID '{data.get('user_id')}' आधी अस्तित्वात आहे!")
                        return redirect('L01:membership_form_create')  # go back to the form

                    # === Create CustomUser ===

                    custom_user = CustomUser.objects.create(
                        first_name=data.get("first_name"),
                        last_name=data.get("last_name"),
                        full_name=f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
                        email=data.get("email"),
                        phone=data.get("mobile_no"),
                        role_id=3,  # ✅ Hardcoded
                        address=data.get("local_address"),
                        govt_id=data.get("aadhar_no"),
                        is_active=False,
                        is_staff=True,
                        user_type="member",
                        username=data.get("user_id"),
                        password=make_password(raw_password),  # ✅ Hash password
                        first_time_login=1,
                    )

                    # === Save Raw Password in password_storage ===
                    password_storage.objects.create(
                        user=custom_user,
                        passwordText=raw_password
                    )
                    
                    messages.success(request, "✅ नवीन सदस्य यशस्वीरित्या तयार झाला!")
                    return redirect('L01:membership_approval')

            except Exception as e:
                tb = traceback.extract_tb(e.__traceback__)
                fun = tb[0].name if tb else "registration"
                cursor.callproc("stp_error_log", [fun, str(e), request.session.get("library_db", "default_lib")])
                messages.error(request, f"❌ Error while saving: {str(e)}")
                return HttpResponse(f"Error: {str(e)}", status=500)
          
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "library_list"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, "Oops...! Something went wrong!")
        return render(request, "L01/membership_approval/membership_approval.html", {})

@login_required
def membership_form_edit(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        library_code = request.session.get('library_db', None)
        user_code = request.session["user_id"]

        if request.method == "GET":
            
            if library_code:
                
                    membershipid = dec(request.GET.get('membershipid'))
                
                    library_details = tbl_librarymasterL01.objects.filter(is_active=1, library_code=library_code)
                    for lilo in library_details:
                        encrypted_library_code = enc(lilo.library_code)
                        lilo.libraries = encrypted_library_code
                    print(library_details)
                
                    membership_options = parameter_master_L01.objects.filter(isactive=1,parameter_name='MembershipForm').order_by('parameter_id')
                    
                    # ward_details = WardMaster.objects.using('default').filter(is_active=1)
                    ward_details = WardMaster.objects.filter(is_active=1)
                    
                    membership_Master = MembershipMaster.objects.filter(isactive=1)
                    if membership_Master.exists():
                        for mema in membership_Master:
                            encrypted_membership_id = enc(str(mema.id))
                            mema.membership_id_enc = encrypted_membership_id
                        print(membership_Master)
                        
                    membership = get_object_or_404(MembershipDetails, id=membershipid, isactive=1)    
                    
                    document_details = DocumentDetails.objects.filter(
                        membership=membership, isactive=1
                    ).select_related("document")

                    documents_map = {}
                    for doc in document_details:
                        doc.doc_id_enc = enc(str(doc.id))  # encrypt document ID
                        documents_map[doc.document.id] = doc
                        
            template = "L01/membership_approval/membership_form_edit.html"
            context = {
                'MEDIA_URL': settings.MEDIA_URL,
                'library_details': library_details,
                'membership_options': membership_options,
                'ward_details': ward_details,
                'membership_Master': membership_Master,
                'library_code': library_code,
                'membership': membership,
                'documents_map': documents_map,
            }
                
            return render(request, template, context)
        
        # notepad++ 352 
        
        if request.method == "POST":
            with transaction.atomic():
                membership_id = request.POST.get("membership_id")
                user_id = request.POST.get("user_id")
                
                membertype = request.POST.get("membertype")

                # Default: get education value
                education_value = request.POST.get("education") or None

                # Override education if membertype == 5
                if membertype == "5":
                    education_value = request.POST.get("practitioner_details") or None
                    
                membership = get_object_or_404(MembershipDetails, id=membership_id)

                updated = False  # flag to track changes

                # Mapping of form fields -> model fields
                field_map = {
                    "first_name": "first_name",
                    "first_name_mar": "first_name_mar",
                    "middle_name": "middle_name",
                    "middle_name_mar": "middle_name_mar",
                    "last_name": "last_name",
                    "last_name_mar": "last_name_mar",
                    "occupation_details": "occupation",
                    "phone_number": "office_phone",
                    "education": "education",
                    "institution_name": "institute_name",
                    "referrer_details": "recommender_details",
                    "is_nmmc": "is_resident_of_nmmc",
                    "same_aadhar": "address_same_as_aadhar",
                    "local_address": "local_address",
                    "membershiptype": "membership_id",  # FK to MembershipMaster
                    "membertype": "member_type_id",     # FK to parameter_master_L01
                    "months": "membership_duration",
                    "fromDate": "from_date",
                    "toDate": "to_date",  
                    "dob": "dob",
                }
                
                # --- Handle Ward & Pincode (including "Other") ---
                ward_name = request.POST.get("ward_name")
                custom_ward = request.POST.get("custom_ward_name")
                custom_pincode = request.POST.get("custom_pincode")
                dropdown_pincode = request.POST.get("pincode")

                membership.ward = ward_name or None
                membership.other_ward = custom_ward if ward_name == "Other" else None
                membership.pincode = (custom_pincode if ward_name == "Other" else dropdown_pincode) or None
                updated = True  # mark as updated if any ward-related change
                
                date_fields = ["dob", "from_date", "to_date"]

                for form_field, model_field in field_map.items():
                    if form_field == "education":
                        new_value = education_value
                    else:
                        new_value = request.POST.get(form_field)
                        
                    old_value = getattr(membership, model_field, None)

                    # --- Boolean-like fields ---
                    if model_field in ["is_resident_of_nmmc", "address_same_as_aadhar"]:
                        if new_value is not None:
                            new_value = 1 if str(new_value).lower() in ["1", "yes", "true", "on"] else 0
                        else:
                            new_value = None

                    # --- Integer fields ---
                    elif model_field == "membership_duration":
                        new_value = int(new_value) if new_value else None

                    # --- ForeignKey fields ---
                    elif model_field in ["membership_id", "member_type_id"]:
                        new_value = int(new_value) if new_value else None

                    # --- Date fields ---
                    elif model_field in date_fields:
                        if new_value:
                            try:
                                # Ensure we only keep date part
                                new_value = datetime.strptime(new_value.split(" ")[0], "%Y-%m-%d").date()
                            except Exception as e:
                                print("❌ Date parse error for", form_field, ":", new_value, "->", e)
                                new_value = None
                        else:
                            new_value = None

                    # --- Default ---
                    else:
                        new_value = new_value.strip() if isinstance(new_value, str) else new_value

                    if new_value != old_value:
                        setattr(membership, model_field, new_value)
                        updated = True
                        
                # --- Handle monetary fields (FIXED VERSION) ---
                money_fields = {
                    "deposit": "deposit",
                    "entry_fees": "entry_fees",
                    "subscription": "subscription"
                }

                for field_name, model_field in money_fields.items():
                    if field_name in request.POST:
                        raw_value = request.POST.get(field_name, '').strip()
                        
                        # ✅ ADD THIS: Skip if empty
                        if not raw_value:  # This catches '', None, empty string
                            continue
                        
                        try:
                            new_value = float(raw_value)
                            old_value = getattr(membership, model_field, 0.0)
                            
                            if abs(new_value - old_value) > 0.001:
                                setattr(membership, model_field, new_value)
                                updated = True
                        except ValueError:
                            pass  # Skip invalid values

                # Handle file uploads
                file_fields = {
                    "photo_upload": 1,
                    "id_upload": 2,
                    "employee_letter": 3,
                    "nagarsevak_letter": 4,
                    "agreement_copy": 5
                }

                for field_name, doc_type in file_fields.items():
                    uploaded_file = request.FILES.get(field_name)
                    if uploaded_file:
                        print(f"\n=== Processing file upload: {field_name} ===")
                        
                        # Get existing document
                        old_doc = DocumentDetails.objects.filter(
                            membership=membership, 
                            document_id=doc_type,
                            isactive=1
                        ).first()

                        # Delete old file if exists
                        if old_doc and old_doc.file_path:
                            print(f"Deleting old document: {old_doc.file_path}")
                            try:
                                file_storage_service.delete_file(old_doc.file_path)
                            except Exception as e:
                                print(f"Error deleting old file: {e}")
                            
                            # Soft delete the old document record
                            old_doc.isactive = 0
                            old_doc.updated_by = user_id
                            old_doc.save()
                            print(f"Soft deleted old document record")

                        # Generate unique filename
                        timestamp = timezone.now().strftime("%Y%m%dT%H%M%S")
                        short_uuid = str(uuid.uuid4())[:8]
                        filename, ext = os.path.splitext(uploaded_file.name)
                        unique_filename = f"{filename}_{timestamp}_{short_uuid}{ext}"

                        # Build save path -> library_code/username/Document - {doc_type}/
                        username = membership.user_id  # or get username from membership
                        save_dir = f"{library_code}/{username}/Document - {doc_type}"
                        save_path = os.path.join(save_dir, unique_filename)
                        
                        print(f"New file path: {save_path}")
                        print(f"Environment: {file_storage_service.environment}")
                        
                        # ✅ USE FILE STORAGE SERVICE TO SAVE FILE
                        saved_file_path = file_storage_service.save_file(uploaded_file, save_path)
                        print(f"Saved file path returned: {saved_file_path}")

                        # ✅ Save document record with normalized path
                        DocumentDetails.objects.create(
                            membership=membership,
                            document_id=doc_type,
                            file_name=unique_filename,
                            file_path=saved_file_path,  # Already normalized by service
                            created_by=user_id,
                            created_at=timezone.now()
                        )
                        
                        print(f"Created new document record")
                        updated = True

                if updated:
                    membership.updated_by = user_id
                    membership.save()
                    messages.success(request, "Membership updated successfully!")
                else:
                    messages.info(request, "No changes detected. Membership not updated.")

            return redirect("L01:membership_approval")
         
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "library_list"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, "Oops...! Something went wrong!")
        return redirect("L01:membership_approval")

def secure_document_view(request, doc_id_enc):
    """
    Stream documents through Django (hides S3 URLs from users)
    """
    try:
        # Decrypt the document ID
        doc_id = dec(doc_id_enc)
        document = get_object_or_404(DocumentDetails, id=doc_id, isactive=1)
        
        environment = getattr(settings, 'ENVIRONMENT', 'local')
        
        if environment == 'production':
            # ========== PRODUCTION: Stream from S3 through Django ==========
            # Get AWS credentials from settings
            aws_access_key = getattr(settings, 'AWS_ACCESS_KEY_ID', '')
            aws_secret_key = getattr(settings, 'AWS_SECRET_ACCESS_KEY', '')
            bucket_name = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', '')
            region = getattr(settings, 'AWS_S3_REGION_NAME', 'ap-south-1')
            
            if not all([aws_access_key, aws_secret_key, bucket_name]):
                raise Http404("S3 configuration missing")
            
            # Prepare S3 path
            s3_path = file_storage_service._prepare_path(document.file_path, add_base=True)
            
            print(f"[secure_document_view] Streaming from S3: {s3_path}")
            
            # Create S3 client
            s3_client = boto3.client(
                's3',
                aws_access_key_id=aws_access_key,
                aws_secret_access_key=aws_secret_key,
                region_name=region
            )
            
            try:
                # Get file from S3
                s3_response = s3_client.get_object(Bucket=bucket_name, Key=s3_path)
                
                # Get content type
                content_type = s3_response.get('ContentType', 'application/octet-stream')
                if content_type == 'application/octet-stream':
                    # Try to guess from filename
                    guessed_type, _ = mimetypes.guess_type(document.file_name)
                    if guessed_type:
                        content_type = guessed_type
                
                # ✅ Use StreamingHttpResponse for better performance
                def file_iterator(file_obj, chunk_size=8192):
                    while True:
                        chunk = file_obj.read(chunk_size)
                        if not chunk:
                            break
                        yield chunk
                
                # Create streaming response
                response = StreamingHttpResponse(
                    file_iterator(s3_response['Body']),
                    content_type=content_type
                )
                
                # Set headers
                from urllib.parse import quote
                response['Content-Disposition'] = f'inline; filename="{quote(document.file_name)}"'
                response['Content-Length'] = str(s3_response['ContentLength'])
                response['X-Content-Type-Options'] = 'nosniff'
                
                return response
                
            except ClientError as e:
                print(f"[secure_document_view] S3 error: {e}")
                error_code = e.response.get('Error', {}).get('Code', 'Unknown')
                print(f"Error code: {error_code}")
                
                if error_code == 'NoSuchKey':
                    raise Http404("Document not found on S3")
                elif error_code == 'AccessDenied':
                    raise Http404("Access denied to document")
                else:
                    raise Http404("Error retrieving document")
                
        else:
            # ========== LOCAL/TEST: Serve from local filesystem ==========
            from pathlib import Path
            
            file_path = Path(settings.MEDIA_ROOT) / Path(document.file_path.replace("\\", "/"))
            
            if not file_path.exists():
                raise Http404(f"Document file not found at: {file_path}")
            
            # Determine content type
            content_type, encoding = mimetypes.guess_type(str(file_path))
            if content_type is None:
                content_type = 'application/octet-stream'
            
            # Serve the file
            from django.http import FileResponse
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type,
                as_attachment=False,
                filename=document.file_name
            )
            
            response['X-Content-Type-Options'] = 'nosniff'
            
            return response
            
    except Http404:
        raise  # Re-raise Http404
    except Exception as e:
        print(f"[secure_document_view] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise Http404("Document not found")
    
@login_required
def membership_form_view(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        library_code = request.session.get('library_db', None)
        user_code = request.session["user_id"]
        role_id = request.session["role_id"]
        
        if request.method == "GET":
            
            if library_code:
                
                    membershipid = dec(request.GET.get('membershipid'))
                
                    library_details = tbl_librarymasterL01.objects.filter(is_active=1, library_code=library_code)
                    for lilo in library_details:
                        encrypted_library_code = enc(lilo.library_code)
                        lilo.libraries = encrypted_library_code
                
                    membership_options = parameter_master_L01.objects.filter(isactive=1,parameter_name='MembershipForm').order_by('parameter_id')
                    
                    # ward_details = WardMaster.objects.using('default').filter(is_active=1)
                    ward_details = WardMaster.objects.filter(is_active=1)
                    
                    membership = get_object_or_404(MembershipDetails, id=membershipid, isactive=1)    
                    
                    membership_Master = MembershipMaster.objects.filter(isactive=1)
                    if membership_Master.exists():
                        for mema in membership_Master:
                            mema.membership_id_enc = enc(str(mema.id))
                            mema.membership_code_enc = enc(str(mema.membership_code)) if mema.membership_code else None
                    
                    document_details = DocumentDetails.objects.filter(
                        membership=membership, isactive=1
                    ).select_related("document")

                    documents_map = {}
                    for doc in document_details:
                        doc.doc_id_enc = enc(str(doc.id))  # encrypt document ID
                        documents_map[doc.document.id] = doc
                        
            template = "L01/membership_approval/membership_form_view.html"
            context = {
                'MEDIA_URL': settings.MEDIA_URL,
                'library_details': library_details,
                'membership_options': membership_options,
                'ward_details': ward_details,
                'membership_Master': membership_Master,
                'library_code': library_code,
                'membership': membership,
                'documents_map': documents_map,
                'role_id': role_id,
            }
                
            return render(request, template, context)
        
        if request.method == "POST":
            # Step 1: Get POST data
            membership_id = request.POST.get("membership_id")
            action = request.POST.get("action")  # approve / reject
            user_id = request.POST.get("user_id")
            
            # Get the parameter record
            offline_flag_record = parameter_master_L01.objects.filter(
                parameter_name="MembershipPaymentOfflineFlag",
                isactive=1
            ).first()
            
            # Check if record exists and get its value
            if offline_flag_record and offline_flag_record.parameter_value:
                offline_flag = offline_flag_record.parameter_value
            else:
                offline_flag = "0"  # Default value if not found
            
            # Step 2: Get the membership object
            membership = get_object_or_404(MembershipDetails, id=membership_id, isactive=1)

            if action == "payment_received":
                try:
                    with transaction.atomic():

                        # ----------------------------------
                        # 1. Decide payment type
                        # ----------------------------------
                        payment_type = (
                            "Membership Renewed"
                            if membership.membership_renew == 1
                            else "Membership"
                        )

                        status_received = StatusMaster.objects.get(id=5)
                        membership_master = membership.membership

                        # ----------------------------------
                        # 2. Create PaymentDetails (SOURCE OF TRUTH)
                        # ----------------------------------
                        PaymentDetails.objects.create(
                            membership=membership,
                            payment_mode="Offline",
                            payment_method="Cash/Cheque/Other",
                            payment_type=payment_type,
                            deposit_amount=membership.deposit,
                            entry_fee_amount=membership.entry_fees,
                            monthly_subscription_amount=membership_master.subscription_fees,
                            total_subscription_amount=membership.subscription,
                            subscription_from=membership.from_date,
                            subscription_to=membership.to_date,
                            status=status_received,
                            membership_code=membership.membership_code,
                            user_id=user_id,
                            created_by=user_code,
                            updated_by=user_code,
                            payment_date=timezone.now().date(),
                        )

                        # ----------------------------------
                        # 3. Generate NEW membership code
                        # ----------------------------------
                        increment_master = get_object_or_404(
                            IncrementMaster,
                            id=1,
                            isactive=1
                        )

                        current_number = int(increment_master.incrementFieldNumber)
                        new_number = current_number + 1
                        new_number_str = str(new_number).zfill(3)

                        new_membership_code = f"{increment_master.incrementFieldName}{new_number_str}"

                        increment_master.incrementFieldNumber = new_number_str
                        increment_master.save()

                        # ----------------------------------
                        # 4. Update MembershipDetails
                        # ----------------------------------
                        membership.status = status_received
                        membership.membership_renew = 0
                        membership.membership_code = new_membership_code
                        membership.membership_start_date = timezone.now().date()  # ✅ Set payment date
                        membership.updated_by = user_code
                        membership.save()

                        # ----------------------------------
                        # 5. Activate User
                        # ----------------------------------
                        user = CustomUser.objects.get(username=user_id)
                        user_was_inactive = not user.is_active  # Check if user was inactive before

                        if not user.is_active:
                            user.is_active = True
                            user.save()
                            user_activated = True
                        else:
                            user_activated = False

                        # ----------------------------------
                        # 6. Fetch password from password_storage
                        # ----------------------------------
                        password_entry = get_object_or_404(password_storage, user=user)
                        user_password = password_entry.passwordText

                        # ----------------------------------
                        # 7. Prepare SMS message using SmsTemplate
                        # ----------------------------------
                        try:
                            # Get the fees success template
                            sms_template = SmsTemplate.objects.get(template_id='1107176880334605433')
                            
                            # Get user's first name or full name
                            user_name = membership.first_name if membership.first_name else user.full_name
                            
                            # Create the comprehensive message with ALL details
                            user_details = f"{user.username} and {user_password} and Membership Code: {new_membership_code}"
                            
                            # Replace the {#var#} placeholder with all details
                            # The template has only one placeholder, so include everything there
                            message = sms_template.template_message.replace('{#var#}', user_details)
                            
                            print(f"Generated message: {message}")
                            
                        except SmsTemplate.DoesNotExist:
                            # Fallback to OTP template
                            print("Fees success template not found, using OTP template")
                            otp_template = get_object_or_404(OTPMessage, OTPIDNumber=1)
                            message = otp_template.OTPText.replace(
                                '@UserId',
                                f"Username: {user.username}, Password: {user_password}, Membership Code: {new_membership_code}"
                            )

                        # ----------------------------------
                        # 8. Send SMS (LAST STEP)
                        # ----------------------------------
                        # Use membership.mobile_no if available, otherwise user.phone
                        mobile_no = membership.mobile_no if membership.mobile_no else user.phone
                        
                        if mobile_no:
                            # Send SMS
                            sms_sent = send_sms(mobile_no, message)
                            
                            # Generate unique ID for logging
                            unique_id = f"PAYMENT_RECEIVED_{membership.id}_{int(time.time())}"
                            
                            # Log the SMS
                            if sms_sent:
                                log_sms(user, mobile_no, message, 'Success', unique_id)
                                sms_status = "sent"
                            else:
                                log_sms(user, mobile_no, message, 'Failed', unique_id)
                                sms_status = "failed"
                                
                            print(f"Payment confirmation SMS {sms_status} to {user_name}")
                        else:
                            print(f"No mobile number found for {user_name}")
                            sms_sent = False

                        # ----------------------------------
                        # 9. Success Message
                        # ----------------------------------
                        if user_activated:
                            if sms_sent:
                                messages.success(
                                    request,
                                    f"User {user.username} has been activated successfully. Payment has been recorded and SMS sent."
                                )
                            else:
                                messages.warning(
                                    request,
                                    f"User {user.username} has been activated successfully. Payment has been recorded but SMS could not be sent (no mobile number or sending failed)."
                                )
                        else:
                            if sms_sent:
                                messages.success(
                                    request,
                                    f"User {user.username} is already active. Payment has been recorded and SMS sent."
                                )
                            else:
                                messages.warning(
                                    request,
                                    f"User {user.username} is already active. Payment has been recorded but SMS could not be sent."
                                )

                    return redirect("L01:membership_approval")

                except Exception as e:
                    messages.error(
                        request,
                        f"An error occurred while processing the payment: {str(e)}"
                    )
                    return redirect("L01:membership_approval")
                
            elif action == "renewal_approved":
                try:
                    with transaction.atomic():
                        membership = get_object_or_404(MembershipDetails, id=membership_id)
                        
                        # Get adjusted amount from form
                        adjusted_amount = request.POST.get('adjusted_amount')
                        
                        # ✅ STEP 1: Find and delete staging record
                        staging_record = MembershipRenewalStaging.objects.filter(
                            membership=membership,
                            status='pending'
                        ).first()
                        
                        if staging_record:
                            staging_record.delete()
                        
                        # ✅ STEP 2: Create payment using values from MembershipDetails
                        # Convert all to float for calculation
                        deposit = float(membership.deposit or 0)
                        entry_fees = float(membership.entry_fees or 0)
                        total_subscription = float(membership.subscription or 0)
                        
                        # ORIGINAL FINE COMPONENTS (from calculation)
                        gap_subscription_delay = float(membership.gap_subscription_delay or 0)
                        gap_fine = float(membership.gap_fine or 0)
                        late_fee = float(membership.late_fee or 0)
                        original_total_fine = float(membership.total_fine_membership or 0)
                        
                        # Get monthly rate (already float)
                        monthly_subscription = float(membership.membership.subscription_fees or 0)
                        
                        # Process adjusted amount if provided
                        adjusted_amount_float = None
                        if adjusted_amount and adjusted_amount.strip() != '':
                            try:
                                adjusted_amount_float = float(adjusted_amount)
                            except ValueError:
                                # If invalid number, ignore adjustment
                                pass
                        
                        # Determine which fine amount to use for calculation
                        fine_to_use = original_total_fine
                        if adjusted_amount_float is not None:
                            fine_to_use = adjusted_amount_float
                        
                        # Calculate total amount
                        total_amount = deposit + entry_fees + total_subscription + fine_to_use
                        
                        # ✅ STEP 3: MODIFY ACTION if gap fine exists
                        # Check if any fine components exist
                        has_gap_fine = gap_fine > 0 or gap_subscription_delay > 0 or late_fee > 0
                        
                        # Modify the action variable
                        if has_gap_fine and action == "renewal_approved":
                            action = "renewal_gap_approved"
                        
                        # Create payment record
                        payment = PaymentDetails.objects.create(
                            membership=membership,
                            payment_mode="Offline",
                            payment_type="Membership Renewal",
                            payment_method="Cash/Cheque/Other",
                            
                            # Amounts from MembershipDetails table
                            deposit_amount=deposit,
                            entry_fee_amount=entry_fees,
                            
                            # Subscription from MembershipDetails
                            monthly_subscription_amount=monthly_subscription,
                            total_subscription_amount=total_subscription,
                            
                            # Total fine amount
                            fine_amount=original_total_fine,
                            book_fine_amount=0,
                            
                            # ADJUSTED AMOUNT: Store adjusted amount if provided
                            adjusted_amount=adjusted_amount_float,
                            
                            # Dates from MembershipDetails
                            subscription_from=membership.from_date,
                            subscription_to=membership.to_date,
                            
                            # Status and tracking
                            status_id=5,
                            user_id=membership.user_id,
                            membership_code=membership.membership_code,
                            created_by=user_code,
                            updated_by=user_code,
                            payment_date=timezone.now().date(),
                            
                            # Payment breakdown with detailed fine components
                            remarks=(
                                f"Renewal approved by {user_code}\n"
                                f"Action: {action}\n"
                                f"Subscription: ₹{monthly_subscription}/month × {membership.membership_duration or 1} months = ₹{total_subscription}\n"
                                f"Deposit: ₹{deposit} | Entry Fees: ₹{entry_fees}\n"
                                f"\n--- FINE BREAKDOWN ---\n"
                                f"Gap Subscription Delay: ₹{gap_subscription_delay:.2f}\n"
                                f"Gap Fine (Penalty): ₹{gap_fine:.2f}\n"
                                f"Late Fee: ₹{late_fee:.2f}\n"
                                f"Original Total Fine: ₹{original_total_fine:.2f}\n"
                                f"{'Adjusted Fine: ₹' + str(adjusted_amount_float) if adjusted_amount_float is not None else ''}\n"
                                f"Fine Charged: ₹{fine_to_use:.2f}\n"
                                f"\n--- TOTAL ---\n"
                                f"Total Amount: ₹{total_amount:.2f}"
                            )
                        )
                        
                        # ✅ STEP 4: Update membership status
                        membership.status_id = 5  # RENEWED status
                        membership.actionperformed = action  # Use the (possibly modified) action
                        membership.reviewed = user_code
                        membership.reviewed_at = timezone.now()
                        membership.updated_by = user_code
                        
                        # ✅ STEP 5: Add approval note with detailed breakdown
                        adjustment_note = ""
                        if adjusted_amount_float is not None and adjusted_amount_float != original_total_fine:
                            adjustment_note = f" | Fine adjusted: ₹{original_total_fine:.2f} → ₹{adjusted_amount_float:.2f}"
                        
                        approval_note = (
                            f"[Renewal Approved: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}] "
                            f"Approved by: {user_code} | Payment ID: {payment.id} | Action: {action}{adjustment_note}\n"
                            f"Fine Components: Gap Sub: ₹{gap_subscription_delay:.2f} + Gap Fine: ₹{gap_fine:.2f} + Late Fee: ₹{late_fee:.2f}"
                        )
                        
                        if membership.remarks:
                            membership.remarks = f"{membership.remarks}\n{approval_note}"
                        else:
                            membership.remarks = approval_note
                        
                        membership.save()
                        
                        # ✅ STEP 7: SEND SMS FOR MEMBERSHIP RENEWAL
                        try:
                            # Get the membership renewal template
                            sms_template = SmsTemplate.objects.get(template_id='1107176880347608387')
                            
                            # Get user's first name or full name
                            user_name = membership.first_name if membership.first_name else ""
                            
                            # Get CustomUser for logging
                            try:
                                user = CustomUser.objects.get(username=membership.user_id)
                            except CustomUser.DoesNotExist:
                                user = None
                            
                            # Create renewal details message
                            renewal_details = (
                                f"Renewal ID: {payment.id}, "
                                f"From: {membership.from_date.strftime('%d-%m-%Y')} to {membership.to_date.strftime('%d-%m-%Y')}, "
                                f"Total: ₹{total_amount:.2f}"
                            )
                            
                            # Replace the placeholder with user's name and renewal details
                            message = sms_template.template_message.replace('{#var#}', user_name)
                            
                            # Send SMS - Use membership.mobile_no
                            mobile_no = membership.mobile_no
                            
                            if mobile_no:
                                # Send SMS
                                sms_sent = send_sms(mobile_no, message)
                                
                                # Generate unique ID for logging
                                unique_id = f"RENEWAL_APPROVED_{membership.id}_{payment.id}_{int(time.time())}"
                                
                                # Log the SMS
                                if sms_sent:
                                    if user:
                                        log_sms(user, mobile_no, message, 'Success', unique_id)
                                    else:
                                        log_sms(None, mobile_no, message, 'Success', unique_id)
                                    sms_status = "sent"
                                else:
                                    if user:
                                        log_sms(user, mobile_no, message, 'Failed', unique_id)
                                    else:
                                        log_sms(None, mobile_no, message, 'Failed', unique_id)
                                    sms_status = "failed"
                                    
                                print(f"Renewal confirmation SMS {sms_status} to {user_name}")
                            else:
                                print(f"No mobile number found for {user_name}")
                                sms_sent = False
                                
                        except SmsTemplate.DoesNotExist:
                            print("Membership renewal template not found")
                            sms_sent = False
                        except Exception as e:
                            print(f"Error sending renewal SMS: {str(e)}")
                            sms_sent = False
                        
                        # ✅ STEP 6: Show appropriate success message with breakdown
                        success_message = ""
                        if adjusted_amount_float is not None and adjusted_amount_float != original_total_fine:
                            if has_gap_fine:
                                success_message = (
                                    f"Renewal with gap fine approved with adjustment!\n"
                                    f"Action: {action}\n"
                                    f"Original Fine: ₹{original_total_fine:.2f} (Gap Sub: ₹{gap_subscription_delay:.2f} + Gap Fine: ₹{gap_fine:.2f} + Late: ₹{late_fee:.2f})\n"
                                    f"Adjusted Fine: ₹{adjusted_amount_float:.2f}\n"
                                    f"Payment record #{payment.id} created."
                                )
                            else:
                                success_message = (
                                    f"Renewal approved with adjustment!\n"
                                    f"Action: {action}\n"
                                    f"Original Fine: ₹{original_total_fine:.2f}\n"
                                    f"Adjusted Fine: ₹{adjusted_amount_float:.2f}\n"
                                    f"Payment record #{payment.id} created."
                                )
                        else:
                            if has_gap_fine:
                                success_message = (
                                    f"Renewal with gap fine approved!\n"
                                    f"Action: {action}\n"
                                    f"Fine Details: Gap Subscription Delay: ₹{gap_subscription_delay:.2f} + Gap Fine: ₹{gap_fine:.2f} + Late Fee: ₹{late_fee:.2f}\n"
                                    f"Payment record #{payment.id} created."
                                )
                            else:
                                success_message = (
                                    f"Renewal approved!\n"
                                    f"Action: {action}\n"
                                    f"No gap fine applied.\n"
                                    f"Payment record #{payment.id} created."
                                )
                        
                        # Add SMS status to the success message
                        if sms_sent:
                            success_message += " SMS notification sent to member."
                        else:
                            success_message += " (SMS notification could not be sent)"
                        
                        messages.success(request, success_message)
                        
                except Exception as e:
                    messages.error(request, f"Error: {str(e)}")
            
            elif action == "renewal_rejected":
                try:
                    with transaction.atomic():
                        membership = get_object_or_404(MembershipDetails, id=membership_id)
                        
                        # ✅ STEP 1: Find the pending staging record for this membership
                        staging_record = MembershipRenewalStaging.objects.filter(
                            membership=membership,
                            status='pending'
                        ).first()
                        
                        if staging_record:
                            # ✅ STEP 2: REVERT membership data to OLD values from staging
                            membership.from_date = staging_record.old_from_date
                            membership.to_date = staging_record.old_to_date
                            membership.membership_duration = staging_record.old_duration
                            membership.deposit = staging_record.old_deposit
                            membership.entry_fees = staging_record.old_entry_fees
                            membership.subscription = staging_record.old_subscription
                            
                            # Revert membership type if changed
                            if staging_record.old_membership_id:
                                try:
                                    old_membership_type = MembershipMaster.objects.get(id=staging_record.old_membership_id)
                                    membership.membership = old_membership_type
                                except MembershipMaster.DoesNotExist:
                                    pass  # Keep current if old not found
                            
                            # ✅ STEP 3: Reset fine calculation fields if they were changed
                            # (Optional: You might want to revert these too if they were part of renewal)
                            membership.gap_months = 0
                            membership.gap_subscription_delay = 0.00  # NEW: Reset subscription delay
                            membership.gap_fine = 0.00
                            membership.late_fee = 0.00
                            membership.total_fine_membership = 0.00
                            
                            # NEW: Reset gap period fields
                            membership.gap_period_from = None
                            membership.gap_period_to = None
                            
                            # ✅ STEP 5: DELETE the staging record
                            staging_record.delete()
                        
                        # ✅ STEP 6: Update membership status and tracking fields (your existing code)
                        membership.status_id = 11  # RENEWAL REJECTED
                        membership.actionperformed = action
                        membership.reviewed = user_code
                        membership.reviewed_at = timezone.now()
                        membership.updated_by = user_code
                        membership.membership_renew = 0  # Reset renewal flag
                        
                        # ✅ STEP 7: Add rejection note to remarks
                        rejection_note = f"[Renewal Rejected: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}] Rejected by: {user_code}"
                        if request.POST.get('rejection_reason'):
                            rejection_note += f", Reason: {request.POST.get('rejection_reason')}"
                        
                        # NEW: Add note about reverting fine calculations
                        rejection_note += f"\n[Fine Calculation Reverted: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}] All fine calculations reset to zero."
                        
                        if membership.remarks:
                            membership.remarks = f"{membership.remarks}\n{rejection_note}"
                        else:
                            membership.remarks = rejection_note
                        
                        # ✅ STEP 8: Save membership with reverted data
                        membership.save()
                        
                        # ✅ STEP 9: SEND SMS FOR RENEWAL REJECTION
                        try:
                            # Get the renewal rejection template
                            sms_template = SmsTemplate.objects.get(template_id='1107176880362727317')
                            
                            # Get user's first name or full name
                            user_name = membership.first_name if membership.first_name else ""
                            
                            # Get CustomUser for logging
                            try:
                                user = CustomUser.objects.get(username=membership.user_id)
                            except CustomUser.DoesNotExist:
                                user = None
                            
                            # Replace the placeholder with user's name
                            message = sms_template.template_message.replace('{#var#}', user_name)
                            
                            # Send SMS - Use membership.mobile_no
                            mobile_no = membership.mobile_no
                            
                            if mobile_no:
                                # Send SMS
                                sms_sent = send_sms(mobile_no, message)
                                
                                # Generate unique ID for logging
                                unique_id = f"RENEWAL_REJECTED_{membership.id}_{int(time.time())}"
                                
                                # Log the SMS
                                if sms_sent:
                                    if user:
                                        log_sms(user, mobile_no, message, 'Success', unique_id)
                                    else:
                                        log_sms(None, mobile_no, message, 'Success', unique_id)
                                    sms_status = "sent"
                                else:
                                    if user:
                                        log_sms(user, mobile_no, message, 'Failed', unique_id)
                                    else:
                                        log_sms(None, mobile_no, message, 'Failed', unique_id)
                                    sms_status = "failed"
                                    
                                print(f"Renewal rejection SMS {sms_status} to {user_name}")
                            else:
                                print(f"No mobile number found for {user_name}")
                                sms_sent = False
                                
                        except SmsTemplate.DoesNotExist:
                            print("Renewal rejection template not found")
                            sms_sent = False
                        except Exception as e:
                            print(f"Error sending renewal rejection SMS: {str(e)}")
                            sms_sent = False
                        
                        # ✅ STEP 10: Show success message
                        success_message = "Renewal rejected successfully!"
                        if sms_sent:
                            success_message += " SMS notification sent to member."
                        else:
                            success_message += " (SMS notification could not be sent)"
                        
                        messages.warning(request, success_message)
                        
                except Exception as e:
                    messages.error(request, f"Error: {str(e)}")
                    
            else:
                # Step 3: Map actions to status IDs
                status_map = {
                    "approved": 2,
                    "rejected": 3,
                }
                new_status_id = status_map.get(action)

                if new_status_id:
                    
                    if new_status_id == 2:  # APPROVED
                        
                        if membership.membership_code:
                            messages.warning(request, "Membership code already exists. No new code generated.")

                        if offline_flag == "1":
                            try:
                                pay_offline_status = StatusMaster.objects.get(
                                    status_code="PAY_OFFLINE",
                                    isactive=1
                                )
                                new_status_id = pay_offline_status.id
                                
                                messages.success(request, f"User {membership.first_name} is requested to kindly visit the library to complete the payment process.")
                                
                                # --- SMS Logic for Payment Reminder ---
                                try:
                                    # Get the payment reminder template
                                    sms_template = SmsTemplate.objects.get(template_id='1107176880188300865')
                                    
                                    # Replace the placeholder with user's name
                                    user_name = membership.first_name if membership.first_name else ""
                                    message = sms_template.template_message.replace('{#var#}', user_name)
                                    
                                    # Send SMS - Get phone from MembershipDetails.mobile_no
                                    if membership.mobile_no:
                                        send_sms(membership.mobile_no, message)
                                        
                                        # Generate unique ID for logging
                                        unique_id = f"PAY_OFFLINE_{membership.id}_{int(time.time())}"
                                        
                                        # Log the SMS - need to get user from CustomUser if exists
                                        try:
                                            user = CustomUser.objects.get(username=membership.user_id)
                                            log_sms(user, membership.mobile_no, message, 'Success', unique_id)
                                        except CustomUser.DoesNotExist:
                                            # Create a dummy user for logging if needed
                                            print(f"No CustomUser found for {membership.user_id}")
                                            log_sms(None, membership.mobile_no, message, 'Success', unique_id)
                                        
                                        print(f"Payment reminder SMS sent to {user_name} at {membership.mobile_no}")
                                    else:
                                        print(f"No mobile number found for {membership.first_name}")
                                        
                                except SmsTemplate.DoesNotExist:
                                    print("Payment reminder template not found")
                                except Exception as e:
                                    print(f"Error sending payment reminder SMS: {str(e)}")
                                
                                # --- Update membership fields ---
                                membership.actionperformed = action
                                membership.reviewed = user_code
                                membership.reviewed_at = timezone.now()
                                membership.status_id = new_status_id
                                membership.save()
                                
                                print(f"Membership updated with PAY_OFFLINE status (ID: {new_status_id})")
                                
                            except StatusMaster.DoesNotExist:
                                messages.error(request, "PAY_OFFLINE status not found in StatusMaster")
                                return
                            except Exception as e:
                                messages.error(request, f"Error fetching status: {str(e)}")
                                return
                            
                        else:
                            # --- Update user status to active ---
                            user = CustomUser.objects.get(username=user_id)
                            user.is_active = True
                            user.save()
                            
                            # --- Get password from password_storage ---
                            password_entry = get_object_or_404(password_storage, user=user)
                            user_password = password_entry.passwordText  # Password fetched from password_storage

                            # --- SMS Logic for Login Credentials ---
                            try:
                                # Get the login credentials template
                                sms_template = SmsTemplate.objects.get(template_id='1107176880213056412')
                                
                                # Replace first placeholder with user's first name
                                user_name = membership.first_name if membership.first_name else user.full_name
                                message = sms_template.template_message.replace('{#var#}', user_name, 1)
                                
                                # Replace second placeholder with username and password
                                credentials = f"{user.username} and {user_password}"
                                message = message.replace('{#var#}', credentials, 1)
                                
                                # Send SMS - Use membership.mobile_no or user.phone
                                mobile_no = membership.mobile_no if membership.mobile_no else user.phone
                                if mobile_no:
                                    send_sms(mobile_no, message)
                                    
                                    # Generate unique ID for logging
                                    unique_id = f"APPROVED_{membership.id}_{int(time.time())}"
                                    
                                    # Log the SMS
                                    log_sms(user, mobile_no, message, 'Success', unique_id)
                                    
                                    print(f"Login credentials SMS sent to {user_name} at {mobile_no}")
                                else:
                                    print(f"No mobile number found for {user_name}")
                                
                            except SmsTemplate.DoesNotExist:
                                print("Login credentials template not found")
                                # Fallback to old OTP message method
                                try:
                                    mobile_no = membership.mobile_no if membership.mobile_no else user.phone
                                    if mobile_no:
                                        otp_template = get_object_or_404(OTPMessage, OTPIDNumber=1)
                                        message = otp_template.OTPText.replace('@UserId', f"{user.username} and {user_password}")
                                        send_sms(mobile_no, message)
                                        
                                        unique_id = f"APPROVED_{membership.id}_{int(time.time())}"
                                        log_sms(user, mobile_no, message, 'Success', unique_id)
                                except OTPMessage.DoesNotExist:
                                    print("OTP template also not found")
                            except Exception as e:
                                print(f"Error sending credentials SMS: {str(e)}")

                            messages.success(request, f"User {user.username} activated successfully and SMS sent.")

                            # --- Update membership fields ---
                            membership.actionperformed = action
                            membership.reviewed = user_code
                            membership.reviewed_at = timezone.now()
                            membership.status_id = new_status_id
                            membership.save()
                        
                    else:
                        # --- SMS Logic for Application Rejection ---
                        # Get phone from MembershipDetails.mobile_no
                        if membership.mobile_no:
                            try:
                                # Get the application rejection template
                                sms_template = SmsTemplate.objects.get(template_id='1107176880318206587')
                                
                                # Replace the placeholder with user's name
                                user_name = membership.first_name if membership.first_name else ""
                                message = sms_template.template_message.replace('{#var#}', user_name)
                                
                                # Send SMS
                                send_sms(membership.mobile_no, message)
                                
                                # Generate unique ID for logging
                                unique_id = f"REJECTED_{membership.id}_{int(time.time())}"
                                
                                # Log the SMS - try to get user from CustomUser
                                try:
                                    user = CustomUser.objects.get(username=membership.user_id)
                                    log_sms(user, membership.mobile_no, message, 'Success', unique_id)
                                except CustomUser.DoesNotExist:
                                    # Log without user if not found
                                    log_sms(None, membership.mobile_no, message, 'Success', unique_id)
                                
                                print(f"Application rejection SMS sent to {user_name} at {membership.mobile_no}")
                                
                            except SmsTemplate.DoesNotExist:
                                print("Application rejection template not found")
                            except Exception as e:
                                print(f"Error sending rejection SMS: {str(e)}")
                        else:
                            print(f"No mobile number found for {membership.first_name}")
                        
                        # --- Update membership fields ---
                        membership.actionperformed = action
                        membership.reviewed = user_code
                        membership.reviewed_at = timezone.now()
                        membership.status_id = new_status_id
                        membership.save()

                        messages.success(request, f"Membership {action.capitalize()}d successfully.")
                    
                else:
                    messages.error(request, "Invalid action!")

            return redirect("L01:membership_approval")
        
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "library_list"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, "Oops...! Something went wrong!")
        return redirect("L01:membership_approval")

def send_sms(mobile, message):
    try:
        api_url = f"https://api.pinnacle.in/index.php/sms/send/NMMCSV/{mobile}/{message}/TXT?apikey=6ac137-a52260-71ad83-05eea9-422b6e"

        response = requests.get(api_url)
        if response.status_code == 200:
            return True
        else:
            logger.error(f"Error sending SMS to {mobile}: {response.text}")
            return False
    except Exception as e:
        logger.error(f"Error sending SMS to {mobile}: {e}")
        return False

def log_sms(user, mobile, message, status, unique_id):
    try:
        SMSLog.objects.create(
            user=user,
            mobile=mobile,
            message=message,
            status=status,
            unique_id=unique_id,
            content_type="text",
            response_status="Success" if status == 'Success' else "Failed",
            response_uri="API_URL",
            status_description="Message Sent" if status == 'Success' else "Failed",
            error_exception="N/A",
            error_message="N/A",
        )
    except Exception as e:
        logger.error(f"Error logging SMS for user {user.username} (ID: {user.id}): {e}")
        SMSLog.objects.create(
            user=user,
            mobile=mobile,
            message=message,
            status="Failed",
            unique_id=unique_id,
            content_type="text",
            response_status="Failed",
            response_uri="API_URL",
            status_description="Logging Failed",
            error_exception=str(e),
            error_message="Failed to log SMS due to an exception",
        )

# Membership Management Views by Member

@login_required
def membership_payment_index(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        library_code = request.session.get('library_db', None)
        username = request.session.get('username', None)
        user_id = request.session["user_id"]
        role_id = request.session["role_id"]

        if request.method == "GET":
            today = date.today()
            
            # Get all memberships
            memberships = (
                MembershipDetails.objects
                .select_related("status", "membership")
                .filter(user_id=username)
            )
            
            renewal_candidates = []
            
            for mem in memberships:
                mem.membership_id_enc = enc(str(mem.id))
                mem.per_month_subscription = mem.membership.subscription_fees if mem.membership else 0
                
                # Check for pending books
                pending_books_query = CirculationTransaction.objects.filter(
                    member=mem,
                    return_date__isnull=True
                )
                mem.has_pending_books = pending_books_query.exists()
                mem.pending_books_count = pending_books_query.count()
                
                # Check if membership is expired (EXCLUDE LIFETIME MEMBERSHIPS - IDs 5 and 6)
                mem.membership_id = mem.membership.id if mem.membership else None
                
                # Skip lifetime memberships from expiration check (IDs 5 and 6)
                if mem.membership_id in [5, 6]:  # LB and PB - Lifetime memberships
                    mem.is_expired = False  # Lifetime memberships never expire
                elif mem.to_date:
                    mem.is_expired = today >= mem.to_date
                else:
                    mem.is_expired = False
                
                # Check if eligible for renewal (expired AND no pending books AND not lifetime)
                if mem.is_expired and not mem.has_pending_books:
                    renewal_candidates.append(mem)
            
            # Get first renewal candidate
            renew_membership = renewal_candidates[0] if renewal_candidates else None
            
            # Set variables for template
            show_renew_notice = renew_membership is not None
            membership_id_enc_for_renew = enc(str(renew_membership.id)) if renew_membership else None
            has_pending_books_for_renew = False  # Already filtered out
            
            # Check for rejected renewals (exclude lifetime memberships)
            renew_membership_rejected = memberships.filter(
                status__status_code="APP_RENEW_REJECT"
            ).exclude(
                membership__id__in=[5, 6]  # Exclude lifetime memberships by ID
            ).first()
            
            membership_id_enc_for_reject = (
                enc(str(renew_membership_rejected.id)) if renew_membership_rejected else None
            )
            
            # Library Name
            library = tbl_librarymasterL01.objects.filter(library_code=library_code, is_active=1).first()
            library_name_mar = library.library_name_mar if library else ""
            
            return render(
                request,
                "L01/Member Payment/member_payment.html",
                {
                    "MEDIA_URL": settings.MEDIA_URL,
                    "library_code": library_code,
                    "memberships": memberships,
                    "library_name_mar": library_name_mar,
                    "show_renew_notice": show_renew_notice,
                    "membership_id_enc_for_renew": membership_id_enc_for_renew,
                    "has_pending_books_for_renew": has_pending_books_for_renew,
                    "renew_membership_rejected": renew_membership_rejected,
                    "membership_id_enc_for_reject": membership_id_enc_for_reject,
                }
            )
           
        if request.method == "POST":
            membership_id = dec(request.POST.get("membership_id"))
            payment_type = request.POST.get("payment_type")
            
            # Update membership details as offline payment pending
            mem = MembershipDetails.objects.filter(id=membership_id, user_id=username).first()
            if mem:
                # Set status to Payment Pending
                pay_pending_status = StatusMaster.objects.get(status_code="PAY_OFFLINE")
                mem.status = pay_pending_status
                mem.updated_by = user_id
                mem.save()
                messages.success(request, "ऑफलाइन देयक प्रक्रिया सुरू आहे. स्थिती 'Payment Offline' मध्ये बदलली गेली आहे.")
                return redirect("L01:membership_payment_index")

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "library_list"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, "Oops...! Something went wrong!")
        return render(request, "L01/Member Payment/member_payment.html", {})

@login_required
def membership_paymentreceipt_download(request):
    try:
        membership_id = dec(request.GET.get("membershipid"))
        receipt_type = request.GET.get("type", "initial")  # Default to initial
        
        membership = get_object_or_404(MembershipDetails, id=membership_id)
        membership_master = membership.membership
        
        # ✅ Conditional payment query based on type parameter
        if receipt_type == "renewal":
            # Get RENEWAL payments only
            payments = PaymentDetails.objects.filter(
                membership=membership,
                payment_type='Membership Renewal'
            ).order_by('-id')[:1] 
        else:
            # Get INITIAL membership payments only (default)
            payments = PaymentDetails.objects.filter(
                membership=membership,
                payment_type='Membership'
            ).order_by('-id')[:1] 

        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        regular_font_path = os.path.join(settings.BASE_DIR, 'static/fonts/Merriweather_120pt-Regular.ttf')
        bold_font_path = os.path.join(settings.BASE_DIR, 'static/fonts/Merriweather_120pt-Bold.ttf')
        pdfmetrics.registerFont(TTFont('Merriweather', regular_font_path))
        pdfmetrics.registerFont(TTFont('Merriweather-Bold', bold_font_path))

        # Page margins
        left_margin = 20*mm
        right_margin = width - 20*mm
        top_margin = height - 10*mm
        bottom_margin = 20*mm
        y = top_margin

        # ✅ Full-page border
        c.setLineWidth(1)
        c.rect(10, 10, width - 20, height - 20)

        # Logo
        logo_path = os.path.join(settings.BASE_DIR, 'static/images/administrative/nmmc-logo.jpeg')
        logo_width = 30*mm
        logo_height = 30*mm
        if os.path.exists(logo_path):
            c.drawImage(logo_path, left_margin, y - logo_height, width=logo_width, height=logo_height, mask='auto')

        # Header text - side by side with logo
        text_x = left_margin + logo_width + 10*mm
        text_y = y - logo_height/2 + 12

        c.setFont("Merriweather", 24)
        c.drawString(text_x, text_y, "Navi Mumbai Municipal Corporation")

        # Different title based on receipt type
        if receipt_type == "renewal":
            c.setFont("Merriweather", 14)
            c.drawString(text_x, text_y - 30, "Membership Renewal Payment Receipt")
        else:
            c.setFont("Merriweather", 14)
            c.drawString(text_x, text_y - 30, "Membership Payment Receipt")

        # Move y down for next content
        y = y - logo_height - 30

        # ---------------- Member Details Table ----------------
        c.setFont("Merriweather-Bold", 14)
        c.drawString(left_margin, y, "Member Details:")
        y -= 60

        table_data = [
            ["Field", "Details"],
            ["Full Name", f"{membership.first_name} {membership.middle_name or ''} {membership.last_name}"],
            ["User ID", membership.user_id],
            ["Library", membership.library_name],
            ["Membership Name", membership_master.membership_type_en],
        ]

        table = Table(table_data, colWidths=[60*mm, 100*mm], hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Merriweather-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Merriweather'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ]))
        table.wrapOn(c, width, height)
        table.drawOn(c, left_margin, y - (len(table_data) * 20))
        y = y - (len(table_data) * 20) - 20

        # ---------------- Payment Details Table ----------------
        y -= 10
        c.setFont("Merriweather-Bold", 14)
        c.drawString(left_margin, y, "Payment Details:")
        y -= 60

        payment_data_list = []
        for idx, p in enumerate(payments, start=1):
            payment_data_list.append(["Payment Date", str(p.payment_date)])
            payment_data_list.append(["Payment Mode", str(p.payment_mode)])
            payment_data_list.append(["Payment Status", str(p.status.status_name if p.status else "")])
            
            # ✅ Use dates from Payment table, format as DD-MM-YYYY
            if p.subscription_from and p.subscription_to:
                # Format dates as DD-MM-YYYY
                from_date = p.subscription_from.strftime("%d-%m-%Y")
                to_date = p.subscription_to.strftime("%d-%m-%Y")
                payment_data_list.append(["Membership Duration", f"{from_date} to {to_date}"])
            
            if receipt_type == "renewal":
                # ✅ RENEWAL receipt
                payment_data_list.append(["Monthly Subscription (₹)", f"{p.monthly_subscription_amount or 0:.2f}"])
                payment_data_list.append(["Total Subscription (₹)", f"{p.total_subscription_amount or 0:.2f}"])
                
                # ✅ Fine calculations for renewal
                if p.fine_amount or p.adjusted_amount:
                    if p.adjusted_amount and p.adjusted_amount > 0 and p.adjusted_amount != p.fine_amount:
                        payment_data_list.append(["Original Fine (₹)", f"{p.fine_amount or 0:.2f}"])
                        payment_data_list.append(["Adjusted Fine (₹)", f"{p.adjusted_amount or 0:.2f}"])
                        total_paid = (p.total_subscription_amount or 0) + (p.adjusted_amount or 0)
                    else:
                        payment_data_list.append(["Fine Amount (₹)", f"{p.fine_amount or 0:.2f}"])
                        total_paid = (p.total_subscription_amount or 0) + (p.fine_amount or 0)
                else:
                    total_paid = p.total_subscription_amount or 0
                
                # Add empty row for separation
                payment_data_list.append(["", ""])
                payment_data_list.append(["Total Amount Paid (₹)", f"{total_paid:.2f}"])
                
            else:
                # ✅ INITIAL membership receipt
                payment_data_list.append(["Deposit (₹)", f"{p.deposit_amount or 0:.2f}"])
                payment_data_list.append(["Entry Fee (₹)", f"{p.entry_fee_amount or 0:.2f}"])
                payment_data_list.append(["Monthly Subscription (₹)", f"{p.monthly_subscription_amount or 0:.2f}"])
                payment_data_list.append(["Total Subscription (₹)", f"{p.total_subscription_amount or 0:.2f}"])
                
                # ✅ Calculate Total Amount Paid for initial receipt
                total_paid = (p.deposit_amount or 0) + (p.entry_fee_amount or 0) + (p.total_subscription_amount or 0)
                
                # Add empty row for separation
                payment_data_list.append(["", ""])
                payment_data_list.append(["Total Amount Paid (₹)", f"{total_paid:.2f}"])

        if not payment_data_list:
            payment_data_list = [["No payments found", ""]]

        payment_table = Table(payment_data_list, colWidths=[60*mm, 100*mm], hAlign='LEFT')
        payment_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Merriweather'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -3), 0.5, colors.black),  # Grid for all except last 2 rows
            ('GRID', (0, -2), (-1, -1), 0.5, colors.black),  # Grid for total row
            ('LINEABOVE', (0, -2), (-1, -2), 1, colors.black),  # Line above total
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Highlight total row
            ('FONTNAME', (0, -1), (-1, -1), 'Merriweather-Bold'),  # Bold total
            ('FONTSIZE', (0, -1), (-1, -1), 14),  # Larger font for total
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        payment_table.wrapOn(c, width, height)
        payment_table.drawOn(c, left_margin, y - (len(payment_data_list) * 20))
        y = y - (len(payment_data_list) * 20) - 40  # Extra space after payment table

        # ---------------- Notes Section ----------------
        if receipt_type == "renewal":
            # ✅ Note for renewal receipt
            note_text = "Note: Deposit and Entry Fee were paid during initial membership and are not payable again."
            c.setFont("Merriweather", 10)
            c.setFillColor(colors.darkblue)
            c.drawString(left_margin, y, note_text)
            y -= 20

        # Footer
        footer_y = bottom_margin + 15
        
        # ✅ Right side: Page number
        c.drawString(right_margin - 50, footer_y, "Page 1 of 1")
        
        c.save()
        buffer.seek(0)
        
        # Different filename based on receipt type
        if receipt_type == "renewal":
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="renewal_receipt_{membership.user_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        else:
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="payment_receipt_{membership.user_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response

    except Exception as e:
        print(f"Error generating PDF: {e}")
        messages.error(request, "Oops! Something went wrong!")
        return render(request, "L01/Member Payment/member_payment.html", {})

@login_required
def membership_form_renew(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        library_code = request.session.get('library_db', None)
        user_code = request.session["user_id"]

        if request.method == "GET":
            
            if library_code:
                
                    membershipid = dec(request.GET.get('membershipid'))
                
                    library_details = tbl_librarymasterL01.objects.filter(is_active=1, library_code=library_code)
                    for lilo in library_details:
                        encrypted_library_code = enc(lilo.library_code)
                        lilo.libraries = encrypted_library_code
                    print(library_details)
                
                    membership_options = parameter_master_L01.objects.filter(isactive=1,parameter_name='MembershipForm').order_by('parameter_id')
                    
                    # ward_details = WardMaster.objects.using('default').filter(is_active=1)
                    ward_details = WardMaster.objects.filter(is_active=1)
                    
                    membership_Master = MembershipMaster.objects.filter(isactive=1)
                    if membership_Master.exists():
                        for mema in membership_Master:
                            encrypted_membership_id = enc(str(mema.id))
                            mema.membership_id_enc = encrypted_membership_id
                        print(membership_Master)
                        
                    membership = get_object_or_404(MembershipDetails, id=membershipid, isactive=1)   
                    
                    # Get the membership type details directly
                    current_membership_type = membership.membership  # This is the ForeignKey to MembershipMaster 
                    
                    document_details = DocumentDetails.objects.filter(
                        membership=membership, isactive=1
                    ).select_related("document")

                    documents_map = {}
                    for doc in document_details:
                        doc.doc_id_enc = enc(str(doc.id))  # encrypt document ID
                        documents_map[doc.document.id] = doc
                        
            template = "L01/membership_approval/membership_form_renew.html"
            context = {
                'MEDIA_URL': settings.MEDIA_URL,
                'library_details': library_details,
                'membership_options': membership_options,
                'ward_details': ward_details,
                'membership_Master': membership_Master,
                'library_code': library_code,
                'membership': membership,
                'documents_map': documents_map,
                'current_membership_type': current_membership_type,
                'subscription_fees': membership.membership.subscription_fees,
                'fine_membership': membership.membership.fine_membership,
            }
                
            return render(request, template, context)
        
        if request.method == "POST":
            with transaction.atomic():
                membership_id = request.POST.get("membership_id")
                membership = get_object_or_404(MembershipDetails, id=membership_id)
                
                # ✅ FIRST: Check for existing pending renewal
                existing_pending = MembershipRenewalStaging.objects.filter(
                    membership=membership,
                    status='pending'
                ).first()
             
                # Parse dates and convert types
                new_from_date_str = request.POST.get("fromDate")
                new_to_date_str = request.POST.get("toDate")
                new_duration_str = request.POST.get("months")
                new_membership_type_str = request.POST.get("membershiptype")
                
                # Convert NEW values
                new_from_date = None
                new_to_date = None
                new_duration = None
                new_membership_id = None
                
                if new_from_date_str:
                    new_from_date = datetime.strptime(new_from_date_str, "%Y-%m-%d").date()
                if new_to_date_str:
                    new_to_date = datetime.strptime(new_to_date_str, "%Y-%m-%d").date()
                if new_duration_str:
                    new_duration = int(new_duration_str)
                if new_membership_type_str:
                    new_membership_id = int(new_membership_type_str)
                
                # Get monetary NEW values
                new_deposit = float(request.POST.get("deposit", 0))
                new_entry_fees = float(request.POST.get("entry_fees", 0))
                new_subscription = float(request.POST.get("subscription", 0))
                
                # ✅ THIRD: Create staging record with OLD (from DB) and NEW (from POST)
                staging_record = MembershipRenewalStaging.objects.create(
                    membership=membership,
                    
                    # Store OLD data (current membership values from database)
                    old_from_date=membership.from_date,  # ← From DB
                    old_to_date=membership.to_date,      # ← From DB
                    old_duration=membership.membership_duration,  # ← From DB
                    old_deposit=membership.deposit,      # ← From DB
                    old_entry_fees=membership.entry_fees, # ← From DB
                    old_subscription=membership.subscription, # ← From DB
                    old_membership_id=membership.membership.id if membership.membership else None,  # ← From DB
                    
                    # Store NEW data (from form submission directly)
                    new_from_date=new_from_date,  # ← From POST request
                    new_to_date=new_to_date,      # ← From POST request
                    new_duration=new_duration,    # ← From POST request
                    new_deposit=new_deposit,      # ← From POST request
                    new_entry_fees=new_entry_fees, # ← From POST request
                    new_subscription=new_subscription, # ← From POST request
                    new_membership_id=new_membership_id,  # ← From POST request
                    
                    status='pending',
                    created_by=user_code,
                    remarks="Renewal submitted - awaiting approval"
                )

                # 1️⃣ Field map (only relevant fields)
                field_map = {
                    "membershiptype": "membership_id",
                    "months": "membership_duration",
                    "fromDate": "from_date",
                    "toDate": "to_date",
                }

                # 2️⃣ Detect field changes
                updated = False
                for form_field, model_field in field_map.items():
                    new_value = request.POST.get(form_field)

                    if new_value in ["", "None", None]:
                        new_value = None

                    if model_field == "membership_duration" and new_value:
                        new_value = int(new_value)

                    if model_field == "membership_id" and new_value:
                        new_value = int(new_value)

                    if model_field in ["from_date", "to_date"] and new_value:
                        new_value = datetime.strptime(new_value, "%Y-%m-%d").date()

                    old_value = getattr(membership, model_field, None)
                    if new_value != old_value:
                        setattr(membership, model_field, new_value)
                        updated = True

                # 3️⃣ Monetary fields
                money_fields = ["deposit", "entry_fees", "subscription"]
                for field in money_fields:
                    new_value = request.POST.get(field)
                    new_value = float(new_value) if new_value else 0.0
                    old_value = getattr(membership, field, 0.0)
                    if new_value != old_value:
                        setattr(membership, field, new_value)
                        updated = True

                # 4️⃣ Store fine calculation details in NEW SEPARATE COLUMNS
                # 4️⃣ Store fine calculation details in NEW SEPARATE COLUMNS
                gap_months = request.POST.get("gap_months")
                subscription_delay = request.POST.get("subscription_delay")  # NEW
                gap_fine = request.POST.get("gap_fine")
                late_fee = request.POST.get("late_fee")
                gap_period_from_date_str = request.POST.get("gap_period_from_date")  # "2026-01-01" (from frontend)
                gap_period_to_date_str = request.POST.get("gap_period_to_date")      # "2026-02-28" (from frontend)
                gap_period_display = request.POST.get("gap_period_display")          # "01-01-2026 to 28-02-2026"

                # Convert to appropriate types
                gap_months_int = int(gap_months) if gap_months and gap_months.strip() else 0
                subscription_delay_float = float(subscription_delay) if subscription_delay and subscription_delay.strip() else 0.00  # NEW
                gap_fine_float = float(gap_fine) if gap_fine and gap_fine.strip() else 0.00
                late_fee_float = float(late_fee) if late_fee and late_fee.strip() else 0.00
                total_fine_calculated = subscription_delay_float + gap_fine_float + late_fee_float  # Updated

                # Convert YYYY-MM-DD to DD-MM-YYYY format
                gap_period_from_formatted = None
                gap_period_to_formatted = None

                if gap_period_from_date_str and gap_period_from_date_str.strip():
                    # Convert from "2026-01-01" to "01-01-2026"
                    date_obj = datetime.strptime(gap_period_from_date_str, "%Y-%m-%d").date()
                    gap_period_from_formatted = date_obj.strftime("%d-%m-%Y")

                if gap_period_to_date_str and gap_period_to_date_str.strip():
                    # Convert from "2026-02-28" to "28-02-2026"
                    date_obj = datetime.strptime(gap_period_to_date_str, "%Y-%m-%d").date()
                    gap_period_to_formatted = date_obj.strftime("%d-%m-%Y")

                # Check if fine values have changed
                if (gap_months_int != (membership.gap_months or 0) or
                    subscription_delay_float != (membership.gap_subscription_delay or 0.00) or  # NEW
                    gap_fine_float != (membership.gap_fine or 0.00) or
                    late_fee_float != (membership.late_fee or 0.00) or
                    gap_period_from_formatted != (membership.gap_period_from or "") or  # Compare DD-MM-YYYY strings
                    gap_period_to_formatted != (membership.gap_period_to or "")):       # Compare DD-MM-YYYY strings
                    
                    # Update fine columns
                    membership.gap_months = gap_months_int
                    membership.gap_subscription_delay = subscription_delay_float  # NEW: Store subscription part
                    membership.gap_fine = gap_fine_float                          # Store only fine part (₹5 per month)
                    membership.late_fee = late_fee_float
                    membership.total_fine_membership = total_fine_calculated
                    membership.fine_calculated_at = datetime.now()
                    
                    # Update display strings as DD-MM-YYYY format
                    membership.gap_period_from = gap_period_from_formatted   # "01-01-2026"
                    membership.gap_period_to = gap_period_to_formatted       # "28-02-2026"
                    
                    updated = True
                    
                    # Add note to remarks about fine calculation
                    fine_note = f"[Fine Calculated: {datetime.now().strftime('%Y-%m-%d %H:%M')}] "
                    if gap_months_int > 0:
                        if gap_period_display:
                            fine_note += f"Gap: {gap_months_int} month(s) ({gap_period_display})"
                        elif gap_period_from_formatted and gap_period_to_formatted:
                            fine_note += f"Gap: {gap_months_int} month(s) ({gap_period_from_formatted} to {gap_period_to_formatted})"
                        else:
                            fine_note += f"Gap: {gap_months_int} month(s)"
                        
                        # Show breakdown in remarks
                        fine_note += f", Subscription Delay: ₹{subscription_delay_float:.2f}"
                        fine_note += f", Gap Fine: ₹{gap_fine_float:.2f}"
                        if late_fee_float > 0:
                            fine_note += f", Late Fee: ₹{late_fee_float:.2f}"
                        fine_note += f", Total: ₹{total_fine_calculated:.2f}"
                    else:
                        fine_note += "No gap period"
                    
                    if membership.remarks:
                        import re
                        remarks = re.sub(r'\[Fine Calculated:.*?\].*?(?=\n|$)', '', membership.remarks, flags=re.DOTALL)
                        remarks = remarks.strip()
                        if remarks:
                            membership.remarks = f"{remarks}\n{fine_note}"
                        else:
                            membership.remarks = fine_note
                    else:
                        membership.remarks = fine_note
                elif gap_months_int == 0:
                    # Clear gap period if no gap
                    membership.gap_period_from = None
                    membership.gap_period_to = None
                    membership.gap_subscription_delay = 0.00  # Reset to 0
                    updated = True

                # 5️⃣ Save if updated
                if updated:
                    membership.updated_by = user_code
                    membership.status_id = 9  # Renewed status
                    membership.membership_renew = 1
                    membership.actionperformed = f"Renewal request submitted"
                    membership.reviewed = None
                    membership.reviewed_at = None
                    
                    # ✅ TRIGGER WILL AUTOMATICALLY CREATE HISTORY RECORD!
                    membership.save()
                    
                    messages.success(request, "Membership renewed successfully!")
                    
                    # Optional: Log success
                    print(f"Membership {membership_id} renewed. Trigger created history record.")
                    
                else:
                    messages.info(request, "No changes detected. Membership not updated.")

                return redirect("L01:membership_payment_index")
        
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "library_list"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, "Oops...! Something went wrong!")
        return redirect("L01:membership_payment_index")

@login_required
def membership_form_cancellation(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        library_code = request.session.get('library_db', None)
        user_id = request.session.get("user_id")
        
        if request.method == "GET":
            try:
                membership_id_enc = request.GET.get("membership_id")
                membership_id = dec(membership_id_enc)

                membership = MembershipDetails.objects.select_related("membership", "status").get(id=membership_id)

                data = {
                    "id": membership.id,
                    "full_name": f"{membership.first_name or ''} {membership.middle_name or ''} {membership.last_name or ''}".strip(),
                    "membership_type": membership.membership.membership_type,
                    "membership_type_en": membership.membership.membership_type_en,
                    "deposit": membership.deposit,
                    "entry_fees": membership.entry_fees,
                    "subscription": membership.subscription,
                    "duration": membership.membership_duration,
                    "from_date": membership.from_date.strftime("%d-%m-%Y") if membership.from_date else None,
                    "to_date": membership.to_date.strftime("%d-%m-%Y") if membership.to_date else None,
                    "status_name": membership.status.status_name,
                }
                return JsonResponse({"success": True, "data": data})

            except MembershipDetails.DoesNotExist:
                return JsonResponse({"success": False, "message": "सदस्य माहिती सापडली नाही."})
            except Exception as e:
                print("Error fetching membership details:", e)
                return JsonResponse({"success": False, "message": "सर्व्हर त्रुटी आली."})
        
        if request.method == "POST":
            membership_id = dec(request.POST.get("membership_id"))
            membership = get_object_or_404(MembershipDetails, id=membership_id)
            
            action_by = request.POST.get("action_by")  # 'member' or 'librarian'

            if action_by == "member":
                # ✅ Check if member has any pending books
                has_pending_books = CirculationTransaction.objects.filter(
                    member=membership,
                    return_date__isnull=True
                ).exists()
                
                if has_pending_books:
                    messages.error(
                        request,
                        "सदस्यत्व रद्द करणे शक्य नाही. कारण आपल्या नावावर अजूनही पुस्तके ग्रंथालयात परत केलेली नाहीत."
                        " कृपया सर्व पुस्तके परत केल्यानंतरच सदस्यत्व रद्द करा."
                    )
                    return redirect("L01:membership_payment_index")
                
                # ✅ Step 1: Member requests cancellation (only if no pending books)
                cancelled_status = StatusMaster.objects.filter(status_code__iexact="APP_CANCEL_REQ").first()
                if not cancelled_status:
                    raise Exception("Cancelled status not found in StatusMaster.")

                membership.status = cancelled_status
                membership.updated_by = user_id
                membership.remarks = "Cancelled by Member (Pending Librarian Approval)"
                
                # ✅ Update audit trail for member request
                membership.actionperformed = "Membership Cancellation Requested"  # What action was performed
                membership.reviewed = None  # Not reviewed yet
                membership.reviewed_at = None  # Not reviewed yet
                
                membership.save()
                
                messages.success(
                    request,
                    "सदस्यत्व रद्द करण्याची विनंती यशस्वीरित्या सबमिट केली आहे."
                    " ग्रंथालय प्रशासकाच्या मंजुरीनंतर तुमची ठेव परत केली जाईल."
                )
                return redirect("L01:membership_payment_index")

            elif action_by == "librarian":
                # ✅ Check if member has any pending books before approving cancellation
                has_pending_books = CirculationTransaction.objects.filter(
                    member=membership,
                    return_date__isnull=True
                ).exists()
                
                if has_pending_books:
                    # Count pending books for better error message
                    pending_count = CirculationTransaction.objects.filter(
                        member=membership,
                        return_date__isnull=True
                    ).count()
                    
                    return JsonResponse({
                        "success": False, 
                        "message": f"सदस्यत्व रद्द करणे शक्य नाही. सदस्याच्या नावावर अजूनही {pending_count} पुस्तक/पुस्तके परत केलेली नाहीत. कृपया प्रथम सर्व पुस्तके परत करा."
                    })
                
                # ✅ Check if already cancelled
                if membership.status.status_code == "APP_CANCEL":
                    return JsonResponse({
                        "success": False, 
                        "message": "हे सदस्यत्व आधीच रद्द केले गेले आहे."
                    })
                
                # ✅ Get the correct cancelled status
                cancelled_status = StatusMaster.objects.filter(status_code__iexact="APP_CANCEL").first()
                if not cancelled_status:
                    raise Exception("APP_CANCEL status not found in StatusMaster.")
                
                current_time = timezone.now()
                
                # ✅ Step 2: Librarian confirms cancellation & deactivates user + membership
                membership.status = cancelled_status
                membership.updated_by = user_id  # librarian's ID (for audit)
                membership.remarks = "Cancellation Approved by Librarian"
                
                # ✅ Update audit trail fields
                membership.actionperformed = "Membership Cancelled"  # What action was performed
                membership.reviewed = user_id  # Who reviewed/approved the cancellation
                membership.reviewed_at = current_time  # When it was reviewed/approved
                
                membership.save()
                
                # ✅ Deactivate the member's user account (not librarian's)
                member_user = CustomUser.objects.filter(username=membership.user_id).first()
                if member_user:
                    member_user.is_active = False
                    member_user.save()
                
                return JsonResponse({
                    "success": True, 
                    "message": "सदस्यत्व यशस्वीपणे रद्द केले गेले आहे. ठेव रकमेची परतावा प्रक्रिया सुरू केली गेली आहे."
                })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "membership_form_cancellation"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"Error: {e}")
        messages.error(request, "Oops...! Something went wrong!")
        return redirect("L01:membership_payment_index")

# circulation
@login_required
def bar_code_index(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        library_code = request.session.get('library_db', None)
        username = request.session.get('username', None)
        user_id = request.session["user_id"]
        role_id = request.session["role_id"]

        if request.method == "GET":
            
            accessions = BookAccession.objects.select_related(
                'catalogue',  # 👈 pulls in BookCatalog
                'supplier', 'currency', 'funding_source',
                'condition_at_entry', 'location', 'status'
            ).filter(status__status_id=4).order_by("-accession_id")
            
            for a in accessions:
                a.encrypted_id = enc(str(a.accession_id))
            
            # barcode for bulk download
            
            # circulation_accessions = list(
            #     CirculationCopyStatus.objects
            #     .filter(accession_no__isnull=False)
            #     .values('id', 'accession_no', 'accession_id')
            # )
            
            circulation_accessions = list(
                CirculationCopyStatus.objects
                .filter(accession_no__isnull=False)
                .annotate(
                    accession_no_int=Cast('accession_no', IntegerField())
                )
                .order_by('accession_no_int')
                .values('id', 'accession_no', 'accession_id')
            )

            for circ in circulation_accessions:
                circ['accession_encrypted_no'] = enc(str(circ['accession_no']))

            for circ in circulation_accessions:
                circ['accession_encrypted_no'] = enc(str(circ['accession_no']))
                
            subjectNames = SubjectTypeMaster.objects.filter(is_active=1)
            for sub in subjectNames:
                sub.subjectIdEnc = enc(str(sub.id))
                
            location_list = ResourceLocationMaster.objects.filter(is_active=1)

            for loc in location_list:
                loc.locationEnc = enc(str(loc.location_id))
                loc.full_name = f"{loc.location_code} - {loc.location_name}"
        
            return render(
                request,
                "L01/Circulation/barcodeindex.html",
                {
                    "MEDIA_URL": settings.MEDIA_URL,
                    "library_code": library_code,
                    "accessions": accessions,
                    "circulation_accessions": circulation_accessions,
                    "subjectNames": subjectNames,
                    "location_list": location_list,
                    
                }
            )
        
        # In your barcode generation code, replace Marathi text with English:
        if request.method == "POST":
            if library_code:
                library_details = tbl_librarymasterL01.objects.filter(library_code=library_code).first()
                if not library_details:
                    return HttpResponseBadRequest("Invalid library code")
                library_location = library_details.location
            else:
                return HttpResponseBadRequest("Missing library code")

            # Decode accession numbers
            from_acc = dec(str(request.POST.get("from_accession")))
            to_acc = dec(str(request.POST.get("to_accession")))
            subject_id = request.POST.get("subject_id")
            location_id = request.POST.get("location_id")

            if subject_id:
                subject_id = dec(str(subject_id))
            if location_id:
                location_id = dec(str(location_id))

            if not from_acc or not to_acc:
                return JsonResponse({"error": "Missing fields"}, status=400)

            # Get circulation records
            circulation_qs = (
                CirculationCopyStatus.objects
                .select_related("bookcatalog", "bookcatalog__subject", "shelf_location")
                .annotate(barcode_int=Cast("barcode", IntegerField()))
                .filter(barcode_int__range=(int(from_acc), int(to_acc)))
                .order_by("barcode_int")
            )
            if subject_id:
                circulation_qs = circulation_qs.filter(bookcatalog__subject_id=subject_id)
            if location_id:
                circulation_qs = circulation_qs.filter(shelf_location=location_id)
            if not circulation_qs.exists():
                return JsonResponse({"error": "No matching records found"}, status=404)

            # PDF setup
            pdf_buffer = BytesIO()
            # TSC TE244: 50mm × 25mm, 203 DPI → 141.73 x 70.87 points
            label_width_pt = 50 * 2.83465
            label_height_pt = 25 * 2.83465

            p = canvas.Canvas(pdf_buffer, pagesize=(label_width_pt, label_height_pt))

            # Font path for Marathi
            font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSerifDevanagari-Bold.ttf")

            for idx, record in enumerate(circulation_qs):
                accession_no = record.accession_no or "UNKNOWN"
                
                # Get subject name and location
                subject_name = ""
                if record.bookcatalog and record.bookcatalog.subject:
                    subject_name = record.bookcatalog.subject.subjectNameMarathi or ""
                
                location_name = ""
                if record.shelf_location:
                    location_name = record.shelf_location.location_name or ""

                # Create label image (50x25 mm → 400x200 px at 203 DPI)
                img_width = 400
                img_height = 200
                label_img = Image.new("RGB", (img_width, img_height), "white")
                draw = ImageDraw.Draw(label_img)

                # Load fonts with larger sizes
                try:
                    font_large = ImageFont.truetype(font_path, 28)
                    font_medium = ImageFont.truetype(font_path, 28)
                    font_small = ImageFont.truetype(font_path, 22)
                except:
                    font_large = ImageFont.load_default()
                    font_medium = ImageFont.load_default()
                    font_small = ImageFont.load_default()

                # Draw text content
                line1 = "नवी मुंबई महानगरपालिका"  # NMMC in Marathi
                line2 = f"{library_location} - {library_code}"
                
                # Generate barcode with larger size
                barcode_buffer = BytesIO()
                barcode_obj = Code128(accession_no, writer=ImageWriter())
                barcode_obj.write(barcode_buffer, {
                    "module_width": 0.32,
                    "module_height": 35,
                    "quiet_zone": 1.5,
                    "write_text": False,
                    "dpi": 203,
                })
                barcode_buffer.seek(0)
                barcode_img = Image.open(barcode_buffer)
                if barcode_img.mode != "RGB":
                    barcode_img = barcode_img.convert("RGB")
                barcode_img = barcode_img.resize((350, 50), Image.Resampling.LANCZOS)

                # Line 4: Accession no - Subject - Location
                line4 = f"{accession_no} - {subject_name} - {location_name}"

                # Calculate text dimensions
                bbox1 = draw.textbbox((0, 0), line1, font=font_large)
                text_height1 = bbox1[3] - bbox1[1]
                
                bbox2 = draw.textbbox((0, 0), line2, font=font_medium)
                text_height2 = bbox2[3] - bbox2[1]
                
                bbox4 = draw.textbbox((0, 0), line4, font=font_small)
                text_height4 = bbox4[3] - bbox4[1]

                # Calculate total content height and starting position for perfect centering
                total_content_height = text_height1 + 8 + text_height2 + 10 + 50 + 8 + text_height4
                start_y = (img_height - total_content_height) // 2

                # Line 1: Center aligned horizontally
                text_width1 = bbox1[2] - bbox1[0]
                line1_x = (img_width - text_width1) // 2
                line1_y = start_y
                draw.text((line1_x, line1_y), line1, font=font_large, fill="black")

                # Line 2: Center aligned horizontally
                text_width2 = bbox2[2] - bbox2[0]
                line2_x = (img_width - text_width2) // 2
                line2_y = line1_y + text_height1 + 8
                draw.text((line2_x, line2_y), line2, font=font_medium, fill="black")

                # Line 3: Barcode - perfectly centered horizontally
                barcode_x = (img_width - 350) // 2
                barcode_y = line2_y + text_height2 + 10
                label_img.paste(barcode_img, (barcode_x, barcode_y))

                # Line 4: Center aligned horizontally
                text_width4 = bbox4[2] - bbox4[0]
                line4_x = (img_width - text_width4) // 2
                line4_y = barcode_y + 50 + 8
                draw.text((line4_x, line4_y), line4, font=font_small, fill="black")

                # Convert label image to PDF page
                temp_buffer = BytesIO()
                label_img.save(temp_buffer, format="PNG", dpi=(203, 203))
                temp_buffer.seek(0)
                p.drawImage(ImageReader(temp_buffer), 0, 0, width=label_width_pt, height=label_height_pt)

                if idx < len(circulation_qs) - 1:
                    p.showPage()

            p.save()
            pdf_buffer.seek(0)
            filename = f"{from_acc}_to_{to_acc}_barcodes.pdf"
            response = HttpResponse(pdf_buffer, content_type="application/pdf")
            response['Content-Disposition'] = f'inline; filename="{filename}"'
            return response
        
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "library_list"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, "Oops...! Something went wrong!")
        return render(request, "L01/Circulation/barcodeindex.html", {})

@login_required
def generate_barcode(request):
    try:
        # Database connection
        Db.closeConnection()
        m = Db.get_connection()
        cursor = m.cursor()

        # Session variables
        library_code = request.session.get('library_db', None)
        username = request.session.get('username', None)
        user_id = request.session.get("user_id")
        role_id = request.session.get("role_id")

        # 🔹 GET: Generate barcode
        if request.method == "GET":
            
            place = request.GET.get("place")
            
            if place:
                
                books_qs = CirculationCopyStatus.objects.filter(shelf_location__isnull=True).select_related('accession', 'accession__catalogue')
                books = []
                for b in books_qs:
                    accession_no = b.accession.accession_no if b.accession else ''
                    title = b.accession.catalogue.title if b.accession and b.accession.catalogue else ''
                    label = f"{accession_no} - {title}"
                    books.append({'accession_id': b.accession_id, 'accession_no': b.accession_no, 'label': label})

                # Fetch active locations
                locations_qs = ResourceLocationMaster.objects.filter(is_active=1)
                locations = [{'location_id': loc.location_id, 'location_name': loc.location_name} for loc in locations_qs]

                return JsonResponse({'books': books, 'locations': locations})
            
            else:
                
                if library_code: 
                    library_details = tbl_librarymasterL01.objects.filter(library_code=library_code).first() 
                    library_location = library_details.location 
                else:
                    library_location = None
                    return HttpResponseBadRequest("Invalid library code")
                
                accession_id = request.GET.get("accessionid")
                if not accession_id:
                    return HttpResponseBadRequest("Missing accession ID")

                try:
                    decrypted_id = dec(accession_id)
                except Exception:
                    return HttpResponseBadRequest("Invalid accession ID")

                accession = get_object_or_404(BookAccession, accession_id=decrypted_id)
                accession_number = accession.accession_no or "UNKNOWN"
                subject_name = accession.catalogue.subject.subjectNameMarathi or ""

                # ✅ Generate barcode
                CODE = barcode.get_barcode_class('code128')
                writer_options = {
                    "module_width": 0.4,
                    "module_height": 8,
                    "quiet_zone": 2,
                    "font_size": 6,
                    "text_distance": 2.5,
                    "write_text": False,  # we’ll handle text manually
                }

                buffer = BytesIO()
                CODE(accession_number, writer=ImageWriter()).write(buffer, options=writer_options)
                buffer.seek(0)
                barcode_image = Image.open(buffer)

                # ✅ Font path
                font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSansDevanagari-Regular.ttf")

                # ✅ Load fonts with proper sizes
                try:
                    marathi_font = ImageFont.truetype(font_path, 18)
                    number_font = ImageFont.truetype(font_path, 16)
                except:
                    marathi_font = ImageFont.load_default()
                    number_font = ImageFont.load_default()

                # ✅ Text setup
                marathi_text = "नवी मुंबई महानगरपालिका"
                marathi_text = f"{marathi_text} - {library_location}"
                
                number_text = f"{accession_number} - {subject_name}"

                # ✅ Measure text widths
                draw_temp = ImageDraw.Draw(barcode_image)
                marathi_width = draw_temp.textlength(marathi_text, font=marathi_font)
                number_width = draw_temp.textlength(number_text, font=number_font)

                # ✅ Calculate total height (Marathi + small gap + barcode + small gap + number)
                top_margin = 5
                gap_above_barcode = 6
                gap_below_barcode = 6
                new_height = top_margin + marathi_font.size + gap_above_barcode + barcode_image.height + gap_below_barcode + number_font.size

                new_image = Image.new("RGB", (barcode_image.width, new_height), "white")
                draw = ImageDraw.Draw(new_image)

                # ✅ Marathi text (top)
                text_x = (barcode_image.width - marathi_width) // 2
                text_y = top_margin
                draw.text((text_x, text_y), marathi_text, fill="black", font=marathi_font)

                # ✅ Barcode (middle)
                barcode_y = text_y + marathi_font.size + gap_above_barcode
                new_image.paste(barcode_image, (0, barcode_y))

                # ✅ Accession number + subject (bottom)
                number_x = (barcode_image.width - number_width) // 2
                number_y = barcode_y + barcode_image.height + gap_below_barcode
                draw.text((number_x, number_y), number_text, fill="black", font=number_font)

                # ✅ Encode to Base64
                output = BytesIO()
                new_image.save(output, format="PNG")
                barcode_image_b64 = base64.b64encode(output.getvalue()).decode("utf-8")

                return JsonResponse({
                    "accession_no": accession_number,
                    "subject_name": subject_name,
                    "barcode_image": barcode_image_b64,
                })

        # 🔹 POST: Save barcode
        elif request.method == "POST":
            
            place = request.POST.get("place")
            
            if place:
                
                accession_no = request.POST.get("accession_id")
                shelf_location = request.POST.get("shelf_location")
                
                # Validate input
                if not accession_no or not shelf_location:
                    messages.error(request, "Accession आणि Shelf Location आवश्यक आहेत.")
                    return redirect("L01:bar_code_index")

                try:
                    with transaction.atomic():
                        # Fetch the circulation copy by accession_no
                        ccs = CirculationCopyStatus.objects.get(accession_no=accession_no)
                        
                        # Update shelf location and audit fields
                        ccs.shelf_location_id = shelf_location
                        ccs.updated_by = user_id
                        ccs.save(update_fields=['shelf_location', 'updated_at', 'updated_by'])

                    messages.success(request, "Shelf location यशस्वीरित्या जतन झाली आहे.")
                    return redirect("L01:bar_code_index")

                except CirculationCopyStatus.DoesNotExist:
                    messages.error(request, "संबंधित Circulation copy सापडली नाही.")
                    return redirect("L01:bar_code_index")

                except Exception as e:
                    messages.error(request, f"Shelf location जतन करताना त्रुटी: {str(e)}")
                    return redirect("L01:bar_code_index")
            
            else:
                
                # Form fields
                accession_id_enc = request.POST.get("accession_id")
                accession_no = request.POST.get("accession_no")
                barcode_no = request.POST.get("barcode_no")
                user = request.session.get("user_id")

                if not accession_id_enc or not barcode_no:
                    messages.error(request, "काही आवश्यक माहिती गायब आहे!")
                    return redirect("L01:bar_code_index")

                try:
                    # Decrypt accession ID
                    accession_id = dec(str(accession_id_enc))
                    accession = get_object_or_404(BookAccession, accession_id=accession_id)

                    # Get status objects
                    processing_status = status_master.objects.get(status_id=7)  # Inventory: Ready
                    current_status = status_master.objects.get(status_id=13)    # Circulation: On-Shelf

                    # Wrap in transaction
                    with transaction.atomic():
                        # 1️⃣ Insert into CirculationCopyStatus
                        CirculationCopyStatus.objects.create(
                            accession=accession,
                            bookcatalog=accession.catalogue,
                            accession_no=accession_no,
                            barcode=barcode_no,
                            processing_status=processing_status,
                            current_status=current_status,
                            date_processed=timezone.now().date(),
                            created_by=user,
                        )

                        # 2️⃣ Update BookAccession status
                        accession.status_id = 5
                        accession.updated_by = user
                        accession.save(update_fields=['status_id', 'updated_at', 'updated_by'])

                    messages.success(request, "बारकोड यशस्वीरित्या जतन झाला!")
                    return redirect("L01:bar_code_index")

                except Exception as e_post:
                    # Optional: log error in DB
                    messages.error(request, f"त्रुटी आली: {str(e_post)}")
                    return redirect("L01:bar_code_index")

        # Unsupported method
        else:
            return HttpResponseBadRequest("Invalid request method")

    except Exception as e:
        # Global exception catch
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "generate_barcode"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, "Oops...! Something went wrong!")
        return redirect("L01:bar_code_index")

# Issue Book / Return Book
@login_required
def issue_return_book_create(request):
    try:
        Db.closeConnection()
        m = Db.get_connection()
        cursor = m.cursor()
        library_code = request.session.get('library_db', None)
        username = request.session.get('username', None)
        user_id = request.session["user_id"]
        role_id = request.session["role_id"]

        if request.method == "GET":
            # Get only active MEMBERS (role_id=3)
            active_member_usernames = CustomUser.objects.filter(
                is_active=True,
                role_id=3  # Only Member role
            ).values_list("username", flat=True)

            # Status IDs to exclude: Cancellation Requested (10), Membership cancelled (12)
            EXCLUDED_STATUS_IDS = [10, 12]
            
            members = (
                MembershipDetails.objects
                .filter(isactive=1)                       # active members only
                .exclude(membership__id=8)               # exclude Practitioner Branch
                .filter(user_id__in=active_member_usernames)  # only Member role users
                .exclude(status__id__in=EXCLUDED_STATUS_IDS)  # FIXED: status__status_id
            )
            
            for m in members:
                m.member_encrypted_id = enc(str(m.id))
                
            circulation = CirculationCopyStatus.objects.filter(shelf_location__isnull=False)
            
            for circ in circulation:
                circ.circ_encrypted_barcode = enc(str(circ.barcode))
                circ.circ_encrypted_id = enc(str(circ.id))
            
            # barcode for returned books
            barcodes_qs = CirculationCopyStatus.objects.filter(current_status=14).values_list('barcode', flat=True)
            circulationTran = CirculationTransaction.objects.select_related(
                'catalog', 'member', 'circulation', 'return_condition'
            ).filter(
                barcode__in=barcodes_qs,
                return_condition_id=14  # assuming 14 is the PK for return_condition
            )
            
            for tran in circulationTran:
                tran.circulation_encrypted_barcode = enc(str(tran.barcode))
            
            # book Info
            bookcatelog = BookCatalog.objects.all()
            
            for bc in bookcatelog:
                bc.bc_encryp_id = enc(str(bc.cat_ref_num))
                
            # status dropdown for return condition
            status = status_master.objects.filter(status_type='circulation transaction', is_active=1)
            for st in status:
                st.status_encrypted_id = enc(str(st.status_id))
                
            return render(
                request,
                "L01/Circulation/issue_return_book_create.html",
                {
                    "MEDIA_URL": settings.MEDIA_URL,
                    "members": members,
                    "circulation": circulation,
                    "bookcatelog": bookcatelog,
                    "circulationTran": circulationTran,
                    "status_list": status,
                }
            )
        
        if request.method == "POST":

            # Check if it is issue or return
            if request.POST.get("issue") == "1":
                # ------------------- ISSUE LOGIC -------------------
                member_id = dec(str(request.POST.get("member_id")))
                barcode_id = dec(str(request.POST.get("barcode_id")))

                circ = CirculationCopyStatus.objects.select_related('accession', 'bookcatalog').get(barcode=barcode_id)
                member = MembershipDetails.objects.get(id=member_id)

                accession = circ.accession
                circulation = circ
                bookcatalog = circ.bookcatalog
                membercode = member.membership_code

                issue_date = request.POST.get("issue_date")
                due_date = request.POST.get("due_date")
                notes = request.POST.get("notes")

                return_condition_status = status_master.objects.get(status_id=14)  # default issued condition

                # Save transaction
                transaction = CirculationTransaction.objects.create(
                    catalog=bookcatalog,
                    accession=accession,
                    circulation=circulation,
                    member=member,
                    barcode=circ.barcode,
                    issue_date=issue_date,
                    due_date=due_date,
                    remarks=notes,
                    membership_code=membercode,
                    return_condition=return_condition_status,
                    issued_by=user_id,
                    created_by=user_id
                )

                # Update copy status
                circ.current_status = return_condition_status
                circ.save()

                messages.success(request, "ग्रंथ यशस्वीरित्या निर्गम केला गेला आहे!")

            elif request.POST.get("return") == "1":
                barcode_id = dec(str(request.POST.get("barcode_id")))

                trans_obj = CirculationTransaction.objects.filter(
                    barcode=barcode_id,
                    return_date__isnull=True
                ).first()

                if not trans_obj:
                    messages.error(request, "सक्षम पुस्तके परत करण्यासाठी रेकॉर्ड सापडले नाही!")
                    return redirect('L01:issue_return_book_create')

                # Book condition and fine
                condition_name = request.POST.get("book_condition")
                fine_amount = float(request.POST.get("fine_amount", 0) or 0)
                adjusted_fine = float(request.POST.get("adjusted_amount", 0) or 0)  # <-- new
                book_price_amount = float(request.POST.get("book_price_amount", 0) or 0)
                total_amount = float(request.POST.get("total_amount", 0))
                fine_Breakdown = float(request.POST.get("fine_Breakdown", 0))
                notes = request.POST.get("notes", "")
                payment_method = request.POST.get("payment_method", "Cash")

                # Determine new circulation status
                if condition_name == '17':  # Good
                    new_copy_status = status_master.objects.get(status_id=13)
                elif condition_name in ['18', '19']:  # Torn or Lost
                    new_copy_status = status_master.objects.get(status_id=19)
                else:
                    new_copy_status = status_master.objects.get(status_id=13)

                # Wrap all DB operations in a transaction
                try:
                    with db_transaction.atomic():
                        # Update CirculationTransaction
                        
                        final_fine = adjusted_fine if adjusted_fine > 0 else fine_amount
                        
                        trans_obj.return_date = timezone.now().date()
                        trans_obj.return_condition = new_copy_status
                        trans_obj.fine_amount = fine_amount
                        trans_obj.adjusted_fine = adjusted_fine  # <-- save adjusted fine
                        trans_obj.book_fine_amount = book_price_amount
                        trans_obj.total_fine = total_amount
                        trans_obj.days_overdue_count = fine_Breakdown
                        trans_obj.fine_status = "Paid" if fine_amount > 0 else "None"
                        trans_obj.fine_paid_date = timezone.now().date()
                        trans_obj.transaction_status = "Success"
                        trans_obj.transaction_type = "Offline"
                        trans_obj.received_by = user_id
                        trans_obj.remarks = notes
                        trans_obj.updated_by = user_id
                        trans_obj.save()

                        # Update book copy status
                        circ = trans_obj.circulation
                        circ.current_status = new_copy_status
                        circ.save()

                        # Payment logic
                        fine_nonzero = fine_amount > 0.001
                        book_nonzero = book_price_amount > 0.001

                        if fine_nonzero and not book_nonzero:
                            payment_type = "Fine"
                        elif fine_nonzero and book_nonzero:
                            payment_type = "Fine and Book Price"
                        elif not fine_nonzero and book_nonzero:
                            payment_type = "Loss"
                        else:
                            payment_type = None

                        if total_amount > 0 and payment_type:
                            PaymentDetails.objects.create(
                                membership=trans_obj.member,
                                circulation_transaction=trans_obj,
                                payment_mode="Offline",
                                payment_method=payment_method,
                                payment_type=payment_type,
                                fine_amount=final_fine,              
                                book_fine_amount=book_price_amount,  
                                user_id=trans_obj.member.user_id,
                                membership_code=trans_obj.membership_code,
                                payment_date=timezone.now().date(),
                                created_by=user_id,
                                updated_by=user_id,
                                status=StatusMaster.objects.get(id=5),
                            )

                    messages.success(request, "पुस्तक यशस्वीरित्या परत झाले आहे!")

                except Exception as e:
                    # Rollback automatically on any error
                    messages.error(request, f"त्रुटी: {str(e)}. कृपया पुन्हा प्रयत्न करा!")
                    return redirect('L01:issue_return_book_create')
            
            return redirect('L01:issue_return_book_create')
        
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "library_list"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, "Oops...! Something went wrong!")
        return render(request, "L01/Circulation/issue_return_book_create.html", {})

@csrf_exempt
def get_member_details(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method"})

    member_id = dec(str(request.POST.get("member_id")))
    if not member_id:
        return JsonResponse({"success": False, "error": "Member ID missing"})

    try:
        Db.closeConnection()
        m = Db.get_connection()
        cursor = m.cursor()
        library_code = request.session.get('library_db', None)
        username = request.session.get('username', None)
        user_id = request.session["user_id"]
        role_id = request.session["role_id"]

        # 🔴 FIRST: Check if member has cancellation status
        try:
            membership = MembershipDetails.objects.select_related('status').get(id=member_id)
            
            # FIXED: Use status.id instead of status.status_id
            if membership.status and membership.status.id in [10, 12]:  # 10=Cancellation Requested, 12=Membership cancelled
                return JsonResponse({
                    "success": False,
                    "error": f"Member has '{membership.status.status_name}'. Cannot issue books."
                })
        except MembershipDetails.DoesNotExist:
            return JsonResponse({"success": False, "error": "Member not found"})

        # Fetch DocumentDetails for this member where document_id=1
        documents = DocumentDetails.objects.filter(
            membership_id=member_id,
            document_id=1,
            isactive=1
        ).select_related(
            'membership__membership',       # MembershipMaster
            'membership__member_type',      # parameter_master_L01
            'membership__status',           # Include status
            'document'
        )

        document_list = []
        for doc in documents:
            membership = doc.membership
            membership_master = membership.membership
            member_type_param = membership.member_type
            
            # Format dates
            from_date = membership.from_date.strftime("%d %b %Y") if membership.from_date else '-'
            to_date = membership.to_date.strftime("%d %b %Y") if membership.to_date else '-'
            
            # Get file URL using FileStorageService
            file_url = "#"
            if doc.file_path:
                file_url = file_storage_service.get_file_url(doc.file_path)

            document_list.append({
                "membership_id": membership.id,
                "full_name": f"{membership.first_name} {membership.middle_name or ''} {membership.last_name}",
                "email": membership.email,
                "mobile_no": membership.mobile_no,
                "membership_code": membership_master.membership_code if membership_master else '',
                "membership_type": membership_master.membership_type_en if membership_master else '',
                "membership_item": membership_master.item if membership_master else '',
                "membership_days": membership_master.days if membership_master else '',
                "member_type_name": member_type_param.parameter_name if member_type_param else '',
                "membership_fromDate": from_date,
                "membership_toDate": to_date,
                "membership_duration": membership.membership_duration,
                "membership_status": membership.status.id if membership.status else None,  # FIXED: status.id
                "membership_status_name": membership.status.status_name if membership.status else None,
                "document_id": doc.document.id,
                "document_name": doc.document.document_name if doc.document else '',
                "file_name": doc.file_name,
                "file_path": doc.file_path,
                "file_url": file_url,
            })
            
        # Count issued books (only books not returned yet)
        issued_books_count = CirculationTransaction.objects.filter(
            member_id=member_id,
            return_date__isnull=True
        ).count()

        return JsonResponse({
            "success": True,
            "documents": document_list,
            "MEDIA_URL": settings.MEDIA_URL,
            "issued_books_count": issued_books_count,
        })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else "library_list"
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        return JsonResponse({"success": False, "error": f"Internal server error: {str(e)}"})
    

@csrf_exempt  # Remove this if you’re using CSRF token in AJAX
def get_book_details(request):
    if request.method == "POST":
        try:
            # --- Session Variables ---
            library_code = request.session.get('library_db', None)
            username = request.session.get('username', None)
            user_id = request.session.get("user_id")
            role_id = request.session.get("role_id")

            # --- Barcode / Encrypted ID ---
            barcode_id = request.POST.get("barcode_id")
            returnProcess = request.POST.get("returnProcess")
            
            if not barcode_id:
                return JsonResponse({"success": False, "message": "Barcode ID missing."}, status=400)
            
            if not returnProcess:

                # --- Decrypt if needed ---
                try:
                    barcode_id = dec(barcode_id)
                except Exception:
                    # If not encrypted, just use it as-is
                    pass

                # --- Fetch Book Details ---
                circ = (
                    CirculationCopyStatus.objects
                    .select_related("bookcatalog", "shelf_location", "current_status")
                    .get(barcode=barcode_id)  # <--- match by barcode
                )
                
                # Get front page URL using FileStorageService
                front_page_url = "#"
                if circ.bookcatalog and circ.bookcatalog.front_page_photo:
                    front_page_url = file_storage_service.get_file_url(circ.bookcatalog.front_page_photo)

                data = {
                    "success": True,
                    "MEDIA_URL": settings.MEDIA_URL,
                    "book": {
                        "title": circ.bookcatalog.title if circ.bookcatalog else "-",
                        "author": circ.bookcatalog.author if circ.bookcatalog else "-",
                        "isbn_issn": circ.bookcatalog.isbn_issn if circ.bookcatalog else "-",
                        "pages": circ.bookcatalog.pages if circ.bookcatalog else "-",
                        "language": circ.bookcatalog.language if circ.bookcatalog else "-",
                        "front_page_photo": front_page_url,  # Only this line
                        "location_name": circ.shelf_location.location_name if circ.shelf_location else "-",
                        "status_name": circ.current_status.status_name if circ.current_status else "Unknown",
                        "status_color": circ.current_status.status_color if circ.current_status else "#999",
                    },
                }
                return JsonResponse(data)
            
            else:   
                
                barcode_id = dec(barcode_id)
                
                try:
                    circ = CirculationTransaction.objects.select_related(
                        "catalog", "member", "circulation", "return_condition","accession"
                    ).get(barcode=barcode_id,
                          return_condition__status_id=14)

                    book = circ.catalog
                    member = circ.member
                    membership_master = member.membership
                    price = circ.accession.price if circ.accession else 0
                    today = timezone.now().date()

                    # Format due date nicely (e.g., 10 Jan 2025)
                    due_date_str = (
                        circ.due_date.strftime("%d %b %Y") if circ.due_date else "-"
                    )

                    # Calculate overdue days
                    days_overdue = 0
                    status_name = (
                        circ.return_condition.status_name if circ.circulation else "Unknown"
                    )

                    if circ.due_date and today > circ.due_date:
                        days_overdue = (today - circ.due_date).days
                        status_name = "Overdue"
                        
                    # Calculate fine amount
                    fine_per_day = float(membership_master.fine or 0)  # default 0 if None
                    total_fine = days_overdue * fine_per_day

                    # Fetch member profile photo (Document ID = 1)
                    profile_doc = (
                        DocumentDetails.objects.filter(
                            membership=member, document_id=1, isactive=1
                        )
                        .order_by("-created_at")
                        .first()
                    )
                    # Get URLs using FileStorageService
                    front_page_url = "#"
                    profile_photo_url = "#"
                    
                    if book and book.front_page_photo:
                        front_page_url = file_storage_service.get_file_url(book.front_page_photo)
                    
                    if profile_doc and profile_doc.file_path:
                        profile_photo_url = file_storage_service.get_file_url(profile_doc.file_path)

                    # Member full name
                    member_name = " ".join(
                        filter(None, [member.first_name, member.middle_name, member.last_name])
                    )

                    # Construct return data
                    data = {
                        "success": True,
                        "MEDIA_URL": settings.MEDIA_URL,
                        "book": {
                            "title": book.title if book else "-",
                            "author": book.author if book else "-",
                            "isbn_issn": book.isbn_issn if book else "-",
                            "front_page_photo": front_page_url,  # Only this line
                            "status_name": status_name,
                            "due_date": due_date_str,
                            "issue_date": (
                                circ.issue_date.strftime("%d %b %Y")
                                if circ.issue_date
                                else "-"
                            ),
                            "days_overdue": days_overdue,
                            "fine_amount": total_fine,
                            "price": price,
                        },
                        "member": {
                            "member_name": member_name,
                            "member_code": member.membership_code or "-",
                            "profile_photo": profile_photo_url,  # Only this line
                            "mobile_no": member.mobile_no or "-",
                        },
                    }

                    return JsonResponse(data)
                
                except ObjectDoesNotExist:
                    return JsonResponse({
                        "success": False,
                        "error": "No circulation transaction found for this barcode."
                    })

        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name if tb else "get_book_details"

            try:
                # Log error in DB (if your stored procedure exists)
                Db.closeConnection()
                m = Db.get_connection()
                cursor = m.cursor()
                cursor.callproc("stp_error_log", [fun, str(e), library_code])
            except Exception:
                pass

            print(f"Error in {fun}: {e}")
            messages.error(request, "Oops...! Something went wrong!")
            return JsonResponse({"success": False, "message": "Internal Server Error"}, status=500)

    else:
        return JsonResponse({"success": False, "message": "Invalid request method"}, status=405)

@csrf_exempt
def get_book_circulation_status(request):
    if request.method == "POST":
        book_id = dec(str(request.POST.get("book_id")))
        if not book_id:
            return JsonResponse({"error": "Missing book_id"}, status=400)

        records = (
            CirculationCopyStatus.objects.filter(bookcatalog_id=book_id)
            .select_related("bookcatalog", "current_status", "accession")
            .values(
                "id",
                "barcode",
                "bookcatalog__cat_ref_num",
                "bookcatalog__title",
                "bookcatalog__author",
                "bookcatalog__classification_number",
                "current_status__status_name",
                "accession__accession_no"
            )
        )

        return JsonResponse(list(records), safe=False)

@csrf_exempt
def circulation_transaction_details(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        library_code = request.session.get('library_db', None)
        username = request.session.get('username', None)
        user_id = request.session.get("user_id")
        role_id = request.session.get("role_id")
        
        member_id = request.GET.get("member_id")
        
        barcodeName = request.GET.get("barcodeName")
        
        if not member_id:
            if barcodeName:
                barcode = request.GET.get("barcode")
            else:
                barcode = dec(str(request.GET.get("barcode")))
        
        
        if member_id:

            transactions = (
                CirculationTransaction.objects
                .select_related("catalog", "accession", "return_condition")
                .filter(
                    member_id=member_id,
                    return_condition__status_id=14
                )
                .values(
                    "catalog__title",
                    "accession__accession_no",
                    "issue_date",
                    "due_date",
                    "return_condition__status_name"
                )
            )

            transaction_list = [
                {
                    "title": t["catalog__title"],
                    "accession_no": t["accession__accession_no"],
                    "issue_date": t["issue_date"].strftime("%Y-%m-%d") if t["issue_date"] else None,
                    "due_date": t["due_date"].strftime("%Y-%m-%d") if t["due_date"] else None,
                    "status_name": t["return_condition__status_name"],
                }
                for t in transactions
            ]

            return JsonResponse({"success": True, "transactions": transaction_list})
        
        if barcode:
            
            transactions = (
                CirculationTransaction.objects
                .select_related("catalog", "accession", "member", "return_condition")
                .filter(barcode=barcode,
                        return_condition__status_id=14)
                .values(
                    "catalog__title",
                    "accession__accession_no",
                    "issue_date",
                    "due_date",
                    "return_condition__status_name",
                    "member__first_name",
                    "member__last_name",
                    "member__membership_code"
                )
            )

            transaction_list = [
                {
                    "title": t["catalog__title"],
                    "accession_no": t["accession__accession_no"],
                    "issue_date": t["issue_date"].strftime("%Y-%m-%d") if t["issue_date"] else None,
                    "due_date": t["due_date"].strftime("%Y-%m-%d") if t["due_date"] else None,
                    "status_name": t["return_condition__status_name"],
                    "member_name": f"{t['member__first_name']} {t['member__last_name']}",
                    "membership_code": t["member__membership_code"]
                }
                for t in transactions
            ]

            return JsonResponse({"success": True, "transactions": transaction_list})
        
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else 'library_list'
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, 'Oops...! Something went wrong!')
    
# Palavees work
@login_required
def library_master_index_individual(request):
    library = None
    library_name = ""
    ward_name = ""
    location_name = ""
    accounting_code_name = ""
    library_details = []
    ward_details = []
    location_details = []
    accounting_code_details = []

    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()

    try:
        if request.method == "POST":
            # ------------------------
            # Form fields
            # ------------------------
            library_code = request.POST.get("library_code", "").strip()
            library_name = request.POST.get("library_name", "").strip()
            library_name_mar = request.POST.get("library_name_mar", "").strip()
            location_id = request.POST.get("location_id", "").strip()
            parent_ward_id = request.POST.get("parent_ward_id", "").strip()
            accounting_code_id = request.POST.get("accounting_code_id", "").strip()
            librarian_name = request.POST.get("librarian_name", "").strip()
            contact_email = request.POST.get("contact_email", "").strip()
            contact_phone = request.POST.get("contact_phone", "").strip()
            about_library = request.POST.get("about_library", "").strip()
            is_active = int(request.POST.get("is_active", 1))

            # ------------------------
            # Check if library code exists in L01 DB
            # ------------------------
            try:
                library = tbl_librarymasterL01.objects.get(library_code=library_code)
            except tbl_librarymasterL01.DoesNotExist:
                library = tbl_librarymasterL01()  # create new if not exists

            # ------------------------
            # Additional fields from template
            # ------------------------
            landing_page_link = request.POST.get("landing_page_link", "").strip()
            membership_page_link = request.POST.get("membership_page_link", "").strip()
            library_rules = request.POST.get("library_rules", "").strip()
            membership_rules = request.POST.get("membership_rules", "").strip()
            facebook_url = request.POST.get("facebook_url", "").strip()
            twitter_url = request.POST.get("twitter_url", "").strip()
            instagram_url = request.POST.get("instagram_url", "").strip()
            youtube_url = request.POST.get("youtube_url", "").strip()
            capacity = request.POST.get("capacity", "").strip()
            opening_hours = request.POST.get("opening_hours", "").strip()
            location_url = request.POST.get("location_url", "").strip()
            est_year = request.POST.get("est_year", "").strip()

            # ------------------------
            # Uploaded image (keep old if no new one)
            # ------------------------
            library_image = request.FILES.get("library_photo")
            image_url = getattr(library, "image_url", "")  # keep old image if no new one

            if library_image:
                library_folder = os.path.join(settings.MEDIA_ROOT, library_code, "library_images")
                os.makedirs(library_folder, exist_ok=True)
                image_path = os.path.join(library_folder, library_image.name)
                with open(image_path, "wb+") as f:
                    for chunk in library_image.chunks():
                        f.write(chunk)
                image_url = os.path.join(library_code, "library_images", library_image.name).replace("\\", "/")

            # ------------------------
            # Validation
            # ------------------------
            if not library_name:
                messages.error(request, "Library Name is required!")
            else:
                # ------------------------
                # Save/Update record in tbl_librarymasterL01
                # ------------------------
                library.library_code = library_code
                library.library_name = library_name
                library.library_name_mar = library_name_mar
                library.location = location_id
                library.parent_ward = parent_ward_id
                library.library_accounting_code = accounting_code_id
                library.librarian_name = librarian_name
                library.contact_email = contact_email
                library.contact_phone = contact_phone
                library.about_library = about_library
                library.landing_page_link = landing_page_link
                library.membership_page_link = membership_page_link
                library.library_rules = library_rules
                library.membership_rules = membership_rules
                library.facebook_url = facebook_url
                library.twitter_url = twitter_url
                library.instagram_url = instagram_url
                library.youtube_url = youtube_url
                library.capacity = capacity
                library.opening_hours = opening_hours
                library.location_url = location_url
                library.est_year = int(est_year) if est_year.isdigit() else None
                library.image_url = image_url
                library.is_active = bool(is_active)
                library.updated_by = request.user.id
                library.updated_at = timezone.now()
                library.save()

                # ------------------------
                # Sync with LibraryMaster (default DB)
                # ------------------------
                try:
                    # ------------------------
                    # Get related FK IDs from default DB
                    # ------------------------
                    location_obj_id = None
                    if location_id:
                        if not str(location_id).isdigit():
                            loc_obj = LibraryLocationMaster.objects.using("default").filter(location_name=location_id).first()
                            location_obj_id = loc_obj.location_id if loc_obj else None
                        else:
                            location_obj_id = int(location_id)

                    parent_ward_obj_id = None
                    if parent_ward_id:
                        if not str(parent_ward_id).isdigit():
                            ward_obj = WardMaster.objects.using("default").filter(ward_name=parent_ward_id).first()
                            parent_ward_obj_id = ward_obj.ward_id if ward_obj else None
                        else:
                            parent_ward_obj_id = int(parent_ward_id)

                    accounting_code_obj_id = None
                    if accounting_code_id:
                        if not str(accounting_code_id).isdigit():
                            ac_obj = WardMaster.objects.using("default").filter(accounting_code=accounting_code_id).first()
                            accounting_code_obj_id = ac_obj.ward_id if ac_obj else None
                        else:
                            accounting_code_obj_id = int(accounting_code_id)

                    # ------------------------
                    # Save to LibraryMaster (default DB)
                    # ------------------------
                    library_master, created = LibraryMaster.objects.using("default").get_or_create(
                        library_code=library_code,
                        defaults={
                            "library_name": library_name,
                            "library_name_mar": library_name_mar,
                            "location_id": location_obj_id,
                            "parent_ward_id": parent_ward_obj_id,
                            "library_accounting_code_id": accounting_code_obj_id,
                            "librarian_name": librarian_name,
                            "contact_email": contact_email,
                            "contact_phone": contact_phone,
                            "about_library": about_library,
                            "landing_page_link": landing_page_link,
                            "membership_page_link": membership_page_link,
                            "library_rules": library_rules,
                            "membership_rules": membership_rules,
                            "facebook_url": facebook_url,
                            "twitter_url": twitter_url,
                            "instagram_url": instagram_url,
                            "youtube_url": youtube_url,
                            "capacity": capacity,
                            "opening_hours": opening_hours,
                            "location_url": location_url,
                            "est_year": int(est_year) if est_year.isdigit() else None,
                            "image_url": image_url,
                            "is_active": bool(is_active),
                            "created_by": str(request.user.id),
                            "created_at": timezone.now(),
                            "updated_by": str(request.user.id),
                            "updated_at": timezone.now(),
                        }
                    )

                    if not created:
                        library_master.library_name = library_name
                        library_master.library_name_mar = library_name_mar
                        library_master.location_id = location_obj_id
                        library_master.parent_ward_id = parent_ward_obj_id
                        library_master.library_accounting_code_id = accounting_code_obj_id
                        library_master.librarian_name = librarian_name
                        library_master.contact_email = contact_email
                        library_master.contact_phone = contact_phone
                        library_master.about_library = about_library
                        library_master.landing_page_link = landing_page_link
                        library_master.membership_page_link = membership_page_link
                        library_master.library_rules = library_rules
                        library_master.membership_rules = membership_rules
                        library_master.facebook_url = facebook_url
                        library_master.twitter_url = twitter_url
                        library_master.instagram_url = instagram_url
                        library_master.youtube_url = youtube_url
                        library_master.capacity = capacity
                        library_master.opening_hours = opening_hours
                        library_master.location_url = location_url
                        library_master.est_year = int(est_year) if est_year.isdigit() else None
                        library_master.image_url = image_url
                        library_master.is_active = bool(is_active)
                        library_master.updated_by = str(request.user.id)
                        library_master.updated_at = timezone.now()
                        library_master.save(using="default")

                except Exception as ex:
                    import traceback
                    print("Error syncing LibraryMaster:", ex)
                    print(traceback.format_exc())

                messages.success(request, f"Library '{library_name}' saved successfully!")

            # ------------------------
            # After saving, reload page with updated values
            # ------------------------
            if library_code:
                library = tbl_librarymasterL01.objects.get(library_code=library_code)

        # ------------------------ GET request or after POST ----------------
        library_code = request.session.get("library_db", None)
        if library_code:
            library_details = LibraryMaster.objects.using("default").filter(is_active=1, library_code=library_code)
            for lilo in library_details:
                lilo.libraries = enc(lilo.library_code)
            if library_details.exists() and not library:
                library = library_details.first()
                library_name = library.library_name
        else:
            library_details = LibraryMaster.objects.using("default").filter(is_active=1)

        ward_code = request.session.get("ward_db", None)
        if ward_code:
            ward_details = WardMaster.objects.using("default").filter(is_active=1, ward_code=ward_code)
            for w in ward_details:
                w.wards = enc(w.ward_code)
            ward_name = ward_details.first().ward_name if ward_details.exists() else ""
        else:
            ward_details = WardMaster.objects.using("default").filter(is_active=1)

        location_code = request.session.get("location_db", None)
        if location_code:
            location_details = LibraryLocationMaster.objects.using("default").filter(is_active=1, location_code=location_code)
            for loc in location_details:
                loc.locations = enc(loc.location_code)
            location_name = location_details.first().location_name if location_details.exists() else ""
        else:
            location_details = LibraryLocationMaster.objects.using("default").filter(is_active=1)

        accounting_code = request.session.get("accounting_code_db", None)
        if accounting_code:
            accounting_code_details = WardMaster.objects.using("default").filter(is_active=1, accounting_code=accounting_code)
            for ac in accounting_code_details:
                ac.codes = enc(ac.accounting_code)
            accounting_code_name = accounting_code_details.first().accounting_code if accounting_code_details.exists() else ""
        else:
            accounting_code_details = WardMaster.objects.using("default").filter(is_active=1)

        return render(
            request,
            "L01/Master/library_master_index_individual.html",
            {
                "library": library,
                "libraries": library_details,
                "wards": ward_details,
                "locations": location_details,
                "accounting_codes": accounting_code_details,
                "library_name": library_name,
                "ward_name": ward_name,
                "location_name": location_name,
                "accounting_code_name": accounting_code_name,
                "MEDIA_URL": settings.MEDIA_URL,
            },
        )

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else 'library_list'
        cursor.callproc("stp_error_log", [fun, str(e), library_code])
        print(f"error: {e}")
        messages.error(request, 'Oops...! Something went wrong!')

@login_required
def user_master_index(request):
    try:
        if request.user.is_authenticated:
            global user, role_id
            user = request.user.id
            role_id = request.user.role_id

        if request.method == "GET":
            # 🔹 Fetch all user records with related role
            users = CustomUser.objects.all()

            # 🔹 Attach encrypted_id & role_name for each user
            for usr in users:
                usr.encrypted_id = enc(str(usr.id))

                try:
                    role = roles.objects.get(id=usr.id)
                    usr.role_name = role.role_name
                except roles.DoesNotExist:
                    usr.role_name = "N/A"

            return render(
                request,
                'L01/Master/user_master_index.html',
                {"users": users}
            )

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log", [fun, str(e), user])
        messages.error(request, 'Oops...! Something went wrong!')
        return redirect("user_master_index")

@login_required
def update_user_status(request):
    if request.method == "POST":
        try:
            print("Request body raw:", request.body)
            try:
                data = json.loads(request.body.decode("utf-8"))
            except (json.JSONDecodeError, UnicodeDecodeError) as e:
                print("JSON parse error:", e)
                return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)

            print("Parsed data:", data)

            encrypted_id = data.get("id")
            is_active = data.get("status")

            if encrypted_id is None or is_active is None:
                return JsonResponse({"success": False, "error": "Missing data"}, status=400)

            user_id = int(dec(encrypted_id))
            user = CustomUser.objects.get(id=user_id)
            user.is_active = bool(int(is_active))
            user.save()

            return JsonResponse({"success": True, "status": int(user.is_active)})

        except CustomUser.DoesNotExist:
            return JsonResponse({"success": False, "error": "User not found"}, status=404)
        except Exception as e:
            print("❌ Error:", e)
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

from django.db import transaction

@login_required
def user_create(request):
    try:
        user_id = request.user.id

        if request.method == "POST":
            with transaction.atomic():
                enc_id = request.POST.get("id")  # Hidden encrypted ID (for editing)
                username = request.POST.get("username", "").strip()
                first_name = request.POST.get("first_name", "").strip()
                last_name = request.POST.get("last_name", "").strip()
                full_name = f"{first_name} {last_name}".strip()
                email = request.POST.get("email", "").strip()
                address = request.POST.get("address", "").strip()
                phone = request.POST.get("mobile", "").strip()
                role_id = request.POST.get("role_id", "").strip()
                is_active = request.POST.get("is_active", 1)
                
                # Get password from hidden field (real password) or visible field
                password = request.POST.get("real_password", "").strip()
                if not password:  # Fallback to visible field
                    password = request.POST.get("password", "").strip()
                
                confirm_password = request.POST.get("confirm_password", "").strip()

                # 🔹 Validation
                validation_errors = []
                if not username:
                    validation_errors.append("User Name is required!")
                if not first_name:
                    validation_errors.append("First Name is required!")
                if not last_name:
                    validation_errors.append("Last Name is required!")
                if not email:
                    validation_errors.append("Email is required!")
                if not phone:
                    validation_errors.append("Phone Number is required!")
                if not role_id:
                    validation_errors.append("Role is required!")
                
                # 🔹 CRITICAL: Different validation for NEW vs EDIT users
                if not enc_id:  # NEW user
                    if not password:
                        validation_errors.append("Password is required for new user!")
                    elif password == "********":  # Should not happen with new fix
                        validation_errors.append("Please enter a valid password.")
                # For EDIT users, password can be empty (means don't change password)
                else:
                    # For EDIT: Check if password was provided (means user wants to change it)
                    if password:
                        # Check if password is just asterisks (masked)
                        if password == "********":
                            # Password is masked, treat as empty (no change)
                            password = ""
                        # Validate password rules if password is being changed
                        elif len(password) > 0:
                            # Validate password strength
                            if len(password) < 6:
                                validation_errors.append("Password must be at least 6 characters.")
                            if len(password) > 20:
                                validation_errors.append("Password cannot exceed 20 characters.")
                            if not any(c.isupper() for c in password):
                                validation_errors.append("Password must contain at least one uppercase letter.")
                            if not any(c.isdigit() for c in password):
                                validation_errors.append("Password must contain at least one number.")
                            if not any(c in "_@-" for c in password):
                                validation_errors.append("Password must contain at least one special character (_ @ -).")
                
                if validation_errors:
                    for error in validation_errors:
                        messages.error(request, error)
                    roles_list = roles.objects.exclude(role_type='member')
                    return render(request, "L01/Master/user_create.html", {
                        "roles_list": roles_list,
                        "post_data": request.POST,
                        "user_obj": None
                    })
                
                # Fetch selected role
                selected_role = roles.objects.get(pk=role_id)

                # Extract role_type
                user_type = selected_role.role_type

                # 🔹 Update existing user
                if enc_id:
                    user_pk = int(dec(enc_id))
                    user_obj = CustomUser.objects.get(pk=user_pk)

                    # Check for duplicate username (if changed)
                    if user_obj.username != username and CustomUser.objects.filter(username=username).exists():
                        messages.error(request, f"Username '{username}' already exists!")
                        roles_list = roles.objects.exclude(role_type='member')
                        return render(request, "L01/Master/user_create.html", {
                            "roles_list": roles_list,
                            "post_data": request.POST,
                            "user_obj": user_obj
                        })

                    # Check for duplicate email (if changed)
                    if user_obj.email != email and CustomUser.objects.filter(email=email).exists():
                        messages.error(request, f"Email '{email}' already exists!")
                        roles_list = roles.objects.exclude(role_type='member')
                        return render(request, "L01/Master/user_create.html", {
                            "roles_list": roles_list,
                            "post_data": request.POST,
                            "user_obj": user_obj
                        })

                    # Update user fields
                    user_obj.username = username
                    user_obj.first_name = first_name
                    user_obj.last_name = last_name
                    user_obj.full_name = full_name
                    user_obj.email = email
                    user_obj.address = address
                    user_obj.phone = phone
                    user_obj.is_active = bool(int(is_active))
                    user_obj.role_id = role_id
                    user_obj.user_type = selected_role.role_type 
                    user_obj.updated_by = user_id
                    user_obj.updated_at = timezone.now()

                    # ✅ Update password ONLY if provided and not empty
                    if password:
                        # Store plain password before hashing
                        plain_password = password
                        # Set hashed password for CustomUser
                        user_obj.set_password(password)
                        user_obj.save()
                        
                        # Update REAL password in password_storage
                        password_storage.objects.update_or_create(
                            user=user_obj,
                            defaults={'passwordText': plain_password}
                        )
                        
                        messages.success(request, f"User '{username}' updated with new password!")
                    else:
                        # Save without changing password
                        user_obj.save()
                        messages.success(request, f"User '{username}' updated successfully!")

                    return redirect("L01:user_master_index")

                # 🔹 Create new user (enc_id is empty)
                # Check for duplicate username
                if CustomUser.objects.filter(username=username).exists():
                    messages.error(request, f"Username '{username}' already exists!")
                    roles_list = roles.objects.exclude(role_type='member')
                    return render(request, "L01/Master/user_create.html", {
                        "roles_list": roles_list,
                        "post_data": request.POST,
                        "user_obj": None
                    })
                
                # Check for duplicate email
                if CustomUser.objects.filter(email=email).exists():
                    messages.error(request, f"Email '{email}' already exists!")
                    roles_list = roles.objects.exclude(role_type='member')
                    return render(request, "L01/Master/user_create.html", {
                        "roles_list": roles_list,
                        "post_data": request.POST,
                        "user_obj": None
                    })
                #Check for duplicate phone
                if CustomUser.objects.filter(phone=phone).exists():
                    messages.error(request, f"Phone number '{phone}' already exists!")
                    roles_list = roles.objects.exclude(role_type='member')
                    return render(request, "L01/Master/user_create.html", {
                        "roles_list": roles_list,
                        "post_data": request.POST,
                        "user_obj": None
                    })


                # Store plain password before creating user
                plain_password = password
                
                # Create user with hashed password
                new_user = CustomUser.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    is_active=bool(int(is_active)),
                    user_type=user_type
                )
                
                # Set additional fields
                new_user.first_name = first_name
                new_user.last_name = last_name
                new_user.full_name = full_name
                new_user.phone = phone
                new_user.role_id = role_id
                new_user.address = address
                new_user.created_by = user_id
                new_user.created_at = timezone.now()
                new_user.save()

                # Store plain text password in password_storage
                password_storage.objects.create(
                    user=new_user,
                    passwordText=plain_password
                )

                messages.success(request, f"User '{username}' created successfully!")
                return redirect("L01:user_master_index")

        # GET → Show form
        roles_list = roles.objects.exclude(role_type='member')
        
        # Prepare context for template
        context = {
            "roles_list": roles_list,
            "user_obj": None
        }
        
        # If editing, pass the user object
        if 'id' in request.GET:
            try:
                enc_id = request.GET.get('id')
                user_pk = int(dec(enc_id))
                user_obj = CustomUser.objects.get(pk=user_pk)
                context['user_obj'] = user_obj
                context['user_obj'].encrypted_id = enc_id
            except:
                pass

        return render(request, "L01/Master/user_create.html", context)

    except CustomUser.DoesNotExist:
        messages.error(request, "User not found.")
        return redirect("L01:user_master_index")
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else 'user_create'
        callproc("stp_error_log", [fun, str(e), user_id])
        messages.error(request, 'Oops...! Something went wrong!')
        return redirect("L01:user_master_index")

@login_required
def user_edit(request):
    try:
        user_id = request.user.id

        # 🔹 Read from query string or hidden field
        enc_id = request.GET.get("user_id") or request.POST.get("id")
        if not enc_id:
            messages.error(request, "Invalid user reference.")
            return redirect("L01:user_master_index")

        user_pk = int(dec(enc_id))  # Decrypt the ID

        # 🔹 Fetch user or show error
        try:
            user_obj = CustomUser.objects.get(pk=user_pk)
        except CustomUser.DoesNotExist:
            messages.error(request, "User not found.")
            return redirect("L01:user_master_index")

        if request.method == "POST":
            username = request.POST.get("username", "").strip()
            first_name = request.POST.get("first_name", "").strip()
            last_name = request.POST.get("last_name", "").strip()
            full_name = f"{first_name} {last_name}".strip()
            email = request.POST.get("email", "").strip()
            phone = request.POST.get("mobile", "").strip()
            role_id = request.POST.get("role_id", "").strip()
            is_active = request.POST.get("is_active", 1)
            password = request.POST.get("password", "").strip()
            confirm_password = request.POST.get("confirm_password", "").strip()
            address = request.POST.get("address", "").strip()

            # 🔹 Validation
            if not username:
                messages.error(request, "User Name is required!")
            elif not first_name:
                messages.error(request, "First Name is required!")
            elif not last_name:
                messages.error(request, "Last Name is required!")
            elif not email:
                messages.error(request, "Email is required!")
            elif not phone:
                messages.error(request, "Phone Number is required!")
            elif not role_id:
                messages.error(request, "Role is required!")
            elif password and password != confirm_password:
                messages.error(request, "Passwords do not match!")

            # If errors → re-render with form data
            if messages.get_messages(request):
                roles_list = roles.objects.exclude(role_type="member")
                return render(request, "L01/Master/user_edit.html", {
                    "roles_list": roles_list,
                    "post_data": request.POST,
                    "user_obj": user_obj
                })

            # 🔹 Update user
            user_obj.username = username
            user_obj.first_name = first_name
            user_obj.last_name = last_name
            user_obj.full_name = full_name
            user_obj.email = email
            user_obj.phone = phone
            user_obj.address = address
            user_obj.password = password
            user_obj.confirm_password = confirm_password
            user_obj.is_active = bool(int(is_active))
            user_obj.role_id = role_id
            user_obj.updated_by = user_id
            user_obj.updated_at = timezone.now()

            if password:
                user_obj.set_password(password)

            user_obj.save()

            messages.success(request, f"User '{username}' updated successfully!")
            return redirect("L01:user_master_index")

        # GET → load edit form
        roles_list = roles.objects.exclude(role_type="member")
        return render(request, "L01/Master/user_edit.html", {
            "roles_list": roles_list,
            "user_obj": user_obj
        })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log", [fun, str(e), user_id])
        messages.error(request, "Oops...! Something went wrong!")
        return redirect("L01:user_master_index")

@login_required
def user_view(request):
    try:
        user_id = request.user.id

        enc_id = request.GET.get("user_id")
        if not enc_id:
            messages.error(request, "Invalid user reference.")
            return redirect("L01:user_master_index")

        user_pk = int(dec(enc_id))  # Decrypt ID

        try:
            user_obj = CustomUser.objects.get(pk=user_pk)
        except CustomUser.DoesNotExist:
            messages.error(request, "User not found.")
            return redirect("L01:user_master_index")

        # Fetch the role object using role_id
        role_obj = roles.objects.filter(id=user_obj.role_id).first()

        return render(request, "L01/Master/user_view.html", {
            "user_obj": user_obj,
            "role_obj": role_obj
        })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        callproc("stp_error_log", [fun, str(e), user_id])
        messages.error(request, "Oops...! Something went wrong!")
        return redirect("L01:user_master_index")

def view_catalogue(request):
    return render(request, "L01/view_catalogue.html", {
        "MEDIA_URL": settings.MEDIA_URL
    })
    
def view_catalogue_login_page(request):
    username = request.session.get('username')
    return render(request, "L01/view_catalogue_login_page.html", {
        "MEDIA_URL": settings.MEDIA_URL,
        "username": username
    })

def view_ebook_catalogue(request):
    return render(request, "L01/view_ebook_catalogue.html", {
        "MEDIA_URL": settings.MEDIA_URL
    })


@csrf_exempt
def bookcatalog_search(request):
    try:
        if request.method != "POST":
            return JsonResponse({"error": "Invalid request method. Use POST."}, status=405)

        query = request.POST.get("query", "").strip()
        search_type = request.POST.get("searchType", "").strip().lower()
        language = request.POST.get("language", "").strip()   # ✅ ADD THIS

        if not query or not search_type:
            return JsonResponse(
                {"error": "Please provide both query and search type."},
                status=400
            )

        # Base queryset
        results = BookCatalog.objects.using('L01').all()

        # 🔑 STEP 1: APPLY LANGUAGE FILTER FIRST
        if language:
            results = results.filter(language__iexact=language)

        # 🔍 STEP 2: APPLY SEARCH FILTER
        if search_type in ["title", "books"]:
            results = results.filter(title__istartswith=query)

        elif search_type == "author":
            results = results.filter(author__istartswith=query)

        elif search_type == "publisher":
            results = results.filter(publisher__icontains=query)

        elif search_type == "keyword":
            results = results.filter(keywords__icontains=query)

        elif search_type == "year":
            year_filters = Q()
            if query.isdigit():
                year_filters |= Q(year_of_publication=int(query))
            year_filters |= Q(publication_year__icontains=query)
            results = results.filter(year_filters)

        elif search_type == "call_number":
            results = results.filter(call_number__icontains=query)

        elif search_type == "cutter_number":
            results = results.filter(cutter_number__icontains=query)

        else:
            return JsonResponse(
                {"error": f"Invalid search type: {search_type}"},
                status=400
            )

        # 🔢 LIMIT + FIELDS
        results = results.values(
            "title",
            "author",
            "publisher",
            "language",
            "year_of_publication",
            "publication_year",
            "call_number",
            "cutter_number",
            "front_page_photo",
            "last_page_photo",
            "ebook_available",
            "cat_ref_num",
        )[:50]

        # 📷 IMAGE + ENCRYPT
        updated_results = []

        for r in results:
            front = r.get("front_page_photo")
            back = r.get("last_page_photo")

            # get full file URL from storage service
            r["front_page_photo"] = (
                file_storage_service.get_file_url(front) if front else ""
            )

            r["last_page_photo"] = (
                file_storage_service.get_file_url(back) if back else ""
            )

            r["encrypted_cat_ref_num"] = (
                enc(str(r["cat_ref_num"])) if r["cat_ref_num"] else ""
            )

            updated_results.append(r)

        return JsonResponse(updated_results, safe=False)

    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )
    

@csrf_exempt
def index_book_search(request):
    try:
        if request.method != "POST":
            return JsonResponse(
                {"error": "Invalid request method. Use POST."},
                status=405
            )

        query = request.POST.get("query", "").strip()
        search_type = request.POST.get("searchType", "").strip().lower()
        language = request.POST.get("language", "").strip()   # ✅ ADD THIS

        if not query:
            return JsonResponse(
                {"error": "Please enter a search term."},
                status=400
            )

        # Base queryset
        results = BookCatalog.objects.using('L01').all()

        # 🔑 STEP 1: APPLY LANGUAGE FILTER FIRST
        if language:
            results = results.filter(language__iexact=language)

        # 🔤 STEP 2: INDEX SEARCH → ALWAYS STARTS WITH
        if search_type in ["", "title", "books"]:
            results = results.filter(title__istartswith=query)

        elif search_type == "author":
            results = results.filter(author__istartswith=query)

        elif search_type == "publisher":
            results = results.filter(publisher__istartswith=query)

        elif search_type == "keyword":
            results = results.filter(keywords__istartswith=query)

        elif search_type == "year":
            year_filters = Q()
            if query.isdigit():
                year_filters |= Q(year_of_publication=int(query))
            year_filters |= Q(publication_year__istartswith=query)
            results = results.filter(year_filters)

        else:
            return JsonResponse(
                {"error": f"Invalid search type: {search_type}"},
                status=400
            )

        # 🔢 Limit + required fields
        results = results.values(
            "title",
            "author",
            "publisher",
            "language",
            "year_of_publication",
            "publication_year",
            "call_number",
            "cutter_number",
            "front_page_photo",
            "last_page_photo",
            "ebook_available",
            "cat_ref_num",
        )[:50]

        # 📷 Image URL + encryption
        updated_results = []

        for r in results:
            front = r.get("front_page_photo")
            back = r.get("last_page_photo")

            # get full URL from storage service
            r["front_page_photo"] = (
                file_storage_service.get_file_url(front) if front else ""
            )

            r["last_page_photo"] = (
                file_storage_service.get_file_url(back) if back else ""
            )

            r["encrypted_cat_ref_num"] = (
                enc(str(r["cat_ref_num"])) if r["cat_ref_num"] else ""
            )

            updated_results.append(r)

        return JsonResponse(updated_results, safe=False)

    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )
    

@csrf_exempt
def libraryebook_search(request):
    try:
        if request.method != "POST":
            return JsonResponse(
                {"error": "Invalid request method. Use POST."},
                status=405
            )

        query = request.POST.get("query", "").strip()
        search_type = request.POST.get("searchType", "").strip().lower()
        language = request.POST.get("language", "").strip()   # ✅ ADD THIS

        if not query or not search_type:
            return JsonResponse(
                {"error": "Please provide both query and search type."},
                status=400
            )

        # Base queryset
        results = LibraryEbook.objects.using('L01').all()

        # 🔑 STEP 1: APPLY LANGUAGE FILTER FIRST
        if language:
            results = results.filter(eb_language__iexact=language)

        # 🔍 STEP 2: SEARCH FILTER
        if search_type in ["title", "ebook", "book"]:
            results = results.filter(eb_title__icontains=query)

        elif search_type == "author":
            results = results.filter(
                Q(eb_author__icontains=query) |
                Q(eb_other_authors__icontains=query)
            )

        elif search_type == "publisher":
            results = results.filter(eb_publisher__icontains=query)

        elif search_type == "keyword":
            results = results.filter(eb_keywords__icontains=query)

        elif search_type in ["isbn", "issn"]:
            results = results.filter(eb_isbn_issn__icontains=query)

        elif search_type == "year":
            year_filters = Q()
            if query.isdigit():
                year_filters |= Q(eb_year_of_publication=int(query))
            year_filters |= Q(eb_year_of_publication__icontains=query)
            results = results.filter(year_filters)

        else:
            return JsonResponse(
                {"error": f"Invalid search type: {search_type}"},
                status=400
            )

        # ---- LIMIT AND RETURN FIELDS ----
        results = results.values(
            "ebook_id",
            "eb_title",
            "eb_author",
            "eb_other_authors",
            "eb_publisher",
            "eb_language",
            "eb_year_of_publication",
            "eb_keywords",
            "eb_isbn_issn",
            "eb_pdf_url",
            "eb_front_page_photo",
            "eb_last_page_photo",
        )[:50]

        # ---- FORMAT + ENCRYPT ----
        updated_results = []
        for r in results:
            front = r.get("eb_front_page_photo")
            back = r.get("eb_last_page_photo")

            r["eb_front_page_photo"] = (
                request.build_absolute_uri(f"/media/{front}") if front else ""
            )

            r["eb_last_page_photo"] = (                              # ✅ ADD
                request.build_absolute_uri(f"/media/{back}") if back else ""
            )

            r["encrypted_ebook_id"] = enc(str(r["ebook_id"]))
            r["encrypted_pdf_url"] = (
                enc(r["eb_pdf_url"]) if r["eb_pdf_url"] else ""
            )

            updated_results.append(r)

        return JsonResponse(updated_results, safe=False)

    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )



@csrf_exempt
def index_ebook_search(request):
    try:
        if request.method != "POST":
            return JsonResponse(
                {"error": "Invalid request method. Use POST."},
                status=405
            )

        query = request.POST.get("query", "").strip()
        search_type = request.POST.get("searchType", "").strip().lower()
        language = request.POST.get("language", "").strip()   # ✅ ADD THIS

        if not query:
            return JsonResponse(
                {"error": "Please enter a search term."},
                status=400
            )

        # Base queryset
        results = LibraryEbook.objects.using('L01').all()

        # 🔑 STEP 1: APPLY LANGUAGE FILTER FIRST
        if language:
            results = results.filter(eb_language__iexact=language)

        # 🔤 STEP 2: FIRST-LETTER SEARCH
        if search_type in ["title", "book", "ebook", ""]:
            results = results.filter(eb_title__istartswith=query)

        elif search_type == "author":
            results = results.filter(
                Q(eb_author__istartswith=query) |
                Q(eb_other_authors__istartswith=query)
            )

        elif search_type == "publisher":
            results = results.filter(eb_publisher__istartswith=query)

        elif search_type == "keyword":
            results = results.filter(eb_keywords__istartswith=query)

        elif search_type in ["isbn", "issn"]:
            results = results.filter(eb_isbn_issn__istartswith=query)

        elif search_type == "year":
            results = results.filter(eb_year_of_publication__istartswith=query)

        else:
            return JsonResponse(
                {"error": f"Invalid search type: {search_type}"},
                status=400
            )

        # ---- SELECT FIELDS ----
        results = results.values(
            "ebook_id",
            "eb_title",
            "eb_author",
            "eb_other_authors",
            "eb_publisher",
            "eb_language",
            "eb_year_of_publication",
            "eb_keywords",
            "eb_isbn_issn",
            "eb_pdf_url",
            "eb_front_page_photo",
            "eb_last_page_photo",
        )[:50]

        # ---- FORMAT + ENCRYPT ----
        updated_results = []
        for r in results:
            front = r.get("eb_front_page_photo")
            back = r.get("eb_last_page_photo")

            r["eb_front_page_photo"] = (
                request.build_absolute_uri(f"/media/{front}") if front else ""
            )

            r["eb_last_page_photo"] = (                              # ✅ ADD
                request.build_absolute_uri(f"/media/{back}") if back else ""
            )

            r["encrypted_ebook_id"] = enc(str(r["ebook_id"]))
            r["encrypted_pdf_url"] = (
                enc(r["eb_pdf_url"]) if r["eb_pdf_url"] else ""
            )

            updated_results.append(r)

        return JsonResponse(updated_results, safe=False)

    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )
    
@csrf_exempt
def open_pdf(request):
    encrypted_pdf = request.GET.get("pdf")

    if not encrypted_pdf:
        return redirect("L01:membership_dashboard")

    try:
        pdf_path = dec(encrypted_pdf)
    except Exception:
        return redirect("L01:membership_dashboard")

    if ".." in pdf_path or not pdf_path.endswith(".pdf"):
        return redirect("L01:membership_dashboard")

    file_url = request.build_absolute_uri(
        settings.MEDIA_URL + pdf_path
    )

    return redirect(file_url)

def upsc_ebook_index(request):
    try:
        # Fetch the competitive exam details for UPSC (id = 1) from L01 DB
        competitive_exam = get_object_or_404(
            CompetitiveExamMaster.objects.using('L01'),
            competitive_id=1
        )

        # Fetch all related sections linked to UPSC (from L01 DB)
        sections = Sections.objects.using('L01').filter(
            competitive_id=1
        ).order_by('section_no')

        # ✅ Fetch subjects for each section (order by subject_id)
        # Keep as list of tuples (section_no, subjects)
        subjects_by_section = []
        for section in sections:
            section.encrypted_section_no = enc(str(section.section_no))
            subjects = Subjects.objects.using('L01').filter(
                section_no=section.section_no
            ).order_by('subject_id')  # use subject_id since subject_no doesn't exist
            # subjects_by_section.append((section.section_no, subjects))
            subjects_by_section.append((section.encrypted_section_no, subjects))

        # Pass everything to template without converting to dict
        return render(request, "L01/UPSC/upsc_ebook_index.html", {
            "MEDIA_URL": settings.MEDIA_URL,
            "competitive_exam": competitive_exam,
            "sections": sections,
            "subjects_by_section": subjects_by_section,  # list of tuples
        })
    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )

def topic_index(request, section_no):
    try:
        # ✅ Get the competitive exam (UPSC)
        competitive_exam = get_object_or_404(
            CompetitiveExamMaster.objects.using('L01'),
            competitive_id=1
        )

        # ✅ Fetch the specific section using the passed section_no
        section_no = int(dec(section_no))
        section = get_object_or_404(
            Sections.objects.using('L01'),
            section_no=section_no
        )

        # ✅ Fetch all subjects under this section
        subjects = Subjects.objects.using('L01').filter(
            section_no=section.section_no
        ).order_by('subject_id')

        subjects_data = []

        # ✅ For each subject, fetch topics under the same section
        for subject in subjects:
            topics = Topics.objects.using('L01').filter(
                subject_id=subject.subject_id,
                section_no=section.section_no
            ).order_by('topic_id')

            # Use get_file_url for the topic image URL
            for topic in topics:
                if topic.topic_image_url:
                    # Get the full URL for the topic image
                    topic.topic_image_url = file_storage_service.get_file_url(topic.topic_image_url)

            subjects_data.append({
                "subject": subject,
                "topics": topics
            })

        # ✅ Pass the structured data to the template
        return render(request, "L01/UPSC/topics_index.html", {
            "competitive_exam": competitive_exam,
            "section": section,
            "subjects_data": subjects_data,
        })

    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )

def chapters_index(request, topic_id):
    try:
        
        library_code = request.session.get('library_db', None)
        
        # ✅ Fetch the selected topic from L01
        topic_id = int(dec(topic_id))
        topic = get_object_or_404(
            Topics.objects.using('L01'),
            topic_id=topic_id
        )

        # ✅ Fetch the parent section for breadcrumb
        section = get_object_or_404(
            Sections.objects.using('L01'),
            section_no=topic.section_no_id
        )

        # ✅ Fetch all chapters under this topic (sorted numerically)
        chapters = Chapters.objects.using('L01') \
            .filter(topic_id=topic_id) \
            .annotate(
                chapter_no_int=Cast('chapter_no', IntegerField())
            ) \
            .order_by('chapter_no_int')

        # ✅ Prepare structured data (if needed for template)
        # chapters_data = []
        # for chapter in chapters:
        #     chapters_data.append({
        #         "chapter": chapter
        #     })
        
        chapters_data = [
            {
                "chapter": chapter,
                "chapter_pdf_url": file_storage_service.get_file_url(chapter.chapter_pdf_url)  # Call the function here
            }
            for chapter in chapters
        ]

        # ✅ Render the chapters page
        return render(request, "L01/UPSC/chapters_index.html", {
            "MEDIA_URL": settings.MEDIA_URL,
            "section": section,
            "topic": topic,
            "chapters_data": chapters_data,
        })

    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )
    
def get_membership_code(request):
    username = request.session.get("username")
    # email = get_object_or_404(CustomUser, id = user_id).email
    member = MembershipDetails.objects.get(user_id=username)

    return JsonResponse({
        "membership_code": member.membership_code
    })

def mpsc_ebook_index(request):
    try:
        # Fetch the competitive exam details for MPSC (id = 2) from L01 DB
        competitive_exam = get_object_or_404(
            CompetitiveExamMaster.objects.using('L01'),
            competitive_id=2
        )

        # Fetch all related sections linked to MPSC (from L01 DB)
        sections = Sections.objects.using('L01').filter(
            competitive_id=2
        ).order_by('section_no')

        # ✅ Fetch subjects for each section (order by subject_id)
        subjects_by_section = []

        for section in sections:
            section.encrypted_section_no = enc(str(section.section_no))
            subjects = Subjects.objects.using('L01').filter(
                section_no=section.section_no
            ).order_by('subject_id')  # using subject_id same as UPSC
            # subjects_by_section.append((section.section_no, subjects))
            # encrypted_section_no = enc(str(section.section_no))

            # 🔐 send encrypted section_no to template
            subjects_by_section.append((section.encrypted_section_no, subjects))

        # Pass everything to template without converting to dict
        return render(request, "L01/MPSC/mpsc_ebooks_index.html", {
            "MEDIA_URL": settings.MEDIA_URL,
            "competitive_exam": competitive_exam,
            "sections": sections,
            "subjects_by_section": subjects_by_section,
        })

    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )

def mpsc_topics_index(request, section_no):
    try:
        # ✅ Get the competitive exam (MPSC)
        competitive_exam = get_object_or_404(
            CompetitiveExamMaster.objects.using('L01'),
            competitive_id=2   # ← MPSC ID
        )

        # ✅ Fetch the specific section using the passed section_no
        section_no = int(dec(section_no))
        section = get_object_or_404(
            Sections.objects.using('L01'),
            section_no=section_no
        )

        # ✅ Fetch all subjects under this section
        subjects = Subjects.objects.using('L01').filter(
            section_no=section.section_no
        ).order_by('subject_id')

        subjects_data = []

        # ✅ For each subject, fetch topics under the same section
        for subject in subjects:
            topics = Topics.objects.using('L01').filter(
                subject_id=subject.subject_id,
                section_no=section.section_no
            ).order_by('topic_id')

            # Use get_file_url for the topic image URL
            for topic in topics:
                if topic.topic_image_url:
                    topic.topic_image_url = file_storage_service.get_file_url(topic.topic_image_url)

            subjects_data.append({
                "subject": subject,
                "topics": topics
            })

        # ✅ Render the MPSC topics page
        return render(request, "L01/MPSC/mpsc_topics_index.html", {   # ← Change template if needed
            "MEDIA_URL": settings.MEDIA_URL,
            "competitive_exam": competitive_exam,
            "section": section,
            "subjects_data": subjects_data,
        })

    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )

def mpsc_chapters_index(request, topic_id):
    try:
        # Fetch competitive exam (MPSC)
        competitive_exam = get_object_or_404(
            CompetitiveExamMaster.objects.using('L01'),
            competitive_id=2
        )

        # Fetch the topic
        topic_id = int(dec(topic_id))
        topic = get_object_or_404(
            Topics.objects.using('L01'),
            topic_id=topic_id
        )

        # Fetch parent section
        section = get_object_or_404(
            Sections.objects.using('L01'),
            section_no=topic.section_no_id
        )

        # Fetch chapters under this topic
        chapters = Chapters.objects.using('L01') \
            .filter(topic_id=topic_id) \
            .annotate(chapter_no_int=Cast('chapter_no', IntegerField())) \
            .order_by('chapter_no_int')

        # chapters_data = [{"chapter": chapter} for chapter in chapters]
        
        chapters_data = [
            {
                "chapter": chapter,
                "chapter_pdf_url": file_storage_service.get_file_url(chapter.chapter_pdf_url)  # Call the function here
            }
            for chapter in chapters
        ]

        return render(request, "L01/MPSC/mpsc_chapters_index.html", {
            "MEDIA_URL": settings.MEDIA_URL,
            "competitive_exam": competitive_exam,
            "section": section,
            "topic": topic,
            "chapters_data": chapters_data,
        })

    except Exception as e:
        return JsonResponse(
            {"error": "An unexpected error occurred.", "details": str(e)},
            status=500
        )

def member_entry_exit(request):
    from django.utils.timezone import localdate
    active_user_ids = CustomUser.objects.filter(
        is_active=1
    ).values_list('username', flat=True)   # <-- important


    # 2️⃣ Filter MembershipDetails for those active user_ids
    members = MembershipDetails.objects.using('L01').filter(
        isactive=1,
        user_id__in=active_user_ids
    ).annotate(
        display_name=Concat(
            'first_name', Value(' '),
            'middle_name', Value(' '),
            'last_name', Value(' - '),
            'membership_code',
            output_field=CharField()
        )
    ).values('membership_code', 'display_name')
    active_members = MembershipDetails.objects.using('L01').filter(isactive=True).count()
    today = date.today()

    today = localdate()

    scan_today = MemberEntryExit.objects.using('L01').filter(
        entry_time__date=today
    ).count()



    context = {
        'members': list(members),
        'active_members': active_members,
        'scan_today': scan_today
    }

    return render(request, 'L01/Member Payment/member_log.html', context)

def get_member_detail(request, membership_code):
    try:
        # Query from L01 database
        member = MembershipDetails.objects.using('L01').get(membership_code=membership_code)
        # membership_typemember_type = f"{member.membership.membership_type '+' (member.membership.membership_code)})"
        
        # Check if membership is valid based on to_date
        current_date = date.today()
        is_valid = (member.from_date <= current_date <= member.to_date if (member.from_date and member.to_date) else False)

        
        profile_pic = '/static/images/user.png'

        try:
            # Fetch the document
            document = DocumentDetails.objects.using('L01').get(membership=member.id, document_id=1)
            
            # Check if the document has a file path
            if document.file_path:
                # file_path = Path(settings.MEDIA_ROOT) / Path(document.file_path.replace("\\", "/"))
                
                # # Only update profile_pic if the file exists
                # if file_path.exists():
                    profile_pic = file_storage_service.get_file_url(document.file_path)
        except DocumentDetails.DoesNotExist:
            pass  # Keep default profile_pic
        member_data = {
            'membership_code': member.membership_code,
            'full_name': f"{member.first_name} {member.middle_name} {member.last_name}".strip(),
            'ward': member.ward or 'Not specified',
            'member_type': member.membership.membership_type,
            'is_active': member.isactive,
            'status': 'valid' if is_valid else 'invalid',
            'profile_pic': profile_pic,
            'to_date': member.to_date.strftime('%Y-%m-%d') if member.to_date else 'N/A',
        }
        
        # AUTO CREATE ENTRY ONLY IF MEMBERSHIP IS VALID (to_date not passed)
        if is_valid and member.isactive:
            current_time = datetime.now()
            
            # Check for existing open entry today
            today_start = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
            today_end = current_time.replace(hour=23, minute=59, second=59, microsecond=999999)
            
            existing_entry = MemberEntryExit.objects.using('L01').filter(
                membership_code=membership_code,
                entry_time__gte=today_start,
                entry_time__lte=today_end,
                exit_time__isnull=True
            ).first()
            
            if not existing_entry:
                # Create new entry record
                entry_record = MemberEntryExit.objects.using('L01').create(
                    membership_code=membership_code,
                    entry_time=current_time,
                    exit_time=None,
                    role_id=getattr(member, 'role_id', None),
                    created_by=request.user if request.user.is_authenticated else None,
                    updated_by=request.user if request.user.is_authenticated else None
                )
                
                member_data['auto_entry_created'] = True
                member_data['entry_time'] = current_time.strftime('%Y-%m-%d %H:%M:%S')
                member_data['entry_id'] = entry_record.id
                member_data['log_message'] = f"Entry log successfully created for {member.first_name} {member.last_name}"
            else:
                member_data['auto_entry_created'] = False
                member_data['log_message'] = f"Active entry already exists for {member.first_name} {member.last_name}"
        else:
            member_data['auto_entry_created'] = False
            if not is_valid:
                member_data['log_message'] = f"Membership expired for {member.first_name} {member.last_name}."
            elif not member.isactive:
                member_data['log_message'] = f"Member account is inactive for {member.first_name} {member.last_name}."
        
        return JsonResponse(member_data)
    
    except MembershipDetails.DoesNotExist:
        return JsonResponse({"error": "Member not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": "An unexpected error occurred.", "details": str(e)}, status=500)

# @login_required
# def membership_dashboard(request):
#     # Get username from session
#     username = request.session.get('username')
    
#     try:
#         breadcrumb = request.POST.get("breadcrumb")
#         if not breadcrumb:
#             breadcrumb = 0
#         member_id = get_object_or_404(MembershipDetails, user_id =  username)
#         today = timezone.now().date()
        
#         # Calculate due soon threshold (2 days from now)
#         due_soon_threshold = today + timedelta(days=2)
#         currently_borrowed = CirculationTransaction.objects.using('L01').filter(
#             member=member_id,
#             return_date__isnull=True
#         ).count()
        
#         # Get due soon books (due date within 2 days and not returned)
#         due_soon = CirculationTransaction.objects.using('L01').filter(
#             member=member_id,
#             due_date__lte=due_soon_threshold,
#             due_date__gte=today,
#             return_date__isnull=True
#         ).count()
        
#         # Get overdue books (due date passed and not returned)
#         overdue = CirculationTransaction.objects.using('L01').filter(
#             member=member_id,
#             due_date__lt=today,
#             return_date__isnull=True
#         ).count()
        
#         # Get latest 3 borrowed books with catalog details
#         latest_books = []
#         transactions = CirculationTransaction.objects.filter(
#             member=member_id,
#             return_date__isnull=True
#         ).select_related('catalog').order_by('-issue_date')[:3]
#         pending_action = request.session.get("pending_action")
#         sweet_alert = request.session.get("sweet_alert")
#         for transaction in transactions:
#             book_data = {
#                 'transaction': transaction,
#                 'title': transaction.catalog.title if transaction.catalog else 'Unknown Title',
#                 'author': transaction.catalog.author if transaction.catalog and transaction.catalog.author else 'Unknown Author',
#                 'issue_date': transaction.issue_date,
#                 'due_date': transaction.due_date,
#                 'cover_image': get_book_cover(transaction.catalog) if transaction.catalog else None
#             }
#             latest_books.append(book_data)
#         request.session.pop("sweet_alert", None)
#         context = {
#             'username': username,
#             # 'membership_code': membership_code,
#             'currently_borrowed': currently_borrowed,
#             'due_soon': due_soon,
#             'overdue': overdue,
#             'latest_books': latest_books,
#             'today': today,
#             'pending_action': pending_action,    # send to template
#             'sweet_alert': sweet_alert,
#             'breadcrumb':breadcrumb
#         }
        
#         return render(request, 'L01/Dashboard/member_dashboard.html', context)
        
#     except Exception as e:
#         return render(request, 'error.html', {'error': str(e)})
#     except Exception as e:
#         return render(request, 'error.html', {'error': str(e)})

def membership_dashboard(request):
    # Generate or retrieve your key (ensure it's secure and private)

    # Get username from session
    username = request.session.get('username')
    
    try:
        breadcrumb = request.POST.get("breadcrumb")
        if not breadcrumb:
            breadcrumb = 0
        member_id = get_object_or_404(MembershipDetails, user_id = username)
        today = timezone.now().date()
        
        # Calculate due soon threshold (2 days from now)
        due_soon_threshold = today + timedelta(days=2)
        
        # Get latest 3 borrowed books with catalog details
        latest_books = []
        transactions = CirculationTransaction.objects.filter(
            member=member_id,
            return_date__isnull=True
        ).select_related('catalog').order_by('-issue_date')[:3]
        
        for transaction in transactions:
            # Encrypt the cat_ref_num (catalog reference number)
            cat_ref_num = str(transaction.catalog.cat_ref_num)
            encrypted_cat_ref_num = enc(cat_ref_num)  # Encrypt and convert to string

            book_data = {
                'transaction': transaction,
                'title': transaction.catalog.title if transaction.catalog else 'Unknown Title',
                'author': transaction.catalog.author if transaction.catalog and transaction.catalog.author else 'Unknown Author',
                'issue_date': transaction.issue_date,
                'due_date': transaction.due_date,
                'cover_image': get_book_cover(transaction.catalog) if transaction.catalog else None,
                'encrypted_cat_ref_num': encrypted_cat_ref_num  # Send the encrypted value to the template
            }
            latest_books.append(book_data)
        
        context = {
            'username': username,
            'latest_books': latest_books,
            'today': today,
            'breadcrumb': breadcrumb
        }
        
        return render(request, 'L01/Dashboard/member_dashboard.html', context)
        
    except Exception as e:
        return render(request, 'error.html', {'error': str(e)})

def get_book_cover(catalog):
    """
    Returns front page image URL using file_storage_service if available
    or falls back to default media handling.
    """
    try:
        if catalog and catalog.front_page_photo:
            # Assuming file_storage_service is defined correctly
            if hasattr(file_storage_service, 'get_file_url'):
                return file_storage_service.get_file_url(catalog.front_page_photo)

            # Fallback to default Django file storage
            file_path = default_storage.url(catalog.front_page_photo)
            return file_path

        # Return default image if no cover found

    except Exception as e:
        print("Error fetching book cover:", e)
    
@login_required
def get_borrowing_history(request):
    """AJAX view to get member's borrowing history"""
    username = request.session.get('username')
    
    if not username:
        return JsonResponse({'success': False, 'error': 'User not logged in'})
    
    try:
        # Get membership details
        membership = MembershipDetails.objects.get(user_id=username)
        membership_code = membership.membership_code
        member = get_object_or_404(MembershipDetails, membership_code=membership_code )
        
        # Get all circulation transactions for this member
        transactions = CirculationTransaction.objects.filter(
            member=member
        ).select_related('catalog', 'return_condition').order_by('-issue_date')
        
        history_data = []
        for transaction in transactions:
            # Determine status based on dates
            status = "Issued"
            if transaction.return_date:
                status = "Returned"
            elif transaction.due_date and transaction.due_date < timezone.now().date():
                status = "Overdue"
            elif transaction.due_date and transaction.due_date <= timezone.now().date() + timedelta(days=2):
                status = "Due Soon"
            
            history_data.append({
                'book_title': transaction.catalog.title if transaction.catalog else 'Unknown Title',
                'author': transaction.catalog.author if transaction.catalog and transaction.catalog.author else 'Unknown Author',
                'issue_date': transaction.issue_date.strftime('%b %d, %Y') if transaction.issue_date else 'N/A',
                'return_date': transaction.return_date.strftime('%b %d, %Y') if transaction.return_date else None,
                'status': status,
                'return_condition': transaction.return_condition.status_name if transaction.return_condition else 'N/A'
            })
        
        return JsonResponse({
            'success': True,
            'history': history_data
        })
        
    except MembershipDetails.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Membership not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    

        
def book_info_login(request):
    try:
        # Get cat_ref_num from GET params
        cat_ref_num = request.GET.get("cat_ref_num")
        if not cat_ref_num:
            return render(request, "L01/error.html", {"message": "cat_ref_num not provided"})

        # Fetch the book object
        book = get_object_or_404(BookCatalog.objects.using('L01'), cat_ref_num=cat_ref_num)

        # Build URLs for front and last pages
        front_page_url = request.build_absolute_uri(f"/media/{book.front_page_photo}") if book.front_page_photo else ""
        last_page_url = request.build_absolute_uri(f"/media/{book.last_page_photo}") if book.last_page_photo else ""

        # Pass data to template
        context = {
            "book": book,
            "front_page_url": front_page_url,
            "last_page_url": last_page_url,
        }

        return render(request, "L01/book_info_login.html", context)

    except Exception as e:
        print(f"Error in book_info_login: {e}")
        return render(request, "L01/error.html", {"message": "Something went wrong"})
    
@login_required
def save_eod_log(request):
    try:
        if request.method == "POST":

            # Step 1: Get all returned transaction IDs
            all_ids_raw = request.POST.get("all_ids", "")
            all_ids = [int(x) for x in all_ids_raw.split(",") if x]

            # Step 2: Validation: every book MUST be checked
            for tid in all_ids:
                if not request.POST.get(f"shelved_{tid}"):
                    return messages.error(("Please confirm that all books are put on shelf and then try again."))
                    # return JsonResponse({
                    #     "success": False,
                    #     "error": "Please confirm that all books are put on shelf and then try again."
                    # })

            # Step 3: Create master EOD entry
            today = timezone.now().date()
            eod_log = EODLog.objects.create(
                date=today,
                is_active=True,
                created_by=request.user,
                updated_by=request.user,
            )

            # Step 4: Insert details for each returned book 
            for tid in all_ids:
                ciculation_id = get_object_or_404(CirculationTransaction,id =tid)
                catalog_id = ciculation_id.catalog.cat_ref_num
                barcode = request.POST.get(f"barcode_{tid}", "")
                is_shelved = 1  # all must be checked to reach here

                cat_obj = BookCatalog.objects.get(cat_ref_num=catalog_id)

                BookReturnLog.objects.create(
                    eod_log=eod_log,
                    cat_rem_num=cat_obj,
                    barcode=barcode,
                    is_shelved=is_shelved,
                    created_by=request.user,
                    updated_by=request.user,
                )

            return redirect("payment_report")

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

# Cate list notepad view 440

@login_required
def show_Library_catalogue(request):
    try:
        # --- SESSION CHECKS ---
        library_code = request.session.get('library_db')
        username = request.session.get('username')
        user_id = request.session.get('user_id')
        role_id = request.session.get('role_id')

        if library_code != 'L01':
            messages.error(request, "Invalid library access.")
            request.session.flush()
            return redirect('library_list')

        if not all([library_code, username, user_id, role_id]):
            messages.warning(request, "Session expired or invalid. Please login again.")
            return render(request, "L01/LibraryCateVisit/visit_library_Cate.html", {})
        
        membership_detail = MembershipDetails.objects.filter(
            user_id=username,
            isactive=1
        ).select_related('membership').first()
        
        membership_id = membership_detail.membership.id if membership_detail else None

        # --- CONTEXT ---
        context = {
            'MEDIA_URL': settings.MEDIA_URL,
            'membership_id': membership_id,
        }

        return render(request, "L01/LibraryCateVisit/show_Library_catalogue.html", context)

    except Exception as e:
        print(f"Error: {e}")
        return render(request, "L01/LibraryCateVisit/show_Library_catalogue.html", {})

@login_required
def visit_Library_catalogue(request):
    try:
        # --- SESSION CHECKS ---
        library_code = request.session.get('library_db')
        username = request.session.get('username')
        user_id = request.session.get('user_id')
        role_id = request.session.get('role_id')

        if library_code != 'L01':
            messages.error(request, "Invalid library access.")
            request.session.flush()
            return redirect('library_list')

        if not all([library_code, username, user_id, role_id]):
            messages.warning(request, "Session expired or invalid. Please login again.")
            return render(request, "L01/LibraryCateVisit/visit_library_Cate.html", {})

        # --- SUBJECTS ---
        subjects_qs = SubjectTypeMaster.objects.filter(is_active=1)
        subjects = []
        for s in subjects_qs:
            s.id_enc = enc(str(s.id))  # encoding function
            subjects.append(s)

        first_subject = subjects[0] if subjects else None

        # --- BOOKS BY FIRST SUBJECT (use indexed subject_id) ---
        if first_subject:
            books = BookCatalog.objects.filter(subject_id=first_subject.id)\
                                       .select_related('subject', 'material', 'ebook')
        else:
            books = BookCatalog.objects.none()
            
        # --- ADD FILE URLs TO BOOKS ---
        for book in books:
            book.bookIdEnc = enc(str(book.cat_ref_num)) if 'enc' in globals() else book.cat_ref_num
            
            # ✅ Use FileStorageService for book cover image
            if book.front_page_photo:
                book.front_page_url = file_storage_service.get_file_url(book.front_page_photo)
            else:
                book.front_page_url = None
            
            # Add ebook encrypted ID if available
            if book.ebook:
                book.ebookEnc = enc(str(book.ebook.ebook_id))
                
                # ✅ Also get ebook front page URL if available
                if book.ebook.eb_front_page_photo:
                    book.ebook_front_page_url = file_storage_service.get_file_url(book.ebook.eb_front_page_photo)
                else:
                    book.ebook_front_page_url = None

        # --- PAGINATION ---
        paginator = Paginator(books, 8)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        for b in page_obj:
            b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num
            
            if b.ebook:
                b.ebookEnc = enc(str(b.ebook.ebook_id))

        # --- UPCOMING / LATEST BOOKS (use indexed cat_ref_num & created_at if indexed) ---
        latest_created_at = BookCatalog.objects.aggregate(latest=Max('created_at'))['latest']
        recent_books = BookCatalog.objects.filter(created_at=latest_created_at) if latest_created_at else BookCatalog.objects.none()

        remaining_count = 10 - recent_books.count()
        fallback_books = BookCatalog.objects.exclude(cat_ref_num__in=recent_books.values_list('cat_ref_num', flat=True))\
                                           .order_by('-cat_ref_num')[:remaining_count] if remaining_count > 0 else BookCatalog.objects.none()

        new_books = list(recent_books) + list(fallback_books)
        for b in new_books:
            b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num
            
            # ✅ Add file URLs for new books
            if b.front_page_photo:
                b.front_page_url = file_storage_service.get_file_url(b.front_page_photo)
            else:
                b.front_page_url = None

        # --- MOST REVIEWED BOOKS (use indexed cat_ref_id in BookReview) ---
        most_reviewed_books = (
            BookCatalog.objects.annotate(
                review_count=Count('cat_ref_id_reviews'),
                avg_rating=Avg('cat_ref_id_reviews__rating')
            )
            .filter(review_count__gt=0)
            .order_by('-avg_rating', '-review_count')[:10]
        )

        for b in most_reviewed_books:
            b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num
            avg = b.avg_rating or 0
            b.avg_rating = round(avg, 1)
            b.stars = [True if i < round(avg) else False for i in range(5)]
            
            # ✅ Add file URLs for reviewed books
            if b.front_page_photo:
                b.front_page_url = file_storage_service.get_file_url(b.front_page_photo)
            else:
                b.front_page_url = None

        # --- NEW ARRIVALS IN CIRCULATION (use indexed bookcatalog_id in circulation) ---
        new_arrivals_qs = BookCatalog.objects.annotate(
            in_circulation=Exists(
                CirculationCopyStatus.objects.filter(bookcatalog_id=OuterRef('cat_ref_num'))
            )
        ).filter(in_circulation=True).order_by('-cat_ref_num')[:10]

        for b in new_arrivals_qs:
            b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num
            
            # ✅ Add file URLs for new arrivals
            if b.front_page_photo:
                b.front_page_url = file_storage_service.get_file_url(b.front_page_photo)
            else:
                b.front_page_url = None

        # --- CONTEXT ---
        context = {
            'subjects': subjects,
            'books': page_obj,
            'paginator': paginator,
            'page_number': int(page_number),
            'first_subject_id_enc': first_subject.id_enc if first_subject else None,
            'MEDIA_URL': settings.MEDIA_URL,
            'new_books': new_books,
            'most_reviewed_books': most_reviewed_books,
            'new_arrivals_qs': new_arrivals_qs,
        }

        return render(request, "L01/LibraryCateVisit/visit_library_Cate.html", context)

    except Exception as e:
        print(f"Error: {e}")
        return render(request, "L01/LibraryCateVisit/visit_library_Cate.html", {})

@login_required
def get_books_by_subject(request):
    try:
        subject_id_enc = request.GET.get('subject_id')
        subject_id = dec(subject_id_enc) if subject_id_enc else None

        search = request.GET.get('search', '').strip()
        searching = request.GET.get('searching', '').strip()
        
        if searching:
            # Global search mode (across all subjects)
            all_books = BookCatalog.objects.all().select_related('subject', 'material', 'ebook')

            if search:
                # Search across all subjects
                all_books = all_books.filter(
                    Q(title__icontains=search) |
                    Q(author__icontains=search)
                )
            elif subject_id:
                # Filter by subject only if no search
                all_books = all_books.filter(subject_id=subject_id)

            # ✅ ADD FILE URLS FOR ALL BOOKS
            for b in all_books:
                b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num
                
                # Get front page URL using FileStorageService
                if b.front_page_photo:
                    b.front_page_url = file_storage_service.get_file_url(b.front_page_photo)
                else:
                    b.front_page_url = None
                
                # Get ebook info
                if b.ebook:
                    b.ebookEnc = enc(str(b.ebook.ebook_id))
                    
                    # Also get ebook front page if available
                    if b.ebook.eb_front_page_photo:
                        b.ebook_front_page_url = file_storage_service.get_file_url(b.ebook.eb_front_page_photo)
                    else:
                        b.ebook_front_page_url = None

            # Pagination
            paginator = Paginator(all_books, 8)
            page_number = request.GET.get('page', 1)
            books_page = paginator.get_page(page_number)

        else:
            # Filter by specific subject
            if not subject_id:
                return JsonResponse({'error': 'Subject ID required'}, status=400)
                
            all_books = BookCatalog.objects.filter(subject_id=subject_id)\
                                           .select_related('subject', 'material', 'ebook')

            # Search filter
            if search:
                all_books = all_books.filter(
                    Q(title__icontains=search) |
                    Q(author__icontains=search)
                )

            # ✅ ADD FILE URLS FOR BOOKS
            for b in all_books:
                b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num
                
                # Get front page URL using FileStorageService
                if b.front_page_photo:
                    b.front_page_url = file_storage_service.get_file_url(b.front_page_photo)
                else:
                    b.front_page_url = None
                
                # Get ebook info
                if b.ebook:
                    b.ebookEnc = enc(str(b.ebook.ebook_id))
                    
                    # Also get ebook front page if available
                    if b.ebook.eb_front_page_photo:
                        b.ebook_front_page_url = file_storage_service.get_file_url(b.ebook.eb_front_page_photo)
                    else:
                        b.ebook_front_page_url = None

            # Pagination
            paginator = Paginator(all_books, 8)
            page_number = request.GET.get('page', 1)
            books_page = paginator.get_page(page_number)

        context = {
            'books': books_page,
            'MEDIA_URL': settings.MEDIA_URL,  # Keep for backward compatibility
            'subject_id_enc': subject_id_enc
        }
        
        return render(request, "L01/LibraryCateVisit/book_list_partial.html", context)

    except Exception as e:
        print("Error fetching books for subject:", e)
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': 'Failed to fetch books'}, status=500)

def membership_card(request):
    username = request.session.get('username')
    member_id = None
    member_details = None
    library_name_from_session = None
    library_address = None
    library_phone = None
    user_image_url = None
    transaction_details = []

    # Get library_code from session
    library_code_from_session = request.session.get('library_db', None)

    # Fetch library details
    if library_code_from_session:
        try:
            library = tbl_librarymasterL01.objects.filter(library_code__iexact=library_code_from_session).first()
            if library:
                library_name_from_session = library.library_name
                library_address = library.location
                library_phone = library.contact_phone
        except Exception:
            pass

    # Fetch member details
    if username:
        try:
            member = MembershipDetails.objects.get(user_id=username)
            member_id = member.id
            request.session['member_id'] = member_id

            # Fetch user image
            try:
                document = DocumentDetails.objects.filter(
                    membership_id=member_id, 
                    isactive=1
                ).first()

                if document and document.file_path:
                    file_path_lower = document.file_path.lower()
                    
                    if file_path_lower.endswith('.pdf'):
                        # If PDF, set user_image_url to None to use default image
                        user_image_url = None
                    else:
                        # Not a PDF, use the file path
                        if document.file_path.startswith(settings.MEDIA_URL):
                            user_image_url = document.file_path
                        else:
                            user_image_url = f"{settings.MEDIA_URL}{document.file_path}"
            except Exception:
                pass

            # Fetch full transaction list (NO PAGINATION)
            try:
                transactions = (
                    CirculationTransaction.objects
                    .filter(member_id=member_id)
                    .select_related('catalog')
                    .order_by('-issue_date')
                )

                for transaction in transactions:
                    is_overdue = False
                    if transaction.due_date and transaction.return_date is None:
                        is_overdue = transaction.due_date < date.today()

                    transaction_details.append({
                        'book_title': transaction.catalog.title if transaction.catalog else 'Not available',
                        'author': transaction.catalog.author if transaction.catalog else 'Not available',
                        'issue_date': transaction.issue_date.strftime('%d-%b-%Y') if transaction.issue_date else 'Not set',
                        'due_date': transaction.due_date.strftime('%d-%b-%Y') if transaction.due_date else 'Not set',
                        'return_date': transaction.return_date.strftime('%d-%b-%Y') if transaction.return_date else 'Not returned',
                        'due_date_raw': transaction.due_date,
                        'is_overdue': is_overdue
                    })

            except Exception as e:
                print(f"Error fetching transactions: {e}")

            # Prepare member details
            member_details = {
                'id': member.id,
                'name': f"{member.first_name_mar or ''} {member.middle_name_mar or ''} {member.last_name_mar or ''}".strip(),
                'address': member.local_address or 'Not available',
                "member_type":member.membership.membership_type,
                'mobile_no': member.mobile_no or 'Not available',
                'membership_from': member.from_date.strftime('%d-%b-%Y') if member.from_date else 'Not set',
                'membership_to': member.to_date.strftime('%d-%b-%Y') if member.to_date else 'Not set',
                'membership_code': member.membership_code or 'Not available',
                'library_name':  member.library_name_mar or 'Not available',
                'email': member.email or 'Not available',
                'user_id': member.user_id,
                'library_address': library_address or 'Not available',
                'library_phone': library_phone or 'Not available',
                'user_image_url': user_image_url
            }

        except MembershipDetails.DoesNotExist:
            member_details = None

    # Context WITHOUT pagination
    context = {
        "MEDIA_URL": settings.MEDIA_URL,
        "username": username,
        "member_id": member_id,
        "member_details": member_details,
        "library_name_from_session": library_name_from_session,
        "transaction_details": transaction_details,
        "current_date": date.today().strftime('%d-%b-%Y'),
        "today": date.today(),
    }

    return render(request, "Master/membership_card.html", context)

@login_required
def view_book_detail(request):
    try:
        library_code = request.session.get('library_db')
        username = request.session.get('username')
        user_id = request.session.get('user_id')
        role_id = request.session.get('role_id')

        if not all([library_code, username, user_id, role_id]):
            messages.warning(request, "Session expired or invalid. Please login again.")
            return render(request, "L01/LibraryCateVisit/visit_library_Cate.html", {})
        
        cat_ref_num_enc = request.GET.get('cat_ref_num')
        if not cat_ref_num_enc:
            messages.error(request, "Invalid book request.")
            return redirect("visit_library_catalogue")

        cat_ref_num = dec(cat_ref_num_enc)

        # Fetch the book
        book = get_object_or_404(BookCatalog, cat_ref_num=cat_ref_num)
        
        for b in [book]:
            b.bookIdEnc = enc(str(b.cat_ref_num))
            
        if book.front_page_photo:
            book.front_page_url = file_storage_service.get_file_url(book.front_page_photo)
        else:
            book.front_page_url = None
            
        # Get last page URL
        if book.last_page_photo:
            book.last_page_url = file_storage_service.get_file_url(book.last_page_photo)
        else:
            book.last_page_url = None

        # Collect images
        images = []
        if book.front_page_photo:
            images.append(book.front_page_photo)
        if book.last_page_photo:
            images.append(book.last_page_photo)

        # Fetch all circulation rows for this book
        circulation_qs = CirculationCopyStatus.objects.filter(bookcatalog_id=book.cat_ref_num)
        total_qty = circulation_qs.count()

        # Get status master entries for current and processing statuses
        status_on_shelf = status_master.objects.filter(status_name__iexact="On-Shelf").first()
        status_unknown = "Not Available"

        # Count how many copies are currently "On-Shelf"
        current_on_shelf_count = circulation_qs.filter(current_status=status_on_shelf).count() if status_on_shelf else 0

        # Decide display values for availability
        if total_qty > 0:
            if current_on_shelf_count > 0:
                current_status_display = "On-Shelf"
                availability_text = f"{current_on_shelf_count} of {total_qty} available"
            else:
                current_status_display = status_unknown
                availability_text = f"0 of {total_qty} available"
        else:
            current_status_display = status_unknown
            availability_text = "-"

        # Calculate location display **only for On-Shelf copies**
        location_counts = {}
        if status_on_shelf:
            on_shelf_copies = circulation_qs.filter(current_status=status_on_shelf)
            for circ in on_shelf_copies:
                if circ.shelf_location:
                    loc_name = circ.shelf_location.location_name
                    location_counts[loc_name] = location_counts.get(loc_name, 0) + 1

        if location_counts:
            # Join multiple locations with count if multiple exist
            location_display = ", ".join([f"{loc} ({count})" for loc, count in location_counts.items()])
        else:
            location_display = "Unknown"
            
        # --- REVIEWS WITH USER JOIN ---
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # Get all reviews for this book ordered by latest
        all_reviews = BookReview.objects.filter(book=book).order_by('-created_at')
        
        # Get distinct user IDs from reviews
        user_ids = all_reviews.values_list('user_id', flat=True).distinct()
        
        # Fetch all users in one query
        users = User.objects.filter(id__in=user_ids)
        
        # Create a dictionary for quick lookup: user_id -> user object
        user_dict = {user.id: user for user in users}
        
        # Add user objects to reviews
        reviews_with_users = []
        for review in all_reviews:
            # Create a copy of the review with added user information
            review.user_obj = user_dict.get(review.user_id)
            reviews_with_users.append(review)
        
        # Get current user from session
        current_user_id = request.session.get('user_id')
        
        # Check if current user has reviewed
        user_review = None
        if current_user_id:
            try:
                user_review = all_reviews.filter(user_id=int(current_user_id)).first()
                if user_review:
                    user_review.user_obj = user_dict.get(int(current_user_id))
            except (ValueError, TypeError):
                user_review = None
        
        # Calculate average rating
        avg_rating = all_reviews.aggregate(avg=models.Avg('rating'))['avg'] or 0
        total_reviews = all_reviews.count()
        
        # Pagination - Show only 5 reviews per page
        reviews_per_page = 5
        show_pagination = total_reviews > reviews_per_page
        
        if show_pagination:
            # Paginate the reviews_with_users list
            paginator = Paginator(reviews_with_users, reviews_per_page)
            page_number = request.GET.get('page', 1)
            
            try:
                reviews_page = paginator.page(page_number)
            except PageNotAnInteger:
                reviews_page = paginator.page(1)
            except EmptyPage:
                reviews_page = paginator.page(paginator.num_pages)
        else:
            # If 5 or fewer reviews, show all
            reviews_page = reviews_with_users

        context = {
            'book': book,
            'images': images,
            'MEDIA_URL': settings.MEDIA_URL,
            'total_qty': total_qty,
            'current_status_display': current_status_display,
            'availability_text': availability_text,
            'location_display': location_display,
            'reviews': reviews_page,  # Now includes user objects
            'user_review': user_review,
            'avg_rating': round(avg_rating, 1),
            'total_reviews': total_reviews,
            'show_pagination': show_pagination,
            'reviews_per_page': reviews_per_page,
            'current_params': request.GET.copy(),
            'cat_ref_num_enc': cat_ref_num_enc,
            'front_page_url': book.front_page_url,
            'last_page_url': book.last_page_url,
        }

        return render(request, "L01/LibraryCateVisit/view_book_detail.html", context)

    except Exception as e:
        print("Error in view_book_detail:", e)
        messages.error(request, "Unable to load book details.")
        return redirect("visit_library_catalogue")

@csrf_exempt
@login_required
def submit_review(request):

    if request.method == "POST":
        try:
            # Get data from request
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                book_id = dec(data.get('book_id'))
                rating = int(data.get('rating'))
                review_text = data.get('review')
            else:
                book_id = dec(request.POST.get('book_id'))
                rating = int(request.POST.get('rating'))
                review_text = request.POST.get('review')
            
            user_id = request.session.get('user_id')
            library_code = request.session.get('library_db')

            # Validate required fields
            if not all([book_id, rating, review_text, user_id]):
                return JsonResponse({
                    "success": False, 
                    "message": "Missing required fields"
                }, status=400)

            # Get book
            try:
                book = BookCatalog.objects.get(cat_ref_num=book_id)
            except BookCatalog.DoesNotExist:
                return JsonResponse({
                    "success": False, 
                    "message": "Book not found"
                }, status=404)

            # Prevent multiple reviews
            if BookReview.objects.filter(book=book, user_id=user_id).exists():
                return JsonResponse({
                    "success": False, 
                    "message": "You have already reviewed this book."
                })

            # Create review
            BookReview.objects.create(
                book=book,
                user_id=user_id,
                rating=rating,
                review=review_text,
                library_code=library_code
            )

            # Get latest reviews with user information
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            all_reviews = BookReview.objects.filter(book=book).order_by('-created_at')
            
            # Get user IDs from reviews
            user_ids = all_reviews.values_list('user_id', flat=True).distinct()
            
            # Fetch users
            users = User.objects.filter(id__in=user_ids)
            user_dict = {user.id: user for user in users}
            
            # Add user objects to reviews
            reviews_with_users = []
            for review in all_reviews:
                review.user_obj = user_dict.get(review.user_id)
                reviews_with_users.append(review)
            
            # Calculate average rating
            avg_rating = all_reviews.aggregate(avg=models.Avg('rating'))['avg'] or 0
            total_reviews = all_reviews.count()
            
            # Get only FIRST 5 reviews for display
            from django.core.paginator import Paginator
            paginator = Paginator(reviews_with_users, 5)
            first_page_reviews = paginator.page(1)
            
            # Render reviews HTML
            reviews_html = render_to_string(
                "L01/LibraryCateVisit/review_list.html", 
                {
                    "reviews": first_page_reviews,
                    "MEDIA_URL": settings.MEDIA_URL
                }
            )

            return JsonResponse({
                "success": True,
                "message": "Review submitted successfully!",
                "reviews_html": reviews_html,
                "avg_rating": round(avg_rating, 1),
                "total_reviews": total_reviews,
                "has_next_page": first_page_reviews.has_next(),
                "total_pages": paginator.num_pages,
                "current_page": 1
            })
            
        except ValueError as e:
            print(f"ValueError: {e}")
            return JsonResponse({
                "success": False, 
                "message": "Invalid rating value"
            }, status=400)
        except Exception as e:
            print(f"Error in submit_review: {e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                "success": False, 
                "message": f"Server error: {str(e)}"
            }, status=500)
    
    return JsonResponse({
        "success": False, 
        "message": "Invalid request method"
    }, status=405)
    
@csrf_exempt
def clear_pdf_session(request):
    request.session.pop("open_pdf_url", None)
    return JsonResponse({"status": "ok"})

from django.http import FileResponse, Http404
from django.conf import settings
import os

@login_required
def read_ebook_secure(request):
    """Stream ebook through Django using FileStorageService"""
    try:
        token = request.GET.get('token')
        if not token:
            raise Http404("Invalid access token")
        
        ebook_id = dec(str(token))
        ebook = get_object_or_404(LibraryEbook, pk=ebook_id)

        if not ebook.eb_pdf_url:
            raise Http404("E-Book file not found")

        # ✅ Stream through Django (no redirect to direct URLs)
        filename = f"{ebook.eb_title or 'ebook'}.pdf"
        return file_storage_service.get_secure_stream_response(
            ebook.eb_pdf_url,
            filename=filename,
            content_type='application/pdf'
        )
            
    except FileNotFoundError:
        raise Http404("E-Book file not found")
    except ValueError as e:
        raise Http404(str(e))
    except Exception as e:
        print(f"[read_ebook_secure] Error: {str(e)}")
        raise Http404("Invalid access or file not found")

# for tv display view and api

def book_to_dict(b):
    # Construct the full image URL
    photo_url = ""
    if b.front_page_photo:
        # Clean the photo path
        photo_path = b.front_page_photo
        if photo_path.startswith('/'):
            photo_path = photo_path[1:]
        photo_url = f"{settings.MEDIA_URL}{photo_path}"
    
    return {
        "cat_ref_num": b.cat_ref_num,
        "title": b.title or "",
        "subtitle": b.subtitle or "",
        "author": b.author or "",
        "publisher": b.publisher or "",
        "publication_year": b.publication_year or "",
        "year_of_publication": b.year_of_publication or "",
        "subject": getattr(b.subject, "subjectNameMarathi", "") if b.subject else "",
        "front_page_photo": photo_url,  # Return full URL
        "isbn_issn": b.isbn_issn or "",
        "edition": b.edition or "",
        "language": b.language or "",
        "publication_place": b.publication_place or "",
        "call_number": b.call_number or "",
        "cutter_number": b.cutter_number or "",
        "pages": b.pages or "",
        "keywords": b.keywords or "",
        "remarks": b.remarks or "",
        "classification_number": b.classification_number or "",
        "other_authors": b.other_authors or "",
    }

def tv_dashboard_page(request):
    library_code = request.session.get('library_db')
    username = request.session.get('username')
    user_id = request.session.get('user_id')
    role_id = request.session.get('role_id')
    
    if library_code != 'L01':
        messages.error(request, "Invalid library access.")
        request.session.flush()
        return redirect('library_list')
    
    libraryDetails = tbl_librarymasterL01.objects.filter(library_code__iexact=library_code).first()
    library_name = libraryDetails.library_name_mar if libraryDetails else "लायब्ररी"
    
    context = {
        "library_name_marathi": library_name,
        "MEDIA_URL": settings.MEDIA_URL,  # Add this line
    }
    return render(request, "L01/tv_display.html", context)

def tv_popular_books_api(request):
    qs = (
        BookCatalog.objects.annotate(
            review_count=Count('cat_ref_id_reviews'),
            avg_rating=Avg('cat_ref_id_reviews__rating')
        )
        .filter(review_count__gt=0)
        .order_by('-avg_rating', '-review_count')[:10]
    )

    data = []
    for b in qs:
        data.append({
            "cat_ref_num": b.cat_ref_num,
            "title": b.title or "",
            "author": b.author or "",
            "avg_rating": round(b.avg_rating or 0, 1),
            "review_count": b.review_count or 0,
            "subject": getattr(b.subject, "subjectNameMarathi", "") if b.subject else "",
            "front_page_photo": b.front_page_photo or "",
        })

    return JsonResponse({"popular_books": data}, safe=True)

def tv_categories_api(request):
    qs = SubjectTypeMaster.objects.filter(is_active=1).order_by('subjectNameMarathi')
    data = [{"id": s.id, "name_marathi": s.subjectNameMarathi or s.subjectNameEnglish or ""} for s in qs]
    return JsonResponse({"categories": data}, safe=True)

def tv_new_arrivals_api(request):
    latest_created_at = BookCatalog.objects.aggregate(latest=Max('created_at'))['latest']
    recent_books = BookCatalog.objects.none()
    if latest_created_at:
        recent_books = BookCatalog.objects.filter(created_at=latest_created_at)[:10]

    remaining = 10 - recent_books.count()
    fallback = BookCatalog.objects.exclude(cat_ref_num__in=recent_books.values_list('cat_ref_num', flat=True)) \
                                  .order_by('-cat_ref_num')[:remaining] if remaining > 0 else BookCatalog.objects.none()

    books = list(recent_books) + list(fallback)
    data = [book_to_dict(b) for b in books]
    return JsonResponse({"new_arrivals": data}, safe=True)

def tv_all_books_api(request):
    qs = BookCatalog.objects.select_related('subject', 'material').order_by('-cat_ref_num')[:50]
    data = []
    for b in qs:
        # Include all necessary fields for book details display
        data.append({
            "cat_ref_num": b.cat_ref_num,
            "title": b.title or "",
            "subtitle": b.subtitle or "",  # Add subtitle
            "author": b.author or "",
            "publisher": b.publisher or "",
            "publication_year": b.publication_year or "",
            "year_of_publication": b.year_of_publication or "",
            "subject": getattr(b.subject, "subjectNameMarathi", "") if b.subject else "",
            "front_page_photo": b.front_page_photo or "",  # ADD THIS LINE - CRITICAL!
            "isbn_issn": b.isbn_issn or "",
            "edition": b.edition or "",
            "language": b.language or "",
            "publication_place": b.publication_place or "",
            "call_number": b.call_number or "",
            "cutter_number": b.cutter_number or "",
            "pages": b.pages or "",
            "keywords": b.keywords or "",
            "remarks": b.remarks or "",
            "classification_number": b.classification_number or "",
            "other_authors": b.other_authors or "",
        })
    return JsonResponse({"all_books": data}, safe=True)

def insert_book_by_isbn(request, isbn):
    try:
        # 1. Call Google Books API
        url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
        response = requests.get(url).json()

        if response.get("totalItems", 0) == 0:
            return JsonResponse({"status": "error", "message": "Book not found in Google Books API"})

        book_data = response["items"][0]["volumeInfo"]

        # 2. Extract identifiers
        isbn_10 = isbn_13 = None
        for iden in book_data.get("industryIdentifiers", []):
            if iden["type"] == "ISBN_10":
                isbn_10 = iden["identifier"]
            elif iden["type"] == "ISBN_13":
                isbn_13 = iden["identifier"]

        title = book_data.get("title", "")
        description = book_data.get("description", "")
        language = book_data.get("language", "")

        # 3. Save or update BookMaster in L01
        with transaction.atomic(using="L01"):
            master_qs = BookMaster.objects.using("L01").filter(isbn_13=isbn_13)
            if master_qs.exists():
                master = master_qs.first()
                master.title = title or master.title
                master.description = description or master.description
                master.language = language or master.language
                master.updated_by = "System"
                master.save(using="L01")
            else:
                master = BookMaster.objects.using("L01").create(
                    isbn_10=isbn_10,
                    isbn_13=isbn_13,
                    title=title,
                    description=description,
                    language=language,
                    created_by="System"
                )

        # 4. Save BookDetails in L01
        authors = book_data.get("authors", ["Unknown Author"])
        publisher = book_data.get("publisher", "")
        published_date = book_data.get("publishedDate", "")
        page_count = book_data.get("pageCount", None)
        image_links = book_data.get("imageLinks", {})
        preview_link = image_links.get("thumbnail", "") 
        categories = book_data.get("categories", [])
        category = categories[0] if categories else None

        # Use the master's primary key instead of the object
        master_pk = master.pk

        for author in authors:
            # Check if this author already exists for this master
            detail_qs = BookDetails.objects.using("L01").filter(master_id=master_pk, author=author)
            if detail_qs.exists():
                detail = detail_qs.first()
                detail.publisher = publisher or detail.publisher
                detail.published_date = published_date or detail.published_date
                detail.page_count = page_count or detail.page_count
                detail.category = category or detail.category
                detail.preview_link = preview_link or detail.preview_link
                detail.updated_by = "System"
                detail.save(using="L01")
            else:
                # Create using the foreign key ID directly
                BookDetails.objects.using("L01").create(
                    master_id=master_pk,  # Use the ID instead of the object
                    author=author,
                    publisher=publisher,
                    published_date=published_date,
                    page_count=page_count,
                    category=category,
                    preview_link=preview_link,
                    edition=None,
                    created_by="System"
                )

        return JsonResponse({
            "status": "success",
            "message": "Book inserted successfully",
            "book_title": title,
            "authors": authors
        })

    except Exception as e:
        print("Error in insert_book_by_isbn:", e)
        return JsonResponse({"status": "error", "message": str(e)})

# Stock Checking by Imran

@csrf_exempt
def scan_barcode(request):
    try:
        if request.method == "GET":
            stock_year = StockYearMaster.objects.filter(is_active = 1)

            return render(request, "L01/Stock/stock_check.html", {"stock_year":stock_year})

        if request.method == "POST":
            barcode = request.POST.get("barcode")
            stock_year_id = request.POST.get("stock_year_id")

            # Validate stock year
            try:
                stock_year = StockYearMaster.objects.get(id=stock_year_id)
            except StockYearMaster.DoesNotExist:
                return JsonResponse({"error": "Invalid Stock Year"}, status=400)

            stock_master, created = StockMaster.objects.get_or_create(
                stock_year=stock_year,
                defaults={"is_completed": False}             
            )
            try:
                circulation = CirculationCopyStatus.objects.get(barcode=barcode)
            except CirculationCopyStatus.DoesNotExist:
                return JsonResponse(
                    {"error": f"Barcode {barcode} not found in circulation"},
                    status=404
                )

            if StockDetail.objects.filter(stock=stock_master, barcode=barcode).exists():
                return JsonResponse(
                    {"error": f"Barcode {barcode} already scanned"},
                    status=400
                )

            StockDetail.objects.create(
                stock=stock_master,
                stock_year=stock_year,
                circulation=circulation,
                barcode=barcode
            )

            return JsonResponse({
                "message": f"Barcode {barcode} scanned successfully",
                "stock_master_id": stock_master.id
            })

        return JsonResponse({"error": "Invalid request"}, status=400)
    except Exception as e:
        print("Error in insert_book_by_isbn:", e)
        return JsonResponse({"status": "error", "message": str(e)})

def get_recent_scans(request):
    year_id = request.GET.get("stock_year_id")

    scans = StockDetail.objects.filter(stock_year_id=year_id).order_by('-id')[:20]

    data = []
    for s in scans:
        data.append({
            "barcode": s.barcode,
            "time": s.scanned_at.strftime("%Y-%m-%d %H:%M:%S")
        })

    return JsonResponse({"data": data})

# Stock Report Imran

def complete_stock_batch(request):
    try:
        data = json.loads(request.body)
        year_id = data.get('year_id')
        
        if not year_id:
            return JsonResponse({'success': False, 'error': 'Year ID is required'})
        
        
        try:
            stock_year = StockMaster.objects.get(id=year_id)
            stock_year.is_completed = True
            stock_year.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Stock batch has been marked as completed.'
            })
            
        except StockYearMaster.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Stock year not found'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
class StockReportView(View):
    def get(self, request):
        """
        Handle GET request - Display stock year selection form
        """
        try:
            # Stock years for dropdown
            stock_years = StockYearMaster.objects.all().order_by('-year_name')

            # Get active year safely (NO 404 break)
            active_year = StockYearMaster.objects.filter(is_active=1).first()

            if active_year:
                stock_data = StockMaster.objects.filter(stock_year=active_year)
                active_year_name = active_year.year_name
            else:
                stock_data = []
                active_year_name = "Select a year"

            context = {
                'stock_years': stock_years,
                'active_year_name': active_year_name,
                'stock_data': stock_data,
                'report_data': None,
            }

            return render(request, 'L01/Stock/stock_report.html', context)

        except Exception as e:
            context = {
                'stock_years': [],
                'active_year_name': 'Error',
                'stock_data': [],
                'report_data': self.create_empty_report_data(f"System error: {str(e)}"),
            }
            return render(request, 'L01/Stock/stock_report.html', context)
    
    def post(self, request):
        """
        Handle POST request - Generate report for selected year
        """
        try:
            year_id = request.POST.get('stock_year')
            
            if not year_id:
                return redirect('L01:stock_report')
            
            # Get selected stock year
            stock_year = StockYearMaster.objects.get(id=year_id)
            
            # Get all stock years for dropdown (to keep dropdown populated)
            stock_years = StockYearMaster.objects.all().order_by('-year_name')
            
            # Generate report for selected stock year
            report_data = self.generate_report_data(year_id)

            save_data = self.save_report_to_database(report_data, request.user, stock_year)
            
            context = {
                'stock_years': stock_years,
                'selected_year_id': int(year_id),
                'active_year_name': stock_year.year_name,
                'report_data': report_data,
            }
            
            return render(request, 'L01/Stock/stock_report.html', context)
            
        except StockYearMaster.DoesNotExist:
            return redirect('L01:stock_report')
        except Exception as e:
            # Get all stock years for dropdown
            stock_years = StockYearMaster.objects.all().order_by('-year_name')
            
            context = {
                'stock_years': stock_years,
                'active_year_name': 'Error',
                'report_data': self.create_empty_report_data(f"Error: {str(e)}"),
            }
            return render(request, 'L01/Stock/stock_report.html', context)
    
    def generate_report_data(self, year_id):
        """
        Generate stock verification report data
        Compare CirculationCopyStatus (all books) with StockDetail (scanned books)
        """
        # Get stock year
        stock_year = StockYearMaster.objects.get(id=year_id)
        
        # Get ALL books from CirculationCopyStatus
        all_books = CirculationCopyStatus.objects.select_related(
            'bookcatalog',
            'shelf_location',
            'current_status'
        ).order_by('barcode')
        
        # Get books that were scanned in stock verification for this year
        scanned_book_ids = StockDetail.objects.filter(
            stock_year_id=year_id
        ).values_list('circulation_id', flat=True)
        
        # Convert to set for faster lookup
        scanned_books_set = set(scanned_book_ids)
        
        # Prepare report data
        report_data = []
        status_counts = {
            'scanned': 0,
            'unknown': 0,
            'total': 0
        }
        
        for book in all_books:
            # Check if this book was scanned during stock verification
            is_scanned = book.id in scanned_books_set
            
            # Get book catalog details
            title = 'N/A'
            if book.bookcatalog:
                title = book.bookcatalog.title
            
            # Get shelf location
            shelf_location = 'N/A'
            if book.shelf_location:
                shelf_location = book.shelf_location.location_name
            
            # Get current status from StatusMaster
            current_status = 'N/A'
            if book.current_status:
                current_status = book.current_status.status_name
            
            # Determine Stock Status
            if is_scanned:
                stock_status = 'Scanned'
                status_key = 'scanned'
            else:
                stock_status = 'Unknown'
                status_key = 'unknown'
            
            # Update counts
            status_counts[status_key] += 1
            status_counts['total'] += 1
            
            # Format date processed
            date_processed = 'N/A'
            if book.date_processed:
                date_processed = book.date_processed.strftime('%Y-%m-%d')
            
            book_data = {
                'barcode': book.barcode,
                'title': title,
                'shelf_location': shelf_location,
                'current_status': current_status,
                'stock_status': stock_status,
                'date_processed': date_processed,
            }
            
            report_data.append(book_data)

            
        
        return {
            'books': report_data,
            'summary': status_counts,
            'stock_year': stock_year.year_name,
            'stock_year_id': year_id,
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def create_empty_report_data(self, message="No data available"):
        """
        Create empty report data structure
        """
        return {
            'books': [],
            'summary': {
                'total': 0,
                'scanned': 0,
                'unknown': 0
            },
            'stock_year': message,
            'stock_year_id': None,
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def save_report_to_database(self, report_data, user, stock_year):
        """
        Save generated report to StockReport table (optimized for speed)
        """
        try:
            # Delete old reports for this stock year
            StockReport.objects.filter(stock_year=stock_year).delete()

            username = user.username if user else "system"

            # 🔥 Load ALL locations in one query
            all_locations = ResourceLocationMaster.objects.all()

            # 🔥 Create dictionary: location_name → object
            location_map = {loc.location_name: loc for loc in all_locations}

            report_objects = []

            for book in report_data["books"]:
                loc_name = book["shelf_location"]

                # Lookup location without DB query
                shelf_location = location_map.get(loc_name)

                if not shelf_location:
                    print(f"Warning: Shelf location not found: {loc_name}")
                    continue

                report_objects.append(
                    StockReport(
                        stock_year=stock_year,
                        barcode=book["barcode"],
                        title=book["title"],
                        shelf_location=shelf_location,
                        current_status=book["current_status"],
                        stock_status=book["stock_status"],
                        date_processed=book["date_processed"],
                        created_by=username,
                        updated_by=username,
                    )
                )

            # 🔥 Bulk insert (FAST!)
            StockReport.objects.bulk_create(report_objects, batch_size=500)

            print(
                f"Successfully saved {len(report_objects)} records to StockReport "
                f"for stock year {stock_year.year_name}"
            )

        except Exception as e:
            print(f"Error saving report: {str(e)}")
            print(traceback.format_exc())
class GenerateStockReportAPI(View):
    """
    API endpoint for AJAX report generation
    """
    def post(self, request):
        try:
            data = json.loads(request.body)
            year_id = data.get('year_id')
            
            if not year_id:
                return JsonResponse({'error': 'Year ID is required'}, status=400)
            
            stock_year = get_object_or_404(StockYearMaster, id = year_id )
            
            # Create an instance of StockReportView to use its methods
            report_view = StockReportView()
            report_data = report_view.generate_report_data(year_id)
            save_data = report_view.save_report_to_database(report_data, request.user, stock_year)
            
            
            return JsonResponse({
                'success': True,
                'report_data': report_data
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
class ExportStockReportView(View):
    """
    View for exporting stock reports in different formats
    """
    def post(self, request):
        try:
            # Parse JSON data
            data = json.loads(request.body)
            format_type = data.get('format', 'excel')
            year_id = data.get('year_id')
            
            if not year_id:
                return JsonResponse({'error': 'Year ID is required'}, status=400)
            
            # Get stock year
            try:
                stock_year = StockYearMaster.objects.get(id=year_id)
            except StockYearMaster.DoesNotExist:
                return JsonResponse({'error': 'Stock year not found'}, status=404)
            
            # Generate report data for selected stock year
            report_view = StockReportView()
            report_data = report_view.generate_report_data(year_id)
            
            # Export based on format
            if format_type == 'excel':
                return self.export_to_excel(report_data)
            elif format_type == 'pdf':
                return self.export_to_pdf(report_data)
            elif format_type == 'csv':
                return self.export_to_csv(report_data)
            else:
                return JsonResponse({'error': 'Invalid export format'}, status=400)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    def export_to_excel(self, report_data):
        """
        Export report to Excel format
        """
        # Create DataFrame from report data
        try:
            df = pd.DataFrame(report_data['books'])
            columns_order = ['barcode', 'title', 'shelf_location', 'current_status', 'stock_status', 'date_processed']
            df = df[columns_order]
            
            # Rename columns for better display
            df.columns = ['बारकोड', 'शीर्षक', 'शेल्फ स्थान', 'सध्याची स्थिती', 'स्टॉक स्थिती', 'प्रक्रिया दिनांक']
            
            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f"स्टॉक_सत्यापन_अहवाल_{report_data['stock_year'].replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Create workbook directly
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "स्टॉक सत्यापन"
            
            # Header rows
            # Row 1: Logo space (we'll leave it empty or try to add later)
            ws.merge_cells('A1:F1')
            
            # Row 2: Title
            ws.merge_cells('A2:F2')
            ws['A2'] = 'विष्णुदास भावे नाट्य ग्रंथालय'
            ws['A2'].font = Font(size=16, bold=True, name='Arial')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Row 3: Report Title
            ws.merge_cells('A3:F3')
            ws['A3'] = 'स्टॉक सत्यापन अहवाल'
            ws['A3'].font = Font(size=14, bold=True, name='Arial')
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Row 4-7: Report Info
            ws['A4'] = f'स्टॉक वर्ष: {report_data["stock_year"]}'
            ws['A5'] = f'निर्मिती दिनांक: {report_data["generated_at"]}'
            ws['A6'] = f'एकूण पुस्तके: {report_data["summary"]["total"]}'
            ws['A7'] = f'स्कॅन केलेली: {report_data["summary"]["scanned"]}'
            ws['D6'] = f'अज्ञात स्थिती: {report_data["summary"]["unknown"]}'
            
            # Apply styling to info cells
            for row in range(4, 8):
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        cell.font = Font(name='Arial', size=11)
            
            # Row 8: Empty row
            ws.row_dimensions[8].height = 15
            
            # Row 9: Table Headers (starting from row 9)
            headers = ['बारकोड', 'शीर्षक', 'शेल्फ स्थान', 'सध्याची स्थिती', 'स्टॉक स्थिती', 'प्रक्रिया दिनांक']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, name='Arial')
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Row 10 onwards: Data
            start_row = 10
            for idx, row in df.iterrows():
                for col_num, (col_name, value) in enumerate(row.items(), 1):
                    cell = ws.cell(row=start_row, column=col_num)
                    cell.value = value
                    cell.font = Font(name='Arial')
                
                # Alternate row coloring
                if start_row % 2 == 0:
                    for col_num in range(1, 7):
                        ws.cell(row=start_row, column=col_num).fill = PatternFill(
                            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'
                        )
                
                start_row += 1
            
            # Adjust column widths
            for col_num, column_title in enumerate(headers, 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                # Check header length
                if len(str(column_title)) > max_length:
                    max_length = len(str(column_title))
                
                # Check data length in this column
                for row in range(10, start_row):
                    cell_value = ws.cell(row=row, column=col_num).value
                    if cell_value:
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                # Set column width (with some padding)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook to response
            wb.save(response)
        
            return response
        except Exception as e:
            return JsonResponse(
                {"error": "An unexpected error occurred.", "details": str(e)},
                status=500
                )
    
    def export_to_csv(self, report_data):
        """
        Export report to CSV format
        """
        try:
        # Create DataFrame from report data
            df = pd.DataFrame(report_data['books'])
            
            # Reorder and rename columns
            columns_order = ['barcode', 'title', 'shelf_location', 'current_status', 'stock_status', 'date_processed']
            df = df[columns_order]
            df.columns = ['बारकोड', 'शीर्षक', 'शेल्फ स्थान', 'सध्याची स्थिती', 'स्टॉक स्थिती', 'प्रक्रिया दिनांक']
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            filename = f"स्टॉक_सत्यापन_अहवाल_{report_data['stock_year'].replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Write CSV with BOM for UTF-8
            response.write('\ufeff')  # UTF-8 BOM
            
            # Write header
            header_lines = [
                'विष्णुदास भावे नाट्य ग्रंथालय',
                'स्टॉक सत्यापन अहवाल',
                '',
                f'स्टॉक वर्ष: {report_data["stock_year"]}',
                f'निर्मिती दिनांक: {report_data["generated_at"]}',
                f'एकूण पुस्तके: {report_data["summary"]["total"]}',
                f'स्कॅन केलेली: {report_data["summary"]["scanned"]}',
                f'अज्ञात स्थिती: {report_data["summary"]["unknown"]}',
                '',
                'तपशीलवार अहवाल',
                ''
            ]
            
            for line in header_lines:
                response.write(line + '\n')
            
            # Write data
            df.to_csv(response, index=False, encoding='utf-8-sig', mode='a')
            
            return response
        except Exception as e:
            return JsonResponse(
                {"error": "An unexpected error occurred.", "details": str(e)},
                status=500
            )
    
    def export_to_pdf(self, report_data):
        """
        Export report to PDF format using WeasyPrint for better Marathi support
        """
        try:
            from weasyprint import HTML, CSS
        except ImportError:
            # Fallback to ReportLab if WeasyPrint is not installed
            return self.export_to_pdf_fallback(report_data)
        
        import io
        import html
        
        # Create HTML content with Marathi text
        html_content = f"""
        <!DOCTYPE html>
        <html lang="mr">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>स्टॉक सत्यापन अहवाल - {report_data['stock_year']}</title>
            <style>
                @page {{
                    size: A4;
                    margin: 1.5cm;
                    
                    @top-center {{
                        content: "विष्णुदास भावे नाट्य ग्रंथालय";
                        font-size: 14px;
                        color: #2c3e50;
                        margin-top: 0.5cm;
                    }}
                    
                    @bottom-center {{
                        content: "पृष्ठ " counter(page) " / " counter(pages);
                        font-size: 10px;
                        margin-bottom: 0.5cm;
                    }}
                }}
                
                body {{
                    font-family: 'Noto Sans Devanagari', 'Arial Unicode MS', 'Nirmala UI', sans-serif;
                    line-height: 1.4;
                    color: #333;
                }}
                
                .header-container {{
                    text-align: center;
                    margin-bottom: 25px;
                }}
                
                .library-name {{
                    font-size: 22px;
                    font-weight: bold;
                    color: #2c3e50;
                    margin: 10px 0;
                }}
                
                .report-title {{
                    font-size: 18px;
                    font-weight: bold;
                    color: #2980b9;
                    margin: 5px 0 20px 0;
                }}
                
                .report-info {{
                    background-color: #f8f9fa;
                    padding: 15px;
                    border-radius: 5px;
                    border-left: 4px solid #3498db;
                    margin-bottom: 25px;
                }}
                
                .info-row {{
                    margin-bottom: 8px;
                    font-size: 13px;
                }}
                
                .info-label {{
                    font-weight: bold;
                    color: #2c3e50;
                    display: inline-block;
                    width: 140px;
                }}
                
                .info-value {{
                    color: #555;
                }}
                
                .data-table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                    font-size: 11px;
                }}
                
                .data-table th {{
                    background-color: #2c3e50;
                    color: white;
                    padding: 10px 8px;
                    text-align: left;
                    font-weight: bold;
                    border: 1px solid #ddd;
                }}
                
                .data-table td {{
                    padding: 8px;
                    border: 1px solid #ddd;
                    text-align: left;
                }}
                
                .data-table tr:nth-child(even) {{
                    background-color: #f9f9f9;
                }}
                
                .data-table tr:hover {{
                    background-color: #f5f5f5;
                }}
                
                .footer {{
                    margin-top: 30px;
                    padding-top: 15px;
                    border-top: 1px solid #eee;
                    font-size: 12px;
                    text-align: center;
                }}
                
                .summary {{
                    font-weight: bold;
                    color: #2c3e50;
                    margin-bottom: 5px;
                }}
                
                .timestamp {{
                    color: #7f8c8d;
                    font-size: 11px;
                }}
                
                /* Print styles */
                @media print {{
                    body {{
                        margin: 0;
                        padding: 0;
                    }}
                    
                    .no-print {{
                        display: none;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="header-container">
                <!-- Logo would go here if we could embed images -->
                <div class="library-name">विष्णुदास भावे नाट्य ग्रंथालय</div>
                <div class="report-title">स्टॉक सत्यापन अहवाल</div>
            </div>
            
            <div class="report-info">
                <div class="info-row">
                    <span class="info-label">स्टॉक वर्ष:</span>
                    <span class="info-value">{html.escape(str(report_data['stock_year']))}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">निर्मिती दिनांक:</span>
                    <span class="info-value">{html.escape(str(report_data['generated_at']))}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">एकूण पुस्तके:</span>
                    <span class="info-value">{html.escape(str(report_data['summary']['total']))}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">स्कॅन केलेली:</span>
                    <span class="info-value">{html.escape(str(report_data['summary']['scanned']))}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">अज्ञात स्थिती:</span>
                    <span class="info-value">{html.escape(str(report_data['summary']['unknown']))}</span>
                </div>
            </div>
            
            <table class="data-table">
                <thead>
                    <tr>
                        <th width="5%">क्र.</th>
                        <th width="15%">बारकोड</th>
                        <th width="35%">शीर्षक</th>
                        <th width="15%">शेल्फ स्थान</th>
                        <th width="15%">सध्याची स्थिती</th>
                        <th width="15%">स्टॉक स्थिती</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        # Add table rows
        books = report_data['books']
        for i, book in enumerate(books):
            barcode = html.escape(str(book.get('barcode', '')))
            title = html.escape(str(book.get('title', '')))
            shelf_location = html.escape(str(book.get('shelf_location', '')))
            current_status = html.escape(str(book.get('current_status', '')))
            stock_status = html.escape(str(book.get('stock_status', '')))
            
            html_content += f"""
                    <tr>
                        <td>{i + 1}</td>
                        <td>{barcode}</td>
                        <td>{title}</td>
                        <td>{shelf_location}</td>
                        <td>{current_status}</td>
                        <td>{stock_status}</td>
                    </tr>
            """
        
        # Close HTML
        html_content += f"""
                </tbody>
            </table>
            
            <div class="footer">
                <div class="summary">एकूण नोंदी: {len(books)}</div>
                <div class="timestamp">हा अहवाल {timezone.now().strftime('%d/%m/%Y %H:%M:%S')} रोजी निर्माण करण्यात आला</div>
            </div>
        </body>
        </html>
        """
        
        try:
            # Generate PDF with WeasyPrint
            html = HTML(string=html_content)

            font_regular = os.path.join(settings.BASE_DIR,"static","fonts","NotoSansDevanagari-Regular.ttf")

            font_bold = os.path.join(settings.BASE_DIR,"static","fonts","NotoSerifDevanagari-Bold.ttf")

            
            # Add CSS for better typography
            css_string = f"""
                @font-face {{
                    font-family: 'Noto Sans Devanagari';
                    src: url('file://{font_regular}');
                    font-weight: normal;
                    font-style: normal;
                }}

                @font-face {{
                    font-family: 'Noto Sans Devanagari';
                    src: url('file://{font_bold}');
                    font-weight: bold;
                    font-style: normal;
                }}

                body {{
                    font-family: 'Noto Sans Devanagari', 'Arial Unicode MS', sans-serif;
                    text-rendering: optimizeLegibility;
                }}
                """

            
            css = CSS(string=css_string)
            pdf_data = html.write_pdf(stylesheets=[css])
            
        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name if tb else 'library_list'
            callproc("stp_error_log", [fun, str(e), ''])
            # Fallback to basic HTML to PDF without CSS
            try:
                html = HTML(string=html_content)
                pdf_data = html.write_pdf()
            except Exception as e2:
                print(f"Even basic WeasyPrint failed: {e2}")
                return self.export_to_pdf_fallback(report_data)
        
        # Create HTTP response
        response = HttpResponse(pdf_data, content_type='application/pdf')
        
        # Create filename
        from django.utils.text import slugify
        safe_year = slugify(report_data['stock_year'])
        filename = f"stock_verification_report_{safe_year}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    def export_to_pdf_fallback(self, report_data):
        """
        Fallback PDF export using ReportLab if WeasyPrint fails
        """
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        import io
        
        buffer = io.BytesIO()
        
        # Create simple PDF with ReportLab
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title in English
        title = Paragraph("Vishnudas Bhave Natya Granthalay", styles['Heading1'])
        story.append(title)
        
        subtitle = Paragraph("Stock Verification Report", styles['Heading2'])
        story.append(subtitle)
        
        # Add spacer
        from reportlab.platypus import Spacer
        story.append(Spacer(1, 20))
        
        # Report info
        info_text = f"""
        <b>Stock Year:</b> {report_data['stock_year']}<br/>
        <b>Generated:</b> {report_data['generated_at']}<br/>
        <b>Total Books:</b> {report_data['summary']['total']}<br/>
        <b>Scanned:</b> {report_data['summary']['scanned']}<br/>
        <b>Unknown:</b> {report_data['summary']['unknown']}<br/>
        """
        info = Paragraph(info_text, styles['Normal'])
        story.append(info)
        
        story.append(Spacer(1, 20))
        
        # Create table
        table_data = []
        headers = ['Sr', 'Barcode', 'Title', 'Shelf Location', 'Current Status', 'Stock Status']
        table_data.append(headers)
        
        books = report_data['books'][:50]  # Limit rows for PDF
        for i, book in enumerate(books):
            row = [
                str(i + 1),
                str(book.get('barcode', ''))[:15],
                str(book.get('title', ''))[:30],
                str(book.get('shelf_location', ''))[:15],
                str(book.get('current_status', ''))[:15],
                str(book.get('stock_status', ''))[:15]
            ]
            table_data.append(row)
        
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(table)
        
        # Footer
        story.append(Spacer(1, 20))
        footer = Paragraph(f"<b>Total Records:</b> {len(report_data['books'])}", styles['Normal'])
        story.append(footer)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        pdf_data = buffer.getvalue()
        
        # Create response
        response = HttpResponse(pdf_data, content_type='application/pdf')
        filename = f"stock_report_{report_data['stock_year'].replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

def generate_final_report(request):
    """
    Generate final PDF report from StockReport table data
    """
    try:
        data = json.loads(request.body)
        year_id = data.get('year_id')
         
        if not year_id:
            return JsonResponse({'success': False, 'error': 'Year ID is required'})
        
        # Get the stock year
        stock_year = StockYearMaster.objects.using('L01').get(id=year_id)
        
        stock_reports = StockReport.objects.using('L01').filter(
            stock_year=stock_year
        ).select_related('shelf_location').order_by('barcode')
        
        if not stock_reports.exists():
            return JsonResponse({'success': False, 'error': 'No report data found for this stock year'})
        
        # Prepare report data from StockReport table
        report_data = prepare_report_data_from_stock_reports(stock_reports, stock_year)
        
        # Generate PDF
        response = export_final_report_to_pdf(report_data)
        return response
            
    except Exception as e:
        error_trace = traceback.format_exc()
        error_log.objects.create(
            method='generate_final_report',
            error=str(e),
            error_date=timezone.now()
        )

        return JsonResponse({
            'success': False,
            'error': 'Internal server error while generating report'
        }, status=500)

def prepare_report_data_from_stock_reports(stock_reports, stock_year):
    try:
        """
        Prepare report data from StockReport queryset
        """
        books = []
        status_counts = {
            'scanned': 0,
            'unknown': 0,
            'total': 0
        }
        
        for report in stock_reports:
            # Get shelf location name
            shelf_location = 'N/A'
            if report.shelf_location:
                shelf_location = report.shelf_location.location_name
            
            # Get date processed as string
            date_processed = 'N/A'
            if report.date_processed:
                date_processed = report.date_processed.strftime('%Y-%m-%d')
            
            # Update status counts
            stock_status_lower = report.stock_status.lower()
            if 'scanned' in stock_status_lower:
                status_counts['scanned'] += 1
            elif 'unknown' in stock_status_lower:
                status_counts['unknown'] += 1
            status_counts['total'] += 1
            
            book_data = {
                'barcode': report.barcode,
                'title': report.title,
                'shelf_location': shelf_location,
                'current_status': report.current_status,
                'stock_status': report.stock_status,
                'date_processed': date_processed,
            }
            books.append(book_data)
        
        return {
            'books': books,
            'summary': status_counts,
            'stock_year': stock_year.year_name,
            'stock_year_id': stock_year.id,
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    except Exception as e:
        error_trace = traceback.format_exc()
        error_log.objects.create(
            method='generate_final_report',
            error=str(e),
            error_date=timezone.now()
        )

def export_final_report_to_pdf(report_data):
    try:
        """
        Export final report to PDF format from StockReport data
        """
        try:
            from weasyprint import HTML, CSS
            use_weasyprint = True
        except ImportError:
            use_weasyprint = False
        
        import io
        import html as html_module  # Import html module with alias to avoid conflict
        
        # Create HTML content with Marathi text
        html_content = f"""
        <!DOCTYPE html>
        <html lang="mr">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>स्टॉक सत्यापन अंतिम अहवाल - {report_data['stock_year']}</title>
            <style>
                @page {{
                    size: A4;
                    margin: 1.5cm;
                    
                    @top-center {{
                        content: "विष्णुदास भावे नाट्य ग्रंथालय - अंतिम स्टॉक अहवाल";
                        font-size: 14px;
                        color: #2c3e50;
                        margin-top: 0.5cm;
                    }}
                    
                    @bottom-center {{
                        content: "पृष्ठ " counter(page) " / " counter(pages);
                        font-size: 10px;
                        margin-bottom: 0.5cm;
                    }}
                }}
                
                body {{
                    font-family: 'Noto Sans Devanagari', 'Arial Unicode MS', 'Nirmala UI', sans-serif;
                    line-height: 1.4;
                    color: #333;
                }}
                
                .header-container {{
                    text-align: center;
                    margin-bottom: 25px;
                }}
                
                .library-name {{
                    font-size: 22px;
                    font-weight: bold;
                    color: #2c3e50;
                    margin: 10px 0;
                }}
                
                .report-title {{
                    font-size: 18px;
                    font-weight: bold;
                    color: #d35400;
                    margin: 5px 0 20px 0;
                    border-bottom: 2px solid #d35400;
                    padding-bottom: 10px;
                }}
                
                .report-subtitle {{
                    font-size: 14px;
                    color: #7f8c8d;
                    margin-bottom: 25px;
                    font-style: italic;
                }}
                
                .report-info {{
                    background-color: #f8f9fa;
                    padding: 15px;
                    border-radius: 5px;
                    border-left: 4px solid #3498db;
                    margin-bottom: 25px;
                }}
                
                .info-row {{
                    margin-bottom: 8px;
                    font-size: 13px;
                }}
                
                .info-label {{
                    font-weight: bold;
                    color: #2c3e50;
                    display: inline-block;
                    width: 150px;
                }}
                
                .info-value {{
                    color: #555;
                }}
                
                .summary-stats {{
                    display: flex;
                    justify-content: space-around;
                    margin: 25px 0;
                    flex-wrap: wrap;
                }}
                
                .stat-box {{
                    background: linear-gradient(45deg, #3498db, #2980b9);
                    color: white;
                    padding: 15px;
                    border-radius: 8px;
                    text-align: center;
                    min-width: 150px;
                    margin: 5px;
                }}
                
                .stat-value {{
                    font-size: 24px;
                    font-weight: bold;
                    margin: 5px 0;
                }}
                
                .stat-label {{
                    font-size: 14px;
                }}
                
                .data-table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                    font-size: 11px;
                }}
                
                .data-table th {{
                    background-color: #2c3e50;
                    color: white;
                    padding: 10px 8px;
                    text-align: left;
                    font-weight: bold;
                    border: 1px solid #ddd;
                }}
                
                .data-table td {{
                    padding: 8px;
                    border: 1px solid #ddd;
                    text-align: left;
                }}
                
                .data-table tr:nth-child(even) {{
                    background-color: #f9f9f9;
                }}
                
                .status-scanned {{
                    color: #27ae60;
                    font-weight: bold;
                }}
                
                .status-unknown {{
                    color: #e74c3c;
                    font-weight: bold;
                }}
                
                .footer {{
                    margin-top: 30px;
                    padding-top: 15px;
                    border-top: 1px solid #eee;
                    font-size: 12px;
                    text-align: center;
                }}
                
                .summary {{
                    font-weight: bold;
                    color: #2c3e50;
                    margin-bottom: 5px;
                }}
                
                .timestamp {{
                    color: #7f8c8d;
                    font-size: 11px;
                }}
                
                .completed-badge {{
                    background-color: #27ae60;
                    color: white;
                    padding: 3px 10px;
                    border-radius: 12px;
                    font-size: 11px;
                    margin-left: 10px;
                }}
                
                /* Print styles */
                @media print {{
                    body {{
                        margin: 0;
                        padding: 0;
                    }}
                    
                    .no-print {{
                        display: none;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="header-container">
                <div class="library-name">विष्णुदास भावे नाट्य ग्रंथालय</div>
                <div class="report-title">स्टॉक सत्यापन अंतिम अहवाल</div>
                <div class="report-subtitle">(Stock Verification Final Report)</div>
            </div>
            
            <div class="report-info">
                <div class="info-row">
                    <span class="info-label">स्टॉक वर्ष:</span>
                    <span class="info-value">{html_module.escape(str(report_data['stock_year']))}</span>
                    <span class="completed-badge">✓ पूर्ण</span>
                </div>
                <div class="info-row">
                    <span class="info-label">अहवाल क्र.:</span>
                    <span class="info-value">SR/{report_data['stock_year_id']}/{timezone.now().strftime('%Y%m%d')}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">निर्मिती दिनांक:</span>
                    <span class="info-value">{html_module.escape(str(report_data['generated_at']))}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">डेटा स्रोत:</span>
                    <span class="info-value">स्टॉक अहवाल डेटाबेस</span>
                </div>
            </div>
            
            <div class="summary-stats">
                <div class="stat-box">
                    <div class="stat-value">{report_data['summary']['total']}</div>
                    <div class="stat-label">एकूण पुस्तके</div>
                </div>
                <div class="stat-box" style="background: linear-gradient(45deg, #27ae60, #229954);">
                    <div class="stat-value">{report_data['summary']['scanned']}</div>
                    <div class="stat-label">स्कॅन केलेली</div>
                </div>
                <div class="stat-box" style="background: linear-gradient(45deg, #e74c3c, #c0392b);">
                    <div class="stat-value">{report_data['summary']['unknown']}</div>
                    <div class="stat-label">अज्ञात स्थिती</div>
                </div>
                <div class="stat-box" style="background: linear-gradient(45deg, #9b59b6, #8e44ad);">
                    <div class="stat-value">{report_data['summary']['scanned'] / report_data['summary']['total'] * 100 if report_data['summary']['total'] > 0 else 0:.1f}%</div>
                    <div class="stat-label">स्कॅन दर</div>
                </div>
            </div>
            
            <table class="data-table">
                <thead>
                    <tr>
                        <th width="5%">क्र.</th>
                        <th width="15%">बारकोड</th>
                        <th width="30%">शीर्षक</th>
                        <th width="15%">शेल्फ स्थान</th>
                        <th width="15%">सध्याची स्थिती</th>
                        <th width="15%">स्टॉक स्थिती</th>
                        <th width="10%">तारीख</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        # Add table rows
        books = report_data['books']
        for i, book in enumerate(books):
            barcode = html_module.escape(str(book.get('barcode', '')))
            title = html_module.escape(str(book.get('title', '')))
            shelf_location = html_module.escape(str(book.get('shelf_location', '')))
            current_status = html_module.escape(str(book.get('current_status', '')))
            stock_status = html_module.escape(str(book.get('stock_status', '')))
            date_processed = html_module.escape(str(book.get('date_processed', '')))
            
            status_class = 'status-scanned' if 'scanned' in book.get('stock_status', '').lower() else 'status-unknown'
            
            html_content += f"""
                    <tr>
                        <td>{i + 1}</td>
                        <td>{barcode}</td>
                        <td>{title}</td>
                        <td>{shelf_location}</td>
                        <td>{current_status}</td>
                        <td class="{status_class}">{stock_status}</td>
                        <td>{date_processed}</td>
                    </tr>
            """
        
        # Close HTML
        html_content += f"""
                </tbody>
            </table>
            
            <div class="footer">
                <div class="summary">एकूण नोंदी: {len(books)}</div>
                <div class="summary">स्टॉक सत्यापन पूर्ण झाले: {timezone.now().strftime('%d/%m/%Y')}</div>
                <div class="timestamp">हा अंतिम अहवाल {timezone.now().strftime('%d/%m/%Y %H:%M:%S')} रोजी निर्माण करण्यात आला</div>
                <div class="timestamp" style="margin-top: 10px;">
                    ** हा एक अधिकृत अहवाल आहे जो स्टॉक सत्यापन प्रक्रिया पूर्ण झाल्यानंतर निर्माण करण्यात आला आहे **
                </div>
            </div>
        </body>
        </html>
        """
        
        pdf_data = None
        
        if use_weasyprint:
            try:
                # Generate PDF with WeasyPrint
                weasyprint_html = HTML(string=html_content)  # Use different variable name

                font_regular = os.path.join(settings.BASE_DIR,"static","fonts","NotoSansDevanagari-Regular.ttf")

                font_bold = os.path.join(settings.BASE_DIR,"static","fonts","NotoSerifDevanagari-Bold.ttf")

                
                # Add CSS for better typography
                css_string = f"""
                    @font-face {{
                        font-family: 'Noto Sans Devanagari';
                        src: url('file://{font_regular}');
                        font-weight: normal;
                        font-style: normal;
                    }}

                    @font-face {{
                        font-family: 'Noto Sans Devanagari';
                        src: url('file://{font_bold}');
                        font-weight: bold;
                        font-style: normal;
                    }}

                    body {{
                        font-family: 'Noto Sans Devanagari', 'Arial Unicode MS', sans-serif;
                        text-rendering: optimizeLegibility;
                    }}
                """

                
                css = CSS(string=css_string)
                pdf_data = weasyprint_html.write_pdf(stylesheets=[css])
                
            except Exception as e:
                print(f"WeasyPrint generation error: {e}")
                # Fallback to basic HTML to PDF without CSS
                try:
                    weasyprint_html = HTML(string=html_content)
                    pdf_data = weasyprint_html.write_pdf()
                except Exception as e2:
                    print(f"Even basic WeasyPrint failed: {e2}")
                    use_weasyprint = False
        
        if not use_weasyprint or pdf_data is None:
            # Fallback to ReportLab
            pdf_data = export_final_report_fallback(report_data)
        
        # Create HTTP response
        response = HttpResponse(pdf_data, content_type='application/pdf')
        
        # Create filename
        from django.utils.text import slugify
        safe_year = slugify(report_data['stock_year'])
        filename = f"final_stock_report_{safe_year}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    except Exception as e:
        error_trace = traceback.format_exc()
        error_log.objects.create(
            method='generate_final_report',
            error=str(e),
            error_date=timezone.now()
        )

def export_final_report_fallback(report_data):
    """
    Fallback PDF export using ReportLab for final report
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io
    
    buffer = io.BytesIO()
    
    # Use landscape orientation for more columns
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        alignment=TA_CENTER,
        spaceAfter=20,
        textColor=colors.HexColor('#2c3e50')
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        alignment=TA_CENTER,
        spaceAfter=30,
        textColor=colors.HexColor('#2980b9')
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20
    )
    
    # Create story elements
    story = []
    
    # Title
    title = Paragraph("VISHNUDAS BHAVE NATYA GRANTHALAYA", title_style)
    story.append(title)
    
    subtitle = Paragraph("STOCK VERIFICATION FINAL REPORT", subtitle_style)
    story.append(subtitle)
    
    # Report info
    info_text = f"""
    <b>Stock Year:</b> {report_data['stock_year']} &nbsp;&nbsp; | &nbsp;&nbsp;
    <b>Report No:</b> SR/{report_data['stock_year_id']}/{timezone.now().strftime('%Y%m%d')} &nbsp;&nbsp; | &nbsp;&nbsp;
    <b>Generated:</b> {report_data['generated_at']}<br/>
    <b>Total Books:</b> {report_data['summary']['total']} &nbsp;&nbsp; | &nbsp;&nbsp;
    <b>Scanned:</b> {report_data['summary']['scanned']} &nbsp;&nbsp; | &nbsp;&nbsp;
    <b>Unknown:</b> {report_data['summary']['unknown']} &nbsp;&nbsp; | &nbsp;&nbsp;
    <b>Scan Rate:</b> {report_data['summary']['scanned'] / report_data['summary']['total'] * 100 if report_data['summary']['total'] > 0 else 0:.1f}%
    """
    info = Paragraph(info_text, info_style)
    story.append(info)
    
    story.append(Spacer(1, 20))
    
    # Create table
    table_data = []
    headers = ['Sr', 'Barcode', 'Title', 'Shelf Location', 'Current Status', 'Stock Status', 'Date']
    table_data.append(headers)
    
    books = report_data['books'][:100]  # Limit rows for PDF performance
    for i, book in enumerate(books):
        row = [
            str(i + 1),
            str(book.get('barcode', ''))[:12],
            str(book.get('title', ''))[:40],
            str(book.get('shelf_location', ''))[:15],
            str(book.get('current_status', ''))[:15],
            str(book.get('stock_status', ''))[:15],
            str(book.get('date_processed', ''))[:10]
        ]
        table_data.append(row)
    
    # Create table with style
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
    ]))
    
    story.append(table)
    
    # Footer
    story.append(Spacer(1, 20))
    footer_text = f"""
    <b>Total Records:</b> {len(report_data['books'])} &nbsp;&nbsp; | &nbsp;&nbsp;
    <b>Stock Verification Completed On:</b> {timezone.now().strftime('%d/%m/%Y')}<br/>
    <i>This is an official final report generated after completion of stock verification process.</i>
    """
    footer = Paragraph(footer_text, info_style)
    story.append(footer)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
    
def set_book_image_urls(books):
    for b in books:
        if b.front_page_photo:
            b.front_page_photo = file_storage_service.get_file_url(
                b.front_page_photo
            )
        else:
            b.front_page_photo = ""

        if b.last_page_photo:
            b.last_page_photo = file_storage_service.get_file_url(
                b.last_page_photo
            )
        else:
            b.last_page_photo = ""


@login_required
def led_tv_index(request):
    
    libraries = (
        LibraryMaster.objects.using('default')
        .filter(is_active=1)
        .order_by('id')
    )

    library_list = []

    for lib in libraries:
        first_image = ""

        if lib.image_url:
            image_array = lib.image_url.split(",")
            if image_array:
                first_image = file_storage_service.get_file_url(image_array[0].strip())

        library_list.append({
            "library_name": lib.library_name,
            "library_name_mar": lib.library_name_mar,
            "description": lib.about_library,
            "address": lib.location.address,
            "email": lib.contact_email,
            "phone": lib.contact_phone,
            "library_image": first_image,
            "location_url": lib.location_url,
            "opening_hours": lib.opening_hours,
            "est_year": lib.est_year
        })


    # ============================================
    # 1️⃣ POPULAR BOOKS (with rating and reviews)
    # ============================================
    popular_books = (
        BookCatalog.objects
        .annotate(
            avg_rating=Avg('cat_ref_id_reviews__rating'),
            review_count=Count('cat_ref_id_reviews')
        )
        .filter(avg_rating__isnull=False)
        .order_by('-avg_rating', '-created_at')[:10]
    )

    # ============================================
    # 2️⃣ NEW ARRIVALS
    # ============================================
    new_arrivals = (
        BookCatalog.objects
        .annotate(
            avg_rating=Avg('cat_ref_id_reviews__rating'),
            review_count=Count('cat_ref_id_reviews')
        )
        .filter(transactions__isnull=False)
        .order_by('-avg_rating')[:10]
    )

    # ============================================
    # 3️⃣ UPCOMING BOOKS
    # ============================================
    upcoming_books = (
        BookCatalog.objects
        .annotate(
            avg_rating=Avg('cat_ref_id_reviews__rating'),
            review_count=Count('cat_ref_id_reviews')
        )
        .filter(transactions__isnull=True)
        .order_by('-avg_rating')[:10]
    )
    
    set_book_image_urls(popular_books)
    set_book_image_urls(new_arrivals)
    set_book_image_urls(upcoming_books)

    # ============================================
    # ADS + EVENTS (only status=1)
    # ============================================
    advertisements = Advertisement.objects.filter(status=1).order_by('-created_at')[:5]

    # ============================================
    # FIX video_path → add MEDIA_URL
    # ============================================
    for ad in advertisements:
        if ad.video_path:
            ad.video_path = settings.MEDIA_URL + ad.video_path.replace("\\", "/")

    events = EventAnnouncement.objects.filter(status=1).order_by('-event_from_date')[:10]

    # ============================================
    # USER REVIEWS HANDLING
    # ============================================
    user_ids = set()
    for book_list in [popular_books, new_arrivals, upcoming_books]:
        for book in book_list:
            reviews_for_book = BookReview.objects.filter(
                book__cat_ref_num=book.cat_ref_num
            ).order_by('-created_at')[:2]

            for review in reviews_for_book:
                user_ids.add(review.user_id)

            book.recent_reviews = list(reviews_for_book)

    users = CustomUser.objects.filter(id__in=user_ids)
    user_dict = {user.id: user.username for user in users}

    for book_list in [popular_books, new_arrivals, upcoming_books]:
        for book in book_list:
            for review in getattr(book, 'recent_reviews', []):
                review.username = user_dict.get(review.user_id, "Unknown")

    return render(request, "L01/led_tv_index.html", {
        "popular_books": popular_books,
        "new_arrivals": new_arrivals,
        "upcoming_books": upcoming_books,
        "advertisements": advertisements,
        "events": events,
        "libraries": library_list, 
    })
    
def advertisement_index(request):
    # Fetch all ads ordered by newest first
        advertisements = Advertisement.objects.all().order_by('-created_at')

        # --- Stats ---
        total_count = advertisements.count()
        active_count = advertisements.filter(status=True).count()
        inactive_count = advertisements.filter(status=False).count()
        today_count = advertisements.filter(created_at=date.today()).count()

        # --- Attach encrypted id dynamically (if your template uses ad.encrypted_id) ---
        # Remove this section if you store encrypted_id directly in DB
        for ad in advertisements:
            if not hasattr(ad, 'encrypted_id'):
                try:
                    ad.encrypted_id = enc(str(ad.id))   # If you use encryption
                except:
                    ad.encrypted_id = ad.id               # fallback (no encryption)

        context = {
            "advertisements": advertisements,
            "active_count": active_count,
            "inactive_count": inactive_count,
            "today_count": today_count,
        }

        return render(request, "L01/Display/advertisement_index.html", context)
    
@login_required
def advertisement_toggle_status(request, encrypted_id):
    if request.method == 'POST':

        # Decrypt the encrypted ID
        try:
            pk = int(dec(encrypted_id))   # dec() returns string → convert to int
        except Exception:
            messages.error(request, "Invalid advertisement ID.")
            return redirect('advertisement_index')

        # Fetch advertisement using decrypted primary key
        advertisement = get_object_or_404(Advertisement, pk=pk)

        # Toggle the status
        advertisement.status = not advertisement.status
        advertisement.save()

        # Success messages
        if advertisement.status:
            messages.success(request, f'"{advertisement.adv_title}" activated successfully.')
        else:
            messages.success(request, f'"{advertisement.adv_title}" deactivated successfully.')

        return redirect('L01:advertisement_index')

    return redirect('L01:advertisement_index')

@login_required
def advertisement_create(request):
    library_code = request.session.get('library_db', None)
    if request.method == 'POST':
        try:
            # Get form data
            adv_title = request.POST.get('adv_title')
            
            # Handle file upload
            video_file = request.FILES.get('video_file')
            video_path = None
            
            if video_file:
                # Validate file size (max 200MB)
                if video_file.size > 200 * 1024 * 1024:
                    messages.error(request, 'File size exceeds 100MB limit.')
                    return render(request, 'L01/Display/advertisement_create.html')
                
                # Validate file extension
                allowed_extensions = ['.mp4', '.webm', '.ogg', '.mov', '.avi', '.wmv', '.flv', '.mkv']
                file_extension = os.path.splitext(video_file.name)[1].lower()
                
                if file_extension not in allowed_extensions:
                    messages.error(request, f'Invalid file format. Allowed formats: {", ".join([ext.replace(".", "") for ext in allowed_extensions])}')
                    return render(request, 'L01/Display/advertisement_create.html')
                
                # Generate unique filename
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                
                clean_filename = video_file.name.replace(' ', '_')
                filename = f"adv_{timestamp}_{clean_filename}"
                
                # Save the file using default_storage
                file_path = f'{library_code}/Advertisements/{filename}'
                saved_path = default_storage.save(file_path, video_file)
                video_path = saved_path  # Store the relative path in the database
                
            else:
                messages.error(request, 'Please select a video file.')
                return render(request, 'L01/Display/advertisement_create.html')
            
            # Parse dates
            campaign_date_from = request.POST.get('campaign_date_from')
            campaign_date_to = request.POST.get('campaign_date_to')
            reported_at = request.POST.get('reported_at')
            
            campaign_date_from = parse_date(campaign_date_from) if campaign_date_from else None
            campaign_date_to = parse_date(campaign_date_to) if campaign_date_to else None
            reported_at = parse_datetime(reported_at) if reported_at else None
            
            # Validate campaign dates
            if campaign_date_from and campaign_date_to:
                if campaign_date_to < campaign_date_from:
                    messages.error(request, 'End date cannot be earlier than start date.')
                    return render(request, 'L01/Display/advertisement_create.html')
            
            # Get organization name
            organization_name = request.POST.get('organization_name', '').strip()
            
            # Get status (checkbox returns 'on' when checked)
            status = request.POST.get('status') == 'on'
            
            # Get reported by (default to current user if not provided)
            reported_by = request.POST.get('reported_by', '').strip()
            if not reported_by and request.user.is_authenticated:
                reported_by = request.user.username
            
            # Create advertisement
            advertisement = Advertisement.objects.create(
                adv_title=adv_title,
                video_path=video_path,  # Store the saved file path
                organization_name=organization_name if organization_name else None,
                status=status,
                campaign_date_from=campaign_date_from,
                campaign_date_to=campaign_date_to,
                reported_at=reported_at,
                reported_by=reported_by if reported_by else None,
                created_by=request.user.username if request.user.is_authenticated else 'Anonymous'
            )
            
            messages.success(request, f'Advertisement "{adv_title}" created successfully!')
            return redirect('L01:advertisement_index')
            
        except Exception as e:
            # Log the error for debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error creating advertisement: {str(e)}')
            
            messages.error(request, f'Error creating advertisement: {str(e)}')
            return render(request, 'L01/Display/advertisement_create.html')
    
    # GET request - show the form
    return render(request, 'L01/Display/advertisement_create.html')

@login_required
def event_announcement_index(request):
    # Fetch all events ordered by newest first
    announcements = EventAnnouncement.objects.all().order_by('-created_at')

    # --- Stats ---
    total_count = announcements.count()
    active_count = announcements.filter(status=True).count()
    inactive_count = announcements.filter(status=False).count()
    today_count = announcements.filter(created_at__date=date.today()).count()
    
    # Count by date status
    today = date.today()
    upcoming_count = announcements.filter(
        event_from_date__gt=today, 
        status=True
    ).count()
    ongoing_count = announcements.filter(
        event_from_date__lte=today,
        event_to_date__gte=today,
        status=True
    ).count()
    past_count = announcements.filter(
        event_to_date__lt=today
    ).count()

    # --- Attach encrypted id dynamically ---
    for announcement in announcements:
        if not hasattr(announcement, 'encrypted_id'):
            try:
                announcement.encrypted_id = enc(str(announcement.id))
            except:
                announcement.encrypted_id = announcement.id

    context = {
        "announcements": announcements,
        "total_count": total_count,
        "active_count": active_count,
        "inactive_count": inactive_count,
        "today_count": today_count,
        "upcoming_count": upcoming_count,
        "ongoing_count": ongoing_count,
        "past_count": past_count,
    }

    return render(request, "L01/Display/event_announcement_index.html", context)

@login_required
def event_announcement_toggle_status(request, encrypted_id):
    if request.method == 'POST':
        # Decrypt the encrypted ID
        try:
            pk = int(dec(encrypted_id))
        except Exception:
            messages.error(request, "Invalid event announcement ID.")
            return redirect('L01:event_announcement_index')

        # Fetch announcement using decrypted primary key
        announcement = get_object_or_404(EventAnnouncement, pk=pk)

        # Toggle the status
        announcement.status = not announcement.status
        announcement.save()

        # Success messages
        if announcement.status:
            messages.success(request, f'"{announcement.event_title}" activated successfully.')
        else:
            messages.success(request, f'"{announcement.event_title}" deactivated successfully.')

        return redirect('L01:event_announcement_index')

    return redirect('L01:event_announcement_index')

@login_required
def event_announcement_create(request):
    if request.method == 'POST':
        try:
            # Get form data
            event_title = request.POST.get('event_title')
            description = request.POST.get('description', '').strip()
            
            # Parse dates
            event_from_date = request.POST.get('event_from_date')
            event_to_date = request.POST.get('event_to_date')
            reported_at = request.POST.get('reported_at')
            
            # Convert date strings to date objects
            from datetime import datetime
            
            if event_from_date:
                event_from_date = datetime.strptime(event_from_date, '%Y-%m-%d').date()
            if event_to_date:
                event_to_date = datetime.strptime(event_to_date, '%Y-%m-%d').date()
            if reported_at:
                reported_at = datetime.strptime(reported_at, '%Y-%m-%dT%H:%M')
            
            # Validate dates
            if event_from_date and event_to_date:
                if event_to_date < event_from_date:
                    messages.error(request, 'Event end date cannot be earlier than start date.')
                    return render(request, 'L01/Display/event_announcement_create.html')
            
            # Get status (checkbox returns 'on' when checked)
            status = request.POST.get('status') == 'on'
            
            # Get reported by (default to current user if not provided)
            reported_by = request.POST.get('reported_by', '').strip()
            if not reported_by and request.user.is_authenticated:
                reported_by = request.user.username
            
            # Create event announcement
            announcement = EventAnnouncement.objects.create(
                event_title=event_title,
                description=description if description else None,
                status=status,
                event_from_date=event_from_date,
                event_to_date=event_to_date,
                reported_at=reported_at,
                reported_by=reported_by if reported_by else None,
                created_by=request.user.username if request.user.is_authenticated else 'Anonymous'
            )
            
            messages.success(request, f'Event announcement "{event_title}" created successfully!')
            return redirect('L01:event_announcement_index')
            
        except Exception as e:
            # Log the error for debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error creating event announcement: {str(e)}')
            
            messages.error(request, f'Error creating event announcement: {str(e)}')
            return render(request, 'L01/Display/event_announcement_create.html')
    
    # GET request - show the form
    return render(request, 'L01/Display/event_announcement_create.html')

@login_required
def advertisement_edit(request, encrypted_id):
    try:
        # Decrypt advertisement ID
        adv_id = int(dec(encrypted_id))

        library_code = request.session.get('library_db', None)

        advertisement = get_object_or_404(Advertisement, pk=adv_id)

        if request.method == 'POST':
            # Get basic fields
            adv_title = request.POST.get('adv_title')
            organization_name = request.POST.get('organization_name', '').strip()
            status = request.POST.get('status') == 'on'
            reported_by = request.POST.get('reported_by', '').strip()

            if not reported_by and request.user.is_authenticated:
                reported_by = request.user.username

            # ----- VIDEO FILE HANDLING -----
            video_file = request.FILES.get('video_file')
            

            if video_file:
                # Validate size (max 200 MB)
                if video_file.size > 200 * 1024 * 1024:
                    messages.error(request, 'File size exceeds 100MB limit.')
                    return render(request, 'L01/Display/advertisement_edit.html', {'data': advertisement})

                # Validate extension
                allowed_extensions = ['.mp4', '.webm', '.ogg', '.mov', '.avi', '.wmv', '.flv', '.mkv']
                file_extension = os.path.splitext(video_file.name)[1].lower()

                if file_extension not in allowed_extensions:
                    messages.error(request, f'Invalid file format. Allowed: {", ".join([ext.replace(".", "") for ext in allowed_extensions])}')
                    return render(request, 'L01/Display/advertisement_edit.html', {'data': advertisement})

                # Generate unique file name
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                clean_filename = video_file.name.replace(' ', '_')
                filename = f"adv_{timestamp}_{clean_filename}"

                file_path = f'{library_code}/Advertisements/{filename}'
                saved_path = default_storage.save(file_path, video_file)

                # OPTIONAL: Delete old file if exists
                if advertisement.video_path and default_storage.exists(advertisement.video_path):
                    default_storage.delete(advertisement.video_path)

                # Set new video path
                advertisement.video_path = saved_path

            # ----- DATE HANDLING -----
            campaign_date_from = request.POST.get('campaign_date_from')
            campaign_date_to = request.POST.get('campaign_date_to')
            reported_at = request.POST.get('reported_at')

            campaign_date_from = parse_date(campaign_date_from) if campaign_date_from else None
            campaign_date_to = parse_date(campaign_date_to) if campaign_date_to else None
            reported_at = parse_datetime(reported_at) if reported_at else None

            if campaign_date_from and campaign_date_to:
                if campaign_date_to < campaign_date_from:
                    messages.error(request, 'End date cannot be earlier than start date.')
                    return render(request, 'L01/Display/advertisement_edit.html', {'data': advertisement})

            # ----- UPDATE FIELDS -----
            advertisement.adv_title = adv_title
            advertisement.organization_name = organization_name if organization_name else None
            advertisement.status = status
            advertisement.campaign_date_from = campaign_date_from
            advertisement.campaign_date_to = campaign_date_to
            advertisement.reported_at = reported_at
            advertisement.reported_by = reported_by
            advertisement.updated_by = request.user.username if request.user.is_authenticated else 'Anonymous'
            advertisement.save()

            messages.success(request, f'Advertisement "{adv_title}" updated successfully!')
            return redirect('L01:advertisement_index')

        # GET request → load edit page
        return render(request, 'L01/Display/advertisement_edit.html', {'data': advertisement})

    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Error editing advertisement: {str(e)}")
        messages.error(request, f"Error editing advertisement: {str(e)}")
        return redirect('L01:advertisement_index')

@login_required
def event_announcement_edit(request, encrypted_id):
    try:
        # Decrypt event announcement ID
        announcement_id = int(dec(encrypted_id))
        announcement = get_object_or_404(EventAnnouncement, pk=announcement_id)

        if request.method == 'POST':
            # Get form data
            event_title = request.POST.get('event_title')
            description = request.POST.get('description', '').strip()
            
            # Parse dates
            event_from_date = request.POST.get('event_from_date')
            event_to_date = request.POST.get('event_to_date')
            reported_at = request.POST.get('reported_at')
            
            # Convert date strings to date objects
            from datetime import datetime
            
            if event_from_date:
                event_from_date = datetime.strptime(event_from_date, '%Y-%m-%d').date()
            else:
                event_from_date = None
                
            if event_to_date:
                event_to_date = datetime.strptime(event_to_date, '%Y-%m-%d').date()
            else:
                event_to_date = None
                
            if reported_at:
                reported_at = datetime.strptime(reported_at, '%Y-%m-%dT%H:%M')
            else:
                reported_at = None
            
            # Validate dates
            if event_from_date and event_to_date:
                if event_to_date < event_from_date:
                    messages.error(request, 'Event end date cannot be earlier than start date.')
                    return render(request, 'L01/Display/event_announcement_edit.html', {'data': announcement})
            
            # Get status (checkbox returns 'on' when checked)
            status = request.POST.get('status') == 'on'
            
            # Get reported by
            reported_by = request.POST.get('reported_by', '').strip()
            if not reported_by:
                reported_by = None
            
            # Update event announcement
            announcement.event_title = event_title
            announcement.description = description if description else None
            announcement.status = status
            announcement.event_from_date = event_from_date
            announcement.event_to_date = event_to_date
            announcement.reported_at = reported_at
            announcement.reported_by = reported_by
            announcement.updated_by = request.user.username if request.user.is_authenticated else 'Anonymous'
            announcement.save()
            
            messages.success(request, f'Event announcement "{event_title}" updated successfully!')
            return redirect('L01:event_announcement_index')
            
        # GET request - show the edit form with existing data
        return render(request, 'L01/Display/event_announcement_edit.html', {'data': announcement})
        
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error editing event announcement: {str(e)}')
        
        messages.error(request, f'Error editing event announcement: {str(e)}')
        return redirect('L01:event_announcement_index')

# Librarian Dashboard
@login_required
def dashboard_view(request):
    """
    Main dashboard view with 4 tabs
    """
    try:
        breadcrumb = request.POST.get("breadcrumb")
        if not breadcrumb:
            breadcrumb = 0
        context = {
            'dashboard_title': 'Library Dashboard',
            'active_tab': 'dashboard1',
            'breadcrumb':breadcrumb
        }
        return render(request, 'L01/Dashboard/library_dashboard.html', context)
    except Exception as e:
        # Log the error here if you have logging configured
        return JsonResponse({
            'error': f'Error loading dashboard: {str(e)}'
        }, status=500)

def get_dashboard1_data(request):
    try:
        from django.utils.timezone import now
        from datetime import datetime, time
        from django.utils import timezone
        from django.db.models import Count
        from django.db.models.expressions import RawSQL
        start_date = request.GET.get('from_date')
        end_date   = request.GET.get('to_date')

        if start_date and end_date:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date   = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            start_date = end_date = now().date()

       

        # ---------- DAILY ----------
        daily = {
            "issued": CirculationTransaction.objects.using("L01")
                        .filter(issue_date__range=[start_date, end_date]).count(),

            "returned": CirculationTransaction.objects.using("L01")
                        .filter(return_date__range=[start_date, end_date]).count(),

            "due": CirculationTransaction.objects.using("L01")
                        .filter(due_date__range=[start_date, end_date],
                                return_date__isnull=True).count(),

            "damaged": CirculationTransaction.objects.using("L01")
                        .filter(return_condition_id__in=[17, 18],
                                return_date__range=[start_date, end_date]).count(),
        }
        
        today = date.today()
        returned_books = CirculationTransaction.objects.using("L01").filter(
            return_date=today
        ).values(
            "id",
            "barcode",
            "catalog__cat_ref_num",
            "catalog__title"
        )

        # ---------- EOD ----------
        payments = PaymentDetails.objects.using("L01")\
                    .filter(payment_date__range=[start_date, end_date])

        eod = payments.aggregate(
            monthly=Sum('monthly_subscription_amount'),
            total=Sum('total_subscription_amount'),
            fine=Sum('fine_amount'),
            book_fine=Sum('book_fine_amount')
        )

        eod = {k: v or 0 for k, v in eod.items()}
        from django.db.models.functions import Trim, Lower

        
        start_dt = timezone.make_aware(datetime.combine(start_date, time.min))
        end_dt   = timezone.make_aware(datetime.combine(end_date, time.max))

        
        footfall_qs = (
            MemberScreenActivity.objects.using("L01")
            .filter(visited_at__range=(start_dt, end_dt))
            .annotate(menu=Lower(Trim('screen_name')))  # ✅ normalize
            .values('menu')
            .annotate(count=Count('session__member', distinct=True))
            .order_by('menu')
        )

        footfall = {
            "labels": [f['menu'] for f in footfall_qs],
            "counts": [f['count'] for f in footfall_qs]
        }

        physical_qs = (
            MemberEntryExit.objects.using("L01")
            .filter(entry_time__range=(start_dt, end_dt))
            .annotate(day=RawSQL("DATE(entry_time)", []))
            .values('day')
            .annotate(count=Count('id'))
        )

        circulation_qs = (
            CirculationTransaction.objects.using("L01")
            .filter(created_at__range=(start_dt, end_dt))
            .annotate(day=RawSQL("DATE(created_at)", []))
            .values('day')
            .annotate(count=Count('id'))
        )


        merged_data = defaultdict(lambda: {
            "physical": 0,
            "circulation": 0,
            "total": 0
        })

        # Physical counts
        for row in physical_qs:
            if row['day']:
                merged_data[row['day']]["physical"] += row['count']
                merged_data[row['day']]["total"] += row['count']

        # Circulation counts
        for row in circulation_qs:
            if row['day']:
                merged_data[row['day']]["circulation"] += row['count']
                merged_data[row['day']]["total"] += row['count']

        physical_footfall = {
            "labels": [
                day.strftime('%d-%b')
                for day in sorted(merged_data.keys())
            ],
            "counts": [
                merged_data[day]["total"]
                for day in sorted(merged_data.keys())
            ],
            "physical_counts": [
                merged_data[day]["physical"]
                for day in sorted(merged_data.keys())
            ],
            "circulation_counts": [
                merged_data[day]["circulation"]
                for day in sorted(merged_data.keys())
            ]
        }

        visitor_qs = (
            VisitorActivity.objects.using('default')   # ✅ DEFAULT DB
            .filter(created_at__range=(start_dt, end_dt))
            .annotate(day=RawSQL("DATE(created_at)", []))
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )

        visitor_footfall = {
            "labels": [
                row['day'].strftime('%d-%b')
                for row in visitor_qs if row['day']
            ],
            "counts": [
                row['count']
                for row in visitor_qs if row['day']
            ]
        }

        # ---------- PARTIAL HTML ----------
        html = render_to_string(
            "L01/Dashboard/dashboard1_content.html",
            request=request
        )

        return JsonResponse({
            "html": html,
            "daily": daily,
            "visitor_footfall": visitor_footfall,
            "eod": eod,
            "footfall": footfall,
            "physical_footfall": physical_footfall,
            "start_date":start_date,
            "end_date":end_date,
            "returned_books":list(returned_books)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_transaction_details(request):
    try:
        from django.utils.timezone import now
        from datetime import datetime
        from django.db.models import Q
        import traceback
        
        # Get parameters from request
        start_date = request.GET.get('from_date')
        end_date = request.GET.get('to_date')
        label = request.GET.get('label')
        detail_type = request.GET.get('detail_type')
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        
        # Get search value - handle both formats
        search_value = request.GET.get('search[value]', '')
        if not search_value:
            # Try alternative format
            search_value = request.GET.get('search_value', '')
        
        print(f"DEBUG: detail_type={detail_type}, search_value='{search_value}'")
        print(f"DEBUG: from_date={start_date}, to_date={end_date}")
        
        if start_date and end_date:
            try:
                start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            except ValueError:
                start_date = end_date = now().date()
        else:
            start_date = end_date = now().date()
        
        # Base queryset based on detail_type
        data = []
        total_records = 0
        
        if detail_type == "issued":
            try:
                qs = CirculationTransaction.objects.using("L01").filter(
                    issue_date__range=[start_date, end_date]
                ).select_related('member', 'catalog')
                
                if search_value:
                    qs = qs.filter(
                        Q(member__membership_code__icontains=search_value) |
                        Q(barcode__icontains=search_value) |
                        Q(catalog__title__icontains=search_value)
                    )
                
                total_records = qs.count()
                qs = qs.order_by('-issue_date')[start:start + length]
                
                for obj in qs:
                    data.append({
                        "member_name": f"{obj.member.first_name_mar or ''} {obj.member.last_name_mar or ''}".strip(),
                        "book_name": obj.catalog.title if obj.catalog else "",
                        "member_code": obj.member.membership_code if obj.member else "",
                        "book_barcode": obj.barcode,
                        "date": obj.issue_date.isoformat() if obj.issue_date else ""
                    })
                    
            except Exception as e:
                print(f"Error in issued section: {str(e)}")
                traceback.print_exc()
        
        elif detail_type == "returned":
            try:
                qs = CirculationTransaction.objects.using("L01").filter(
                    return_date__range=[start_date, end_date]
                ).select_related('member', 'catalog')
                
                if search_value:
                    qs = qs.filter(
                        Q(member__membership_code__icontains=search_value) |
                        Q(barcode__icontains=search_value) |
                        Q(catalog__title__icontains=search_value)
                    )
                
                total_records = qs.count()
                qs = qs.order_by('-return_date')[start:start + length]
                
                for obj in qs:
                    data.append({
                        "member_name": f"{obj.member.first_name_mar or ''} {obj.member.last_name_mar or ''}".strip(),
                        "book_name": obj.catalog.title if obj.catalog else "",
                        "member_code": obj.member.membership_code if obj.member else "",
                        "book_barcode": obj.barcode,
                        "date": obj.return_date.isoformat() if obj.return_date else ""
                    })
                    
            except Exception as e:
                print(f"Error in returned section: {str(e)}")
                traceback.print_exc()
        
        elif detail_type == "due":
            try:
                qs = CirculationTransaction.objects.using("L01").filter(
                    due_date__range=[start_date, end_date],
                    return_date__isnull=True
                ).select_related('member', 'catalog')
                
                if search_value:
                    qs = qs.filter(
                        Q(member__membership_code__icontains=search_value) |
                        Q(barcode__icontains=search_value) |
                        Q(catalog__title__icontains=search_value)
                    )
                
                total_records = qs.count()
                qs = qs.order_by('-due_date')[start:start + length]
                
                for obj in qs:
                    data.append({
                        "member_name": f"{obj.member.first_name_mar or ''} {obj.member.last_name_mar or ''}".strip(),
                        "book_name": obj.catalog.title if obj.catalog else "",
                        "member_code": obj.member.membership_code if obj.member else "",
                        "book_barcode": obj.barcode,
                        "date": obj.due_date.isoformat() if obj.due_date else ""
                    })
                    
            except Exception as e:
                print(f"Error in due section: {str(e)}")
                traceback.print_exc()
        
        elif detail_type == "damaged":
            try:
                qs = CirculationTransaction.objects.using("L01").filter(
                    return_condition_id__in=[17, 18],
                    return_date__range=[start_date, end_date]
                ).select_related('member', 'catalog')
                
                if search_value:
                    qs = qs.filter(
                        Q(member__membership_code__icontains=search_value) |
                        Q(barcode__icontains=search_value) |
                        Q(catalog__title__icontains=search_value)
                    )
                
                total_records = qs.count()
                qs = qs.order_by('-return_date')[start:start + length]
                
                for obj in qs:
                    data.append({
                        "member_name": f"{obj.member.first_name_mar or ''} {obj.member.last_name_mar or ''}".strip(),
                        "book_name": obj.catalog.title if obj.catalog else "",
                        "member_code": obj.member.membership_code if obj.member else "",
                        "book_barcode": obj.barcode,
                        "date": obj.return_date.isoformat() if obj.return_date else ""
                    })
                    
            except Exception as e:
                print(f"Error in damaged section: {str(e)}")
                traceback.print_exc()
        
        elif detail_type in ["monthly", "total", "fine", "book_fine"]:
            try:
                # Determine amount field
                amount_field_map = {
                    "monthly": "monthly_subscription_amount",
                    "total": "total_subscription_amount",
                    "fine": "fine_amount",
                    "book_fine": "book_fine_amount"
                }
                amount_field = amount_field_map.get(detail_type)
                
                qs = PaymentDetails.objects.using("L01").filter(
                    payment_date__range=[start_date, end_date]
                ).select_related('membership')
                
                # Filter for non-zero amounts in the specific field
                if amount_field:
                    qs = qs.filter(**{
                        f"{amount_field}__isnull": False,
                        f"{amount_field}__gt": 0
                    })
                
                if search_value:
                    qs = qs.filter(
                        Q(membership__membership_code__icontains=search_value)
                    )
                
                total_records = qs.count()
                qs = qs.order_by("-payment_date")[start:start + length]
                
                for obj in qs:
                    amount = 0
                    if amount_field:
                        amount = getattr(obj, amount_field, 0) or 0
                    
                    data.append({
                        "member_name": f"{obj.membership.first_name_mar or ''} {obj.membership.last_name_mar or ''}".strip() if obj.membership else "",
                        "member_code": obj.membership.membership_code if obj.membership else "",
                        "amount": float(amount),
                        "payment_date": obj.payment_date.isoformat() if obj.payment_date else ""
                    })
                    
            except Exception as e:
                print(f"Error in payment section: {str(e)}")
                traceback.print_exc()

        elif detail_type == "footfall_online":
            from datetime import datetime, time
            from django.db.models import OuterRef, Subquery, Q

            start_datetime = datetime.combine(start_date, time.min)
            end_datetime = datetime.combine(end_date, time.max)

            # 🔑 Subquery: latest screen activity per MEMBER
            latest_activity = (
                MemberScreenActivity.objects.using("L01")
                .filter(
                    session__member_id=OuterRef('session__member_id'),
                    visited_at__range=[start_datetime, end_datetime],
                    screen_name=label
                )
                .order_by('-visited_at')
                .values('visited_at')[:1]
            )

            qs = (
                MemberScreenActivity.objects.using("L01")
                .select_related(
                    'session',
                    'session__member',
                    'session__member__membership'
                )
                .filter(
                    visited_at=Subquery(latest_activity),
                    screen_name=label  # 🔑 important for MySQL correctness
                )
            )

            # 🔍 Search filter
            if search_value:
                qs = qs.filter(
                    Q(session__member__membership_code__icontains=search_value) |
                    Q(session__member__first_name_mar__icontains=search_value) |
                    Q(session__member__last_name_mar__icontains=search_value)
                )

            total_records = qs.count()
            qs = qs.order_by('-visited_at')[start:start + length]

            for obj in qs:
                data.append({
                    "membership_type": obj.session.member.membership.membership_type if obj.session.member else "",
                    "membership_code": obj.session.member.membership_code if obj.session.member else "",
                    "member_name": f"{obj.session.member.first_name_mar or ''} {obj.session.member.last_name_mar or ''}".strip(),
                    "login_time": obj.session.login_time.isoformat() if obj.session.login_time else "",
                    "visited_at": obj.visited_at.isoformat() if obj.visited_at else "",
                })

        elif detail_type == "visitor_log":
            from datetime import datetime, time ,date
            try:
                if isinstance(start_date, date):
                    from_date_obj = start_date
                else:
                    from_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()

                if isinstance(end_date, date):
                    to_date_obj = end_date
                else:
                    to_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
                label_date = datetime.strptime(label, "%d-%b")

                # Use year from from_date
                entry_date = label_date.replace(year=from_date_obj.year).date()

                entry_start = datetime.combine(entry_date, time.min)
                entry_end = datetime.combine(entry_date, time.max)
                qs = VisitorActivity.objects.using('default').filter(created_at__range=[entry_start, entry_end])
                
                if search_value:
                    # Optional search filter
                    qs = qs.filter(
                        Q(visitor__icontains=search_value) |
                        Q(ip_address__icontains=search_value) |
                        Q(phone__icontains=search_value) |
                        Q(email__icontains=search_value) |
                        Q(remark__icontains=search_value)
                    )

                total_records = qs.count()

                # Pagination
                qs = qs.order_by('-created_at')[start:start + length]

                # Append each row to `data`
                for obj in qs:
                    data.append({
                        "visitor": obj.visitor or "-",
                        "ip": obj.ip_address or "-",
                        "phone": obj.phone or "-",
                        "email": obj.email or "-",
                        "remark": obj.remark or "-",
                        "time": obj.created_at.strftime("%d-%b %I:%M %p") if obj.created_at else "-"
                    })

            except Exception as e:
                print(f"Error in visitor_log section: {str(e)}")
                traceback.print_exc()


        elif detail_type == "footfall_offline":
            from datetime import datetime, date, time
            from django.db.models import Q

            if isinstance(start_date, date):
                from_date_obj = start_date
            else:
                from_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()

            if isinstance(end_date, date):
                to_date_obj = end_date
            else:
                to_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            label_date = datetime.strptime(label, "%d-%b")

            # Use year from from_date
            entry_date = label_date.replace(year=from_date_obj.year).date()

            entry_start = datetime.combine(entry_date, time.min)
            entry_end = datetime.combine(entry_date, time.max)
            qs = (
                MemberEntryExit.objects.using("L01")
                .filter(entry_time__range=[entry_start, entry_end])
                .order_by('-entry_time')
            )
            if search_value:
                qs = qs.filter(
                    Q(membership_code__icontains=search_value)
                )
            qs = list(qs)

            seen = set()
            unique_entries = []

            for obj in qs:
                if obj.membership_code not in seen:
                    seen.add(obj.membership_code)
                    unique_entries.append(obj)

            total_records = len(unique_entries)
            unique_entries = unique_entries[start:start + length]

            # -----------------------------
            # 6️⃣ Fetch member details
            # -----------------------------
            membership_codes = [obj.membership_code for obj in unique_entries]

            member_map = {
                m.membership_code: m
                for m in MembershipDetails.objects.using("L01").filter(
                    membership_code__in=membership_codes
                )
            }

            # -----------------------------
            # 7️⃣ Build response
            # -----------------------------
            for obj in unique_entries:
                member = member_map.get(obj.membership_code)

                data.append({
                    "membership_type": member.membership.membership_type if member else "",
                    "membership_code": obj.membership_code,
                    "member_name": (
                        f"{member.first_name_mar or ''} {member.last_name_mar or ''}".strip()
                        if member else ""
                    ),
                    "login_time": obj.entry_time.isoformat() if obj.entry_time else "",
                })

        
        print(f"DEBUG: Returning {len(data)} records out of {total_records} total")
        
        return JsonResponse({
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": total_records,
            "data": data
        })
        
    except Exception as e:
        print(f"General error in get_transaction_details: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            "draw": 1,
            "recordsTotal": 0,
            "recordsFiltered": 0,
            "data": [],
            "error": str(e)
        }, status=500)
# for Dashboard 2 Librarian
def catalog_data_ajax(request):
    
    subject_id = request.GET.get("subject_id")

    qs = BookCatalog.objects.filter(subject__is_active=1)

    if subject_id:
        qs = qs.filter(subject_id=subject_id)
        subjects = SubjectTypeMaster.objects.filter(id=subject_id, is_active=1)
    else:
        subjects = SubjectTypeMaster.objects.filter(is_active=1)

    data = []

    for sub in subjects:
        titles_count = qs.filter(subject=sub).count()

        copies_count = BookAccession.objects.filter(
            catalogue__subject=sub
        ).count()

        processed_count = CirculationCopyStatus.objects.filter(
            bookcatalog__subject=sub
        ).count()

        if titles_count > 0:
            data.append({
                "subject": sub.subjectNameMarathi or sub.subjectNameEnglish,
                "titles": titles_count,
                "copies": copies_count,
                "processed": processed_count,
            })

    # KPI counts
    if subject_id:
        total_titles = qs.count()
        total_copies = BookAccession.objects.filter(
            catalogue__subject_id=subject_id
        ).count()
        total_processed = CirculationCopyStatus.objects.filter(
            bookcatalog__subject_id=subject_id
        ).count()
    else:
        total_titles = BookCatalog.objects.count()
        total_copies = BookAccession.objects.count()
        total_processed = CirculationCopyStatus.objects.count()

    return JsonResponse({
        "data": data,
        "total_titles": total_titles,
        "total_copies": total_copies,
        "total_processed": total_processed
    })

# for Dashboard 2 Librarian
def build_books_partial_html():
    total_titles = BookCatalog.objects.count()
    total_copies = BookAccession.objects.count()
    
    # Get active subjects for the dropdown
    subjects = SubjectTypeMaster.objects.filter(is_active=1)

    return render_to_string(
        "L01/Dashboard/library_dashboard_catalog.html",
        {
            "total_titles": total_titles,
            "total_copies": total_copies,
            "subjects": subjects,  # pass subjects here
        }
    )
    
def build_membershipType_partial_html():
    total_titles = BookCatalog.objects.count()
    total_copies = BookAccession.objects.count()
    
    # Get active subjects for the dropdown
    subjects = SubjectTypeMaster.objects.filter(is_active=1)

    return render_to_string(
        "L01/Dashboard/dashboard3_content.html",
        {
            "total_titles": total_titles,
            "total_copies": total_copies,
            "subjects": subjects,  # pass subjects here
        }
    )

def get_membershipType_data_dashboardThree(request):
    try:
        memberships = (
            MembershipMaster.objects.using("L01")
            .filter(isactive=1)
            .annotate(member_count=Count('membership_holders'))
            .values(
                'id',
                'membership_code',
                'membership_type',
                'deposit',
                'entry_fees',
                'subscription_fees',
                'fine',
                'outsider',
                'days',
                'item',
                'member_count'
            )
        )

        total_members = sum(m['member_count'] for m in memberships)

        # For pie chart percentage
        pie_data = []
        for m in memberships:
            percent = (
                (m['member_count'] / total_members) * 100
                if total_members > 0 else 0
            )
            pie_data.append({
                'membership_type': m['membership_type'],
                'count': m['member_count'],
                'percent': round(percent, 2)
            })

        return JsonResponse({
            'bar_chart': list(memberships),
            'pie_chart': pie_data,
            'table': list(memberships)
        })
    except Exception as e:
            print(f"WeasyPrint generation error: {e}")

# for Dashboard 2 Librarian
def get_dashboard2_data(request):
    """
    Returns session data and total books for Dashboard 2
    """
    try:
        library_code = request.session.get('library_db', None)
        username = request.session.get('username', None)
        user_id = request.session.get('user_id', None)
        role_id = request.session.get('role_id', None)

        books_html = build_books_partial_html()
        print("Books HTML:", books_html)  # Debugging line  
        return JsonResponse({
            'dashboard': 'dashboard2',
            'title': 'Dashboard 2 - Reports',
            'message': 'Dashboard 2 data will be implemented here',
            'last_updated': '2024-01-01 00:00:00',
            'session_data': {
                'library_code': library_code,
                'username': username,
                'user_id': user_id,
                'role_id': role_id
            },
            'books_html': books_html,
            'status': 'success'
        })

    except Exception as e:
        return JsonResponse({
            'dashboard': 'dashboard2',
            'error': f'Error loading dashboard 2 data: {str(e)}',
            'status': 'error'
        }, status=500)

def get_dashboard3_data(request):
    try:
        library_code = request.session.get('library_db', None)
        username = request.session.get('username', None)
        user_id = request.session.get('user_id', None)
        role_id = request.session.get('role_id', None)

        books_html = build_membershipType_partial_html()
        print("Books HTML:", books_html)  # Debugging line  
        return JsonResponse({
            'dashboard': 'dashboard3',
            'title': 'Dashboard 3 - Reports',
            'message': 'Dashboard 3 data will be implemented here',
            'last_updated': '2024-01-01 00:00:00',
            'session_data': {
                'library_code': library_code,
                'username': username,
                'user_id': user_id,
                'role_id': role_id
            },
            'books_html': books_html,
            'status': 'success'
        })
    except Exception as e:
            print(f"WeasyPrint generation error: {e}")

# for Dashboard 4 Librarian Get DATA
from django.http import JsonResponse
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.utils.dateparse import parse_date
from django.db.models import CharField, F, Value
from dateutil.relativedelta import relativedelta

def library_dashboard_data(request):
    try:
        # -------------------------
        # FILTERS
        # -------------------------
        from_date = parse_date(request.GET.get('from_date'))
        to_date = parse_date(request.GET.get('to_date'))
        subject_id = request.GET.get('subject_id')
        
        subjects = list(
            SubjectTypeMaster.objects
            .filter(is_active=1)
            .values('id', 'subjectNameMarathi', 'subjectNameEnglish')
            .order_by('subjectNameMarathi', 'subjectNameEnglish')
        )

        circulation_qs = CirculationTransaction.objects.all()

        if from_date:
            circulation_qs = circulation_qs.filter(issue_date__gte=from_date)
        if to_date:
            circulation_qs = circulation_qs.filter(issue_date__lte=to_date)
        if subject_id and subject_id.isdigit():
            circulation_qs = circulation_qs.filter(catalog__subject_id=int(subject_id))

        # -------------------------
        # KPI
        # -------------------------
        kpis = {
            "total_members": MembershipDetails.objects.count(),
            "total_titles": BookCatalog.objects.count(),
            "total_copies": BookAccession.objects.count(),
            "total_reviews": BookReview.objects.count(),
        }

        # -------------------------
        # TOP READ BOOKS
        # -------------------------
        top_books_qs = (
            circulation_qs
            .values('catalog__title')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )

        top_books = {
            "labels": [b['catalog__title'] or 'Untitled' for b in top_books_qs],
            "data": [b['count'] for b in top_books_qs]
        }

        # -------------------------
        # TOP READERS
        # -------------------------
        top_readers_qs = (
            circulation_qs
            .values('member__first_name', 'member__last_name')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )

        top_readers = {
            "labels": [
                f"{r['member__first_name'] or ''} {r['member__last_name'] or ''}".strip() or "Unknown"
                for r in top_readers_qs
            ],
            "data": [r['count'] for r in top_readers_qs]
        }

        # -------------------------
        # RATING DISTRIBUTION
        # -------------------------
        rating_qs = (
            BookReview.objects
            .values('rating')
            .annotate(count=Count('id'))
            .order_by('rating')
        )

        rating_distribution = {
            "labels": [f"{r['rating']} Star" for r in rating_qs],
            "data": [r['count'] for r in rating_qs]
        }

        # -------------------------
        # SUBJECT POPULARITY
        # -------------------------
        subject_qs = (
            circulation_qs
            .values(
                'catalog__subject__subjectNameMarathi',
                'catalog__subject__subjectNameEnglish'
            )
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        subject_popularity = {
            "labels": [
                s['catalog__subject__subjectNameMarathi']
                or s['catalog__subject__subjectNameEnglish']
                or 'Unknown'
                for s in subject_qs
            ],
            "data": [s['count'] for s in subject_qs]
        }

        # -------------------------
        # MONTHLY TREND
        # -------------------------
        month_qs = (
            circulation_qs
            .annotate(month=TruncMonth('issue_date'))
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        )

        monthly_trend = {
            "labels": [
                m['month'].strftime('%b %Y') if m['month'] else ''
                for m in month_qs
            ],
            "data": [m['count'] for m in month_qs]
        }

        # -------------------------
        # DRILL DOWN
        # -------------------------
        drill_type = request.GET.get('drill_type')
        drill_value = request.GET.get('drill_value')
        drill_data = []
        drill_headers = []

        if drill_type and drill_value:
            
            if drill_type == 'book':
                qs = (
                    circulation_qs
                    .filter(catalog__title=drill_value)
                    .select_related('member')
                    .values('member__first_name', 'member__last_name', 'issue_date')[:50]
                )
                drill_headers = ["Member Name", "Issue Date"]
                drill_data = [
                    {
                        "Member Name": f"{row['member__first_name'] or ''} {row['member__last_name'] or ''}".strip(),
                        "Issue Date": row['issue_date']
                    }
                    for row in qs
                ]

            elif drill_type == 'reader':
                qs = (
                    circulation_qs
                    .annotate(
                        full_name=Concat(
                            F('member__first_name'),
                            Value(' '),
                            F('member__last_name'),
                            output_field=CharField()   # <--- This fixes your error
                        )
                    )
                    .filter(full_name=drill_value)
                    .select_related('member')
                    .values('catalog__title', 'issue_date')[:50]
                )

                drill_headers = ["Book Title", "Issue Date"]
                drill_data = [
                    {"Book Title": row['catalog__title'], "Issue Date": row['issue_date']}
                    for row in qs
                ]
                
            elif drill_type == 'rating':
                try:
                    numeric_rating = int(''.join(filter(str.isdigit, drill_value)))
                except ValueError:
                    numeric_rating = 0

                qs = BookReview.objects.filter(rating=numeric_rating).values(
                    'user_id', 'book__title', 'rating', 'review', 'created_at'
                )[:50]

                drill_data = [
                    {
                        "User ID": row['user_id'],
                        "Book Title": row['book__title'],
                        "Rating": row['rating'],
                        "Review": row['review'],
                        "Created At": row['created_at'].strftime('%Y-%m-%d %H:%M')
                    }
                    for row in qs
                ]

                drill_headers = ["User ID", "Book Title", "Rating", "Review", "Created At"]

            elif drill_type == 'month':
                try:
                    # Example: "Nov 2025"
                    month_date = datetime.strptime(drill_value, "%b %Y")
                    start_date = month_date.replace(day=1)
                    end_date = start_date + relativedelta(months=1)
                except ValueError:
                    start_date = end_date = None

                if start_date and end_date:
                    qs = (
                        circulation_qs
                        .filter(
                            issue_date__gte=start_date,
                            issue_date__lt=end_date
                        )
                        .select_related(
                            'catalog',
                            'catalog__subject',
                            'accession',
                            'member',
                            'member__member_type'
                        )
                        .annotate(
                            member_name=Concat(
                                F('member__first_name'),
                                Value(' '),
                                F('member__middle_name'),
                                Value(' '),
                                F('member__last_name'),
                                output_field=CharField()
                            )
                        )
                        .values(
                            'issue_date',
                            'catalog__title',
                            'catalog__author',
                            'catalog__subject__subjectNameMarathi',
                            'accession__accession_no',
                            'accession__copy_number',
                            'member_name',
                            'member__member_type__parameter_value',
                            'transaction_type',
                            'fine_amount'
                        )
                        .order_by('issue_date')[:300]
                    )

                    drill_headers = [
                        "Issue Date",
                        "Book Title",
                        "Author",
                        "Subject",
                        "Accession No",
                        "Copy No",
                        "Member Name",
                        "Member Type",
                        "Transaction",
                        "Fine Amount"
                    ]

                    drill_data = []
                    for row in qs:
                        drill_data.append({
                            "Issue Date": row['issue_date'],
                            "Book Title": row['catalog__title'],
                            "Author": row['catalog__author'],
                            "Subject": row['catalog__subject__subjectNameMarathi'],
                            "Accession No": row['accession__accession_no'],
                            "Copy No": row['accession__copy_number'],
                            "Member Name": row['member_name'].strip(),
                            "Member Type": row['member__member_type__parameter_value'],
                            "Transaction": row['transaction_type'],
                            "Fine Amount": row['fine_amount'] or 0
                        })
            
        return JsonResponse({
            "kpis": kpis,
            "top_books": top_books,
            "top_readers": top_readers,
            "rating_distribution": rating_distribution,
            "subject_popularity": subject_popularity,
            "monthly_trend": monthly_trend,
            "drill_data": drill_data,
            "drill_headers": drill_headers,
            "subjects": subjects,
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

# for Dashboard 4 Librarian
def build_memberbookrelation_partial_html():
    total_titles = BookCatalog.objects.count()
    total_copies = BookAccession.objects.count()
    
    # Get active subjects for the dropdown
    subjects = SubjectTypeMaster.objects.filter(is_active=1)

    return render_to_string(
        "L01/Dashboard/library_dashboard_memberBookRelation.html",
        {
            "total_titles": total_titles,
            "total_copies": total_copies,
            "subjects": subjects,  # pass subjects here
        }
    )

# for Dashboard 4 Librarian
def get_dashboard4_data(request):
    """
    Returns placeholder data for Dashboard 4
    """
    try:
        library_code = request.session.get('library_db', None)
        username = request.session.get('username', None)
        user_id = request.session.get('user_id', None)
        role_id = request.session.get('role_id', None)

        books_html = build_memberbookrelation_partial_html()
        print("Books HTML:", books_html)  # Debugging line  
        return JsonResponse({
            'dashboard': 'dashboard4',
            'title': 'Dashboard 4 - Reports',
            'message': 'Dashboard 4 data will be implemented here',
            'last_updated': '2024-01-01 00:00:00',
            'session_data': {
                'library_code': library_code,
                'username': username,
                'user_id': user_id,
                'role_id': role_id
            },
            'books_html': books_html,
            'status': 'success'
        })
    except Exception as e:
        return JsonResponse({
            'dashboard': 'dashboard4',
            'error': f'Error loading dashboard 4 data: {str(e)}',
            'status': 'error'
        }, status=500)

def get_dashboard_data(request, dashboard_id):
    """
    Route to get data for specific dashboard
    """
    try:
        dashboard_handlers = {
            '1': get_dashboard1_data,
            '2': get_dashboard2_data,
            '3': get_dashboard3_data,
            '4': get_dashboard4_data
        }
        
        handler = dashboard_handlers.get(dashboard_id)
        
        if handler:
            return handler(request)
        else:
            return JsonResponse({
                'error': 'Invalid dashboard ID',
                'status': 'error'
            }, status=400)
            
    except Exception as e:
        return JsonResponse({
            'error': f'Error processing dashboard request: {str(e)}',
            'status': 'error'
        }, status=500)

# for Dashboard 2 Data in model Librarian          
def catalog_detail_datatable(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')

        subject_id = request.GET.get('subject_id')
        
        if not subject_id:  # Validate subject selection
            return JsonResponse({
                'error': 'Please Select Categories'
            }, status=400)
        
        metric = request.GET.get('type')

        data = []
        total_records = 0
        
        if not subject_id:  # Validate subject selection
            return JsonResponse({
                'error': 'Please Select Categories'
            }, status=400)

        # ---------------- TITLES ----------------
        if metric == 'titles':
            qs = BookCatalog.objects.filter(subject_id=subject_id)

            if search_value:
                qs = qs.filter(
                    Q(title__icontains=search_value) |
                    Q(author__icontains=search_value)
                )

            total_records = qs.count()
            qs = qs[start:start + length]

            for obj in qs:
                data.append({
                    'title': obj.title,
                    'author': obj.author,
                    'publisher': obj.publisher
                })

        # ---------------- COPIES ----------------
        elif metric == 'copies':
            qs = BookAccession.objects.filter(catalogue__subject_id=subject_id)

            if search_value:
                qs = qs.filter(accession_no__icontains=search_value)

            total_records = qs.count()
            qs = qs[start:start + length]

            for obj in qs:
                data.append({
                    'accession_no': obj.accession_no,
                    'title': obj.catalogue.title,
                    'status': obj.status.status_name if obj.status else ''
                })

        # ---------------- PROCESSED ----------------
        elif metric == 'processed':
            qs = CirculationCopyStatus.objects.filter(
                bookcatalog__subject_id=subject_id
            )

            if search_value:
                qs = qs.filter(
                    Q(barcode__icontains=search_value) |
                    Q(accession_no__icontains=search_value)
                )

            total_records = qs.count()
            qs = qs[start:start + length]

            for obj in qs:
                data.append({
                    'barcode': obj.barcode,
                    'accession_no': obj.accession_no,
                    'current_status': obj.current_status.status_name if obj.current_status else '',
                    'shelf': obj.shelf_location.location_name if obj.shelf_location else ''
                })

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': total_records,
            'data': data
        })
    except Exception as e:
        return JsonResponse({
            'error': f'Error fetching catalog details: {str(e)}'
        }, status=500)

# for Dashboard 2 Excel Librarian
def export_catalog_excel(request):
    try:
        # Get library info
        library_code = request.session.get('library_db', None)
        library_obj = tbl_librarymasterL01.objects.filter(library_code=library_code).first()
        library_name = library_obj.library_name if library_obj else 'Library'

        subject_id = request.GET.get('subject')
        metric = request.GET.get('type')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Define thin border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ---------------- TITLES ----------------
        if metric == 'titles':
            headers = [
                'Title', 'Subtitle', 'Author', 'Other Authors', 'Publisher', 
                'ISBN/ISSN', 'Edition', 'Publication Year', 'Language',
                'Call Number', 'Cutter Number', 'Remarks'
            ]
            num_columns = len(headers)

            # Library Name at top
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
            ws.cell(row=1, column=1, value=library_name).alignment = Alignment(horizontal='center')
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)

            # Headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Data
            qs = BookCatalog.objects.filter(subject_id=subject_id)
            for row_idx, obj in enumerate(qs, start=3):
                row_values = [
                    obj.title, obj.subtitle, obj.author, obj.other_authors, obj.publisher,
                    obj.isbn_issn, obj.edition, obj.year_of_publication, obj.language,
                    obj.call_number, obj.cutter_number, obj.remarks
                ]
                for col_idx, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

        # ---------------- COPIES ----------------
        elif metric == 'copies':
            headers = [
                'Accession No', 'Title', 'Author', 'Publisher',
                'Copy Number', 'Acquisition Date', 'Location', 'Status', 'Price', 'Supplier'
            ]
            num_columns = len(headers)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
            ws.cell(row=1, column=1, value=library_name).alignment = Alignment(horizontal='center')
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            qs = BookAccession.objects.filter(catalogue__subject_id=subject_id)
            for row_idx, obj in enumerate(qs, start=3):
                row_values = [
                    obj.accession_no,
                    obj.catalogue.title if obj.catalogue else '',
                    obj.catalogue.author if obj.catalogue else '',
                    obj.catalogue.publisher if obj.catalogue else '',
                    obj.copy_number,
                    obj.acquisition_date,
                    obj.location.location_name if obj.location else '',
                    obj.status.status_name if obj.status else '',
                    obj.price,
                    obj.supplier.supplier_name if obj.supplier else '',
                ]
                for col_idx, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

        # ---------------- PROCESSED ----------------
        elif metric == 'processed':
            headers = [
                'Barcode', 'Accession No', 'Title', 'Author', 'Current Status',
                'Processing Status', 'Shelf', 'Date Processed', 'Remarks'
            ]
            num_columns = len(headers)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
            ws.cell(row=1, column=1, value=library_name).alignment = Alignment(horizontal='center')
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            qs = CirculationCopyStatus.objects.filter(bookcatalog__subject_id=subject_id)
            for row_idx, obj in enumerate(qs, start=3):
                row_values = [
                    obj.barcode,
                    obj.accession_no,
                    obj.bookcatalog.title if obj.bookcatalog else '',
                    obj.bookcatalog.author if obj.bookcatalog else '',
                    obj.current_status.status_name if obj.current_status else '',
                    obj.processing_status.status_name if obj.processing_status else '',
                    obj.shelf_location.location_name if obj.shelf_location else '',
                    obj.date_processed,
                    obj.remarks,
                ]
                for col_idx, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter

        # Auto-adjust column widths
        for i, col in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in col:
                if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
            
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=report_{metric}.xlsx'
        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(f'Error generating Excel file: {str(e)}', status=500)

# for Dashboard 2 PDF Librarian
def export_catalog_pdf(request):
    try:
        library_code = request.session.get('library_db', None)
        library_obj = tbl_librarymasterL01.objects.filter(library_code=library_code).first()
        library_name = library_obj.library_name if library_obj else 'Library'

        subject_id = request.GET.get('subject')
        metric = request.GET.get('type')

        # Fetch data based on metric
        data = []
        if metric == 'titles':
            headers = [
                'Title', 'Subtitle', 'Author', 'Other Authors', 'Publisher', 
                'ISBN/ISSN', 'Edition', 'Publication Year', 'Language',
                'Call Number', 'Cutter Number', 'Remarks'
            ]
            qs = BookCatalog.objects.filter(subject_id=subject_id)
            for obj in qs:
                data.append([
                    obj.title, obj.subtitle, obj.author, obj.other_authors, obj.publisher,
                    obj.isbn_issn, obj.edition, obj.year_of_publication, obj.language,
                    obj.call_number, obj.cutter_number, obj.remarks
                ])

        elif metric == 'copies':
            headers = [
                'Accession No', 'Title', 'Author', 'Publisher',
                'Copy Number', 'Acquisition Date', 'Location', 'Status', 'Price', 'Supplier'
            ]
            qs = BookAccession.objects.filter(catalogue__subject_id=subject_id)
            for obj in qs:
                data.append([
                    obj.accession_no,
                    obj.catalogue.title if obj.catalogue else '',
                    obj.catalogue.author if obj.catalogue else '',
                    obj.catalogue.publisher if obj.catalogue else '',
                    obj.copy_number,
                    obj.acquisition_date,
                    obj.location.location_name if obj.location else '',
                    obj.status.status_name if obj.status else '',
                    obj.price,
                    obj.supplier.supplier_name if obj.supplier else ''
                ])

        elif metric == 'processed':
            headers = [
                'Barcode', 'Accession No', 'Title', 'Author', 'Current Status',
                'Processing Status', 'Shelf', 'Date Processed', 'Remarks'
            ]
            qs = CirculationCopyStatus.objects.filter(bookcatalog__subject_id=subject_id)
            for obj in qs:
                data.append([
                    obj.barcode,
                    obj.accession_no,
                    obj.bookcatalog.title if obj.bookcatalog else '',
                    obj.bookcatalog.author if obj.bookcatalog else '',
                    obj.current_status.status_name if obj.current_status else '',
                    obj.processing_status.status_name if obj.processing_status else '',
                    obj.shelf_location.location_name if obj.shelf_location else '',
                    obj.date_processed,
                    obj.remarks
                ])

        # ---------------- Build HTML table ----------------
        table_rows = ""
        for row in data:
            row_html = ""
            for cell in row:
                # Replace None or empty string with '-'
                cell_value = cell if cell not in [None, "", "None"] else "-"
                row_html += f"<td>{cell_value}</td>"
            table_rows += f"<tr>{row_html}</tr>"

        headers_html = "".join(f"<th>{header}</th>" for header in headers)

        current_datetime = datetime.now().strftime("%d-%m-%Y")

        html_string = f"""
        <!DOCTYPE html>
        <html lang="mr">
        <head>
            <meta charset="UTF-8">
            <title>{library_name} - Report</title>
        </head>
        <body>
            <h1 style="text-align:center;">{library_name}</h1>
            <table>
                <thead>
                    <tr>
                        {headers_html}
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
        </body>
        </html>
        """

        # ---------------- CSS with page breaks, footer ----------------
        font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSansDevanagari-Regular.ttf")
        css_string = f"""
        @font-face {{
            font-family: 'NotoDeva';
            src: url('file://{font_path}');
        }}
        @page {{
            size: A4;
            margin: 10px 10px 40px 10px; /* small margin for footer */
            @bottom-right {{
                content: "Page " counter(page) " | {current_datetime}";
                font-size: 8pt;
                font-family: 'NotoDeva';
            }}
        }}
        body {{
            font-family: 'NotoDeva', sans-serif;
            font-size: 9pt;
            margin: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed; /* fit table in page */
            page-break-inside: auto;
        }}
        tr {{
            page-break-inside: avoid;  /* do not break row across pages */
            page-break-after: auto;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 4px;
            text-align: center;
            word-wrap: break-word;
        }}
        th {{
            font-weight: bold;
            background-color: #d3d3d3;
        }}
        h1 {{
            font-size: 16pt;
            margin: 5px 0;
        }}
        """

        # Generate PDF
        pdf_file = HTML(string=html_string).write_pdf(stylesheets=[CSS(string=css_string)])

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="report_{metric}.pdf"'
        return response

    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)

# For Kiosk Display and Ebook Catalogue

def kiosk_display(request):
    # Get library_code from session
    library_code_from_session = request.session.get('library_db', None)
    library_name = None  # This will now store Marathi name if available

    # If library_code is provided in session, get library details from tbl_librarymasterL01
    if library_code_from_session:
        try:
            library = tbl_librarymasterL01.objects.filter(
                library_code__iexact=library_code_from_session
            ).first()
            if library:
                # Prefer Marathi name if available, otherwise fallback to English name
                library_name = library.library_name_mar if library.library_name_mar else library.library_name
                
        except Exception as e:
            library_name = None
    
    # Render the template with context
    return render(request, "L01/kiosk_display.html", {
        "MEDIA_URL": settings.MEDIA_URL,
        "library_name": library_name  # Marathi name will go here if present
    })
    
def set_ebook_image_urls(ebooks):
    for b in ebooks:

        if b.eb_front_page_photo:
            b.eb_front_page_photo = file_storage_service.get_file_url(
                b.eb_front_page_photo
            )
        else:
            b.eb_front_page_photo = ""

        if b.eb_last_page_photo:
            b.eb_last_page_photo = file_storage_service.get_file_url(
                b.eb_last_page_photo
            )
        else:
            b.eb_last_page_photo = ""
            
@login_required
def visit_Library_ebook_catalogue(request):
    try:
        # Session checks
        library_code = request.session.get('library_db')
        username = request.session.get('username')
        user_id = request.session.get('user_id')
        role_id = request.session.get('role_id')

        if not all([library_code, username, user_id, role_id]):
            messages.warning(request, "Session expired or invalid. Please login again.")
            return render(request, "L01/LibraryCateVisit/visit_Library_ebook_catalogue.html", {})

        # Load all active subjects
        subjects_qs = SubjectTypeMaster.objects.filter(is_active=1)
        subjects = []
        for s in subjects_qs:
            s.id_enc = enc(str(s.id))
            subjects.append(s)

        first_subject = subjects[0] if subjects else None

        # Fetch ebooks for the first subject
        if first_subject:
            ebooks = LibraryEbook.objects.filter(eb_subject=first_subject).select_related('eb_subject')
        else:
            ebooks = LibraryEbook.objects.none()

        # Pagination – 8 ebooks per page
        paginator = Paginator(ebooks, 8)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        # Encode ebook IDs
        for b in page_obj:
            b.ebookIdEnc = enc(str(b.ebook_id)) if 'enc' in globals() else b.ebook_id
        set_ebook_image_urls(page_obj)

        # --- NEW ARRIVALS LOGIC ---
        thirty_days_ago = timezone.now() - timedelta(days=30)
        new_ebooks = LibraryEbook.objects.filter(created_at__gte=thirty_days_ago).order_by('-created_at')

        if not new_ebooks.exists():
            new_ebooks = LibraryEbook.objects.all().order_by('-ebook_id')[:10]
        set_ebook_image_urls(new_ebooks)

        for b in new_ebooks:
            b.ebookIdEnc = enc(str(b.ebook_id)) if 'enc' in globals() else b.ebook_id

        context = {
            'subjects': subjects,
            'ebooks': page_obj,
            'paginator': paginator,
            'page_number': int(page_number),
            'first_subject_id_enc': first_subject.id_enc if first_subject else None,
            'MEDIA_URL': settings.MEDIA_URL,
            'new_ebooks': new_ebooks,
        }

        return render(request, "L01/LibraryCateVisit/visit_library_Cate_ebooks.html", context)

    except Exception as e:
        print(f"Error Ebook Catalogue: {e}")
        return render(request, "L01/LibraryCateVisit/visit_library_Cate_ebooks.html", {})
    
@login_required
def get_ebooks_by_subject(request):
    try:
        subject_id_enc = request.GET.get('subject_id')
        subject_id = dec(subject_id_enc)

        search = request.GET.get('search', '').strip()

        # Base queryset
        all_ebooks = LibraryEbook.objects.filter(eb_subject_id=subject_id).select_related('eb_subject')

        # Search filter
        if search:
            all_ebooks = all_ebooks.filter(
                Q(eb_title__icontains=search) |
                Q(eb_author__icontains=search)
            )

        # Encode ebook IDs
        for b in all_ebooks:
            b.ebookIdEnc = enc(str(b.ebook_id)) if 'enc' in globals() else b.ebook_id

        # Pagination
        paginator = Paginator(all_ebooks, 8)
        page_number = request.GET.get('page', 1)
        ebooks_page = paginator.get_page(page_number)

        context = {
            'ebooks': ebooks_page,
            'MEDIA_URL': settings.MEDIA_URL,
            'subject_id_enc': subject_id_enc,
        }

        return render(request, "L01/LibraryCateVisit/ebook_list_partial.html", context)

    except Exception as e:
        print("Error fetching ebooks:", e)
        return JsonResponse({'error': 'Failed to fetch ebooks'}, status=500)

# Ebook Kiosk View
def visit_library_Cate_ebooks(request):
    try:
        # --- SESSION CHECKS ---
        library_code = request.session.get('library_db')
        username = request.session.get('username')
        user_id = request.session.get('user_id')
        role_id = request.session.get('role_id')

        if library_code != 'L01':
            messages.error(request, "Invalid library access.")
            request.session.flush()
            return redirect('library_list')

        if not all([library_code, username, user_id, role_id]):
            messages.warning(request, "Session expired or invalid. Please login again.")
            return render(request, "L01/visit_library_Cate_ebooks.html", {})

        # --- SUBJECTS ---
        subjects_qs = SubjectTypeMaster.objects.filter(is_active=1)
        subjects = []
        for s in subjects_qs:
            s.id_enc = enc(str(s.id))
            subjects.append(s)

        print(f"DEBUG: Found {len(subjects)} subjects")
        
        first_subject = subjects[0] if subjects else None
        print(f"DEBUG: First subject ID: {first_subject.id if first_subject else 'None'}")

        # --- EBOOKS BY FIRST SUBJECT ---
        if first_subject:
            ebooks = LibraryEbook.objects.filter(eb_subject_id=first_subject.id)\
                                        .select_related('eb_subject', 'ebook_type')
            print(f"DEBUG: Found {ebooks.count()} ebooks for subject {first_subject.id}")
        else:
            ebooks = LibraryEbook.objects.none()
            print("DEBUG: No first subject, ebooks queryset empty")

        # Debug first few ebooks
        for i, ebook in enumerate(ebooks[:3]):
            print(f"DEBUG: Initial ebook {i+1}: {ebook.eb_title}")

        # --- PAGINATION ---
        paginator = Paginator(ebooks, 8)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        print(f"DEBUG: Paginator count: {paginator.count}")
        print(f"DEBUG: Page object has {len(page_obj)} items")

        # Process ebooks
        processed_ebooks = []
        for ebook in page_obj:
            ebook.ebookIdEnc = enc(str(ebook.ebook_id)) if 'enc' in globals() else ebook.ebook_id
            
            if ebook.eb_front_page_photo:
                ebook.eb_front_page_photo = file_storage_service.get_file_url(
                    ebook.eb_front_page_photo.strip()
                )
                print(f"DEBUG: Set front page for {ebook.eb_title}")
            
            if ebook.eb_last_page_photo:
                ebook.eb_last_page_photo = file_storage_service.get_file_url(
                    ebook.eb_last_page_photo.strip()
                )
            
            processed_ebooks.append(ebook)

        # --- CONTEXT ---
        context = {
            'subjects': subjects,
            'ebooks': page_obj,  # This should be 'ebooks'
            'paginator': paginator,
            'page_number': int(page_number),
            'first_subject_id_enc': first_subject.id_enc if first_subject else None,
            'MEDIA_URL': settings.MEDIA_URL,
        }

        print(f"DEBUG: Final context - ebooks count: {len(context.get('ebooks', []))}")

        return render(
            request,
            "L01/visit_library_Cate_ebooks.html",
            context
        )

    except Exception as e:
        print(f"Error in ebooks kiosk: {e}")
        return render(
            request,
            "L01/visit_library_Cate_ebooks.html",
            {}
        )

def get_ebooks_by_subject_kiosk(request):
    try:
        subject_id_enc = request.GET.get('subject_id')
        subject_id = dec(subject_id_enc) if subject_id_enc else None

        search = request.GET.get('search', '').strip()
        page = request.GET.get('page', 1)

        print(f"DEBUG: Subject ID encrypted: {subject_id_enc}")
        print(f"DEBUG: Subject ID decrypted: {subject_id}")
        print(f"DEBUG: Search term: {search}")
        print(f"DEBUG: Page: {page}")

        # Base queryset
        ebooks = LibraryEbook.objects.all().select_related('eb_subject', 'ebook_type')
        
        print(f"DEBUG: Total ebooks in DB: {ebooks.count()}")

        # Filter by subject if provided
        if subject_id:
            ebooks = ebooks.filter(eb_subject_id=subject_id)
            print(f"DEBUG: After subject filter: {ebooks.count()}")

        # Apply search filter
        if search:
            ebooks = ebooks.filter(
                Q(eb_title__icontains=search) |
                Q(eb_author__icontains=search) |
                Q(eb_keywords__icontains=search)
            )
            print(f"DEBUG: After search filter: {ebooks.count()}")

        # Debug first few ebooks
        for i, ebook in enumerate(ebooks[:3]):
            print(f"DEBUG: Ebook {i+1}: {ebook.eb_title} | Subject: {ebook.eb_subject_id}")

        # Process ebooks
        ebooks_list = []
        for ebook in ebooks:
            ebook.ebookIdEnc = enc(str(ebook.ebook_id)) if 'enc' in globals() else ebook.ebook_id
            
            # Set image URLs
            if ebook.eb_front_page_photo:
                ebook.eb_front_page_photo = file_storage_service.get_file_url(
                    ebook.eb_front_page_photo.strip()
                )
            
            if ebook.eb_last_page_photo:
                ebook.eb_last_page_photo = file_storage_service.get_file_url(
                    ebook.eb_last_page_photo.strip()
                )
            
            ebooks_list.append(ebook)

        print(f"DEBUG: Processed {len(ebooks_list)} ebooks")

        # Pagination
        paginator = Paginator(ebooks_list, 8)
        try:
            ebooks_page = paginator.page(page)
            print(f"DEBUG: Pagination successful. Page {page} of {paginator.num_pages}")
            print(f"DEBUG: Current page has {len(ebooks_page)} ebooks")
        except Exception as e:
            print(f"DEBUG: Pagination error: {e}")
            ebooks_page = paginator.page(1)

        context = {
            'ebooks': ebooks_page,  # Make sure it's 'ebooks' not 'ebooks_page'
            'subject_id_enc': subject_id_enc,
            'MEDIA_URL': settings.MEDIA_URL,
        }

        print(f"DEBUG: Context keys: {context.keys()}")
        if 'ebooks' in context:
            print(f"DEBUG: Ebooks in context: {len(context['ebooks'])}")

        return render(
            request,
            "L01/ebook_details_kiosk.html",
            context
        )

    except Exception as e:
        print(f"DEBUG: Error in get_ebooks_by_subject_kiosk: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': 'Failed to fetch ebooks'}, status=500)

@login_required
def view_ebook_detail(request):
    try:
        # -----------------------
        # SESSION VALIDATION
        # -----------------------
        library_code = request.session.get('library_db')
        username = request.session.get('username')
        user_id = request.session.get('user_id')
        role_id = request.session.get('role_id')

        if not all([library_code, username, user_id, role_id]):
            messages.warning(request, "Session expired or invalid. Please login again.")
            return render(request, "L01/LibraryCateVisit/visit_library_Cate_ebooks.html", {})

        # -----------------------
        # EBOOK ID VALIDATION
        # -----------------------
        ebook_id_enc = request.GET.get('ebook_id')
        if not ebook_id_enc:
            messages.error(request, "Invalid ebook request.")
            return redirect("visit_Library_ebook_catalogue")

        ebook_id = dec(ebook_id_enc)

        # -----------------------
        # FETCH EBOOK DATA
        # -----------------------
        ebook = get_object_or_404(LibraryEbook, ebook_id=ebook_id)

        # Collect images (same style as your physical book)
        images = []
        if ebook.eb_front_page_photo:
            images.append(ebook.eb_front_page_photo)
        if ebook.eb_last_page_photo:
            images.append(ebook.eb_last_page_photo)

        # -----------------------
        # REVIEWS (if needed)
        # -----------------------
        from django.contrib.auth import get_user_model
        User = get_user_model()

        all_reviews = BookReview.objects.filter(ebook=ebook).order_by('-created_at')

        user_ids = all_reviews.values_list('user_id', flat=True).distinct()
        users = User.objects.filter(id__in=user_ids)
        user_dict = {u.id: u for u in users}

        reviews_with_users = []
        for r in all_reviews:
            r.user_obj = user_dict.get(r.user_id)
            reviews_with_users.append(r)

        # Current user review
        current_user_id = request.session.get('user_id')
        user_review = None
        if current_user_id:
            try:
                user_review = all_reviews.filter(user_id=int(current_user_id)).first()
                if user_review:
                    user_review.user_obj = user_dict.get(int(current_user_id))
            except:
                user_review = None

        # Rating summary
        avg_rating = all_reviews.aggregate(avg=models.Avg('rating'))['avg'] or 0
        total_reviews = all_reviews.count()

        # Reviews pagination (same style)
        reviews_per_page = 5
        show_pagination = total_reviews > reviews_per_page

        if show_pagination:
            paginator = Paginator(reviews_with_users, reviews_per_page)
            page_number = request.GET.get('page', 1)
            try:
                reviews_page = paginator.page(page_number)
            except PageNotAnInteger:
                reviews_page = paginator.page(1)
            except EmptyPage:
                reviews_page = paginator.page(paginator.num_pages)
        else:
            reviews_page = reviews_with_users

        # -----------------------
        # FINAL CONTEXT
        # -----------------------
        context = {
            "ebook": ebook,
            "images": images,
            "MEDIA_URL": settings.MEDIA_URL,

            # Ebook Metadata
            "title": ebook.eb_title,
            "subtitle": ebook.eb_subtitle,
            "author": ebook.eb_author,
            "publisher": ebook.eb_publisher,
            "isbn": ebook.eb_isbn_issn,
            "edition": ebook.eb_edition,
            "publication_year": ebook.eb_year_of_publication,
            "language": ebook.eb_language,
            "keywords": ebook.eb_keywords,
            "classification": ebook.eb_classification_number,
            "pages": ebook.eb_pages,
            "remarks": ebook.remarks,
            "pdf_url": ebook.eb_pdf_url,
            "subject": ebook.eb_subject.subject_name if ebook.eb_subject else None,
            "ebook_type": ebook.ebook_type.type_name if ebook.ebook_type else None,
            "call_number": ebook.call_number,
            "cutter_number": ebook.cutter_number,

            # Reviews section
            "reviews": reviews_page,
            "user_review": user_review,
            "avg_rating": round(avg_rating, 1),
            "total_reviews": total_reviews,
            "show_pagination": show_pagination,
            "reviews_per_page": reviews_per_page,

            # Enc ID
            "ebook_id_enc": ebook_id_enc,

            "current_params": request.GET.copy(),
        }

        return render(request, "L01/LibraryCateVisit/view_ebook_detail.html", context)

    except Exception as e:
        print("Error in view_ebook_detail:", e)
        messages.error(request, "Unable to load ebook details.")
        return redirect("visit_Library_ebook_catalogue")

# def kiosk_competitive_exam_type(request):
    
#     competitive_id = request.GET.get('competitive_id', None)
    
#     if competitive_id:
#         # Show exam details
#         exam = get_object_or_404(CompetitiveExamMaster, competitive_id=competitive_id)
#         return render(request, "L01/competitive_exam_details.html", {
#             "MEDIA_URL": settings.MEDIA_URL,
#             "exam": exam
#         })
    
#     # Show exam list
#     competitive_exams = CompetitiveExamMaster.objects.all().order_by('full_name')
#     return render(request, "L01/kiosk_competitive_exam_type.html", {
#         "MEDIA_URL": settings.MEDIA_URL,
#         "competitive_exams": competitive_exams
#     })
    
@login_required
def visit_Library_catalogue_kiosk(request):
    try:
        # --- SESSION CHECKS ---
        library_code = request.session.get('library_db')
        username = request.session.get('username')
        user_id = request.session.get('user_id')
        role_id = request.session.get('role_id')

        if library_code != 'L01':
            messages.error(request, "Invalid library access.")
            request.session.flush()
            return redirect('library_list')

        if not all([library_code, username, user_id, role_id]):
            messages.warning(request, "Session expired or invalid. Please login again.")
            return render(request, "L01/visit_Library_catalogue_kiosk.html", {})

        # --- SUBJECTS ---
        subjects_qs = SubjectTypeMaster.objects.filter(is_active=1)
        subjects = []
        for s in subjects_qs:
            s.id_enc = enc(str(s.id))
            subjects.append(s)

        first_subject = subjects[0] if subjects else None

        # --- BOOKS BY FIRST SUBJECT ---
        if first_subject:
            books = BookCatalog.objects.filter(subject_id=first_subject.id)\
                                       .select_related('subject', 'material')
        else:
            books = BookCatalog.objects.none()

        # --- PAGINATION ---
        paginator = Paginator(books, 8)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        for b in page_obj:
            b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num

        set_book_image_urls(page_obj)

        # --- UPCOMING / LATEST BOOKS ---
        latest_created_at = BookCatalog.objects.aggregate(
            latest=Max('created_at')
        )['latest']

        recent_books = (
            BookCatalog.objects.filter(created_at=latest_created_at)
            if latest_created_at else BookCatalog.objects.none()
        )

        remaining_count = 10 - recent_books.count()
        fallback_books = (
            BookCatalog.objects
            .exclude(cat_ref_num__in=recent_books.values_list('cat_ref_num', flat=True))
            .order_by('-cat_ref_num')[:remaining_count]
            if remaining_count > 0 else BookCatalog.objects.none()
        )

        new_books = list(recent_books) + list(fallback_books)
        for b in new_books:
            b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num
        set_book_image_urls(new_books)

        # --- MOST REVIEWED BOOKS ---
        most_reviewed_books = (
            BookCatalog.objects.annotate(
                review_count=Count('cat_ref_id_reviews'),
                avg_rating=Avg('cat_ref_id_reviews__rating')
            )
            .filter(review_count__gt=0)
            .order_by('-avg_rating', '-review_count')[:10]
        )

        for b in most_reviewed_books:
            b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num
            avg = b.avg_rating or 0
            b.avg_rating = round(avg, 1)
            b.stars = [True if i < round(avg) else False for i in range(5)]

        set_book_image_urls(most_reviewed_books)

        # --- NEW ARRIVALS IN CIRCULATION ---
        new_arrivals_qs = BookCatalog.objects.annotate(
            in_circulation=Exists(
                CirculationCopyStatus.objects.filter(
                    bookcatalog_id=OuterRef('cat_ref_num')
                )
            )
        ).filter(in_circulation=True).order_by('-cat_ref_num')[:10]

        for b in new_arrivals_qs:
            b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num

        set_book_image_urls(new_arrivals_qs)
            

        # --- CONTEXT ---
        context = {
            'subjects': subjects,
            'books': page_obj,
            'paginator': paginator,
            'page_number': int(page_number),
            'first_subject_id_enc': first_subject.id_enc if first_subject else None,
            'MEDIA_URL': settings.MEDIA_URL,
            'new_books': new_books,
            'most_reviewed_books': most_reviewed_books,
            'new_arrivals_qs': new_arrivals_qs,
        }

        return render(
            request,
            "L01/visit_Library_catalogue_kiosk.html",
            context
        )

    except Exception as e:
        print(f"Error: {e}")
        return render(
            request,
            "L01/visit_Library_catalogue_kiosk.html",
            {}
        )

def get_books_by_subject_kiosk(request):
    try:
        subject_id_enc = request.GET.get('subject_id')
        subject_id = dec(subject_id_enc) if subject_id_enc else None

        search = request.GET.get('search', '').strip()
        searching = request.GET.get('searching', '').strip()

        # Base queryset
        all_books = BookCatalog.objects.all().select_related('subject', 'material')

        # Check if we're searching globally (searching parameter)
        if searching:
            # Global search - search across all books regardless of subject
            if search:
                all_books = all_books.filter(
                    Q(title__icontains=search) |
                    Q(author__icontains=search)
                )
            # If no search term but has subject_id, filter by subject
            elif subject_id:
                all_books = all_books.filter(subject_id=subject_id)
            # If no search term and no subject_id, show all books
            else:
                all_books = all_books.all()

        else:
            # Regular browsing mode (not global search)
            if subject_id:
                all_books = all_books.filter(subject_id=subject_id)

            if search:
                all_books = all_books.filter(
                    Q(title__icontains=search) |
                    Q(author__icontains=search)
                )

        # Encode book IDs and get image URLs
        for b in all_books:
            b.bookIdEnc = enc(str(b.cat_ref_num)) if 'enc' in globals() else b.cat_ref_num

            b.front_image_url = ""
            if b.front_page_photo:
                b.front_page_photo = file_storage_service.get_file_url(
                    b.front_page_photo.strip()
                )

            # Last page image
            b.back_image_url = ""
            if b.last_page_photo:
                b.last_page_photo = file_storage_service.get_file_url(
                    b.last_page_photo.strip()
                )

        # Pagination
        paginator = Paginator(all_books, 8)
        page_number = request.GET.get('page', 1)
        books_page = paginator.get_page(page_number)

        context = {
            'books': books_page,
            'MEDIA_URL': settings.MEDIA_URL,
            'subject_id_enc': subject_id_enc,
            'is_kiosk': True,
        }

        return render(
            request,
            "L01/book_details_kiosk.html",
            context
        )

    except Exception as e:
        print("Kiosk error fetching books:", e)
        return JsonResponse({'error': 'Failed to fetch books'}, status=500)

@login_required
def view_book_detail_kiosk(request):
    try:
        # --- SESSION CHECKS ---
        library_code = request.session.get('library_db')
        username = request.session.get('username')
        user_id = request.session.get('user_id')
        role_id = request.session.get('role_id')

        if not all([library_code, username, user_id, role_id]):
            messages.warning(request, "Session expired or invalid.")
            return redirect("visit_library_catalogue")

        # --- GET BOOK ID ---
        cat_ref_num_enc = request.GET.get('cat_ref_num')
        if not cat_ref_num_enc:
            messages.error(request, "Invalid book request.")
            return redirect("visit_library_catalogue")

        cat_ref_num = dec(cat_ref_num_enc)

        # --- FETCH BOOK ---
        book = get_object_or_404(BookCatalog, cat_ref_num=cat_ref_num)
        book.bookIdEnc = enc(str(book.cat_ref_num))

        images = []
        if book.front_page_photo:
            book.front_page_photo_url = file_storage_service.get_file_url(book.front_page_photo)
            images.append(book.front_page_photo_url)

        if book.last_page_photo:
            book.last_page_photo_url = file_storage_service.get_file_url(book.last_page_photo)
            images.append(book.last_page_photo_url)

        # --- CIRCULATION DETAILS ---
        circulation_qs = CirculationCopyStatus.objects.filter(
            bookcatalog_id=book.cat_ref_num
        )
        total_qty = circulation_qs.count()

        status_on_shelf = status_master.objects.filter(
            status_name__iexact="On-Shelf"
        ).first()

        status_unknown = "Not Available"

        current_on_shelf_count = (
            circulation_qs.filter(current_status=status_on_shelf).count()
            if status_on_shelf else 0
        )

        if total_qty > 0:
            if current_on_shelf_count > 0:
                current_status_display = "On-Shelf"
                availability_text = f"{current_on_shelf_count} of {total_qty} available"
            else:
                current_status_display = status_unknown
                availability_text = f"0 of {total_qty} available"
        else:
            current_status_display = status_unknown
            availability_text = "-"

        # --- LOCATION (ONLY ON-SHELF) ---
        location_counts = {}
        if status_on_shelf:
            on_shelf_copies = circulation_qs.filter(current_status=status_on_shelf)
            for circ in on_shelf_copies:
                if circ.shelf_location:
                    loc = circ.shelf_location.location_name
                    location_counts[loc] = location_counts.get(loc, 0) + 1

        if location_counts:
            location_display = ", ".join(
                [f"{loc} ({count})" for loc, count in location_counts.items()]
            )
        else:
            location_display = "Unknown"

        from django.contrib.auth import get_user_model
        User = get_user_model()

        all_reviews = BookReview.objects.filter(book=book).order_by('-created_at')

        user_ids = all_reviews.values_list('user_id', flat=True).distinct()
        users = User.objects.filter(id__in=user_ids)
        user_dict = {u.id: u for u in users}

        reviews_with_users = []
        for review in all_reviews:
            review.user_obj = user_dict.get(review.user_id)
            reviews_with_users.append(review)

        current_user_id = request.session.get('user_id')

        user_review = None
        if current_user_id:
            try:
                user_review = all_reviews.filter(
                    user_id=int(current_user_id)
                ).first()
                if user_review:
                    user_review.user_obj = user_dict.get(int(current_user_id))
            except (ValueError, TypeError):
                user_review = None

        # --- RATINGS ---
        avg_rating = all_reviews.aggregate(
            avg=models.Avg('rating')
        )['avg'] or 0

        total_reviews = all_reviews.count()

        # --- PAGINATION ---
        reviews_per_page = 5
        show_pagination = total_reviews > reviews_per_page

        if show_pagination:
            paginator = Paginator(reviews_with_users, reviews_per_page)
            page_number = request.GET.get('page', 1)

            try:
                reviews_page = paginator.page(page_number)
            except PageNotAnInteger:
                reviews_page = paginator.page(1)
            except EmptyPage:
                reviews_page = paginator.page(paginator.num_pages)
        else:
            reviews_page = reviews_with_users

        # --- CONTEXT ---
        context = {
            'book': book,
            'images': images,
            'MEDIA_URL': settings.MEDIA_URL,
            'total_qty': total_qty,
            'current_status_display': current_status_display,
            'availability_text': availability_text,
            'location_display': location_display,
            'reviews': reviews_page,
            'user_review': user_review,
            'avg_rating': round(avg_rating, 1),
            'total_reviews': total_reviews,
            'show_pagination': show_pagination,
            'reviews_per_page': reviews_per_page,
            'current_params': request.GET.copy(),
            'cat_ref_num_enc': cat_ref_num_enc,
        }

        return render(
            request,
            "L01/view_book_detail_kiosk.html",
            context
        )

    except Exception as e:
        print("Error in view_book_detail_kiosk:", e)
        messages.error(request, "Unable to load book details.")
        return redirect("L01/visit_Library_catalogue_kiosk")
    
def get_dashboard_detail_data(detail_type, from_date, to_date, label):
    from django.db.models.functions import Coalesce
    """
    Returns list of dicts for dashboard export
    Compatible with Excel & Marathi PDF
    """
    from django.utils.timezone import now
    today = now().date()
    # ---------------- BOOK TRANSACTIONS ----------------
    if detail_type == 'issued':
        qs = CirculationTransaction.objects.filter(
            issue_date__range=[from_date, to_date],
        )

    elif detail_type == 'returned':
        qs = CirculationTransaction.objects.filter(
            return_date__range=[from_date, to_date],
        )


    # ---------------- DUE BOOKS ----------------
    elif detail_type == 'due':
        qs = CirculationTransaction.objects.filter(
            due_date__lt=today,                 # due date passed
            return_date__isnull=True,            # not returned yet
            due_date__range=[from_date, to_date] # filter by selected range
        )

    # ---------------- DAMAGED / LOST BOOKS ----------------
    elif detail_type == 'damaged':
        qs = CirculationTransaction.objects.filter(
            return_condition__id__in=[18, 19],   # 18 = damaged, 19 = lost
            return_date__range=[from_date, to_date]
        )

    # ---------- RETURN BOOK DATA ----------
    if detail_type in ['issued', 'returned', 'due', 'damaged']:
        return list(
            qs.select_related('member', 'catalog')
              .annotate(
                  member_code=F('membership_code'),
                  member_name=Concat(
                      Coalesce('member__first_name_mar', Value('')),
                      Value(' '),
                      Coalesce('member__middle_name_mar', Value('')),
                      Value(' '),
                      Coalesce('member__last_name_mar', Value('')),
                      output_field=CharField()
                  ),
                  book_name=F('catalog__title'),
                  book_barcode=Coalesce('barcode', Value('', output_field=CharField())),
                  date=Coalesce(
                      'issue_date',
                      'return_date',
                      'due_date'
                  )
              )
              .values(
                  'member_code',
                  'member_name',
                  'book_name',
                  'book_barcode',
                  'date'
              )
        )

    # ---------------- PAYMENTS ----------------
    elif detail_type in ['monthly', 'total', 'fine', 'book_fine']:
        amount_field_map = {
            'monthly': 'monthly_subscription_amount',
            'total': 'total_subscription_amount',
            'fine': 'fine_amount',
            'book_fine': 'book_fine_amount'
        }

        amount_field = amount_field_map.get(detail_type)

        return list(
            PaymentDetails.objects.filter(
                payment_date__range=[from_date, to_date]
            )
            .exclude(**{f"{amount_field}__isnull": True})
            .select_related('membership')
            .annotate(
                member_code=F('membership_code'),
                member_name=Concat(
                      Coalesce('membership__first_name_mar', Value('')),
                      Value(' '),
                      Coalesce('membership__middle_name_mar', Value('')),
                      Value(' '),
                      Coalesce('membership__last_name_mar', Value('')),
                      output_field=CharField()
                  ),
                amount=F(amount_field)
            )
            .values(
                'member_code',
                'member_name',
                'payment_date',
                'amount'
            )
        )
    elif detail_type == "footfall_online":
        from datetime import datetime, time
        from django.db.models import OuterRef, Subquery

        start_datetime = datetime.combine(from_date, time.min)
        end_datetime = datetime.combine(to_date, time.max)

        # 🔑 Latest screen activity per member (MySQL safe)
        latest_activity = (
            MemberScreenActivity.objects.using("L01")
            .filter(
                session__member_id=OuterRef('session__member_id'),
                visited_at__range=[start_datetime, end_datetime],
                screen_name=label
            )
            .order_by('-visited_at')
            .values('visited_at')[:1]
        )

        qs = (
            MemberScreenActivity.objects.using("L01")
            .select_related(
                'session',
                'session__member',
                'session__member__membership'
            )
            .filter(
                visited_at=Subquery(latest_activity),
                screen_name=label
            )
            .order_by('-visited_at')
        )

        return [
            {
                "membership_type": obj.session.member.membership.membership_type if obj.session.member else "",
                "member_code": obj.session.member.membership_code if obj.session.member else "",
                "member_name": (
                    f"{obj.session.member.first_name_mar or ''} "
                    f"{obj.session.member.last_name_mar or ''}"
                ).strip() if obj.session.member else "",
                "login_time": obj.session.login_time,
                "visited_at": obj.visited_at,
            }
            for obj in qs
        ]

    elif detail_type == "footfall_offline":
        from datetime import datetime, date, time
        from django.db.models import Q

        # -----------------------------
        # 1️⃣ Ensure from_date is date
        # -----------------------------
        if isinstance(from_date, date):
            from_date_obj = from_date
        else:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()

        # -----------------------------
        # 2️⃣ Parse label (format: 18-Nov)
        # -----------------------------
        label_date = datetime.strptime(label, "%d-%b")

        # year from from_date
        entry_date = label_date.replace(year=from_date_obj.year).date()

        entry_start = datetime.combine(entry_date, time.min)
        entry_end = datetime.combine(entry_date, time.max)

        # -----------------------------
        # 3️⃣ Fetch offline entries (that day only)
        # -----------------------------
        qs = (
            MemberEntryExit.objects.using("L01")
            .filter(entry_time__range=[entry_start, entry_end])
            .order_by('-entry_time')
        )

        # -----------------------------
        # 4️⃣ DISTINCT membership_code (MySQL-safe)
        # -----------------------------
        qs = list(qs)

        seen = set()
        unique_entries = []

        for obj in qs:
            if obj.membership_code not in seen:
                seen.add(obj.membership_code)
                unique_entries.append(obj)

        # -----------------------------
        # 5️⃣ Fetch member details
        # -----------------------------
        membership_codes = [obj.membership_code for obj in unique_entries]

        member_map = {
            m.membership_code: m
            for m in MembershipDetails.objects.using("L01")
            .select_related("membership")
            .filter(membership_code__in=membership_codes)
        }

        # -----------------------------
        # 6️⃣ Build response
        # -----------------------------
        return [
            {
                "membership_type": (
                    member_map[obj.membership_code].membership.membership_type
                    if obj.membership_code in member_map else ""
                ),
                "member_code": obj.membership_code,
                "member_name": (
                    f"{member_map[obj.membership_code].first_name_mar or ''} "
                    f"{member_map[obj.membership_code].last_name_mar or ''}"
                ).strip() if obj.membership_code in member_map else "",
                "login_time": obj.entry_time,
            }
            for obj in unique_entries
        ]



    return []
def make_naive(dt):
    if dt and hasattr(dt, 'tzinfo') and dt.tzinfo:
        return dt.replace(tzinfo=None)
    return dt

def export_excel(data, detail_type):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    if detail_type in ['monthly','total','fine','book_fine']:
        headers = ['Member Code', 'Member Name', 'Payment Date', 'Amount']
    elif detail_type in ['footfall_online', 'footfall_offline']:
        headers = ['Member Code','Member Type','Member Name','Login Time']
    else:
        headers = ['Member Code', 'Member Name', 'Book Title', 'Book Barcode', 'Date']

    ws.append(headers)

    for row in data:
        cleaned_row = []
        for value in row.values():
            if isinstance(value, datetime):
                value = make_naive(value)
            cleaned_row.append(value)
        ws.append(cleaned_row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="dashboard_report.xlsx"'
    wb.save(response)
    return response

def dashboard_export(request):
    detail_type = request.GET.get('detail_type')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    format_type = request.GET.get('format')
    label = request.GET.get('label')

    
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
    to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()

    data = get_dashboard_detail_data(detail_type, from_date, to_date, label)

    if format_type == 'excel':
        return export_excel(data, detail_type)

    elif format_type == 'pdf':
        return export_pdf(data, detail_type)

    return HttpResponse("Invalid request", status=400)

from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
from reportlab.pdfgen import canvas

def export_pdf(data, detail_type):
    try:
        # ---------------- Headers based on type ----------------
        if detail_type in ['monthly', 'total', 'fine', 'book_fine']:
            headers = ['सभासद कोड', 'सभासद नाव', 'देयक तारीख', 'रक्कम']
        elif detail_type in ['footfall_online', 'footfall_offline']:
            headers = ['सभासद प्रकार','सभासद कोड','सभासद नाव','लॉगिन वेळ']
        else:
            headers = ['सभासद कोड', 'सभासद नाव', 'पुस्तक नाव', 'बारकोड', 'तारीख']

        


        # ---------------- Build table rows ----------------
        table_rows = ""
        for row in data:
            row_html = ""
            for value in row.values():
                cell_value = value if value not in [None, "", "None"] else "-"
                row_html += f"<td>{cell_value}</td>"
            table_rows += f"<tr>{row_html}</tr>"

        headers_html = "".join(f"<th>{h}</th>" for h in headers)

        current_datetime = datetime.now().strftime("%d-%m-%Y")

        # ---------------- HTML ----------------
        html_string = f"""
        <!DOCTYPE html>
        <html lang="mr">
        <head>
            <meta charset="UTF-8">
            <title>Report</title>
        </head>
        <body>
            <h1 style="text-align:center;">अहवाल</h1>

            <table>
                <thead>
                    <tr>
                        {headers_html}
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
        </body>
        </html>
        """

        # ---------------- CSS (Marathi font + footer) ----------------
        font_path = os.path.join(
            settings.BASE_DIR,
            "static",
            "fonts",
            "NotoSansDevanagari-Regular.ttf"
        )

        css_string = f"""
        @font-face {{
            font-family: 'NotoDeva';
            src: url('file://{font_path}');
        }}

        @page {{
            size: A4;
            margin: 10px 10px 40px 10px;
            @bottom-right {{
                content: "Page " counter(page) " | {current_datetime}";
                font-size: 8pt;
                font-family: 'NotoDeva';
            }}
        }}

        body {{
            font-family: 'NotoDeva', sans-serif;
            font-size: 9pt;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
        }}

        th, td {{
            border: 1px solid #000;
            padding: 4px;
            text-align: center;
            word-wrap: break-word;
        }}

        th {{
            background-color: #d3d3d3;
            font-weight: bold;
        }}

        tr {{
            page-break-inside: avoid;
        }}

        h1 {{
            font-size: 16pt;
            margin-bottom: 10px;
        }}
        """

        # ---------------- Generate PDF ----------------
        pdf_file = HTML(string=html_string).write_pdf(
            stylesheets=[CSS(string=css_string)]
        )

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report.pdf"'
        return response

    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)

@require_POST
def clear_pending_action(request):
    request.session.pop("pending_action", None)
    request.session.modified = True 
    return JsonResponse({"status": "cleared"})

@login_required
def competitive_exams_landing_page(request, competitive_id=None):
    # --- SESSION CHECKS ---
    library_code = request.session.get('library_db')
    username = request.session.get('username')
    user_id = request.session.get('user_id')
    role_id = request.session.get('role_id')
    # if competitive_id:
    #     # Show exam details
    #     exam = get_object_or_404(CompetitiveExamMaster, competitive_id=competitive_id)
    #     return render(request, "L01/competitive_exam_details.html", {
    #         "MEDIA_URL": settings.MEDIA_URL,
    #         "exam": exam
    #     })
    
    # Show exam list
    competitive_exams = CompetitiveExamMaster.objects.all().order_by('full_name')
    return render(request, "L01/competitive_exams_landing_page.html", {
        "MEDIA_URL": settings.MEDIA_URL,
        "competitive_exams": competitive_exams
    })

@login_required
def upsc_index_logged(request):
    try:
        library_code = request.session.get('library_db')

        if not library_code:
            return JsonResponse({"error": "Library database not selected"}, status=400)

        competitive_exam = get_object_or_404(
            CompetitiveExamMaster.objects.using(library_code),
            competitive_id=1
        )

        # Fetch all related sections linked to UPSC (from selected library DB)
        sections = Sections.objects.using(library_code).filter(
            competitive_id=1
        ).order_by('section_no')

        subjects_by_section = []
        for section in sections:
            section.encrypted_section_no = enc(str(section.section_no))

            subjects = Subjects.objects.using(library_code).filter(
                section_no=section.section_no
            ).order_by('subject_id')

            subjects_by_section.append(
                (section.encrypted_section_no, subjects)
            )

        return render(request, "L01/UPSC/upsc_index_logged.html", {
            "MEDIA_URL": settings.MEDIA_URL,
            "competitive_exam": competitive_exam,
            "sections": sections,
            "subjects_by_section": subjects_by_section,
        })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else 'upsc_index_logged'
        callproc("stp_error_log", [fun, str(e), ''])
        print(f"error: {e}")
        return JsonResponse({"error": "Oops... Something went wrong!"}, status=500)

@login_required
def upsc_topics_logged(request, section_no):
    try:
        library_code = request.session.get('library_db')

        if not library_code:
            return JsonResponse({"error": "Library database not selected"}, status=400)
        #  Get the competitive exam (UPSC)
        competitive_exam = get_object_or_404(
            CompetitiveExamMaster.objects.using(library_code),
            competitive_id=1
        )

        # Fetch the specific section using the passed section_no
        section_no = int(dec(section_no))
        section = get_object_or_404(
            Sections.objects.using(library_code),
            section_no=section_no
        )

        #  Fetch all subjects under this section
        subjects = Subjects.objects.using(library_code).filter(
            section_no=section.section_no
        ).order_by('subject_id')

        subjects_data = []

        # For each subject, fetch topics under the same section
        for subject in subjects:
            topics = Topics.objects.using(library_code).filter(
                subject_id=subject.subject_id,
                section_no=section.section_no
            ).order_by('topic_id')
            
            
            for topic in topics:
                if topic.topic_image_url:
                    topic.topic_image_url = file_storage_service.get_file_url(
                        topic.topic_image_url
                    )

            subjects_data.append({
                "subject": subject,
                "topics": topics
            })

        #  Pass the structured data to the template
        return render(request, "L01/UPSC/upsc_topics_logged.html", {
            "MEDIA_URL": settings.MEDIA_URL,
            "competitive_exam": competitive_exam,
            "section": section,
            "subjects_data": subjects_data,
        })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else 'upsc_topics_logged'
        callproc("stp_error_log", [fun, str(e), ''])
        print(f"error: {e}")
        return JsonResponse({"error": "Oops... Something went wrong!"}, status=500)
    
@login_required
def mpsc_index_logged(request):
    try:
        library_code = request.session.get('library_db')

        if not library_code:
            return JsonResponse({"error": "Library database not selected"}, status=400)

        competitive_exam = get_object_or_404(
            CompetitiveExamMaster.objects.using(library_code),
            competitive_id=2
        )

        # Fetch all related sections linked to UPSC (from selected library DB)
        sections = Sections.objects.using(library_code).filter(
            competitive_id=2
        ).order_by('section_no')

        subjects_by_section = []
        for section in sections:
            section.encrypted_section_no = enc(str(section.section_no))

            subjects = Subjects.objects.using(library_code).filter(
                section_no=section.section_no
            ).order_by('subject_id')

            subjects_by_section.append(
                (section.encrypted_section_no, subjects)
            )

        return render(request, "L01/MPSC/mpsc_index_logged.html", {
            "MEDIA_URL": settings.MEDIA_URL,
            "competitive_exam": competitive_exam,
            "sections": sections,
            "subjects_by_section": subjects_by_section,
        })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else 'mpsc_index_logged'
        callproc("stp_error_log", [fun, str(e), ''])
        print(f"error: {e}")
        return JsonResponse({"error": "Oops... Something went wrong!"}, status=500)
    
@login_required
def mpsc_topics_logged(request, section_no):
    try:
        library_code = request.session.get('library_db')

        if not library_code:
            return JsonResponse({"error": "Library database not selected"}, status=400)
        #  Get the competitive exam (UPSC)
        competitive_exam = get_object_or_404(
            CompetitiveExamMaster.objects.using(library_code),
            competitive_id=2
        )

        # Fetch the specific section using the passed section_no
        section_no = int(dec(section_no))
        section = get_object_or_404(
            Sections.objects.using(library_code),
            section_no=section_no
        )

        #  Fetch all subjects under this section
        subjects = Subjects.objects.using(library_code).filter(
            section_no=section.section_no
        ).order_by('subject_id')

        subjects_data = []

        # For each subject, fetch topics under the same section
        for subject in subjects:
            topics = Topics.objects.using(library_code).filter(
                subject_id=subject.subject_id,
                section_no=section.section_no
            ).order_by('topic_id')
            
            for topic in topics:
                if topic.topic_image_url:
                    topic.topic_image_url = file_storage_service.get_file_url(
                        topic.topic_image_url
                    )

            subjects_data.append({
                "subject": subject,
                "topics": topics
            })

        #  Pass the structured data to the template
        return render(request, "L01/MPSC/mpsc_topics_logged.html", {
            "MEDIA_URL": settings.MEDIA_URL,
            "competitive_exam": competitive_exam,
            "section": section,
            "subjects_data": subjects_data,
        })

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name if tb else 'mpsc_topics_logged'
        callproc("stp_error_log", [fun, str(e), ''])
        print(f"error: {e}")
        return JsonResponse({"error": "Oops... Something went wrong!"}, status=500)
    
# views.py
from django.core.files.storage import FileSystemStorage
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, Any, List



def flatten_json(data: Dict[str, Any], parent_key: str = '', separator: str = '_') -> Dict[str, str]:
    """
    Flatten a nested dictionary for storage in key-value pairs.
    
    Args:
        data: The JSON/dictionary to flatten
        parent_key: Parent key for nested structures
        separator: Separator between nested keys
    
    Returns:
        Flattened dictionary with string keys and values
    """
    items: Dict[str, str] = {}
    
    def _flatten(value: Any, key: str = '') -> None:
        """Recursive helper function to flatten nested structures."""
        
        if isinstance(value, dict):
            for k, v in value.items():
                new_key = f"{key}{separator}{k}" if key else k
                _flatten(v, new_key)
        elif isinstance(value, list):
            for i, item in enumerate(value):
                new_key = f"{key}{separator}{i}" if key else str(i)
                _flatten(item, new_key)
        else:
            # Convert value to string
            if value is None:
                items[key] = ''
            elif isinstance(value, bool):
                items[key] = str(value).lower()
            else:
                items[key] = str(value)
    
    _flatten(data, parent_key)
    return items


def flatten_json_iterative(data: Dict[str, Any], separator: str = '_') -> Dict[str, str]:
    """
    Alternative iterative version of flatten_json to avoid recursion depth issues.
    """
    items: Dict[str, str] = {}
    stack: List[tuple] = [(data, '')]
    
    while stack:
        current_data, parent_key = stack.pop()
        
        if isinstance(current_data, dict):
            for key, value in current_data.items():
                new_key = f"{parent_key}{separator}{key}" if parent_key else key
                
                if isinstance(value, dict):
                    stack.append((value, new_key))
                elif isinstance(value, list):
                    for i, item in enumerate(value):
                        list_key = f"{new_key}{separator}{i}"
                        stack.append((item, list_key))
                else:
                    if value is None:
                        items[new_key] = ''
                    elif isinstance(value, bool):
                        items[new_key] = str(value).lower()
                    else:
                        items[new_key] = str(value)
        
        elif isinstance(current_data, list):
            for i, item in enumerate(current_data):
                new_key = f"{parent_key}{separator}{i}" if parent_key else str(i)
                stack.append((item, new_key))
        
        else:
            if current_data is None:
                items[parent_key] = ''
            elif isinstance(current_data, bool):
                items[parent_key] = str(current_data).lower()
            else:
                items[parent_key] = str(current_data)
    
    return items

@login_required
def get_book_data_isbn(request):
    """Render the main search page"""
    return render(request, 'L01/ISBN/book_search_isbn.html')

def validate_isbn(isbn: str) -> str:
    """Validate and clean ISBN"""
    # Remove all non-alphanumeric characters
    clean_isbn = re.sub(r'[^0-9X]', '', isbn.upper())
    
    # Check for valid ISBN-10
    if len(clean_isbn) == 10:
        # Validate check digit for ISBN-10
        check_sum = 0
        for i in range(9):
            check_sum += int(clean_isbn[i]) * (10 - i)
        check_digit = clean_isbn[9]
        if check_digit == 'X':
            check_sum += 10
        else:
            check_sum += int(check_digit)
        if check_sum % 11 == 0:
            return clean_isbn
    
    # Check for valid ISBN-13
    elif len(clean_isbn) == 13 and clean_isbn.isdigit():
        # Validate check digit for ISBN-13
        check_sum = 0
        for i in range(12):
            if i % 2 == 0:
                check_sum += int(clean_isbn[i])
            else:
                check_sum += int(clean_isbn[i]) * 3
        check_digit = (10 - (check_sum % 10)) % 10
        if check_digit == int(clean_isbn[12]):
            return clean_isbn
    
    return None


def save_google_book_data(isbn: str, book_data: Dict[str, Any]) -> bool:
    """Save Google Books data to database using flattened JSON"""
    try:
        items = book_data.get('items', [])
        if not items:
            logger.info(f"No items found in Google Books data for ISBN {isbn}")
            return False
        
        volume_info = items[0].get('volumeInfo', {})
        book_title = volume_info.get('title', f'Unknown Title - {isbn}')
        
        # Extract ISBN from industryIdentifiers if available
        isbn_to_use = isbn
        industry_ids = volume_info.get('industryIdentifiers', [])
        for id_item in industry_ids:
            if id_item.get('type') in ['ISBN_13', 'ISBN_10']:
                isbn_to_use = id_item.get('identifier', isbn)
                break
        
        # Validate ISBN format
        isbn_to_use = validate_isbn(isbn_to_use) or isbn_to_use
        
        # Save to master table
        with transaction.atomic(using='L01'):  # Specify database for transaction
            google_book, created = GoogleBookMaster.objects.using('L01').get_or_create(
                isbn=isbn_to_use,
                defaults={'title': book_title[:200]}  # Limit title length
            )
            
            if not created:
                # Update title if book already exists
                google_book.title = book_title[:200]
                google_book.save(using='L01')
            
            # Get book_id for foreign key
            book_id = google_book.id
            
            # Clear existing details using book_id
            GoogleBookDetail.objects.using('L01').filter(book_id=book_id).delete()
            
            # Flatten and save all data
            flattened_data = flatten_json(book_data)
            
            # Save each key-value pair (batch insert for efficiency)
            details_list = []
            for key, value in flattened_data.items():
                # Skip empty values
                if value is None or value == '':
                    continue
                    
                # Truncate key if too long
                truncated_key = key[:255]
                
                # Convert value to string and truncate if too long
                if isinstance(value, (dict, list)):
                    truncated_value = json.dumps(value, ensure_ascii=False)[:5000]
                else:
                    truncated_value = str(value)[:5000]
                
                details_list.append(GoogleBookDetail(
                    book_id=book_id,  # Use book_id instead of book object
                    key=truncated_key,
                    value=truncated_value
                ))
                
                # Batch insert in chunks to avoid memory issues
                if len(details_list) >= 100:
                    GoogleBookDetail.objects.using('L01').bulk_create(details_list, batch_size=100)
                    details_list = []
            
            # Insert any remaining items
            if details_list:
                GoogleBookDetail.objects.using('L01').bulk_create(details_list, batch_size=100)
            
            logger.info(f"Successfully saved Google Books data for ISBN {isbn_to_use} with {len(flattened_data)} flattened keys")
            return True
        
    except Exception as e:
        logger.error(f"Error saving Google book data for ISBN {isbn}: {str(e)}")
        return False



def save_loc_book_data(isbn: str, loc_data: Dict[str, Any]) -> bool:
    """Save LOC data to database"""
    try:
        # Check if we have actual book data
        results = loc_data.get('results', [])
        total_results = loc_data.get('pagination', {}).get('total', 0)
        
        # IMPORTANT: Check if results array is actually empty
        # Even if total_results > 0, if results array is empty, there's no book data
        if total_results == 0 or not results or len(results) == 0:
            logger.info(f"No book results found in LOC data for ISBN {isbn}. Total: {total_results}, Results count: {len(results)}")
            return False
        
        # Extract title from first result
        first_result = results[0]
        title = first_result.get('title', '') 
        if not title:
            title = first_result.get('item', {}).get('title', f'Unknown Title - {isbn}')
        
        # Save to master table
        with transaction.atomic():
            loc_book, created = LOCBookMaster.objects.using('L01').get_or_create(
                isbn=isbn,
                defaults={'title': title}
            )
            
            if not created:
                # Update title if book already exists
                loc_book.title = title
                loc_book.save()
            
            # Clear existing details
            LOCBookDetail.objects.using('L01').filter(book=loc_book).delete()
            
            # Flatten and save all data
            flattened_data = flatten_json(loc_data)
            
            # Save each key-value pair
            details_list = []
            for key, value in flattened_data.items():
                # Truncate key if too long
                truncated_key = key[:255]
                # Truncate value if too long
                truncated_value = value[:10000] if len(value) > 10000 else value
                
                details_list.append(LOCBookDetail(
                    book=loc_book,
                    key=truncated_key,
                    value=truncated_value
                ))
            
            # Batch insert
            if details_list:
                LOCBookDetail.objects.using('L01').bulk_create(details_list, batch_size=100)
            
            logger.info(f"Successfully saved LOC data for ISBN {isbn}")
            return True
        
    except Exception as e:
        logger.error(f"Error saving LOC book data for ISBN {isbn}: {str(e)}")
        return False


def search_book_by_isbn(isbn: str) -> Dict[str, Any]:
    """Search for a book by ISBN in LOC first, then Google"""
    isbn = isbn.strip()
    results = {
        'isbn': isbn,
        'source': None,
        'success': False,
        'message': '',
        'title': ''
    }
    
    # Try LOC API first
    # try:
    #     logger.info(f"Searching LOC for ISBN: {isbn}")

    #     loc_url = f"https://www.loc.gov/search/?q={isbn}&fo=json"

    #     headers = {
    #         "User-Agent": "BookSearchApp/1.0 (contact@example.com)"
    #     }

    #     response = requests.get(
    #         loc_url,
    #         headers=headers,   # ✅ REQUIRED
    #         timeout=15
    #     )

    #     if response.status_code == 200:
    #         loc_data = response.json()

    #         total_results = loc_data.get('pagination', {}).get('total', 0)
    #         actual_results = loc_data.get('results', [])

    #         logger.info(
    #             f"LOC Response - Total: {total_results}, "
    #             f"Actual results count: {len(actual_results)}"
    #         )

    #         if total_results > 0 and len(actual_results) > 0:
    #             logger.info(f"Found data in LOC for ISBN: {isbn}")

    #             if save_loc_book_data(isbn, loc_data):
    #                 results['source'] = 'Library of Congress'
    #                 results['success'] = True
    #                 results['message'] = 'Data found in Library of Congress'

    #                 first_result = actual_results[0]
    #                 results['title'] = first_result.get('title', 'Unknown Title')

    #                 return results
    #             else:
    #                 results['message'] = 'Failed to save LOC data'
    #         else:
    #             logger.info(f"No actual book results in LOC for ISBN: {isbn}")
    #             results['message'] = 'No book data found in Library of Congress'

    #     else:
    #         logger.warning(
    #             f"LOC API returned status {response.status_code} for ISBN {isbn}"
    #         )
    #         results['message'] = f'LOC API error: HTTP {response.status_code}'

    # except requests.RequestException as e:
    #     logger.warning(
    #         f"LOC API connection error for ISBN {isbn}: {str(e)}"
    #     )
    #     results['message'] = f'LOC API connection error: {str(e)}'

    # except Exception as e:
    #     logger.error(
    #         f"Error processing LOC data for ISBN {isbn}: {str(e)}"
    #     )
    #     results['message'] = f'Error processing LOC data: {str(e)}'

    
    # If LOC fails or has no data, try Google Books API
    try:
        logger.info(f"Searching Google Books for ISBN: {isbn}")
        google_url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
        response = requests.get(google_url, timeout=15)
        
        if response.status_code == 200:
            google_data = response.json()
            
            # Check if we have items
            total_items = google_data.get('totalItems', 0)
            if total_items > 0:
                logger.info(f"Found data in Google Books for ISBN: {isbn}")
                # Save Google data
                if save_google_book_data(isbn, google_data):
                    results['source'] = 'Google Books'
                    results['success'] = True
                    results['message'] = 'Data found in Google Books'
                    
                    # Extract title
                    items = google_data.get('items', [{}])
                    if items:
                        results['title'] = items[0].get('volumeInfo', {}).get('title', 'Unknown Title')
                    
                    return results
                else:
                    results['message'] = 'Failed to save Google Books data'
            else:
                logger.info(f"No results in Google Books for ISBN: {isbn}")
                if not results['message']:  # Only update if LOC also had no message
                    results['message'] = 'No data found in Google Books'
        
        else:
            logger.warning(f"Google Books API returned status {response.status_code} for ISBN {isbn}")
            if not results['message']:
                results['message'] = f'Google Books API error: HTTP {response.status_code}'
                
    except requests.RequestException as e:
        logger.warning(f"Google Books API connection error for ISBN {isbn}: {str(e)}")
        if not results['message']:
            results['message'] = f'Google Books API connection error: {str(e)}'
    except Exception as e:
        logger.error(f"Error processing Google Books data for ISBN {isbn}: {str(e)}")
        if not results['message']:
            results['message'] = f'Error processing Google Books data: {str(e)}'
    
    # If we reach here, no data was found
    if not results['message']:
        results['message'] = 'No data found in both Library of Congress and Google Books'
    
    return results


@csrf_exempt
def search_books(request):
    """Handle ISBN search from frontend"""
    if request.method == 'POST':
        isbns_input = request.POST.get('isbns', '').strip()
        
        if not isbns_input:
            return JsonResponse({'error': 'Please enter ISBN(s)'}, status=400)
        
        # Split by comma and clean up
        raw_isbns = [isbn.strip() for isbn in isbns_input.split(',') if isbn.strip()]
        
        # Validate ISBNs (basic validation)
        valid_isbns = []
        for isbn in raw_isbns:
            # Remove any dashes or spaces
            clean_isbn = isbn.replace('-', '').replace(' ', '')
            if clean_isbn.isdigit() and len(clean_isbn) in [10, 13]:
                valid_isbns.append(clean_isbn)
            else:
                logger.warning(f"Invalid ISBN format: {isbn}")
        
        if not valid_isbns:
            return JsonResponse({'error': 'No valid ISBNs found. Please enter valid 10 or 13 digit ISBNs.'}, status=400)
        
        # Limit to 20 ISBNs per request to prevent timeout
        isbns_to_process = valid_isbns[:2000]
        if len(valid_isbns) > 2000:
            logger.warning(f"Limiting search to first 20 ISBNs out of {len(valid_isbns)}")
        
        # Process ISBNs
        results = []
        
        # For single ISBN, process sequentially
        if len(isbns_to_process) == 1:
            result = search_book_by_isbn(isbns_to_process[0])
            results.append(result)
        else:
            # For multiple ISBNs, use threading with limited workers
            with ThreadPoolExecutor(max_workers=3) as executor:  # Reduced workers to avoid rate limiting
                future_to_isbn = {
                    executor.submit(search_book_by_isbn, isbn): isbn 
                    for isbn in isbns_to_process
                }
                
                for future in as_completed(future_to_isbn):
                    try:
                        result = future.result(timeout=45)  # Increased timeout
                        results.append(result)
                    except Exception as e:
                        isbn = future_to_isbn[future]
                        logger.error(f"Thread error for ISBN {isbn}: {str(e)}")
                        results.append({
                            'isbn': isbn,
                            'source': None,
                            'success': False,
                            'message': f'Processing error: {str(e)}',
                            'title': ''
                        })
        
        # Count successes
        success_count = sum(1 for r in results if r['success'])
        
        return JsonResponse({
            'success': True,
            'results': results,
            'total_searched': len(isbns_to_process),
            'total_found': success_count,
            'summary': f'Found {success_count} out of {len(isbns_to_process)} ISBN(s)',
            'warning': 'Limited to first 20 ISBNs' if len(valid_isbns) > 20 else None
        })
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
def upload_excel(request):
    """Handle Excel file upload"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            # Save the uploaded file temporarily
            fs = FileSystemStorage()
            filename = fs.save(f"temp_{int(time.time())}.xlsx", excel_file)
            file_path = fs.path(filename)
            
            # Read the Excel file
            try:
                df = pd.read_excel(file_path)
            except:
                # Try reading as CSV
                df = pd.read_csv(file_path)
            
            # Find ISBN column (look for columns that might contain ISBNs)
            isbn_column = None
            for col in df.columns:
                # Check if column name suggests ISBN
                col_lower = str(col).lower()
                if 'isbn' in col_lower or 'ean' in col_lower or 'barcode' in col_lower:
                    isbn_column = col
                    break
            
            # If no obvious ISBN column, use first column
            if isbn_column is None:
                isbn_column = df.columns[0]
                logger.info(f"No ISBN column found, using first column: {isbn_column}")
            
            # Extract ISBNs
            isbns = []
            for value in df[isbn_column].dropna().astype(str):
                # Clean the value
                clean_value = value.strip().replace('-', '').replace(' ', '')
                # Basic validation
                if clean_value.isdigit() and len(clean_value) in [10, 13]:
                    isbns.append(clean_value)
                elif len(clean_value) > 0:
                    logger.warning(f"Skipping invalid ISBN format: {value}")
            
            # Clean up
            fs.delete(filename)
            
            # Remove duplicates and limit
            unique_isbns = list(set(isbns))[:500]  # Limit to 100 ISBNs
         
            return JsonResponse({
                'success': True,
                'isbns': unique_isbns,
                'count': len(unique_isbns),
                'message': f'Successfully extracted {len(unique_isbns)} valid ISBN(s) from {isbn_column} column'
            })
            
        except pd.errors.EmptyDataError:
            return JsonResponse({'error': 'The Excel file is empty'}, status=400)
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            return JsonResponse({
                'error': f'Error processing file: {str(e)}. Please ensure it\'s a valid Excel or CSV file.'
            }, status=400)
    
    return JsonResponse({'error': 'No file uploaded'}, status=400)

    
@login_required
def check_old_password(request):
    if request.method == "POST":
        old_password = request.POST.get("old_password")
        user = request.user

        if user.check_password(old_password):
            return JsonResponse({"status": "success"})
        else:
            return JsonResponse({"status": "error"})

from django.contrib.auth import update_session_auth_hash

@login_required
def change_password(request):
    if request.method == "POST":

        old_password = request.POST.get("old_password")
        new_password = request.POST.get("new_password")

        user = request.user

        # Check if the old password is correct
        if not user.check_password(old_password):
            messages.error(request, "Old password is incorrect")
            return redirect(request.META.get("HTTP_REFERER"))
        
        try:

            password = password_storage.objects.using('L01').get(user_id=user.id)
            password.passwordText = new_password # Hash the new password
            password.save()
        
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")

        # Hash the new password using make_password()
        user.password = make_password(new_password)
        user.save()

        # Important: keep the user logged in after password change
        # update_session_auth_hash(request, user)

        messages.success(request, "Password updated successfully")
        return redirect("library_list")
    
# kiosk competitve Logic    

def kiosk_competitive_exam_type(request):
    """Main entry point for competitive exams"""
    try:
        competitive_id = request.GET.get('competitive_id')

        if competitive_id in ['1', '2']:
            # For UPSC (1) and MPSC (2), use existing pages
            if competitive_id == '1':
                return redirect('L01:upsc_index_logged')
            elif competitive_id == '2':
                return redirect('L01:mpsc_index_logged')
        else:
            # For other exams, fetch and show sections
            if competitive_id:
                exam = get_object_or_404(
                    CompetitiveExamMaster,
                    competitive_id=competitive_id
                )
                sections = Sections.objects.filter(
                    competitive_id=competitive_id
                ).order_by('section_no')

                context = {
                    'exam': exam,
                    'sections': sections,
                }
                return render(
                    request,
                    'L01/kiosk_competitive_sections.html',
                    context
                )
            else:
                competitive_exams = CompetitiveExamMaster.objects.all().order_by('full_name')
                return render(
                    request,
                    "L01/kiosk_competitive_exam_type.html",
                    {
                        "MEDIA_URL": settings.MEDIA_URL,
                        "competitive_exams": competitive_exams
                    }
                )

    except Exception as e:
        # Optional: log this properly later
        print("Error in kiosk_competitive_exam_type:", e)
        return redirect('L01:error_page')  # or HttpResponse("Something went wrong", status=500)

def kiosk_competitive_sections(request):
    """Display sections for a competitive exam"""
    competitive_id = request.GET.get('competitive_id')
    exam = get_object_or_404(CompetitiveExamMaster, competitive_id=competitive_id)
    
    # Get all sections for this exam
    sections = Sections.objects.filter(competitive_id=competitive_id).order_by('section_no')
    
    context = {
        'exam': exam,
        'sections': sections,
    }
    
    return render(request, 'L01/kiosk_competitive_sections.html', context)

def kiosk_competitive_subjects(request):
    """Display subjects for a specific section"""
    competitive_id = request.GET.get('competitive_id')
    section_no = request.GET.get('section_no')
    
    exam = get_object_or_404(CompetitiveExamMaster, competitive_id=competitive_id)
    section = get_object_or_404(Sections, section_no=section_no, competitive_id=competitive_id)
    
    # Get subjects for this section
    subjects = Subjects.objects.filter(
        Q(section_no=section_no) | Q(competitive_id=competitive_id)
    ).distinct().order_by('subject_name')
    
    context = {
        'exam': exam,
        'section': section,
        'subjects': subjects,
    }
    
    return render(request, 'L01/kiosk_competitive_subjects.html', context)

def kiosk_competitive_topics(request):
    """Display topics for a specific subject"""
    competitive_id = request.GET.get('competitive_id')
    subject_id = request.GET.get('subject_id')
    
    exam = get_object_or_404(CompetitiveExamMaster, competitive_id=competitive_id)
    subject = get_object_or_404(Subjects, subject_id=subject_id)
    
    # Get topics for this subject
    topics = Topics.objects.filter(
        subject_id=subject_id
    ).order_by('topic_name')
    
    context = {
        'exam': exam,
        'subject': subject,
        'topics': topics,
    }
    
    return render(request, 'L01/kiosk_competitive_topics.html', context)

# AJAX Functions for dynamic loading
def get_competitive_sections(request):
    """AJAX endpoint to get sections"""
    competitive_id = request.GET.get('competitive_id')
    
    sections = Sections.objects.filter(
        competitive_id=competitive_id
    ).order_by('section_no').values('section_no', 'section_name', 'section_description')
    
    return JsonResponse(list(sections), safe=False)

def get_competitive_subjects(request):
    """AJAX endpoint to get subjects"""
    competitive_id = request.GET.get('competitive_id')
    section_no = request.GET.get('section_no')
    
    subjects = Subjects.objects.filter(
        section_no=section_no,
        competitive_id=competitive_id
    ).order_by('subject_name').values('subject_id', 'subject_name', 'subject_description')
    
    return JsonResponse(list(subjects), safe=False)

def get_competitive_topics(request):
    """AJAX endpoint to get topics"""
    subject_id = request.GET.get('subject_id')
    
    topics = Topics.objects.filter(
        subject_id=subject_id
    ).order_by('topic_name').values(
        'topic_id', 'topic_name', 'topic_description', 
        'topic_reference', 'topic_image_url'
    )
    
    return JsonResponse(list(topics), safe=False)