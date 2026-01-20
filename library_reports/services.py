

# reports/services.py
from django.db.models import Count, F, Q, Sum, Case, When, Value, Avg, Max, Min
from django.db.models.functions import TruncMonth, TruncYear, Coalesce, ExtractYear, Concat
from django.utils import timezone
from datetime import timedelta, datetime
import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from decimal import Decimal
import calendar
from django.db import models
from django.db.models import Subquery, OuterRef

# Import your actual models
from L01.models import (
    MembershipDetails, MembershipMaster, StatusMaster, 
    PaymentDetails, MembershipDetailsHistory, DocumentDetails,
    MemberEntryExit, MemberLoginSession, MemberScreenActivity,
    MembershipRenewalStaging, CirculationTransaction, 
    parameter_master_L01, BookCatalog, BookReview
)

import logging

logger = logging.getLogger(__name__)

class ReportService:
    """Generic report service for all report types"""
    
    def __init__(self, report_type, library_code):
        self.report_type = report_type
        self.library_code = library_code
    
    def get_tab_data(self, tab_name, filters=None, selected_ids=None, search='', start=0, length=25, export_mode=False):
        """Get data for specific tab with pagination"""
        if self.report_type == 'member':
            return self._get_member_tab_data(tab_name, filters, selected_ids, search, start, length, export_mode)
        elif self.report_type == 'book':
            return self._get_book_tab_data(tab_name, filters, selected_ids, search, start, length, export_mode)
        else:
            raise ValueError(f"Unknown report type: {self.report_type}")
    
    def _get_member_tab_data(self, tab_name, filters, selected_ids, search, start, length, export_mode):
        """Get member report tab data"""
        service_map = {
            'member_details': self._get_member_details_data,
            'membership_details': self._get_membership_history_data,
            'loan': self._get_loan_data,
            'transactions': self._get_transaction_data,
            'physical_visit': self._get_physical_visit_data,
            'virtual_usage': self._get_virtual_usage_data,
            'payments': self._get_payment_data,
        }
        
        if tab_name not in service_map:
            raise ValueError(f"Unknown tab: {tab_name}")
        
        return service_map[tab_name](filters, selected_ids, search, start, length, export_mode)
    
    def _get_member_details_data(self, filters, selected_ids, search, start, length, export_mode):
        """Tab 1: Member Details"""
        from L01.models import MembershipDetails, MembershipMaster, StatusMaster
        
        queryset = MembershipDetails.objects.using(self.library_code).filter(
            id__in=selected_ids
        ).select_related('membership', 'status', 'member_type')
        
        # Apply filters
        queryset = self._apply_member_details_filters(queryset, filters)
        
        # Apply search
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(membership_code__icontains=search) |
                Q(user_id__icontains=search) |
                Q(mobile_no__icontains=search) |
                Q(email__icontains=search)
            )
        
        total_count = queryset.count()
        
        # Get paginated data for display, all for export
        if not export_mode and length > 0:
            queryset = queryset[start:start + length]
        
        # Prepare data
        data = []
        for member in queryset:
            data.append({
                'id': member.id,
                'first_name': member.first_name or '',
                'middle_name': member.middle_name or '',
                'last_name': member.last_name or '',
                'first_name_mar': member.first_name_mar or '',
                'middle_name_mar': member.middle_name_mar or '',
                'last_name_mar': member.last_name_mar or '',
                'membership_type': member.membership.membership_type if member.membership else '',
                'member_type': member.member_type.parameter_name if member.member_type else '',
                'membership_code': member.membership_code or '',
                'membership_date': member.membership_start_date.strftime('%Y-%m-%d') if member.membership_start_date else '',
                'renewal_from_date': member.from_date.strftime('%Y-%m-%d') if member.from_date else '',
                'renewal_to_date': member.to_date.strftime('%Y-%m-%d') if member.to_date else '',
                'user_id': member.user_id or '',
                'ward': member.ward or '',
                'pincode': member.pincode or '',
                'local_address': member.local_address or '',
                'mobile_no': member.mobile_no or '',
                'email': member.email or '',
                'occupation': member.occupation or '',
                'office_phone': member.office_phone or '',
                'education': member.education or '',
                'institute_name': member.institute_name or '',
                'recommender_details': member.recommender_details or '',
                'dob': member.dob.strftime('%Y-%m-%d') if member.dob else '',
                'aadhar_no': member.aadhar_no or '',
                'address_same_as_aadhar': 'Yes' if member.address_same_as_aadhar == 1 else 'No',
                'is_resident_of_nmmc': 'Yes' if member.is_resident_of_nmmc == 1 else 'No',
                'status': member.status.status_name if member.status else '',
                'isactive': 'Active' if member.isactive == 1 else 'Inactive',
                'created_at': member.created_at.strftime('%Y-%m-%d %H:%M') if member.created_at else '',
                'created_by': member.created_by or '',
                'updated_at': member.updated_at.strftime('%Y-%m-%d %H:%M') if member.updated_at else '',
                'updated_by': member.updated_by or '',
                'approved_by': member.reviewed or '',
                'approved_date': member.reviewed_at.strftime('%Y-%m-%d %H:%M') if member.reviewed_at else '',
                'has_documents': True,  # Flag for document view action
            })
        
        columns = [
            {'data': 'first_name', 'title': 'First Name'},
            {'data': 'middle_name', 'title': 'Middle Name'},
            {'data': 'last_name', 'title': 'Last Name'},
            {'data': 'membership_code', 'title': 'Membership Code'},
            {'data': 'membership_type', 'title': 'Membership Type'},
            {'data': 'member_type', 'title': 'Member Type'},
            {'data': 'user_id', 'title': 'User ID'},
            {'data': 'mobile_no', 'title': 'Mobile'},
            {'data': 'email', 'title': 'Email'},
            {'data': 'ward', 'title': 'Ward'},
            {'data': 'pincode', 'title': 'Pincode'},
            {'data': 'local_address', 'title': 'Address'},
            {'data': 'aadhar_no', 'title': 'Aadhar'},
            {'data': 'dob', 'title': 'DOB'},
            {'data': 'occupation', 'title': 'Occupation'},
            {'data': 'status', 'title': 'Status'},
            {'data': 'created_at', 'title': 'Created At'},
            {'data': 'approved_by', 'title': 'Approved By'},
            {'data': 'approved_date', 'title': 'Approved Date'},
            {
                'data': 'has_documents', 
                'title': 'Documents',
                'render': '''function(data, type, row) {
                    return '<button class="btn btn-sm btn-outline-info view-documents" data-member-id="' + row.id + '">View</button>';
                }''',
                'orderable': False,
                'searchable': False
            }
        ]
        
        return {
            'data': data,
            'columns': columns,
            'total': total_count,
            'filtered_total': total_count if not search else queryset.count()
        }
    
    def _apply_member_details_filters(self, queryset, filters):
        """Apply filters for member details tab"""
        if filters.get('membership_type'):
            queryset = queryset.filter(membership_id__in=filters['membership_type'])
        
        if filters.get('ward'):
            queryset = queryset.filter(ward__in=filters['ward'])
        
        if filters.get('status'):
            queryset = queryset.filter(status_id__in=filters['status'])
        
        if filters.get('member_type'):
            queryset = queryset.filter(member_type_id__in=filters['member_type'])
        
        if filters.get('from_date') and filters.get('to_date'):
            from_date = datetime.strptime(filters['from_date'], '%Y-%m-%d').date()
            to_date = datetime.strptime(filters['to_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__range=[from_date, to_date])
        
        if filters.get('renewal_due_month'):
            # Filter by renewal due month
            pass
        
        return queryset
    
    def _get_membership_history_data(self, filters, selected_ids, search, start, length, export_mode):
        """Tab 2: Membership Details"""
        from L01.models import MembershipDetailsHistory
        
        queryset = MembershipDetailsHistory.objects.using(self.library_code).filter(
            membership_id__in=selected_ids
        ).select_related('membershipmaster', 'status', 'member_type')
        
        # Apply filters
        if filters.get('from_date'):
            from_date = datetime.strptime(filters['from_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(changed_at__date__gte=from_date)
        
        if filters.get('to_date'):
            to_date = datetime.strptime(filters['to_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(changed_at__date__lte=to_date)
        
        if filters.get('membership_type'):
            queryset = queryset.filter(membershipmaster_id__in=filters['membership_type'])
        
        if filters.get('action_performed'):
            queryset = queryset.filter(actionperformed__in=filters['action_performed'])
        
        if filters.get('gap_period'):
            queryset = queryset.filter(gap_months__gt=0)
        
        # Apply search
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(actionperformed__icontains=search)
            )
        
        total_count = queryset.count()
        
        if not export_mode and length > 0:
            queryset = queryset[start:start + length]
        
        data = []
        for history in queryset:
            data.append({
                'id': history.id,
                'member_name': f"{history.first_name or ''} {history.last_name or ''}".strip(),
                'info': history.actionperformed or '',
                'from_date': history.from_date.strftime('%Y-%m-%d') if history.from_date else '',
                'to_date': history.to_date.strftime('%Y-%m-%d') if history.to_date else '',
                'deposit': float(history.deposit) if history.deposit else 0.0,
                'entry_fees': float(history.entry_fees) if history.entry_fees else 0.0,
                'subscription': float(history.subscription) if history.subscription else 0.0,
                'gap_fine_subscription': float(history.gap_fine) if history.gap_fine else 0.0,
                'gap_fine_delay': float(history.gap_subscription_delay) if history.gap_subscription_delay else 0.0,
                'gap_months': history.gap_months or 0,
                'late_fee': float(history.late_fee) if history.late_fee else 0.0,
                'changed_at': history.changed_at.strftime('%Y-%m-%d %H:%M') if history.changed_at else '',
                'changed_by': history.changed_by or '',
                'has_documents': True,
            })
        
        columns = [
            {'data': 'member_name', 'title': 'Member Name'},
            {'data': 'info', 'title': 'Info'},
            {'data': 'from_date', 'title': 'From Date'},
            {'data': 'to_date', 'title': 'To Date'},
            {'data': 'deposit', 'title': 'Deposit', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'entry_fees', 'title': 'Entry Fees', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'subscription', 'title': 'Subscription', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'gap_fine_subscription', 'title': 'Gap Fine', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'gap_fine_delay', 'title': 'Delay Fine', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'gap_months', 'title': 'Gap Months'},
            {'data': 'late_fee', 'title': 'Late Fee', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'changed_at', 'title': 'Changed At'},
            {'data': 'changed_by', 'title': 'Changed By'},
            {
                'data': 'has_documents', 
                'title': 'Documents',
                'render': '''function(data, type, row) {
                    return '<button class="btn btn-sm btn-outline-info view-documents" data-member-id="' + row.id + '">View</button>';
                }''',
                'orderable': False,
                'searchable': False
            }
        ]
        
        return {
            'data': data,
            'columns': columns,
            'total': total_count,
            'filtered_total': total_count if not search else queryset.count()
        }
    
    def _get_loan_data(self, filters, selected_ids, search, start, length, export_mode):
        """Tab 3: Loan Information"""
        from L01.models import CirculationTransaction, BookCatalog, BookAccession
        
        queryset = CirculationTransaction.objects.using(self.library_code).filter(
            member_id__in=selected_ids
        ).select_related('catalog', 'accession', 'member')
        
        # Apply filters
        if filters.get('from_date'):
            from_date = datetime.strptime(filters['from_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(issue_date__gte=from_date)
        
        if filters.get('to_date'):
            to_date = datetime.strptime(filters['to_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(issue_date__lte=to_date)
        
        if filters.get('membership_type'):
            # Filter by member's membership type
            from L01.models import MembershipDetails
            member_ids = MembershipDetails.objects.using(self.library_code).filter(
                membership_id__in=filters['membership_type']
            ).values_list('id', flat=True)
            queryset = queryset.filter(member_id__in=member_ids)
        
        if filters.get('overdue') == 'yes':
            queryset = queryset.filter(days_overdue_count__gt=0)
        
        if filters.get('fine_status'):
            queryset = queryset.filter(fine_status=filters['fine_status'])
        
        if search:
            queryset = queryset.filter(
                Q(barcode__icontains=search) |
                Q(member__membership_code__icontains=search) |
                Q(catalog__title__icontains=search)
            )
        
        total_count = queryset.count()
        
        if not export_mode and length > 0:
            queryset = queryset[start:start + length]
        
        data = []
        for transaction in queryset:
            data.append({
                'membership_code': transaction.member.membership_code if transaction.member else '',
                'barcode': transaction.barcode or '',
                'accession_id': transaction.accession.accession_no if transaction.accession else '',
                'cat_ref_num': transaction.catalog.cat_ref_num if transaction.catalog else '',
                'issue_date': transaction.issue_date.strftime('%Y-%m-%d') if transaction.issue_date else '',
                'due_date': transaction.due_date.strftime('%Y-%m-%d') if transaction.due_date else '',
                'return_date': transaction.return_date.strftime('%Y-%m-%d') if transaction.return_date else '',
                'days_overdue_count': transaction.days_overdue_count or 0,
                'fine_amount': float(transaction.fine_amount) if transaction.fine_amount else 0.0,
                'issued_by': transaction.issued_by or '',
                'remarks': transaction.remarks or '',
                'created_at': transaction.created_at.strftime('%Y-%m-%d %H:%M') if transaction.created_at else '',
            })
        
        columns = [
            {'data': 'membership_code', 'title': 'Membership Code'},
            {'data': 'barcode', 'title': 'Barcode'},
            {'data': 'accession_id', 'title': 'Accession No'},
            {'data': 'cat_ref_num', 'title': 'Catalog Ref'},
            {'data': 'issue_date', 'title': 'Issue Date'},
            {'data': 'due_date', 'title': 'Due Date'},
            {'data': 'return_date', 'title': 'Return Date'},
            {'data': 'days_overdue_count', 'title': 'Overdue Days'},
            {'data': 'fine_amount', 'title': 'Fine Amount', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'issued_by', 'title': 'Issued By'},
            {'data': 'remarks', 'title': 'Remarks'},
            {'data': 'created_at', 'title': 'Created At'},
        ]
        
        return {
            'data': data,
            'columns': columns,
            'total': total_count,
            'filtered_total': total_count if not search else queryset.count()
        }
    
    def _get_transaction_data(self, filters, selected_ids, search, start, length, export_mode):
        """Tab 4: Transactions"""
        from L01.models import CirculationTransaction
        
        queryset = CirculationTransaction.objects.using(self.library_code).filter(
            member_id__in=selected_ids
        ).select_related('member', 'catalog', 'accession')
        
        # Apply filters
        if filters.get('from_month'):
            # Filter by month
            pass
        
        if filters.get('to_month'):
            # Filter by month
            pass
        
        if filters.get('membership_type'):
            from L01.models import MembershipDetails
            member_ids = MembershipDetails.objects.using(self.library_code).filter(
                membership_id__in=filters['membership_type']
            ).values_list('id', flat=True)
            queryset = queryset.filter(member_id__in=member_ids)
        
        if filters.get('late_returns') == 'yes':
            queryset = queryset.filter(return_date__gt=F('due_date'))
        
        if filters.get('fine_status'):
            queryset = queryset.filter(fine_status=filters['fine_status'])
        
        if search:
            queryset = queryset.filter(
                Q(member__first_name__icontains=search) |
                Q(member__last_name__icontains=search) |
                Q(barcode__icontains=search) |
                Q(remarks__icontains=search)
            )
        
        total_count = queryset.count()
        
        if not export_mode and length > 0:
            queryset = queryset[start:start + length]
        
        data = []
        for transaction in queryset:
            data.append({
                'member_name': f"{transaction.member.first_name or ''} {transaction.member.last_name or ''}".strip(),
                'membership_code': transaction.member.membership_code if transaction.member else '',
                'barcode': transaction.barcode or '',
                'accession_id': transaction.accession.accession_no if transaction.accession else '',
                'cat_ref_num': transaction.catalog.cat_ref_num if transaction.catalog else '',
                'issue_date': transaction.issue_date.strftime('%Y-%m-%d') if transaction.issue_date else '',
                'due_date': transaction.due_date.strftime('%Y-%m-%d') if transaction.due_date else '',
                'issued_by': transaction.issued_by or '',
                'return_date': transaction.return_date.strftime('%Y-%m-%d') if transaction.return_date else '',
                'received_by': transaction.received_by or '',
                'days_overdue_count': transaction.days_overdue_count or 0,
                'fine_amount': float(transaction.fine_amount) if transaction.fine_amount else 0.0,
                'book_fine_amount': float(transaction.book_fine_amount) if transaction.book_fine_amount else 0.0,
                'total_fine': float(transaction.total_fine) if transaction.total_fine else 0.0,
                'adjusted_fine': float(transaction.adjusted_fine) if transaction.adjusted_fine else 0.0,
                'fine_status': transaction.fine_status or '',
                'fine_paid_date': transaction.fine_paid_date.strftime('%Y-%m-%d') if transaction.fine_paid_date else '',
                'remarks': transaction.remarks or '',
                'created_at': transaction.created_at.strftime('%Y-%m-%d %H:%M') if transaction.created_at else '',
                'updated_at': transaction.updated_at.strftime('%Y-%m-%d %H:%M') if transaction.updated_at else '',
            })
        
        columns = [
            {'data': 'member_name', 'title': 'Member Name'},
            {'data': 'membership_code', 'title': 'Membership Code'},
            {'data': 'barcode', 'title': 'Barcode'},
            {'data': 'accession_id', 'title': 'Accession No'},
            {'data': 'cat_ref_num', 'title': 'Catalog Ref'},
            {'data': 'issue_date', 'title': 'Issue Date'},
            {'data': 'due_date', 'title': 'Due Date'},
            {'data': 'issued_by', 'title': 'Issued By'},
            {'data': 'return_date', 'title': 'Return Date'},
            {'data': 'received_by', 'title': 'Received By'},
            {'data': 'days_overdue_count', 'title': 'Overdue Days'},
            {'data': 'fine_amount', 'title': 'Fine Amount', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'book_fine_amount', 'title': 'Book Fine', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'total_fine', 'title': 'Total Fine', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'adjusted_fine', 'title': 'Adjusted Fine', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'fine_status', 'title': 'Fine Status'},
            {'data': 'fine_paid_date', 'title': 'Fine Paid Date'},
            {'data': 'remarks', 'title': 'Remarks'},
            {'data': 'created_at', 'title': 'Created At'},
            {'data': 'updated_at', 'title': 'Updated At'},
        ]
        
        return {
            'data': data,
            'columns': columns,
            'total': total_count,
            'filtered_total': total_count if not search else queryset.count()
        }
    
    def _get_physical_visit_data(self, filters, selected_ids, search, start, length, export_mode):
        """Tab 5: Physical Visit"""
        from L01.models import MemberEntryExit, MembershipDetails
        
        # Get membership codes for selected members
        membership_codes = MembershipDetails.objects.using(self.library_code).filter(
            id__in=selected_ids
        ).values_list('membership_code', flat=True)
        
        queryset = MemberEntryExit.objects.using(self.library_code).filter(
            membership_code__in=membership_codes
        )
        
        # Apply filters
        if filters.get('from_date'):
            from_date = datetime.strptime(filters['from_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(entry_time__date__gte=from_date)
        
        if filters.get('to_date'):
            to_date = datetime.strptime(filters['to_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(entry_time__date__lte=to_date)
        
        if filters.get('membership_type'):
            # Get members with specific membership type
            member_codes = MembershipDetails.objects.using(self.library_code).filter(
                membership_id__in=filters['membership_type']
            ).values_list('membership_code', flat=True)
            queryset = queryset.filter(membership_code__in=member_codes)
        
        if filters.get('activity'):
            queryset = queryset.filter(remark__icontains=filters['activity'])
        
        if search:
            queryset = queryset.filter(
                Q(membership_code__icontains=search) |
                Q(remark__icontains=search)
            )
        
        total_count = queryset.count()
        
        if not export_mode and length > 0:
            queryset = queryset[start:start + length]
        
        data = []
        for visit in queryset:
            data.append({
                'member_name': f"Member {visit.membership_code}",  # Could join with member table
                'membership_code': visit.membership_code,
                'entry_time': visit.entry_time.strftime('%Y-%m-%d %H:%M:%S') if visit.entry_time else '',
                'exit_time': visit.exit_time.strftime('%Y-%m-%d %H:%M:%S') if visit.exit_time else '',
                'remark': visit.remark or '',
                'duration': self._calculate_duration(visit.entry_time, visit.exit_time),
            })
        
        columns = [
            {'data': 'member_name', 'title': 'Member Name'},
            {'data': 'membership_code', 'title': 'Membership Code'},
            {'data': 'entry_time', 'title': 'Entry Time'},
            {'data': 'exit_time', 'title': 'Exit Time'},
            {'data': 'duration', 'title': 'Duration'},
            {'data': 'remark', 'title': 'Remark'},
        ]
        
        return {
            'data': data,
            'columns': columns,
            'total': total_count,
            'filtered_total': total_count if not search else queryset.count()
        }
    
    def _calculate_duration(self, entry_time, exit_time):
        """Calculate duration between entry and exit"""
        if entry_time and exit_time:
            duration = exit_time - entry_time
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            return f"{hours}h {minutes}m"
        return ''
    
    def _get_virtual_usage_data(self, filters, selected_ids, search, start, length, export_mode):
        """Tab 6: Virtual Usage"""
        from L01.models import MemberLoginSession, MemberScreenActivity
        
        # Get sessions for selected members
        sessions = MemberLoginSession.objects.using(self.library_code).filter(
            member_id__in=selected_ids
        ).values_list('id', flat=True)
        
        queryset = MemberScreenActivity.objects.using(self.library_code).filter(
            session_id__in=sessions
        ).select_related('session__member')
        
        # Apply filters
        if filters.get('from_date'):
            from_date = datetime.strptime(filters['from_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(visited_at__date__gte=from_date)
        
        if filters.get('to_date'):
            to_date = datetime.strptime(filters['to_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(visited_at__date__lte=to_date)
        
        if filters.get('membership_type'):
            # Filter by membership type
            from L01.models import MembershipDetails
            member_ids = MembershipDetails.objects.using(self.library_code).filter(
                membership_id__in=filters['membership_type']
            ).values_list('id', flat=True)
            session_ids = MemberLoginSession.objects.using(self.library_code).filter(
                member_id__in=member_ids
            ).values_list('id', flat=True)
            queryset = queryset.filter(session_id__in=session_ids)
        
        if filters.get('activity'):
            queryset = queryset.filter(screen_name__icontains=filters['activity'])
        
        if search:
            queryset = queryset.filter(
                Q(screen_name__icontains=search) |
                Q(screen_route__icontains=search) |
                Q(session__member__membership_code__icontains=search)
            )
        
        total_count = queryset.count()
        
        if not export_mode and length > 0:
            queryset = queryset[start:start + length]
        
        data = []
        for activity in queryset:
            data.append({
                'member_name': f"{activity.session.member.first_name or ''} {activity.session.member.last_name or ''}".strip(),
                'membership_code': activity.session.member.membership_code if activity.session.member else '',
                'screen_name': activity.screen_name,
                'screen_route': activity.screen_route,
                'visited_at': activity.visited_at.strftime('%Y-%m-%d %H:%M:%S'),
                'session_start': activity.session.login_time.strftime('%Y-%m-%d %H:%M') if activity.session.login_time else '',
            })
        
        columns = [
            {'data': 'member_name', 'title': 'Member Name'},
            {'data': 'membership_code', 'title': 'Membership Code'},
            {'data': 'screen_name', 'title': 'Screen Name'},
            {'data': 'screen_route', 'title': 'Screen Route'},
            {'data': 'visited_at', 'title': 'Visited At'},
            {'data': 'session_start', 'title': 'Session Start'},
        ]
        
        return {
            'data': data,
            'columns': columns,
            'total': total_count,
            'filtered_total': total_count if not search else queryset.count()
        }
    
    def _get_payment_data(self, filters, selected_ids, search, start, length, export_mode):
        """Tab 7: Payments"""
        from L01.models import PaymentDetails
        
        queryset = PaymentDetails.objects.using(self.library_code).filter(
            membership_id__in=selected_ids
        ).select_related('status', 'membership')
        
        # Apply filters
        if filters.get('from_date'):
            from_date = datetime.strptime(filters['from_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(payment_date__gte=from_date)
        
        if filters.get('to_date'):
            to_date = datetime.strptime(filters['to_date'], '%Y-%m-%d').date()
            queryset = queryset.filter(payment_date__lte=to_date)
        
        if filters.get('membership_type'):
            from L01.models import MembershipDetails
            member_ids = MembershipDetails.objects.using(self.library_code).filter(
                membership_id__in=filters['membership_type']
            ).values_list('id', flat=True)
            queryset = queryset.filter(membership_id__in=member_ids)
        
        if filters.get('payment_type'):
            queryset = queryset.filter(payment_type__in=filters['payment_type'])
        
        if search:
            queryset = queryset.filter(
                Q(transaction_id__icontains=search) |
                Q(remarks__icontains=search) |
                Q(membership__membership_code__icontains=search)
            )
        
        total_count = queryset.count()
        
        if not export_mode and length > 0:
            queryset = queryset[start:start + length]
        
        data = []
        for payment in queryset:
            data.append({
                'payment_mode': payment.payment_mode,
                'payment_type': payment.payment_type or '',
                'payment_method': payment.payment_method or '',
                'deposit_amount': float(payment.deposit_amount) if payment.deposit_amount else 0.0,
                'entry_fee_amount': float(payment.entry_fee_amount) if payment.entry_fee_amount else 0.0,
                'monthly_subscription_amount': float(payment.monthly_subscription_amount) if payment.monthly_subscription_amount else 0.0,
                'total_subscription_amount': float(payment.total_subscription_amount) if payment.total_subscription_amount else 0.0,
                'subscription_from': payment.subscription_from.strftime('%Y-%m-%d') if payment.subscription_from else '',
                'subscription_to': payment.subscription_to.strftime('%Y-%m-%d') if payment.subscription_to else '',
                'transaction_id': payment.transaction_id or '',
                'remarks': payment.remarks or '',
                'user_id': payment.user_id or '',
                'membership_code': payment.membership_code or '',
                'payment_date': payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '',
                'status': payment.status.status_name if payment.status else '',
                'book_fine_amount': float(payment.book_fine_amount) if payment.book_fine_amount else 0.0,
                'fine_amount': float(payment.fine_amount) if payment.fine_amount else 0.0,
                'adjusted_amount': float(payment.adjusted_amount) if payment.adjusted_amount else 0.0,
                'created_at': payment.created_at.strftime('%Y-%m-%d %H:%M') if payment.created_at else '',
                'created_by': payment.created_by or '',
                'updated_at': payment.updated_at.strftime('%Y-%m-%d %H:%M') if payment.updated_at else '',
                'updated_by': payment.updated_by or '',
            })
        
        columns = [
            {'data': 'payment_mode', 'title': 'Payment Mode'},
            {'data': 'payment_type', 'title': 'Payment Type'},
            {'data': 'payment_method', 'title': 'Payment Method'},
            {'data': 'deposit_amount', 'title': 'Deposit', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'entry_fee_amount', 'title': 'Entry Fee', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'monthly_subscription_amount', 'title': 'Monthly Sub', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'total_subscription_amount', 'title': 'Total Sub', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'subscription_from', 'title': 'Sub From'},
            {'data': 'subscription_to', 'title': 'Sub To'},
            {'data': 'transaction_id', 'title': 'Transaction ID'},
            {'data': 'book_fine_amount', 'title': 'Book Fine', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'fine_amount', 'title': 'Fine', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'adjusted_amount', 'title': 'Adjusted', 'render': '$.fn.dataTable.render.number(",", ".", 2, "₹ ")'},
            {'data': 'remarks', 'title': 'Remarks'},
            {'data': 'payment_date', 'title': 'Payment Date'},
            {'data': 'status', 'title': 'Status'},
            {'data': 'created_at', 'title': 'Created At'},
            {'data': 'created_by', 'title': 'Created By'},
        ]
        
        return {
            'data': data,
            'columns': columns,
            'total': total_count,
            'filtered_total': total_count if not search else queryset.count()
        }
    
    def _get_book_tab_data(self, tab_name, filters, selected_ids, search, start, length, export_mode):
        """Get book report tab data (to be implemented)"""
        # Implementation for book report
        return {
            'data': [],
            'columns': [],
            'total': 0,
            'filtered_total': 0
        }

class ExportService:
    """Excel export service with multi-sheet support"""
    
    @staticmethod
    def export_to_excel(data_dict, sheet_names, filename=None):
        """Export multiple dataframes to Excel with different sheets"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        output = BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        if wb.sheetnames:
            wb.remove(wb[wb.sheetnames[0]])
        
        # Create sheets
        for idx, (sheet_key, data) in enumerate(data_dict.items()):
            sheet_name = sheet_names[idx] if idx < len(sheet_names) else f"Sheet{idx+1}"
            sheet_name = sheet_name[:31]  # Excel sheet name limit
            
            ws = wb.create_sheet(title=sheet_name)
            
            if isinstance(data, pd.DataFrame) and not data.empty:
                # Write header
                for col_idx, column in enumerate(data.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                
                # Write data
                for row_idx, row in enumerate(data.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = Alignment(vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin")
                        )
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    


class BaseReportService:
    """Base service with common report functionality"""
    
    @staticmethod
    def get_db_alias(request):
        """Get database alias from session"""
        return request.session.get('library_db', None)
    
    @staticmethod
    def get_queryset_with_db(model, request, *args, **kwargs):
        """Get queryset using appropriate database"""
        db_alias = BaseReportService.get_db_alias(request)
        if db_alias:
            return model.objects.using(db_alias).filter(*args, **kwargs)
        return model.objects.filter(*args, **kwargs)
    
    @staticmethod
    def apply_filters(queryset, filters):
        """Apply filters to queryset dynamically"""
        for field, value in filters.items():
            if value and value not in ['', 'all', 'All', None]:
                if '__in' in field or field.endswith('__in'):
                    queryset = queryset.filter(**{field: value})
                elif isinstance(value, list):
                    queryset = queryset.filter(**{f"{field}__in": value})
                elif field.endswith('__gte'):
                    queryset = queryset.filter(**{field: value})
                elif field.endswith('__lte'):
                    queryset = queryset.filter(**{field: value})
                elif field.endswith('__contains'):
                    queryset = queryset.filter(**{field: value})
                elif field.endswith('__icontains'):
                    queryset = queryset.filter(**{field: value})
                else:
                    queryset = queryset.filter(**{field: value})
        return queryset


class MemberReportDataService(BaseReportService):
    """Service for generating member-related reports"""
    
    # ====================== LEFT PANEL DATA ======================
    
    @classmethod
    def get_member_selector_data(cls, request, filters=None):
        """
        Get data for left-side member selector panel
        
        Args:
            request: Django request object
            filters: Dictionary of filters from left panel
        
        Returns:
            QuerySet: Filtered member list
        """
        db_alias = cls.get_db_alias(request)
        
        # Base queryset
        if db_alias:
            queryset = MembershipDetails.objects.using(db_alias).all()
        else:
            queryset = MembershipDetails.objects.all()
        
        # Apply left panel filters
        if filters:
            # Membership Type filter
            if filters.get('membership_type'):
                queryset = queryset.filter(
                    membership__membership_type=filters['membership_type']
                )
            
            # Status filter
            status_filter = filters.get('status', 'active')
            if status_filter == 'active':
                queryset = queryset.filter(isactive=1)
            elif status_filter == 'inactive':
                queryset = queryset.filter(isactive=0)
            elif status_filter == 'cancelled':
                # Assuming cancelled status has status_id for cancelled
                queryset = queryset.filter(status__status_name__icontains='cancelled')
        
        # Select related for performance
        queryset = queryset.select_related('membership', 'status')
        
        # Annotate with display fields
        queryset = queryset.annotate(
            display_name=Concat(
                F('first_name'), Value(' '), F('last_name'),
                output_field=models.CharField()
            )
        )
        
        return queryset
    
    @classmethod
    def search_members(cls, request, search_term):
        """
        Search members by various criteria (for left panel search)
        
        Args:
            request: Django request object
            search_term: Search string
        
        Returns:
            list: List of matching member IDs
        """
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            queryset = MembershipDetails.objects.using(db_alias)
        else:
            queryset = MembershipDetails.objects
        
        # Search in multiple fields
        members = queryset.filter(
            Q(user_id__icontains=search_term) |
            Q(first_name__icontains=search_term) |
            Q(last_name__icontains=search_term) |
            Q(membership_code__icontains=search_term) |
            Q(aadhar_no__icontains=search_term)
        ).values('id', 'user_id', 'first_name', 'last_name', 'membership_code')
        
        return list(members)
    
    # ====================== TAB 1: MEMBER DETAILS ======================
    
    @classmethod
    def get_member_details_data(cls, request, member_ids=None, filters=None):
        """
        Get data for Tab 1: Member Details
        
        Args:
            request: Django request object
            member_ids: List of selected member IDs (from left panel)
            filters: Tab-specific filters
        
        Returns:
            QuerySet: Member details data
        """
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            queryset = MembershipDetails.objects.using(db_alias)
        else:
            queryset = MembershipDetails.objects
        
        # Filter by selected members
        if member_ids:
            queryset = queryset.filter(id__in=member_ids)
        
        # Apply tab-specific filters
        if filters:
            # Membership Month From
            if filters.get('membership_month_from'):
                month_year = filters['membership_month_from']
                year, month = map(int, month_year.split('-'))
                first_day = datetime(year, month, 1)
                last_day = datetime(year, month, calendar.monthrange(year, month)[1])
                queryset = queryset.filter(from_date__range=[first_day, last_day])
            
            # Membership Month To
            if filters.get('membership_month_to'):
                month_year = filters['membership_month_to']
                year, month = map(int, month_year.split('-'))
                first_day = datetime(year, month, 1)
                last_day = datetime(year, month, calendar.monthrange(year, month)[1])
                queryset = queryset.filter(to_date__range=[first_day, last_day])
            
            # Membership Type
            if filters.get('membership_type'):
                queryset = queryset.filter(
                    membership__membership_type=filters['membership_type']
                )
            
            # Ward
            if filters.get('ward'):
                queryset = queryset.filter(ward=filters['ward'])
            
            # Status
            if filters.get('status'):
                queryset = queryset.filter(status__status_name=filters['status'])
            
            # Renewal Due Month
            if filters.get('renewal_due_month'):
                month_year = filters['renewal_due_month']
                year, month = map(int, month_year.split('-'))
                first_day = datetime(year, month, 1)
                last_day = datetime(year, month, calendar.monthrange(year, month)[1])
                queryset = queryset.filter(to_date__range=[first_day, last_day])
            
            # Member Type
            if filters.get('member_type'):
                queryset = queryset.filter(member_type__parameter_value=filters['member_type'])
        
        # Select related for performance
        queryset = queryset.select_related(
            'membership', 'status', 'member_type'
        )
        
        # Annotate with combined names
        queryset = queryset.annotate(
            full_name=Concat(
                F('first_name'), Value(' '), 
                Case(
                    When(middle_name__isnull=False, then=F('middle_name')),
                    default=Value('')
                ),
                Value(' '), F('last_name'),
                output_field=models.CharField()
            ),
            full_name_mar=Concat(
                F('first_name_mar'), Value(' '), 
                Case(
                    When(middle_name_mar__isnull=False, then=F('middle_name_mar')),
                    default=Value('')
                ),
                Value(' '), F('last_name_mar'),
                output_field=models.CharField()
            ),
            renewal_status=Case(
                When(membership_renew=1, then=Value('Renewed')),
                When(to_date__lt=timezone.now().date(), then=Value('Expired')),
                default=Value('Active')
            )
        )
        
        return queryset
    
    # ====================== TAB 2: MEMBERSHIP DETAILS ======================
    
    @classmethod
    def get_membership_details_data(cls, request, member_ids=None, filters=None):
        """
        Get data for Tab 2: Membership Details
        
        Args:
            request: Django request object
            member_ids: List of selected member IDs
            filters: Tab-specific filters
        
        Returns:
            dict: Contains both current and historical membership data
        """
        db_alias = cls.get_db_alias(request)
        
        # Get current membership details
        if db_alias:
            current_qs = MembershipDetails.objects.using(db_alias)
            history_qs = MembershipDetailsHistory.objects.using(db_alias)
        else:
            current_qs = MembershipDetails.objects
            history_qs = MembershipDetailsHistory.objects
        
        # Filter by selected members
        if member_ids:
            current_qs = current_qs.filter(id__in=member_ids)
            history_qs = history_qs.filter(membership_id__in=member_ids)
        
        # Apply filters
        if filters:
            # From Date
            if filters.get('from_date'):
                current_qs = current_qs.filter(from_date__gte=filters['from_date'])
                history_qs = history_qs.filter(from_date__gte=filters['from_date'])
            
            # To Date
            if filters.get('to_date'):
                current_qs = current_qs.filter(to_date__lte=filters['to_date'])
                history_qs = history_qs.filter(to_date__lte=filters['to_date'])
            
            # Membership Type
            if filters.get('membership_type'):
                current_qs = current_qs.filter(
                    membership__membership_type=filters['membership_type']
                )
                history_qs = history_qs.filter(
                    membershipmaster__membership_type=filters['membership_type']
                )
            
            # Gap Period
            if filters.get('gap_period'):
                if filters['gap_period'] == 'with_gap':
                    current_qs = current_qs.filter(gap_months__gt=0)
                    history_qs = history_qs.filter(gap_months__gt=0)
                elif filters['gap_period'] == 'without_gap':
                    current_qs = current_qs.filter(gap_months=0)
                    history_qs = history_qs.filter(gap_months=0)
            
            # Action Performed
            if filters.get('action_performed'):
                current_qs = current_qs.filter(
                    actionperformed__icontains=filters['action_performed']
                )
                history_qs = history_qs.filter(
                    actionperformed__icontains=filters['action_performed']
                )
        
        # Select related
        current_qs = current_qs.select_related('membership', 'status')
        history_qs = history_qs.select_related('membershipmaster', 'status')
        
        # Annotate with display fields
        current_qs = current_qs.annotate(
            member_name=Concat(F('first_name'), Value(' '), F('last_name')),
            info=F('actionperformed'),
            gap_fine_subscription=F('gap_subscription_delay'),
            gap_fine_delay=F('late_fee')
        )
        
        history_qs = history_qs.annotate(
            member_name=Concat(F('first_name'), Value(' '), F('last_name')),
            info=F('actionperformed'),
            gap_fine_subscription=F('gap_subscription_delay'),
            gap_fine_delay=F('late_fee'),
            changed_by_name=F('changed_by')
        )
        
        return {
            'current_memberships': current_qs,
            'membership_history': history_qs
        }
    
    # ====================== TAB 3: LOAN ======================
    
    @classmethod
    def get_loan_data(cls, request, member_ids=None, filters=None):
        """
        Get data for Tab 3: Loan (Circulation Transactions)
        
        Args:
            request: Django request object
            member_ids: List of selected member IDs
            filters: Tab-specific filters
        
        Returns:
            QuerySet: Loan/transaction data
        """
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            queryset = CirculationTransaction.objects.using(db_alias)
        else:
            queryset = CirculationTransaction.objects
        
        # Filter by selected members
        if member_ids:
            queryset = queryset.filter(member_id__in=member_ids)
        
        # Apply filters
        if filters:
            # From Date
            if filters.get('from_date'):
                queryset = queryset.filter(issue_date__gte=filters['from_date'])
            
            # To Date
            if filters.get('to_date'):
                queryset = queryset.filter(issue_date__lte=filters['to_date'])
            
            # Membership Type
            if filters.get('membership_type'):
                queryset = queryset.filter(
                    member__membership__membership_type=filters['membership_type']
                )
            
            # Overdue
            if filters.get('overdue'):
                if filters['overdue'] == 'overdue':
                    queryset = queryset.filter(
                        return_date__isnull=True,
                        due_date__lt=timezone.now().date()
                    )
                elif filters['overdue'] == 'not_overdue':
                    queryset = queryset.filter(
                        Q(return_date__isnull=False) |
                        Q(due_date__gte=timezone.now().date())
                    )
            
            # Fine Status
            if filters.get('fine_status'):
                if filters['fine_status'] == 'paid':
                    queryset = queryset.filter(fine_status='paid')
                elif filters['fine_status'] == 'not_paid':
                    queryset = queryset.filter(fine_status='unpaid')
        
        # Select related for performance
        queryset = queryset.select_related(
            'member', 'catalog', 'accession'
        )
        
        # Annotate with member name
        queryset = queryset.annotate(
            member_name=Concat(
                F('member__first_name'), Value(' '), F('member__last_name')
            )
        )
        
        return queryset
    
    # ====================== TAB 4: TRANSACTIONS ======================
    
    @classmethod
    def get_transactions_data(cls, request, member_ids=None, filters=None):
        """
        Get data for Tab 4: Transactions (Detailed)
        
        Args:
            request: Django request object
            member_ids: List of selected member IDs
            filters: Tab-specific filters
        
        Returns:
            QuerySet: Transaction data
        """
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            queryset = CirculationTransaction.objects.using(db_alias)
        else:
            queryset = CirculationTransaction.objects
        
        # Filter by selected members
        if member_ids:
            queryset = queryset.filter(member_id__in=member_ids)
        
        # Apply filters
        if filters:
            # From Month
            if filters.get('from_month'):
                month_year = filters['from_month']
                year, month = map(int, month_year.split('-'))
                first_day = datetime(year, month, 1)
                queryset = queryset.filter(issue_date__gte=first_day)
            
            # To Month
            if filters.get('to_month'):
                month_year = filters['to_month']
                year, month = map(int, month_year.split('-'))
                last_day = datetime(year, month, calendar.monthrange(year, month)[1])
                queryset = queryset.filter(issue_date__lte=last_day)
            
            # Membership Type
            if filters.get('membership_type'):
                queryset = queryset.filter(
                    member__membership__membership_type=filters['membership_type']
                )
            
            # Late Returns
            if filters.get('late_returns'):
                if filters['late_returns'] == 'late':
                    queryset = queryset.filter(
                        return_date__gt=F('due_date')
                    )
                elif filters['late_returns'] == 'on_time':
                    queryset = queryset.filter(
                        Q(return_date__lte=F('due_date')) |
                        Q(return_date__isnull=True, due_date__gte=timezone.now().date())
                    )
            
            # Fine Status
            if filters.get('fine_status'):
                queryset = queryset.filter(fine_status=filters['fine_status'])
        
        # Select related
        queryset = queryset.select_related(
            'member', 'catalog', 'accession', 'circulation'
        )
        
        # Annotate with additional fields
        queryset = queryset.annotate(
            member_name=Concat(
                F('member__first_name'), Value(' '), F('member__last_name')
            ),
            is_overdue=Case(
                When(
                    return_date__isnull=True,
                    due_date__lt=timezone.now().date(),
                    then=Value('Yes')
                ),
                When(
                    return_date__gt=F('due_date'),
                    then=Value('Yes')
                ),
                default=Value('No')
            ),
            overdue_days=Case(
                When(
                    return_date__isnull=True,
                    due_date__lt=timezone.now().date(),
                    then=timezone.now().date() - F('due_date')
                ),
                When(
                    return_date__gt=F('due_date'),
                    then=F('return_date') - F('due_date')
                ),
                default=Value(0)
            )
        )
        
        return queryset
    
    # ====================== TAB 5: PHYSICAL VISIT ======================
    
    @classmethod
    def get_physical_visit_data(cls, request, member_ids=None, filters=None):
        """
        Get data for Tab 5: Physical Visit (Member Entry/Exit)
        
        Args:
            request: Django request object
            member_ids: List of selected member IDs
            filters: Tab-specific filters
        
        Returns:
            QuerySet: Physical visit data
        """
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            queryset = MemberEntryExit.objects.using(db_alias)
        else:
            queryset = MemberEntryExit.objects
        
        # Filter by selected members
        if member_ids:
            # Get membership codes for selected members
            membership_codes = MembershipDetails.objects.filter(
                id__in=member_ids
            ).values_list('membership_code', flat=True)
            queryset = queryset.filter(membership_code__in=membership_codes)
        
        # Apply filters
        if filters:
            # From Date
            if filters.get('from_date'):
                queryset = queryset.filter(entry_time__date__gte=filters['from_date'])
            
            # To Date
            if filters.get('to_date'):
                queryset = queryset.filter(entry_time__date__lte=filters['to_date'])
            
            # Membership Type
            if filters.get('membership_type'):
                # Need to join with MembershipDetails
                membership_codes = MembershipDetails.objects.filter(
                    membership__membership_type=filters['membership_type']
                ).values_list('membership_code', flat=True)
                queryset = queryset.filter(membership_code__in=membership_codes)
        
        # Annotate with member name by joining with MembershipDetails
        from django.db import models
        queryset = queryset.annotate(
            member_name=Subquery(
                MembershipDetails.objects.filter(
                    membership_code=models.OuterRef('membership_code')
                ).values('first_name')[:1]
            )
        )
        
        return queryset
    
    # ====================== TAB 6: VIRTUAL USAGE ======================
    
    @classmethod
    def get_virtual_usage_data(cls, request, member_ids=None, filters=None):
        """
        Get data for Tab 6: Virtual Usage (Login & Screen Activity)
        
        Args:
            request: Django request object
            member_ids: List of selected member IDs
            filters: Tab-specific filters
        
        Returns:
            dict: Contains login sessions and screen activity
        """
        db_alias = cls.get_db_alias(request)
        
        # Get login sessions
        if db_alias:
            login_qs = MemberLoginSession.objects.using(db_alias)
            activity_qs = MemberScreenActivity.objects.using(db_alias)
        else:
            login_qs = MemberLoginSession.objects
            activity_qs = MemberScreenActivity.objects
        
        # Filter by selected members
        if member_ids:
            login_qs = login_qs.filter(member_id__in=member_ids)
            # Get session IDs for activity filter
            session_ids = login_qs.values_list('id', flat=True)
            activity_qs = activity_qs.filter(session_id__in=session_ids)
        
        # Apply filters
        if filters:
            # From Date
            if filters.get('from_date'):
                login_qs = login_qs.filter(login_time__date__gte=filters['from_date'])
                activity_qs = activity_qs.filter(visited_at__date__gte=filters['from_date'])
            
            # To Date
            if filters.get('to_date'):
                login_qs = login_qs.filter(login_time__date__lte=filters['to_date'])
                activity_qs = activity_qs.filter(visited_at__date__lte=filters['to_date'])
            
            # Membership Type
            if filters.get('membership_type'):
                # Filter members by membership type first
                members_with_type = MembershipDetails.objects.filter(
                    membership__membership_type=filters['membership_type']
                ).values_list('id', flat=True)
                
                login_qs = login_qs.filter(member_id__in=members_with_type)
                # Re-get session IDs
                session_ids = login_qs.values_list('id', flat=True)
                activity_qs = activity_qs.filter(session_id__in=session_ids)
        
        # Select related
        login_qs = login_qs.select_related('member')
        activity_qs = activity_qs.select_related('session__member')
        
        # Annotate login sessions
        login_qs = login_qs.annotate(
            member_name=Concat(
                F('member__first_name'), Value(' '), F('member__last_name')
            ),
            membership_code=F('member__membership_code'),
            session_duration=Case(
                When(
                    logout_time__isnull=False,
                    then=F('logout_time') - F('login_time')
                ),
                default=None
            )
        )
        
        # Annotate screen activity
        activity_qs = activity_qs.annotate(
            member_name=Concat(
                F('session__member__first_name'), Value(' '), 
                F('session__member__last_name')
            ),
            membership_code=F('session__member__membership_code')
        )
        
        return {
            'login_sessions': login_qs,
            'screen_activity': activity_qs
        }
    
    # ====================== TAB 7: PAYMENTS ======================
    
    @classmethod
    def get_payments_data(cls, request, member_ids=None, filters=None):
        """
        Get data for Tab 7: Payments
        
        Args:
            request: Django request object
            member_ids: List of selected member IDs
            filters: Tab-specific filters
        
        Returns:
            QuerySet: Payment data
        """
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            queryset = PaymentDetails.objects.using(db_alias)
        else:
            queryset = PaymentDetails.objects
        
        # Filter by selected members
        if member_ids:
            queryset = queryset.filter(membership_id__in=member_ids)
        
        # Apply filters
        if filters:
            # From Date
            if filters.get('from_date'):
                queryset = queryset.filter(
                    Q(payment_date__gte=filters['from_date']) |
                    Q(created_at__date__gte=filters['from_date'])
                )
            
            # To Date
            if filters.get('to_date'):
                queryset = queryset.filter(
                    Q(payment_date__lte=filters['to_date']) |
                    Q(created_at__date__lte=filters['to_date'])
                )
            
            # Membership Type
            if filters.get('membership_type'):
                queryset = queryset.filter(
                    membership__membership__membership_type=filters['membership_type']
                )
            
            # Payment Type
            if filters.get('payment_type'):
                queryset = queryset.filter(payment_type=filters['payment_type'])
        
        # Select related
        queryset = queryset.select_related('membership', 'status', 'circulation_transaction')
        
        # Annotate with member name and total amount
        queryset = queryset.annotate(
            member_name=Concat(
                F('membership__first_name'), Value(' '), F('membership__last_name')
            ),
            total_amount=Coalesce(
                F('deposit_amount'), Value(0)
            ) + Coalesce(
                F('entry_fee_amount'), Value(0)
            ) + Coalesce(
                F('monthly_subscription_amount'), Value(0)
            ) + Coalesce(
                F('fine_amount'), Value(0)
            ) + Coalesce(
                F('book_fine_amount'), Value(0)
            )
        )
        
        return queryset
    
    # ====================== EXPORT SERVICES ======================
    
    @classmethod
    def export_member_report(cls, request, member_ids=None, left_filters=None, tab_filters=None):
        """
        Export complete member report to Excel with multiple sheets
        
        Args:
            request: Django request object
            member_ids: List of selected member IDs
            left_filters: Filters from left panel
            tab_filters: Dictionary of filters for each tab
        
        Returns:
            HttpResponse: Excel file response
        """
        # Create Excel writer
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # Tab 1: Member Details
            member_details = cls.get_member_details_data(
                request, member_ids, tab_filters.get('tab1', {})
            )
            df1 = pd.DataFrame(list(member_details.values(
                'user_id', 'first_name', 'last_name', 'first_name_mar', 'last_name_mar',
                'membership__membership_type', 'member_type__parameter_value',
                'membership_code', 'from_date', 'to_date', 'user_id', 'ward',
                'pincode', 'local_address', 'mobile_no', 'email', 'occupation',
                'office_phone', 'education', 'institute_name', 'recommender_details',
                'dob', 'aadhar_no', 'is_resident_of_nmmc', 'status__status_name',
                'isactive', 'created_at', 'created_by', 'updated_at', 'updated_by'
            )))
            df1.to_excel(writer, sheet_name='Member Details', index=False)
            
            # Tab 2: Membership Details
            membership_data = cls.get_membership_details_data(
                request, member_ids, tab_filters.get('tab2', {})
            )
            
            # Current memberships
            df2_current = pd.DataFrame(list(membership_data['current_memberships'].values(
                'member_name', 'actionperformed', 'from_date', 'to_date',
                'deposit', 'entry_fees', 'subscription', 'fine_calculated_at',
                'gap_subscription_delay', 'late_fee', 'gap_months'
            )))
            df2_current.to_excel(writer, sheet_name='Current Memberships', index=False)
            
            # Membership history
            df2_history = pd.DataFrame(list(membership_data['membership_history'].values(
                'member_name', 'actionperformed', 'from_date', 'to_date',
                'deposit', 'entry_fees', 'subscription', 'fine_calculated_at',
                'gap_subscription_delay', 'late_fee', 'gap_months',
                'changed_at', 'changed_by'
            )))
            df2_history.to_excel(writer, sheet_name='Membership History', index=False)
            
            # Tab 3: Loan
            loan_data = cls.get_loan_data(
                request, member_ids, tab_filters.get('tab3', {})
            )
            df3 = pd.DataFrame(list(loan_data.values(
                'member__membership_code', 'barcode', 'accession__accession_no',
                'catalog__title', 'issue_date', 'due_date', 'return_date',
                'days_overdue_count', 'fine_amount', 'issued_by', 'remarks',
                'created_at'
            )))
            df3.to_excel(writer, sheet_name='Loans', index=False)
            
            # Tab 4: Transactions
            transactions_data = cls.get_transactions_data(
                request, member_ids, tab_filters.get('tab4', {})
            )
            df4 = pd.DataFrame(list(transactions_data.values(
                'member_name', 'member__membership_code', 'barcode',
                'accession__accession_no', 'catalog__title', 'issue_date',
                'due_date', 'issued_by', 'return_date', 'received_by',
                'days_overdue_count', 'fine_amount', 'book_fine_amount',
                'total_fine', 'adjusted_fine', 'fine_status', 'fine_paid_date',
                'remarks', 'created_at', 'updated_at'
            )))
            df4.to_excel(writer, sheet_name='Transactions', index=False)
            
            # Tab 5: Physical Visit
            visit_data = cls.get_physical_visit_data(
                request, member_ids, tab_filters.get('tab5', {})
            )
            df5 = pd.DataFrame(list(visit_data.values(
                'member_name', 'membership_code', 'entry_time',
                'exit_time', 'remark'
            )))
            df5.to_excel(writer, sheet_name='Physical Visits', index=False)
            
            # Tab 6: Virtual Usage
            virtual_data = cls.get_virtual_usage_data(
                request, member_ids, tab_filters.get('tab6', {})
            )
            
            # Login sessions
            df6_login = pd.DataFrame(list(virtual_data['login_sessions'].values(
                'member_name', 'member__membership_code', 'login_time',
                'logout_time', 'ip_address', 'device_type', 'session_duration'
            )))
            df6_login.to_excel(writer, sheet_name='Login Sessions', index=False)
            
            # Screen activity
            df6_activity = pd.DataFrame(list(virtual_data['screen_activity'].values(
                'member_name', 'session__member__membership_code', 'screen_name',
                'screen_route', 'visited_at'
            )))
            df6_activity.to_excel(writer, sheet_name='Screen Activity', index=False)
            
            # Tab 7: Payments
            payments_data = cls.get_payments_data(
                request, member_ids, tab_filters.get('tab7', {})
            )
            df7 = pd.DataFrame(list(payments_data.values(
                'member_name', 'payment_mode', 'payment_type', 'payment_method',
                'deposit_amount', 'entry_fee_amount', 'monthly_subscription_amount',
                'total_subscription_amount', 'subscription_from', 'subscription_to',
                'transaction_id', 'remarks', 'user_id', 'membership_code',
                'payment_date', 'status__status_name', 'book_fine_amount',
                'fine_amount', 'adjusted_amount', 'created_at', 'created_by',
                'updated_at', 'updated_by', 'total_amount'
            )))
            df7.to_excel(writer, sheet_name='Payments', index=False)
        
        # Prepare response
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="member_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
    
    # ====================== DOCUMENT VIEW SERVICES ======================
    
    @classmethod
    def get_member_documents(cls, request, member_id):
        """
        Get documents for a specific member
        
        Args:
            request: Django request object
            member_id: Member ID
        
        Returns:
            QuerySet: Member documents
        """
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            return DocumentDetails.objects.using(db_alias).filter(
                membership_id=member_id, isactive=1
            ).select_related('document')
        
        return DocumentDetails.objects.filter(
            membership_id=member_id, isactive=1
        ).select_related('document')
    
    @classmethod
    def get_membership_type_options(cls, request):
        """Get all membership type options for filters"""
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            return MembershipMaster.objects.using(db_alias).filter(
                isactive=1
            ).values_list('membership_type', flat=True).distinct()
        
        return MembershipMaster.objects.filter(
            isactive=1
        ).values_list('membership_type', flat=True).distinct()
    
    @classmethod
    def get_ward_options(cls, request):
        """Get all ward options for filters"""
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            return MembershipDetails.objects.using(db_alias).exclude(
                ward__isnull=True
            ).values_list('ward', flat=True).distinct()
        
        return MembershipDetails.objects.exclude(
            ward__isnull=True
        ).values_list('ward', flat=True).distinct()
    
    @classmethod
    def get_status_options(cls, request):
        """Get all status options for filters"""
        db_alias = cls.get_db_alias(request)
        
        if db_alias:
            return StatusMaster.objects.using(db_alias).filter(
                isactive=1
            ).values_list('status_name', flat=True).distinct()
        
        return StatusMaster.objects.filter(
            isactive=1
        ).values_list('status_name', flat=True).distinct()


class ExportService:
    """Generic export service for reports"""
    
    @staticmethod
    def export_to_excel(dataframes, sheet_names, filename):
        """
        Export multiple dataframes to Excel
        
        Args:
            dataframes: List of pandas DataFrames
            sheet_names: List of sheet names
            filename: Output filename
        
        Returns:
            HttpResponse: Excel file response
        """
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for df, sheet_name in zip(dataframes, sheet_names):
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def export_to_csv(dataframe, filename):
        """
        Export dataframe to CSV
        
        Args:
            dataframe: pandas DataFrame
            filename: Output filename
        
        Returns:
            HttpResponse: CSV file response
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        dataframe.to_csv(response, index=False, encoding='utf-8-sig')
        
        return response


class ReportService:
    """Generic report service with common utilities"""
    
    @staticmethod
    def get_filter_options(request, model_name, field_name):
        """
        Get distinct filter options for a model field
        
        Args:
            request: Django request object
            model_name: Name of the model
            field_name: Name of the field
        
        Returns:
            list: Distinct values
        """
        db_alias = BaseReportService.get_db_alias(request)
        
        # Map model names to actual models
        model_map = {
            'MembershipDetails': MembershipDetails,
            'MembershipMaster': MembershipMaster,
            'StatusMaster': StatusMaster,
            'PaymentDetails': PaymentDetails,
            'CirculationTransaction': CirculationTransaction,
        }
        
        if model_name in model_map:
            model = model_map[model_name]
            if db_alias:
                queryset = model.objects.using(db_alias)
            else:
                queryset = model.objects
            
            return list(queryset.values_list(field_name, flat=True).distinct())
        
        return []
    
    @staticmethod
    def validate_filters(filters, allowed_filters):
        """
        Validate and sanitize filter parameters
        
        Args:
            filters: Dictionary of filters
            allowed_filters: List of allowed filter keys
        
        Returns:
            dict: Sanitized filters
        """
        sanitized = {}
        for key, value in filters.items():
            if key in allowed_filters and value not in ['', None, 'all', 'All']:
                sanitized[key] = value
        return sanitized